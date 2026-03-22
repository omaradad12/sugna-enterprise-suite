from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.translation import gettext as _
from django.views.decorators.http import require_http_methods

from tenant_portal.auth import get_tenant_db_for_request, get_tenant_user, tenant_login, tenant_logout
from rbac.models import user_has_permission
from tenant_portal.decorators import tenant_view


def tenant_login_view(request: HttpRequest) -> HttpResponse:
    tenant = getattr(request, "tenant", None)
    tenant_db = get_tenant_db_for_request(request)
    if not tenant:
        return render(request, "tenant_portal/tenant_missing.html", status=404)
    if not tenant_db:
        return render(request, "tenant_portal/tenant_not_provisioned.html", {"tenant": tenant}, status=503)

    if request.method == "POST":
        email = (request.POST.get("email") or "").strip().lower()
        password = request.POST.get("password") or ""
        from tenant_users.models import TenantUser

        user = TenantUser.objects.using(tenant_db).filter(email=email, is_active=True).first()
        if user and user.check_password(password):
            if getattr(user, "two_factor_enabled", False) and getattr(user, "totp_secret", None):
                request.session["pending_2fa_user_id"] = user.id
                return redirect(reverse("tenant_portal:login_2fa"))
            tenant_login(request, user.id)
            from tenant_users.models import TenantLoginLog
            ip = request.META.get("REMOTE_ADDR") or None
            ua = (request.META.get("HTTP_USER_AGENT") or "")[:500]
            TenantLoginLog.objects.using(tenant_db).create(user=user, ip_address=ip, user_agent=ua)
            user.touch_login()
            return redirect(reverse("tenant_portal:home"))
        messages.error(request, "Invalid email or password.")

    return render(request, "tenant_portal/login.html", {"tenant": tenant})


def tenant_login_2fa_view(request: HttpRequest) -> HttpResponse:
    """Second step of login: verify TOTP code when user has 2FA enabled."""
    tenant = getattr(request, "tenant", None)
    tenant_db = get_tenant_db_for_request(request)
    if not tenant or not tenant_db:
        return redirect(reverse("tenant_portal:login"))
    user_id = request.session.get("pending_2fa_user_id")
    if not user_id:
        return redirect(reverse("tenant_portal:login"))
    from tenant_users.models import TenantUser
    user = TenantUser.objects.using(tenant_db).filter(pk=user_id, is_active=True).first()
    if not user:
        request.session.pop("pending_2fa_user_id", None)
        return redirect(reverse("tenant_portal:login"))

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        from tenant_portal.totp_utils import verify_totp
        if verify_totp(user.totp_secret or "", code):
            request.session.pop("pending_2fa_user_id", None)
            tenant_login(request, user.id)
            from tenant_users.models import TenantLoginLog
            ip = request.META.get("REMOTE_ADDR") or None
            ua = (request.META.get("HTTP_USER_AGENT") or "")[:500]
            TenantLoginLog.objects.using(tenant_db).create(user=user, ip_address=ip, user_agent=ua)
            user.touch_login()
            return redirect(reverse("tenant_portal:home"))
        messages.error(request, "Invalid or expired code. Please try again.")

    return render(request, "tenant_portal/login_2fa.html", {"tenant": tenant})


def tenant_logout_view(request: HttpRequest) -> HttpResponse:
    tenant_logout(request)
    return redirect(reverse("tenant_portal:login"))


def _get_role_display(tenant_user, tenant_db):
    from rbac.models import UserRole
    roles = list(
        UserRole.objects.using(tenant_db)
        .filter(user_id=tenant_user.id)
        .values_list("role__name", flat=True)
    )
    if getattr(tenant_user, "is_tenant_admin", False):
        roles = ["Tenant Admin"] + [r for r in roles if r != "Tenant Admin"]
    return ", ".join(roles) if roles else ("Tenant Admin" if getattr(tenant_user, "is_tenant_admin", False) else "—")


def _get_global_financial_indicators(request):
    """
    Compute organization-wide financial indicators for the global bar.
    Values are calculated from posted journal lines so they stay in sync with transactions.
    """
    from decimal import Decimal

    from django.db.models import Sum

    from tenant_finance.models import BankAccount, ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    indicators = {
        "total_bank_balance": Decimal("0"),
        "total_cash_balance": Decimal("0"),
        "petty_cash_balance": Decimal("0"),
        "outstanding_receivables": Decimal("0"),
        "pending_payables": Decimal("0"),
        "budget_utilization": Decimal("0"),
    }

    # 1–3: Bank, cash, petty cash balances from BankAccount-linked GL accounts
    bank_accounts = list(
        BankAccount.objects.using(tenant_db).select_related("account").all()
    )
    if bank_accounts:
        cash_account_ids = [ba.account_id for ba in bank_accounts if ba.account_id]
        if cash_account_ids:
            bal_rows = (
                JournalLine.objects.using(tenant_db)
                .filter(
                    account_id__in=cash_account_ids,
                    entry__status=JournalEntry.Status.POSTED,
                )
                .values("account_id")
                .annotate(bal=Sum("debit") - Sum("credit"))
            )
            by_account = {r["account_id"]: r.get("bal") or Decimal("0") for r in bal_rows}
            total_bank = Decimal("0")
            total_cash = Decimal("0")
            petty_cash = Decimal("0")
            for ba in bank_accounts:
                bal = by_account.get(ba.account_id, Decimal("0"))
                name = (ba.account_name or "").lower()
                if "petty" in name:
                    petty_cash += bal
                elif "cash" in name:
                    total_cash += bal
                else:
                    total_bank += bal
            indicators["total_bank_balance"] = total_bank
            indicators["total_cash_balance"] = total_cash
            indicators["petty_cash_balance"] = petty_cash

    # 4: Outstanding receivables = asset receivable accounts with positive balance
    recv_q = _receivable_accounts_q()
    recv_ids = list(
        ChartAccount.objects.using(tenant_db)
        .filter(recv_q)
        .values_list("id", flat=True)
    )
    if recv_ids:
        recv_rows = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account_id__in=recv_ids,
                entry__status=JournalEntry.Status.POSTED,
            )
            .values("account_id")
            .annotate(bal=Sum("debit") - Sum("credit"))
        )
        indicators["outstanding_receivables"] = sum(
            (r.get("bal") or Decimal("0")) for r in recv_rows if (r.get("bal") or Decimal("0")) > 0
        )

    # 5: Pending payables: liability payables with credit (negative) balance
    payable_ids = list(
        ChartAccount.objects.using(tenant_db)
        .filter(type=ChartAccount.Type.LIABILITY, code__startswith="2")
        .values_list("id", flat=True)
    )
    if payable_ids:
        pay_rows = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account_id__in=payable_ids,
                entry__status=JournalEntry.Status.POSTED,
            )
            .values("account_id")
            .annotate(bal=Sum("debit") - Sum("credit"))
        )
        indicators["pending_payables"] = sum(
            -(r.get("bal") or Decimal("0")) for r in pay_rows if (r.get("bal") or Decimal("0")) < 0
        )

    # 6: Budget utilization: total spent / total awarded across active grants.
    # Grant schemas may differ across deployments/tenants, so compute award safely:
    # - prefer Grant.award_amount (newer)
    # - fallback to Grant.amount_awarded (older)
    # - fallback to budget lines sum when neither exists
    grants = list(Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE))
    total_award = Decimal("0")
    for g in grants:
        raw = getattr(g, "award_amount", None)
        if raw is None:
            raw = getattr(g, "amount_awarded", None)
        if raw is not None:
            try:
                total_award += Decimal(str(raw or 0))
                continue
            except Exception:
                pass
        # BudgetLine fallback (best-effort)
        try:
            from tenant_grants.models import BudgetLine

            bl_total = (
                BudgetLine.objects.using(tenant_db)
                .filter(grant_id=g.id)
                .aggregate(t=Sum("amount"))
                .get("t")
            )
            total_award += Decimal(str(bl_total or 0))
        except Exception:
            total_award += Decimal("0")
    if total_award > 0:
        spent = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account__type=ChartAccount.Type.EXPENSE,
                entry__status=JournalEntry.Status.POSTED,
                entry__grant__status=Grant.Status.ACTIVE,
            )
            .aggregate(t=Sum("debit"))
            .get("t")
            or Decimal("0")
        )
        indicators["budget_utilization"] = (spent / total_award) * Decimal("100")

    return indicators


@tenant_view()
def profile_view(request: HttpRequest) -> HttpResponse:
    """My Profile: profile info, security (password, 2FA, login history), notifications. Only user can edit own profile."""
    from tenant_users.models import TenantUser, TenantLoginLog

    if request.GET.get("cancel_2fa"):
        request.session.pop("pending_2fa_secret", None)
        return redirect(reverse("tenant_portal:profile"))

    tenant_db = request.tenant_db
    user = request.tenant_user

    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action == "profile_save":
            user.full_name = (request.POST.get("full_name") or "").strip()[:200]
            user.phone_number = (request.POST.get("phone_number") or "").strip()[:30]
            user.position = (request.POST.get("position") or "").strip()[:120]
            user.department = (request.POST.get("department") or "").strip()[:120]
            user.preferred_language = (request.POST.get("preferred_language") or "en").strip()[:10]
            user.time_zone = (request.POST.get("time_zone") or "UTC").strip()[:50]
            photo = request.FILES.get("profile_photo")
            photo_ok = False
            if photo:
                if photo.size > 2 * 1024 * 1024:  # 2MB
                    messages.error(request, "Profile photo must be 2MB or smaller.")
                else:
                    user.profile_photo = photo
                    photo_ok = True
            user.save()
            if not photo or photo_ok:
                messages.success(request, "Profile updated.")
        elif action == "change_password":
            current = request.POST.get("current_password", "")
            new = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")
            if not user.check_password(current):
                messages.error(request, "Current password is incorrect.")
            elif len(new) < 8:
                messages.error(request, "New password must be at least 8 characters.")
            elif new != confirm:
                messages.error(request, "New password and confirmation do not match.")
            else:
                user.set_password(new)
                user.save(update_fields=["password_hash"])
                messages.success(request, "Password changed.")
        elif action == "notifications":
            user.email_notifications = request.POST.get("email_notifications") == "on"
            user.system_alerts = request.POST.get("system_alerts") == "on"
            user.approval_notifications = request.POST.get("approval_notifications") == "on"
            user.save(update_fields=["email_notifications", "system_alerts", "approval_notifications"])
            messages.success(request, "Notification preferences saved.")
        elif action == "enable_2fa_start":
            from tenant_portal.totp_utils import generate_secret
            secret = generate_secret()
            if secret:
                request.session["pending_2fa_secret"] = secret
                messages.success(request, "Scan the QR code or enter the key in your authenticator app, then enter the 6-digit code below.")
            else:
                messages.error(request, "2FA is not available (install pyotp).")
        elif action == "enable_2fa_verify":
            from tenant_portal.totp_utils import verify_totp
            secret = request.session.get("pending_2fa_secret")
            code = (request.POST.get("totp_code") or "").strip()
            if not secret:
                messages.error(request, "2FA setup expired. Please start again.")
            elif verify_totp(secret, code):
                user.totp_secret = secret
                user.two_factor_enabled = True
                user.save(update_fields=["totp_secret", "two_factor_enabled"])
                request.session.pop("pending_2fa_secret", None)
                messages.success(request, "Two-factor authentication is now enabled.")
            else:
                messages.error(request, "Invalid or expired code. Please try again.")
        elif action == "disable_2fa":
            current = request.POST.get("current_password", "")
            from tenant_portal.totp_utils import verify_totp
            if not user.check_password(current) and not (user.totp_secret and verify_totp(user.totp_secret, current)):
                messages.error(request, "Current password or 2FA code is incorrect.")
            else:
                user.totp_secret = ""
                user.two_factor_enabled = False
                user.save(update_fields=["totp_secret", "two_factor_enabled"])
                messages.success(request, "Two-factor authentication has been disabled.")

    # Reload user to get latest from DB (including photo URL)
    user = TenantUser.objects.using(tenant_db).get(pk=user.id)
    role_display = _get_role_display(user, tenant_db)
    login_logs = list(TenantLoginLog.objects.using(tenant_db).filter(user=user).order_by("-created_at")[:20])

    # 2FA setup in progress (show QR + verify form)
    pending_2fa_secret = request.session.get("pending_2fa_secret")
    totp_uri = ""
    totp_qr_data_url = ""
    if pending_2fa_secret:
        from tenant_portal.totp_utils import get_provisioning_uri, get_qr_data_url
        issuer = getattr(request.tenant, "name", "Sugna Tenant")[:32]
        totp_uri = get_provisioning_uri(pending_2fa_secret, user.email, issuer=issuer)
        totp_qr_data_url = get_qr_data_url(totp_uri)

    return render(
        request,
        "tenant_portal/profile.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "role_display": role_display,
            "login_logs": login_logs,
            "pending_2fa_secret": pending_2fa_secret,
            "totp_uri": totp_uri,
            "totp_qr_data_url": totp_qr_data_url,
        },
    )


def _fiscal_span_months(start_month: int, end_month: int) -> int:
    """Return number of months from start_month to end_month (inclusive), cross-year allowed."""
    if start_month <= end_month:
        return end_month - start_month + 1
    return (12 - start_month + 1) + end_month


@tenant_view()
def organization_settings_view(request: HttpRequest) -> HttpResponse:
    """Organization Settings under Account. Only System Admin (tenant admin) can edit."""
    from tenant_finance.models import OrganizationSettings, Currency, JournalEntry, OpeningBalance

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_edit = getattr(user, "is_tenant_admin", False)

    settings = OrganizationSettings.objects.using(tenant_db).first()
    from tenant_finance.models import ensure_default_currencies, Currency

    if not settings:
        # Create initial organization settings
        settings = OrganizationSettings.objects.using(tenant_db).create()

    # Ensure default currencies exist and, if no base currency set yet, default to USD
    ensure_default_currencies(using=tenant_db)
    if not settings.default_currency_id:
        usd = Currency.objects.using(tenant_db).filter(code="USD").first()
        if usd:
            settings.default_currency = usd
            settings.save(using=tenant_db)

    if request.method == "POST" and can_edit:
        # Organization info
        settings.organization_name = (request.POST.get("organization_name") or "").strip()[:255]
        settings.registration_number = (request.POST.get("registration_number") or "").strip()[:80]
        settings.country = (request.POST.get("country") or "").strip()[:100]
        settings.city = (request.POST.get("city") or "").strip()[:100]
        settings.address = (request.POST.get("address") or "").strip()
        settings.phone = (request.POST.get("phone") or "").strip()[:40]
        settings.email = (request.POST.get("email") or "").strip()
        settings.website = (request.POST.get("website") or "").strip()
        # Single brand logo (used everywhere: system, reports, tenant templates)
        if request.FILES.get("organization_logo"):
            settings.organization_logo = request.FILES["organization_logo"]
        # Branding
        settings.primary_color = (request.POST.get("primary_color") or "").strip()[:20]
        settings.secondary_color = (request.POST.get("secondary_color") or "").strip()[:20]
        # Fiscal
        try:
            start_m = max(1, min(12, int(request.POST.get("fiscal_year_start_month") or 1)))
            end_m = max(1, min(12, int(request.POST.get("fiscal_year_end_month") or 12)))
        except (TypeError, ValueError):
            start_m, end_m = settings.fiscal_year_start_month, settings.fiscal_year_end_month
        if _fiscal_span_months(start_m, end_m) != 12:
            messages.error(request, "Fiscal year must span exactly 12 months.")
        else:
            settings.fiscal_year_start_month = start_m
            settings.fiscal_year_end_month = end_m
        settings.currency_format = (request.POST.get("currency_format") or "#,##0.00").strip()[:40]
        settings.time_zone = (request.POST.get("time_zone") or "UTC").strip()[:50]
        new_currency_id = request.POST.get("default_currency")
        if new_currency_id:
            try:
                new_currency_id = int(new_currency_id)
            except (TypeError, ValueError):
                new_currency_id = None
        # Allow tenant to change base currency even after transactions exist
        settings.default_currency_id = new_currency_id or settings.default_currency_id
        # Document prefixes
        settings.invoice_prefix = (request.POST.get("invoice_prefix") or "INV-").strip()[:20]
        settings.payment_voucher_prefix = (request.POST.get("payment_voucher_prefix") or "PV-").strip()[:20]
        settings.receipt_voucher_prefix = (request.POST.get("receipt_voucher_prefix") or "RV-").strip()[:20]
        settings.journal_prefix = (request.POST.get("journal_prefix") or "JV-").strip()[:20]
        if not messages.get_messages(request):
            settings.save()
            messages.success(request, "Organization settings saved.")

    currencies = list(Currency.objects.using(tenant_db).filter(status="active").order_by("code"))
    return render(
        request,
        "tenant_portal/organization_settings.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "settings": settings,
            "can_edit": can_edit,
            "currencies": currencies,
        },
    )


def tenant_home_view(request: HttpRequest) -> HttpResponse:
    tenant = getattr(request, "tenant", None)
    tenant_db = get_tenant_db_for_request(request)
    if not tenant:
        return render(request, "tenant_portal/tenant_missing.html", status=404)
    if not tenant_db:
        return render(request, "tenant_portal/tenant_not_provisioned.html", {"tenant": tenant}, status=503)

    user = get_tenant_user(request)
    if not user:
        return redirect(reverse("tenant_portal:login"))

    enabled = set(tenant.modules.values_list("code", flat=True))

    modules = [
        {"key": "dashboard", "name": "Dashboard", "perm": "platform:dashboard.view", "icon": "bar-chart-2", "requires": None},
        {
            "key": "finance_grants",
            "name": "Financial & Grant Management",
            "perm_any": ["module:finance.view", "module:grants.view"],
            "icon": "layers",
            "requires": "finance_grants",
        },
        {"key": "integrations", "name": "Integrations", "perm_any": ["module:integrations.manage"], "icon": "link", "requires": "integrations"},
        {"key": "users", "name": "Users", "perm_any": ["users:manage"], "icon": "users", "requires": None},
        {"key": "rbac", "name": "Roles & Permissions", "perm_any": ["rbac:roles.manage"], "icon": "shield", "requires": None},
        {"key": "modules", "name": "Modules", "perm_any": ["module:modules.manage"], "icon": "grid", "requires": None},
    ]

    def _is_visible(m):
        if m.get("requires") and m["requires"] not in enabled:
            return False
        perms = m.get("perm_any") or ([m["perm"]] if m.get("perm") else [])
        return any(user_has_permission(user, p, using=tenant_db) for p in perms)

    visible_modules = [m for m in modules if _is_visible(m)]

    return render(
        request,
        "tenant_portal/home.html",
        {"tenant": tenant, "tenant_user": user, "modules": visible_modules},
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_home_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal

    from django.db.models import Sum
    from django.db.models.functions import TruncMonth
    from django.utils import timezone
    from datetime import timedelta

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import BudgetLine, Grant, GrantApproval

    tenant_db = request.tenant_db

    # Date range: current calendar month (dashboard KPI window).
    today = timezone.now().date()
    period_start = today.replace(day=1)

    # Core counts
    accounts_count = ChartAccount.objects.using(tenant_db).count()
    journals_count = JournalEntry.objects.using(tenant_db).count()
    grants_count = Grant.objects.using(tenant_db).count()

    # Total funds from grant award amounts.
    total_funds_available = (
        Grant.objects.using(tenant_db).aggregate(total=Sum("award_amount")).get("total") or Decimal("0")
    )

    # Cash balance approximation: sum of debit-credit on asset accounts.
    cash_balance = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.ASSET)
        .aggregate(total=Sum("debit") - Sum("credit"))
        .get("total")
        or Decimal("0")
    )

    # Total expenses in current period: expense lines in current month.
    expenses_period = (
        JournalLine.objects.using(tenant_db)
        .filter(
            account__type=ChartAccount.Type.EXPENSE,
            entry__entry_date__gte=period_start,
            entry__entry_date__lte=today,
        )
        .aggregate(total=Sum("debit") - Sum("credit"))
        .get("total")
        or Decimal("0")
    )

    # Budget utilisation across all grants.
    budget_total = (
        BudgetLine.objects.using(tenant_db).aggregate(total=Sum("amount")).get("total") or Decimal("0")
    )
    spent_total = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE)
        .aggregate(total=Sum("debit"))
        .get("total")
        or Decimal("0")
    )
    budget_util_pct = Decimal("0")
    if budget_total > 0:
        budget_util_pct = (Decimal(spent_total) / Decimal(budget_total)) * Decimal("100")

    active_grants_count = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).count()

    # Pending financial approvals: reuse grant approvals as a proxy.
    pending_financial_approvals = (
        GrantApproval.objects.using(tenant_db)
        .filter(status=GrantApproval.Status.PENDING)
        .count()
    )

    # Recent journal entries for the transactions table.
    recent_entries_qs = (
        JournalEntry.objects.using(tenant_db)
        .prefetch_related("lines", "lines__account", "grant")
        .order_by("-entry_date", "-id")[:10]
    )
    recent_transactions = []
    for entry in recent_entries_qs:
        lines = list(entry.lines.all())
        total = sum((line.debit - line.credit) for line in lines) if lines else Decimal("0")
        main_account = lines[0].account.name if lines else ""
        recent_transactions.append(
            {
                "date": entry.entry_date,
                "reference": entry.reference or f"JE-{entry.id:05d}",
                "type": "Journal entry",
                "account": main_account,
                "project": entry.grant.title if entry.grant else "",
                "amount": total,
                "status": "Posted",
            }
        )

    # Donor contribution distribution based on grant award amounts.
    donor_rows = (
        Grant.objects.using(tenant_db)
        .values("donor__name")
        .annotate(total=Sum("award_amount"))
        .order_by("-total")[:5]
    )
    donor_total = sum((row["total"] or Decimal("0")) for row in donor_rows)
    donor_contrib = []
    for row in donor_rows:
        amount = row["total"] or Decimal("0")
        pct = (amount / donor_total * Decimal("100")) if donor_total > 0 else Decimal("0")
        donor_contrib.append(
            {"name": row["donor__name"], "amount": amount, "pct": pct}
        )

    # Grant utilisation per top grants (by award amount).
    budgets_by_grant = {
        row["grant_id"]: row["total"] or Decimal("0")
        for row in BudgetLine.objects.using(tenant_db)
        .values("grant_id")
        .annotate(total=Sum("amount"))
    }
    spend_by_grant = {
        row["entry__grant_id"]: row["spent"] or Decimal("0")
        for row in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(spent=Sum("debit"))
    }
    grant_util_rows = []
    top_grants = (
        Grant.objects.using(tenant_db)
        .filter(status=Grant.Status.ACTIVE)
        .order_by("-award_amount")[:3]
    )
    for g in top_grants:
        budget = budgets_by_grant.get(g.id, Decimal("0"))
        spent = spend_by_grant.get(g.id, Decimal("0"))
        ceiling = budget if budget > 0 else Decimal(str(g.award_amount or 0))
        util_pct = (spent / ceiling * Decimal("100")) if ceiling > 0 else Decimal("0")
        grant_util_rows.append(
            {
                "code": g.code,
                "title": g.title,
                "util_pct": util_pct,
            }
        )

    # Monthly expense trend for the last 6 months.
    expense_trend_rows = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE)
        .annotate(month=TruncMonth("entry__entry_date"))
        .values("month")
        .annotate(total=Sum("debit"))
        .order_by("month")
    )
    expense_trend_rows = list(expense_trend_rows)[-6:]
    max_expense = max(
        (row["total"] or Decimal("0") for row in expense_trend_rows), default=Decimal("0")
    )
    expense_trend = []
    for row in expense_trend_rows:
        amount = row["total"] or Decimal("0")
        width = (amount / max_expense * Decimal("100")) if max_expense > 0 else Decimal("0")
        expense_trend.append(
            {
                "label": row["month"],
                "amount": amount,
                "pct": width,
            }
        )

    # Active grants summary metrics.
    avg_utilisation = Decimal("0")
    if grant_util_rows:
        total_util = sum((row["util_pct"] for row in grant_util_rows), Decimal("0"))
        avg_utilisation = total_util / Decimal(len(grant_util_rows))

    # Grants ending in the next 90 days.
    ninety_days = today + timedelta(days=90)
    grants_ending_soon = (
        Grant.objects.using(tenant_db)
        .filter(
            status=Grant.Status.ACTIVE,
            end_date__isnull=False,
            end_date__gte=today,
            end_date__lte=ninety_days,
        )
        .count()
    )

    context = {
        "tenant": request.tenant,
        "tenant_user": request.tenant_user,
        "accounts_count": accounts_count,
        "journals_count": journals_count,
        "grants_count": grants_count,
        "total_funds_available": total_funds_available,
        "cash_balance": cash_balance,
        "expenses_period": expenses_period,
        "budget_util_pct": budget_util_pct,
        "active_grants_count": active_grants_count,
        "pending_financial_approvals": pending_financial_approvals,
        "recent_transactions": recent_transactions,
        "donor_contrib": donor_contrib,
        "grant_util_rows": grant_util_rows,
        "expense_trend": expense_trend,
        "avg_utilisation": avg_utilisation,
        "grants_ending_soon": grants_ending_soon,
        "period_start": period_start,
        "period_end": today,
    }
    context["active_submenu"] = "dashboard"
    context["active_item"] = "dashboard_overview"
    context["global_indicators"] = _get_global_financial_indicators(request)
    return render(request, "tenant_portal/finance/home.html", context)


def _parse_finance_filters(request):
    """Parse common GET filters for finance overview pages."""
    from django.utils.dateparse import parse_date
    from django.utils import timezone

    today = timezone.now().date()
    period_start = parse_date(request.GET.get("period_start") or "")
    period_end = parse_date(request.GET.get("period_end") or "")
    if not period_start:
        period_start = today.replace(day=1)
    if not period_end:
        period_end = today
    if period_end < period_start:
        period_end = period_start
    grant_id = request.GET.get("grant_id") or ""
    donor_id = request.GET.get("donor_id") or ""
    return {
        "period_start": period_start,
        "period_end": period_end,
        "grant_id": grant_id,
        "donor_id": donor_id,
    }


def _parse_cfs_filters(request: HttpRequest) -> dict:
    """Cash flow statement: period, grant, donor, currency, fund dimension, project."""
    f = _parse_finance_filters(request)
    f["currency_id"] = (request.GET.get("currency_id") or "").strip()
    f["dimension_id"] = (request.GET.get("dimension_id") or "").strip()
    f["project_id"] = (request.GET.get("project_id") or "").strip()
    return f


def _cfs_export_urls(request: HttpRequest) -> dict[str, str]:
    from urllib.parse import urlencode

    q = request.GET.copy()
    base = request.path + "?"
    out = {}
    for fmt in ("csv", "xlsx", "pdf"):
        qc = q.copy()
        qc["format"] = fmt
        out[fmt] = base + qc.urlencode()
    return out


def _tenant_legal_name_for_official_reports(request: HttpRequest) -> str:
    """Registered / legal name for print, PDF, Excel (CSV/XLSX), and official report headers only."""
    tenant_db = getattr(request, "tenant_db", None)
    tenant = getattr(request, "tenant", None)
    fallback = (getattr(tenant, "name", None) or "").strip() if tenant else ""
    if not tenant_db:
        return fallback
    try:
        from tenant_finance.models import OrganizationSettings

        s = OrganizationSettings.objects.using(tenant_db).first()
        if s and (s.organization_name or "").strip():
            return (s.organization_name or "").strip()
    except Exception:
        pass
    return fallback


def _official_csv_preamble(
    w,
    request: HttpRequest,
    report_title: str,
    meta_rows: list[tuple[str, str]] | None = None,
) -> None:
    """Write organization legal name and report title at top of official CSV exports."""
    w.writerow([_tenant_legal_name_for_official_reports(request)])
    w.writerow([report_title])
    if meta_rows:
        for label, value in meta_rows:
            w.writerow([label, value])
    w.writerow([])


def _official_report_pdf_logo_flowable(request: HttpRequest):
    """ReportLab Image flowable for organization logo, or None if unavailable."""
    tenant_db = getattr(request, "tenant_db", None)
    if not tenant_db:
        return None
    try:
        import os

        from tenant_finance.models import OrganizationSettings
        from reportlab.lib.units import inch
        from reportlab.platypus import Image as RLImage

        s = OrganizationSettings.objects.using(tenant_db).first()
        if not s or not s.organization_logo:
            return None
        path = s.organization_logo.path
        if not os.path.isfile(path):
            return None
        return RLImage(path, width=1.75 * inch, height=0.52 * inch)
    except Exception:
        return None


def _parse_sfp_filters(request):
    """Statement of Financial Position: single as-at date, grant, optional donor."""
    from django.utils.dateparse import parse_date
    from django.utils import timezone

    today = timezone.now().date()
    as_at = parse_date((request.GET.get("as_at") or "").strip())
    if not as_at:
        as_at = parse_date((request.GET.get("period_end") or "").strip()) or today
    grant_id = (request.GET.get("grant_id") or "").strip()
    donor_id = (request.GET.get("donor_id") or "").strip()
    return {"as_at": as_at, "grant_id": grant_id, "donor_id": donor_id}


def _parse_gl_filters(request: HttpRequest) -> dict:
    """General Ledger: extends finance filters with GL-specific query params."""
    f = _parse_finance_filters(request)
    f["project_id"] = (request.GET.get("project_id") or "").strip()
    f["source_type"] = (request.GET.get("source_type") or "").strip()
    f["journal_no"] = (request.GET.get("journal_no") or "").strip()
    f["budget_line"] = (request.GET.get("budget_line") or "").strip()
    return f


def _gl_journal_number(entry) -> str:
    from tenant_finance.models import JournalEntry

    if not entry:
        return "—"
    doc = (entry.source_document_no or "").strip()
    ref = (entry.reference or "").strip()
    return doc or ref or f"JE-{entry.pk}"


def _gl_source_label(entry) -> str:
    from tenant_finance.models import JournalEntry

    if not entry:
        return "—"
    jt = (entry.journal_type or "").lower()
    st = (entry.source_type or "").strip()
    if "recurring" in jt:
        return "Recurring journal"
    if getattr(entry, "adjustment_type", None):
        return "Adjusting journal"
    if st == JournalEntry.SourceType.INTER_FUND_TRANSFER or jt == "inter_fund_transfer":
        return "Inter-fund transfer"
    if st == JournalEntry.SourceType.PAYMENT_VOUCHER or jt == "payment_voucher":
        return "Payment voucher"
    if st == JournalEntry.SourceType.RECEIPT_VOUCHER or jt == "receipt_voucher":
        return "Receipt voucher"
    if jt in ("adjustment", "adjusting_journal", "adjusting"):
        return "Adjusting journal"
    if st == JournalEntry.SourceType.MANUAL and not entry.is_system_generated:
        if entry.adjustment_type:
            return "Adjusting journal"
        return "Manual journal"
    if st:
        return dict(JournalEntry.SourceType.choices).get(st, st.replace("_", " ").title())
    if jt:
        return jt.replace("_", " ").title()
    if entry.is_system_generated:
        return "System"
    return "Journal entry"


def _finance_export_csv_url(request):
    """Build current page URL with format=csv for export button."""
    from urllib.parse import urlencode

    q = request.GET.copy()
    q["format"] = "csv"
    return request.path + "?" + q.urlencode()


def _journal_list_preset_urls(request: HttpRequest) -> dict[str, str]:
    """Query-string URLs for Today / This week / This accounting period quick filters (journal register)."""
    q = request.GET.copy()
    q.pop("quick", None)
    enc = q.urlencode()
    prefix = request.path + ("?" + enc if enc else "")
    sep = "&" if enc else "?"
    return {
        "today": prefix + sep + "quick=today",
        "week": prefix + sep + "quick=week",
        "period": prefix + sep + "quick=period",
    }


def _apply_journal_quick_date_range(request, tenant_db: str, f: dict) -> None:
    """Mutate f period_start / period_end when ?quick=today|week|period (period = current accounting period)."""
    from datetime import timedelta

    from django.utils import timezone

    from tenant_finance.models import FiscalPeriod

    quick = (request.GET.get("quick") or "").strip()
    if not quick:
        return
    today = timezone.now().date()
    if quick == "today":
        f["period_start"] = f["period_end"] = today
    elif quick == "week":
        start = today - timedelta(days=today.weekday())
        f["period_start"], f["period_end"] = start, today
    elif quick == "period":
        fp = (
            FiscalPeriod.objects.using(tenant_db)
            .filter(start_date__lte=today, end_date__gte=today)
            .order_by("-start_date")
            .first()
        )
        if fp:
            f["period_start"], f["period_end"] = fp.start_date, fp.end_date


def _accounting_period_label_for_date(periods: list, d) -> str:
    """Resolved from Financial Setup calendar: fiscal year + accounting period containing date d."""
    for p in periods:
        if p.start_date <= d <= p.end_date:
            sub = (p.period_name or p.name or "").strip() or f"P{p.period_number}"
            fy = getattr(p, "fiscal_year", None)
            if fy and getattr(fy, "name", None):
                return f"{fy.name} — {sub}"
            return sub
    return "—"


def _journal_source_document_url(entry, je_schema_0034: bool) -> str | None:
    """Relative URL to open originating business document, if known."""
    from django.urls import reverse

    from tenant_finance.models import JournalEntry

    if not je_schema_0034:
        return None
    st = (entry.source_type or "").strip()
    sid = getattr(entry, "source_id", None)
    if not sid:
        return None
    try:
        if st == JournalEntry.SourceType.PAYMENT_VOUCHER:
            return reverse("tenant_portal:pay_payment_voucher_detail", args=[int(sid)])
        if st == JournalEntry.SourceType.RECEIPT_VOUCHER:
            return reverse("tenant_portal:finance_journal_detail", args=[int(sid)])
        if st in (
            JournalEntry.SourceType.FUND_TRANSFER,
            JournalEntry.SourceType.INTER_FUND_TRANSFER,
        ):
            return reverse("tenant_portal:finance_interfund_transfer_detail", args=[int(sid)])
    except Exception:
        return None
    return None


def _journal_audit_trail_url(entry_id: int) -> str:
    from django.urls import reverse
    from urllib.parse import urlencode

    base = reverse("tenant_portal:finance_audit_trail")
    return base + "?" + urlencode({"model_name": "journalentry", "object_id": str(entry_id)})


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_cash_position_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalLine

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    # Balances per asset account (cash/bank) as of period_end
    account_balances = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.ASSET, entry__entry_date__lte=f["period_end"])
        .values("account_id", "account__code", "account__name")
        .annotate(balance=Sum("debit") - Sum("credit"))
    )
    if f["grant_id"]:
        account_balances = account_balances.filter(entry__grant_id=f["grant_id"])

    rows = []
    total_balance = Decimal("0")
    for row in account_balances:
        bal = row.get("balance") or Decimal("0")
        total_balance += bal
        rows.append({
            "code": row.get("account__code"),
            "name": row.get("account__name"),
            "balance": bal,
        })

    # Inflows/outflows in period (asset account movements)
    inflows = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.ASSET, entry__entry_date__gte=f["period_start"], entry__entry_date__lte=f["period_end"])
        .aggregate(total=Sum("debit"))
    )
    outflows = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.ASSET, entry__entry_date__gte=f["period_start"], entry__entry_date__lte=f["period_end"])
        .aggregate(total=Sum("credit"))
    )
    period_inflows = (inflows.get("total") or Decimal("0"))
    period_outflows = (outflows.get("total") or Decimal("0"))

    grants = __import__("tenant_grants.models", fromlist=["Grant"]).Grant.objects.using(tenant_db).filter(status="active").order_by("code")
    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")

    if request.GET.get("format") == "csv":
        import csv
        from io import StringIO
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="cash_position.csv"'
        w = csv.writer(response)
        w.writerow(["Account Code", "Account Name", "Balance"])
        for row in rows:
            w.writerow([row.get("code") or "", row.get("name") or "", str(row.get("balance") or "0")])
        return response

    return render(
        request,
        "tenant_portal/finance/cash_position.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_cash",
            "filters": f,
            "account_balances": rows,
            "total_balance": total_balance,
            "period_inflows": period_inflows,
            "period_outflows": period_outflows,
            "grants": grants,
            "donors": donors,
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_fund_balances_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalLine
    from tenant_finance.services.financial_reporting import filter_grants_for_report_dropdown
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    # Per-grant fund balance: award_amount as ceiling, spent from expense lines (cumulative posted)
    grants_for_filter = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).select_related("donor").filter(status=Grant.Status.ACTIVE).order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if f["donor_id"]:
        grants_for_filter = grants_for_filter.filter(donor_id=f["donor_id"])
    grants_qs = grants_for_filter
    grant_id = (f.get("grant_id") or "").strip()
    if grant_id:
        grants_qs = grants_qs.filter(pk=grant_id)

    fund_rows = []
    for g in grants_qs:
        spent = (
            JournalLine.objects.using(tenant_db)
            .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id=g.id)
            .aggregate(t=Sum("debit"))
        ).get("t") or Decimal("0")
        award = Decimal(str(g.award_amount or 0))
        remaining = award - spent
        fund_rows.append({
            "grant": g,
            "award_amount": award,
            "spent": spent,
            "remaining": remaining,
            "restricted": True,  # grant-funded = restricted in typical NGO usage
        })

    donors = Donor.objects.using(tenant_db).order_by("name")
    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="fund_balances.csv"'
        w = csv.writer(response)
        _official_csv_preamble(
            w,
            request,
            "Fund balance report",
            [
                ("Period context", f"{f['period_start']} — {f['period_end']}"),
                ("Grant filter", grant_id or "All grants"),
                ("Donor filter", f["donor_id"] or "All donors"),
            ],
        )
        w.writerow(["Grant", "Donor", "Award Amount", "Spent", "Remaining", "Type"])
        for row in fund_rows:
            g = row["grant"]
            w.writerow([
                f"{g.code} — {g.title}", g.donor.name, row["award_amount"], row["spent"], row["remaining"],
                "Restricted" if row["restricted"] else "Unrestricted",
            ])
        return response
    return render(
        request,
        "tenant_portal/finance/fund_balances.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_ngo_fund_balance",
            "filters": f,
            "fund_rows": fund_rows,
            "donors": donors,
            "grants": filter_grants_for_report_dropdown(
                Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
                request.tenant_user,
                tenant_db,
            ),
            "export_csv_url": _finance_export_csv_url(request),
            "global_indicators": _get_global_financial_indicators(request),
            "official_report_period_line": f"{f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_grant_utilization_view(request: HttpRequest) -> HttpResponse:
    """
    Grant utilization by budget line: posted expense actuals in the selected period
    vs budget, with variance. Matches budget-line logic used on Budget vs actual.
    """
    from collections import defaultdict
    from decimal import Decimal

    from django.db.models import Q
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
    )
    from tenant_grants.models import BudgetLine, Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    grants_qs = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).select_related("donor").order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if f["donor_id"]:
        grants_qs = grants_qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        grants_qs = grants_qs.filter(pk=f["grant_id"])

    grant_ids = list(grants_qs.values_list("pk", flat=True))
    rows: list[dict] = []

    if grant_ids:
        line_qs = posted_journal_lines(tenant_db).filter(
            gl_date__gte=f["period_start"],
            gl_date__lte=f["period_end"],
            account__type=ChartAccount.Type.EXPENSE,
        ).filter(Q(entry__grant_id__in=grant_ids) | Q(grant_id__in=grant_ids))
        line_qs = restrict_journal_lines_by_grant_scope(line_qs, request.tenant_user, tenant_db)

        spending: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0"))
        for g_line, g_entry, account_id, debit in line_qs.values_list(
            "grant_id", "entry__grant_id", "account_id", "debit"
        ):
            gid = g_line or g_entry
            if gid is None or account_id is None:
                continue
            spending[(int(gid), int(account_id))] += debit or Decimal("0")

        bl_qs = (
            BudgetLine.objects.using(tenant_db)
            .filter(grant_id__in=grant_ids)
            .select_related("grant", "account")
            .order_by("grant__code", "id")
        )
        bl_list = list(bl_qs)

        by_grant_account: dict[tuple[int, int], list] = defaultdict(list)
        for bl in bl_list:
            if bl.account_id:
                by_grant_account[(bl.grant_id, bl.account_id)].append(bl)

        allocated_actual: dict[int, Decimal] = {}
        for (_gid, _aid), grp in by_grant_account.items():
            spent = spending.get((_gid, _aid), Decimal("0"))
            total_budget = sum((x.amount or Decimal("0")) for x in grp)
            if total_budget > 0:
                running = Decimal("0")
                for i, bl in enumerate(grp):
                    amt = bl.amount or Decimal("0")
                    if i == len(grp) - 1:
                        allocated_actual[bl.id] = (spent - running).quantize(Decimal("0.01"))
                    else:
                        part = (spent * amt / total_budget).quantize(Decimal("0.01"))
                        allocated_actual[bl.id] = part
                        running += part
            else:
                for i, bl in enumerate(grp):
                    allocated_actual[bl.id] = spent if i == 0 else Decimal("0")

        for bl in bl_list:
            budget_amt = bl.amount or Decimal("0")
            if bl.account_id:
                actual_amt = allocated_actual.get(bl.id, Decimal("0"))
            else:
                actual_amt = Decimal("0")
            acct = bl.account
            variance = actual_amt - budget_amt
            rows.append(
                {
                    "grant": bl.grant,
                    "account_display": f"{acct.code} — {acct.name}" if acct else "—",
                    "budget_line_label": bl.category,
                    "description": (bl.description or "").strip() or "—",
                    "budget": budget_amt,
                    "actual": actual_amt,
                    "variance": variance,
                }
            )

    donors = Donor.objects.using(tenant_db).order_by("name")
    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="grant_utilization.csv"'
        w = csv.writer(response)
        _official_csv_preamble(
            w,
            request,
            "Grant utilization",
            [("Period", f"{f['period_start']} to {f['period_end']}")],
        )
        w.writerow(["Grant", "Account", "Budget line", "Description", "Budget", "Actual", "Variance"])
        for row in rows:
            g = row["grant"]
            w.writerow(
                [
                    f"{g.code} — {g.title}",
                    row["account_display"],
                    row["budget_line_label"],
                    row["description"],
                    row["budget"],
                    row["actual"],
                    row["variance"],
                ]
            )
        return response
    return render(
        request,
        "tenant_portal/finance/grant_utilization.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_grant_util",
            "filters": f,
            "utilization_rows": rows,
            "grants": filter_grants_for_report_dropdown(
                Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
                request.tenant_user,
                tenant_db,
            ),
            "donors": donors,
            "export_csv_url": _finance_export_csv_url(request),
            "official_report_period_line": f"Period: {f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_project_budget_view(request: HttpRequest) -> HttpResponse:
    """Maintain project budget lines (allocated / remaining) for activity-based planning."""
    from decimal import Decimal, InvalidOperation

    from django.shortcuts import redirect
    from django.urls import reverse

    from tenant_finance.models import ChartAccount
    from tenant_grants.models import Project, ProjectBudget, ProjectBudgetLine

    tenant_db = request.tenant_db
    projects = list(Project.objects.using(tenant_db).order_by("code")[:400])
    pid = (request.GET.get("project_id") or request.POST.get("project_id") or "").strip()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if not pid.isdigit():
            messages.error(request, "Select a project.")
        else:
            proj = Project.objects.using(tenant_db).filter(pk=int(pid)).first()
            if not proj:
                messages.error(request, "Project not found.")
            elif action == "ensure_budget":
                ProjectBudget.objects.using(tenant_db).get_or_create(
                    project=proj, name="Main", defaults={}
                )
                messages.success(request, "Project budget container ready.")
                return redirect(reverse("tenant_portal:grants_project_budget") + f"?project_id={pid}")
            elif action == "add_line":
                budget = (
                    ProjectBudget.objects.using(tenant_db)
                    .filter(project_id=int(pid), name="Main")
                    .first()
                )
                if not budget:
                    messages.error(request, "Create the project budget first (Main).")
                else:
                    cat = (request.POST.get("category") or "").strip()
                    raw_amt = (request.POST.get("allocated_amount") or "0").replace(",", "")
                    acc_id = (request.POST.get("account_id") or "").strip()
                    desc = (request.POST.get("description") or "").strip()
                    try:
                        amt = Decimal(raw_amt)
                    except InvalidOperation:
                        amt = Decimal("0")
                    acct = (
                        ChartAccount.objects.using(tenant_db).filter(pk=int(acc_id)).first()
                        if acc_id.isdigit()
                        else None
                    )
                    if not cat:
                        messages.error(request, "Category is required.")
                    elif amt < 0:
                        messages.error(request, "Allocated amount cannot be negative.")
                    else:
                        ProjectBudgetLine.objects.using(tenant_db).create(
                            project_budget=budget,
                            category=cat,
                            description=desc,
                            account=acct,
                            allocated_amount=amt,
                            remaining_amount=amt,
                        )
                        messages.success(request, "Budget line added.")
                        return redirect(reverse("tenant_portal:grants_project_budget") + f"?project_id={pid}")

    budget = None
    lines: list = []
    selected_project = None
    if pid.isdigit():
        selected_project = Project.objects.using(tenant_db).filter(pk=int(pid)).first()
        if selected_project:
            budget = (
                ProjectBudget.objects.using(tenant_db)
                .filter(project=selected_project, name="Main")
                .first()
            )
            if budget:
                lines = list(
                    ProjectBudgetLine.objects.using(tenant_db)
                    .select_related("account")
                    .filter(project_budget=budget)
                    .order_by("id")
                )

    accounts = list(
        ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")[:500]
    )
    return render(
        request,
        "tenant_portal/grants/project_budget.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_project_budget",
            "projects": projects,
            "selected_project": selected_project,
            "project_budget": budget,
            "budget_lines": lines,
            "accounts": accounts,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_project_budget_activity_report_view(request: HttpRequest) -> HttpResponse:
    """Project budget vs actual by line; activity planned vs actual; grant/donor columns."""
    from decimal import Decimal

    from django.db.models import Sum

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Donor, Grant, Project, ProjectBudgetLine, WorkplanActivity

    tenant_db = request.tenant_db
    pid = (request.GET.get("project_id") or "").strip()
    gid = (request.GET.get("grant_id") or "").strip()
    did = (request.GET.get("donor_id") or "").strip()

    projects = list(Project.objects.using(tenant_db).order_by("code")[:400])
    donors = list(Donor.objects.using(tenant_db).order_by("name"))
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:300])

    line_rows: list[dict] = []
    act_rows: list[dict] = []

    pbl_qs = (
        ProjectBudgetLine.objects.using(tenant_db)
        .select_related("project_budget__project", "account")
        .order_by("project_budget__project__code", "id")
    )
    if pid.isdigit():
        pbl_qs = pbl_qs.filter(project_budget__project_id=int(pid))
    pbl_list = list(pbl_qs[:500])

    for bl in pbl_list:
        actual = (
            JournalLine.objects.using(tenant_db)
            .filter(
                project_budget_line_id=bl.pk,
                entry__status=JournalEntry.Status.POSTED,
                account__type=ChartAccount.Type.EXPENSE,
                debit__gt=0,
            )
            .aggregate(s=Sum("debit"))
            .get("s")
            or Decimal("0")
        )
        alloc = bl.allocated_amount or Decimal("0")
        line_rows.append(
            {
                "project": bl.project_budget.project,
                "category": bl.category,
                "account": bl.account,
                "allocated": alloc,
                "actual": actual,
                "remaining": bl.remaining_amount,
                "variance": actual - alloc,
            }
        )

    wa_qs = (
        WorkplanActivity.objects.using(tenant_db)
        .select_related("grant", "grant__donor", "project_budget_line")
        .order_by("-id")
    )
    if pid.isdigit():
        wa_qs = wa_qs.filter(project_id=int(pid))
    if gid.isdigit():
        wa_qs = wa_qs.filter(grant_id=int(gid))
    if did.isdigit():
        wa_qs = wa_qs.filter(grant__donor_id=int(did))
    for w in wa_qs[:500]:
        act_rows.append(
            {
                "activity_code": w.activity_code or w.workplan_code,
                "name": w.activity,
                "grant": w.grant,
                "donor": w.grant.donor if w.grant_id else None,
                "budget_line": w.project_budget_line.category if w.project_budget_line_id else (w.budget_line or "—"),
                "planned": w.budget_amount or Decimal("0"),
                "actual": w.actual_cost or Decimal("0"),
            }
        )

    if request.GET.get("format") == "csv":
        import csv

        from django.http import HttpResponse

        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="project_budget_activity.csv"'
        w = csv.writer(resp)
        w.writerow(["Section", "Project", "Category / Activity", "Grant", "Donor", "Allocated / Planned", "Actual", "Variance / Remaining"])
        for r in line_rows:
            p = r["project"]
            w.writerow(
                [
                    "Budget line",
                    p.code if p else "",
                    r["category"],
                    "",
                    "",
                    r["allocated"],
                    r["actual"],
                    r["allocated"] - r["actual"],
                ]
            )
        for r in act_rows:
            g = r["grant"]
            d = r["donor"]
            w.writerow(
                [
                    "Activity",
                    "",
                    r["name"],
                    g.code if g else "",
                    d.name if d else "",
                    r["planned"],
                    r["actual"],
                    r["actual"] - r["planned"],
                ]
            )
        return resp

    return render(
        request,
        "tenant_portal/finance/project_budget_activity_report.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_project_budget_activity",
            "projects": projects,
            "grants": grants,
            "donors": donors,
            "line_rows": line_rows,
            "act_rows": act_rows,
            "filter_project_id": pid,
            "filter_grant_id": gid,
            "filter_donor_id": did,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_project_financial_status_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalLine
    from tenant_grants.models import BudgetLine, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    budgets_by_grant = {
        r["grant_id"]: r["total"] or Decimal("0")
        for r in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
    }
    spend_by_grant = {
        r["entry__grant_id"]: r["spent"] or Decimal("0")
        for r in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(spent=Sum("debit"))
    }

    projects_qs = Grant.objects.using(tenant_db).select_related("donor").filter(status=Grant.Status.ACTIVE).order_by("code")
    if f["donor_id"]:
        projects_qs = projects_qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        projects_qs = projects_qs.filter(pk=f["grant_id"])

    rows = []
    for g in projects_qs:
        budget = budgets_by_grant.get(g.id, Decimal("0"))
        spent = spend_by_grant.get(g.id, Decimal("0"))
        remaining = budget - spent
        rows.append({
            "project": g,
            "budget": budget,
            "expenses": spent,
            "remaining": remaining,
        })

    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")
    grants = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")
    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="project_financial_status.csv"'
        w = csv.writer(response)
        w.writerow(["Project/Grant", "Donor", "Budget", "Expenses", "Remaining"])
        for row in rows:
            p = row["project"]
            w.writerow([f"{p.code} — {p.title}", p.donor.name, row["budget"], row["expenses"], row["remaining"]])
        return response
    return render(
        request,
        "tenant_portal/finance/project_financial_status.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_project_status",
            "filters": f,
            "project_rows": rows,
            "grants": grants,
            "donors": donors,
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_donor_funding_status_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalLine
    from tenant_finance.services.financial_reporting import filter_grants_for_report_dropdown
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    donors_qs = Donor.objects.using(tenant_db).order_by("name")
    if f["donor_id"]:
        donors_qs = donors_qs.filter(pk=f["donor_id"])

    spend_by_grant = {
        r["entry__grant_id"]: r["spent"] or Decimal("0")
        for r in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(spent=Sum("debit"))
    }

    grant_id = (f.get("grant_id") or "").strip()

    rows = []
    for d in donors_qs:
        grants = Grant.objects.using(tenant_db).filter(donor=d)
        if grant_id:
            grants = grants.filter(pk=grant_id)
        total_commitment = sum(Decimal(str(g.award_amount or 0)) for g in grants)
        total_spent = sum(spend_by_grant.get(g.id, Decimal("0")) for g in grants)
        remaining = total_commitment - total_spent
        rows.append({
            "donor": d,
            "grants_count": grants.count(),
            "total_commitment": total_commitment,
            "total_spent": total_spent,
            "remaining": remaining,
        })

    if grant_id:
        rows = [r for r in rows if r["grants_count"] > 0]

    donors_list = Donor.objects.using(tenant_db).order_by("name")
    grants_list = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="donor_funding_status.csv"'
        w = csv.writer(response)
        _official_csv_preamble(
            w,
            request,
            "Donor contributions report",
            [
                ("Period", f"{f['period_start']} — {f['period_end']}"),
                ("Grant filter", grant_id or "All grants"),
                ("Donor filter", f["donor_id"] or "All donors"),
            ],
        )
        w.writerow(["Donor", "Total Commitment", "Received/Spent", "Remaining"])
        for row in rows:
            w.writerow([row["donor"].name, row["total_commitment"], row["total_spent"], row["remaining"]])
        return response
    return render(
        request,
        "tenant_portal/finance/donor_funding_status.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_ngo_donor_contributions",
            "filters": f,
            "donor_rows": rows,
            "donors": donors_list,
            "grants": grants_list,
            "export_csv_url": _finance_export_csv_url(request),
            "official_report_period_line": f"{f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_budget_vs_actual_view(request: HttpRequest) -> HttpResponse:
    from collections import defaultdict
    from decimal import Decimal

    from django.db.models import Q
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
    )
    from tenant_grants.models import BudgetLine, Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    grants_qs = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).select_related("donor").order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if f["donor_id"]:
        grants_qs = grants_qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        grants_qs = grants_qs.filter(pk=f["grant_id"])

    grant_ids = list(grants_qs.values_list("pk", flat=True))
    rows: list[dict] = []

    if grant_ids:
        line_qs = posted_journal_lines(tenant_db).filter(
            gl_date__gte=f["period_start"],
            gl_date__lte=f["period_end"],
            account__type=ChartAccount.Type.EXPENSE,
        ).filter(Q(entry__grant_id__in=grant_ids) | Q(grant_id__in=grant_ids))
        line_qs = restrict_journal_lines_by_grant_scope(line_qs, request.tenant_user, tenant_db)

        spending: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0"))
        for g_line, g_entry, account_id, debit in line_qs.values_list(
            "grant_id", "entry__grant_id", "account_id", "debit"
        ):
            gid = g_line or g_entry
            if gid is None or account_id is None:
                continue
            spending[(int(gid), int(account_id))] += debit or Decimal("0")

        bl_qs = (
            BudgetLine.objects.using(tenant_db)
            .filter(grant_id__in=grant_ids)
            .select_related("grant", "account")
            .order_by("grant__code", "id")
        )
        bl_list = list(bl_qs)

        by_grant_account: dict[tuple[int, int], list] = defaultdict(list)
        for bl in bl_list:
            if bl.account_id:
                by_grant_account[(bl.grant_id, bl.account_id)].append(bl)

        allocated_actual: dict[int, Decimal] = {}
        for (_gid, _aid), grp in by_grant_account.items():
            spent = spending.get((_gid, _aid), Decimal("0"))
            total_budget = sum((x.amount or Decimal("0")) for x in grp)
            if total_budget > 0:
                running = Decimal("0")
                for i, bl in enumerate(grp):
                    amt = bl.amount or Decimal("0")
                    if i == len(grp) - 1:
                        allocated_actual[bl.id] = (spent - running).quantize(Decimal("0.01"))
                    else:
                        part = (spent * amt / total_budget).quantize(Decimal("0.01"))
                        allocated_actual[bl.id] = part
                        running += part
            else:
                for i, bl in enumerate(grp):
                    allocated_actual[bl.id] = spent if i == 0 else Decimal("0")

        for bl in bl_list:
            budget_amt = bl.amount or Decimal("0")
            if bl.account_id:
                actual_amt = allocated_actual.get(bl.id, Decimal("0"))
            else:
                actual_amt = Decimal("0")
            acct = bl.account
            rows.append({
                "grant": bl.grant,
                "account_display": f"{acct.code} — {acct.name}" if acct else "—",
                "budget_line_label": bl.category,
                "description": (bl.description or "").strip() or "—",
                "budget": budget_amt,
                "actual": actual_amt,
                "variance": actual_amt - budget_amt,
            })

    donors = Donor.objects.using(tenant_db).order_by("name")
    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="budget_vs_actual.csv"'
        w = csv.writer(response)
        _official_csv_preamble(
            w,
            request,
            "Budget vs actual",
            [("Period", f"{f['period_start']} to {f['period_end']}")],
        )
        w.writerow(["Grant", "Account", "Budget line", "Description", "Budget", "Actual", "Variance"])
        for row in rows:
            g = row["grant"]
            w.writerow(
                [
                    f"{g.code} — {g.title}",
                    row["account_display"],
                    row["budget_line_label"],
                    row["description"],
                    row["budget"],
                    row["actual"],
                    row["variance"],
                ]
            )
        return response
    return render(
        request,
        "tenant_portal/finance/budget_vs_actual.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_ngo_budget_vs_actual",
            "filters": f,
            "bva_rows": rows,
            "grants": filter_grants_for_report_dropdown(
                Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
                request.tenant_user,
                tenant_db,
            ),
            "donors": donors,
            "export_csv_url": _finance_export_csv_url(request),
            "official_report_period_line": f"{f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_expense_trend_view(request: HttpRequest) -> HttpResponse:
    """
    Activity expenses: period expense vs budget lines (account, line, description,
    budget, actual, variance). KPIs: total spend in period and average per calendar month.
    """
    from collections import defaultdict
    from decimal import Decimal

    from django.db.models import Q, Sum
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
    )
    from tenant_grants.models import BudgetLine, Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    ps, pe = f["period_start"], f["period_end"]
    num_months = (pe.year - ps.year) * 12 + (pe.month - ps.month) + 1

    def _scope_lines():
        q = posted_journal_lines(tenant_db).filter(
            gl_date__gte=ps,
            gl_date__lte=pe,
            account__type=ChartAccount.Type.EXPENSE,
        )
        q = restrict_journal_lines_by_grant_scope(q, request.tenant_user, tenant_db)
        gid = (f.get("grant_id") or "").strip()
        if gid.isdigit():
            gid_i = int(gid)
            q = q.filter(Q(entry__grant_id=gid_i) | Q(grant_id=gid_i))
        did = (f.get("donor_id") or "").strip()
        if did.isdigit():
            did_i = int(did)
            q = q.filter(Q(entry__donor_id=did_i) | Q(donor_id=did_i))
        return q

    total_expenses = (_scope_lines().aggregate(t=Sum("debit"))["t"] or Decimal("0"))
    avg_per_month = total_expenses / Decimal(str(max(num_months, 1)))

    grants_qs = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).select_related("donor").order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if f["donor_id"]:
        grants_qs = grants_qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        grants_qs = grants_qs.filter(pk=f["grant_id"])

    grant_ids = list(grants_qs.values_list("pk", flat=True))
    expense_detail_rows: list[dict] = []

    if grant_ids:
        line_qs = posted_journal_lines(tenant_db).filter(
            gl_date__gte=ps,
            gl_date__lte=pe,
            account__type=ChartAccount.Type.EXPENSE,
        ).filter(Q(entry__grant_id__in=grant_ids) | Q(grant_id__in=grant_ids))
        line_qs = restrict_journal_lines_by_grant_scope(line_qs, request.tenant_user, tenant_db)

        spending: dict[tuple[int, int], Decimal] = defaultdict(lambda: Decimal("0"))
        for g_line, g_entry, account_id, debit in line_qs.values_list(
            "grant_id", "entry__grant_id", "account_id", "debit"
        ):
            gid = g_line or g_entry
            if gid is None or account_id is None:
                continue
            spending[(int(gid), int(account_id))] += debit or Decimal("0")

        bl_list = list(
            BudgetLine.objects.using(tenant_db)
            .filter(grant_id__in=grant_ids)
            .select_related("grant", "account")
            .order_by("grant__code", "id")
        )

        by_grant_account: dict[tuple[int, int], list] = defaultdict(list)
        for bl in bl_list:
            if bl.account_id:
                by_grant_account[(bl.grant_id, bl.account_id)].append(bl)

        allocated_actual: dict[int, Decimal] = {}
        for (_gid, _aid), grp in by_grant_account.items():
            spent = spending.get((_gid, _aid), Decimal("0"))
            total_budget = sum((x.amount or Decimal("0")) for x in grp)
            if total_budget > 0:
                running = Decimal("0")
                for i, bl in enumerate(grp):
                    amt = bl.amount or Decimal("0")
                    if i == len(grp) - 1:
                        allocated_actual[bl.id] = (spent - running).quantize(Decimal("0.01"))
                    else:
                        part = (spent * amt / total_budget).quantize(Decimal("0.01"))
                        allocated_actual[bl.id] = part
                        running += part
            else:
                for i, bl in enumerate(grp):
                    allocated_actual[bl.id] = spent if i == 0 else Decimal("0")

        for bl in bl_list:
            budget_amt = bl.amount or Decimal("0")
            if bl.account_id:
                actual_amt = allocated_actual.get(bl.id, Decimal("0"))
            else:
                actual_amt = Decimal("0")
            acct = bl.account
            expense_detail_rows.append(
                {
                    "grant": bl.grant,
                    "account_display": f"{acct.code} — {acct.name}" if acct else "—",
                    "budget_line_label": bl.category,
                    "description": (bl.description or "").strip() or "—",
                    "budget": budget_amt,
                    "actual": actual_amt,
                    "variance": actual_amt - budget_amt,
                }
            )

    donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="activity_expenses.csv"'
        w = csv.writer(response)
        _official_csv_preamble(
            w,
            request,
            "Activity expenses report",
            [
                ("Period", f"{ps} to {pe}"),
                ("Grant", f["grant_id"] or "All grants"),
                ("Donor", f["donor_id"] or "All donors"),
            ],
        )
        w.writerow(["Grant", "Account", "Budget line", "Description", "Budget", "Actual", "Variance"])
        for row in expense_detail_rows:
            g = row["grant"]
            w.writerow(
                [
                    f"{g.code} — {g.title}",
                    row["account_display"],
                    row["budget_line_label"],
                    row["description"],
                    row["budget"],
                    row["actual"],
                    row["variance"],
                ]
            )
        return response
    return render(
        request,
        "tenant_portal/finance/expense_trend.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_ngo_activity_expenses",
            "filters": f,
            "expense_detail_rows": expense_detail_rows,
            "total_expenses": total_expenses,
            "avg_per_month": avg_per_month,
            "num_months": num_months,
            "grants": filter_grants_for_report_dropdown(
                Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
                request.tenant_user,
                tenant_db,
            ),
            "donors": donors,
            "export_csv_url": _finance_export_csv_url(request),
            "official_report_period_line": f"{ps} — {pe}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_pending_approvals_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import GrantApproval, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    pending = (
        GrantApproval.objects.using(tenant_db)
        .filter(status=GrantApproval.Status.PENDING)
        .select_related("grant", "grant__donor", "requested_by")
        .order_by("-created_at")[:50]
    )
    if f["grant_id"]:
        pending = pending.filter(grant_id=f["grant_id"])

    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="pending_approvals.csv"'
        w = csv.writer(response)
        w.writerow(["Grant", "Requested By", "Requested At", "Status"])
        for item in pending:
            w.writerow([
                f"{item.grant.code} — {item.grant.title}",
                item.requested_by.get_full_name() or item.requested_by.username,
                item.created_at.strftime("%Y-%m-%d %H:%M") if item.created_at else "",
                item.get_status_display(),
            ])
        return response
    return render(
        request,
        "tenant_portal/finance/pending_approvals.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_approvals",
            "filters": f,
            "pending_list": list(pending),
            "grants": Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_recent_transactions_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal

    from tenant_finance.models import JournalEntry, JournalLine

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)

    entries_qs = (
        JournalEntry.objects.using(tenant_db)
        .prefetch_related("lines", "lines__account", "grant")
        .filter(entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])
        .order_by("-entry_date", "-id")[:100]
    )
    if f["grant_id"]:
        entries_qs = entries_qs.filter(grant_id=f["grant_id"])

    transactions = []
    for entry in entries_qs:
        lines = list(entry.lines.all())
        total = sum((line.debit - line.credit) for line in lines) if lines else Decimal("0")
        main_account = lines[0].account.name if lines else ""
        transactions.append({
            "id": entry.id,
            "date": entry.entry_date,
            "reference": entry.reference or f"JE-{entry.id:05d}",
            "type": "Journal entry",
            "account": main_account,
            "project": entry.grant.title if entry.grant else "",
            "amount": total,
            "memo": entry.memo,
        })

    grants = __import__("tenant_grants.models", fromlist=["Grant"]).Grant.objects.using(tenant_db).filter(status="active").order_by("code")
    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="recent_transactions.csv"'
        w = csv.writer(response)
        w.writerow(["Date", "Reference", "Account", "Project", "Amount"])
        for t in transactions:
            d = t.get("date")
            w.writerow([
                d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else d,
                t.get("reference"), t.get("account"), t.get("project"), t.get("amount"),
            ])
        return response
    return render(
        request,
        "tenant_portal/finance/recent_transactions.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_recent_txn",
            "filters": f,
            "transactions": transactions,
            "grants": grants,
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_post_transaction_view(request: HttpRequest) -> HttpResponse:
    """
    Post Transaction hub: NGO / humanitarian operational entry points (vouchers, cash & bank, inter-fund).
    Recurring and adjusting journals remain under Core Accounting. Manual journal is optional here for
    controlled adjustments when the user has posting permission.
    """
    from django.urls import reverse

    tenant_db = request.tenant_db
    user = request.tenant_user

    post_txn_links = [
        {
            "code": "payment_voucher",
            "label": "Payment Voucher",
            "description": "Pay vendors, expenses, and payroll from bank or cash.",
            "url": reverse("tenant_portal:pay_payment_vouchers"),
            "icon": "file-minus",
            "section": "vouchers",
        },
        {
            "code": "receipt_voucher",
            "label": "Receipt Voucher",
            "description": "Record income and deposits to bank or cash.",
            "url": reverse("tenant_portal:recv_receipt_vouchers"),
            "icon": "file-plus",
            "section": "vouchers",
        },
        {
            "code": "cash_transfer",
            "label": "Cash Transfer",
            "description": "Move cash between on-hand / petty cash accounts.",
            "url": reverse("tenant_portal:cash_cash_transfers"),
            "icon": "shuffle",
            "section": "cash_bank",
        },
        {
            "code": "bank_transfer",
            "label": "Bank Transfer",
            "description": "Transfer funds between bank accounts.",
            "url": reverse("tenant_portal:cash_bank_transfers"),
            "icon": "repeat",
            "section": "cash_bank",
        },
        {
            "code": "inter_fund_transfer",
            "label": "New inter-fund transfer",
            "description": "Create a project/bank transfer; approval workflow posts to the general ledger.",
            "url": reverse("tenant_portal:finance_interfund_transfer_create") + "?src=post",
            "icon": "arrow-right-circle",
            "section": "fund_transfers",
        },
        {
            "code": "inter_fund_transfers_register",
            "label": "Inter-fund transfers",
            "description": "Register: review, approve, post, and reverse transfers in workflow.",
            "url": reverse("tenant_portal:finance_interfund_transfers"),
            "icon": "list",
            "section": "fund_transfers",
        },
    ]
    if user_has_permission(user, "finance.add_journalentry", using=tenant_db):
        post_txn_links.append(
            {
                "code": "manual_journal",
                "label": "Manual Journal",
                "description": "Controlled manual journal lines (draft → approval → post).",
                "url": reverse("tenant_portal:finance_journals") + "?open_manual=1",
                "icon": "book-open",
                "section": "manual_optional",
            }
        )

    return render(
        request,
        "tenant_portal/finance/post_transaction.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "post_txn_links": post_txn_links,
            "active_submenu": "dashboard",
            "active_item": "dashboard_post_transaction",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def incoming_fund_center_view(request: HttpRequest) -> HttpResponse:
    """Incoming fund module home: receipt KPIs and navigation to receivables workflows."""
    from decimal import Decimal

    from django.db.models import Q, Sum
    from django.utils import timezone
    from django.utils.translation import gettext as _

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    ps, pe = f["period_start"], f["period_end"]
    today = timezone.now().date()
    month_start = today.replace(day=1)

    def _fmt_money(d: Decimal) -> str:
        d = d.quantize(Decimal("0.01"))
        neg = d < 0
        d = abs(d)
        whole_s, _, frac = f"{d:.2f}".partition(".")
        whole_s = "{:,}".format(int(whole_s))
        return ("-" if neg else "") + whole_s + "." + frac

    receipt_entry_q = (
        Q(reference__startswith="RV-")
        | Q(journal_type__iexact="receipt_voucher")
        | Q(source_type=JournalEntry.SourceType.RECEIPT_VOUCHER)
    )
    receipt_line_entry_q = (
        Q(entry__reference__startswith="RV-")
        | Q(entry__journal_type__iexact="receipt_voucher")
        | Q(entry__source_type=JournalEntry.SourceType.RECEIPT_VOUCHER)
    )

    def _income_credit_sum(entry_ids: list) -> Decimal:
        if not entry_ids:
            return Decimal("0")
        t = (
            JournalLine.objects.using(tenant_db)
            .filter(
                entry_id__in=entry_ids,
                account__type=ChartAccount.Type.INCOME,
                credit__gt=0,
            )
            .aggregate(t=Sum("credit"))
            .get("t")
        )
        return t or Decimal("0")

    total_receipts = Decimal("0")
    donor_income = Decimal("0")
    grant_income = Decimal("0")
    received_month = Decimal("0")
    outstanding_ar = Decimal("0")
    pending_n = 0
    posted_receipt_entries_period = 0
    top_donors_display = "—"
    recent_receipts: list[dict] = []

    try:
        posted_period = JournalEntry.objects.using(tenant_db).filter(
            receipt_entry_q,
            status=JournalEntry.Status.POSTED,
            entry_date__gte=ps,
            entry_date__lte=pe,
        )
        posted_receipt_entries_period = posted_period.count()
        ids_period = list(posted_period.values_list("id", flat=True))
        total_receipts = _income_credit_sum(ids_period)

        donor_entries = posted_period.filter(
            Q(donor_id__isnull=False) | Q(grant__donor_id__isnull=False)
        )
        donor_income = _income_credit_sum(list(donor_entries.values_list("id", flat=True)))

        grant_entries = posted_period.filter(grant_id__isnull=False)
        grant_income = _income_credit_sum(list(grant_entries.values_list("id", flat=True)))

        posted_month = JournalEntry.objects.using(tenant_db).filter(
            receipt_entry_q,
            status=JournalEntry.Status.POSTED,
            entry_date__gte=month_start,
            entry_date__lte=today,
        )
        received_month = _income_credit_sum(list(posted_month.values_list("id", flat=True)))

        outstanding_ar = _get_global_financial_indicators(request).get("outstanding_receivables") or Decimal(
            "0"
        )

        pending_n = (
            JournalEntry.objects.using(tenant_db)
            .filter(
                receipt_entry_q,
                status__in=(
                    JournalEntry.Status.DRAFT,
                    JournalEntry.Status.PENDING_APPROVAL,
                ),
            )
            .count()
        )

        donor_rows = (
            JournalLine.objects.using(tenant_db)
            .filter(
                receipt_line_entry_q,
                entry__status=JournalEntry.Status.POSTED,
                entry__entry_date__gte=ps,
                entry__entry_date__lte=pe,
                account__type=ChartAccount.Type.INCOME,
                credit__gt=0,
                entry__grant__donor_id__isnull=False,
            )
            .values("entry__grant__donor__name")
            .annotate(s=Sum("credit"))
            .order_by("-s")[:3]
        )
        parts = []
        for row in donor_rows:
            name = (row.get("entry__grant__donor__name") or "").strip() or "—"
            parts.append(f"{name} ({_fmt_money(row.get('s') or Decimal('0'))})")
        if parts:
            top_donors_display = "; ".join(parts)

        recent_qs = (
            JournalEntry.objects.using(tenant_db)
            .filter(receipt_entry_q, status=JournalEntry.Status.POSTED)
            .select_related("grant", "grant__donor", "donor")
            .order_by("-entry_date", "-id")[:10]
        )
        for e in recent_qs:
            eid = e.id
            amt = _income_credit_sum([eid])
            donor_guess = ""
            if e.donor_id:
                donor_guess = getattr(e.donor, "name", "") or ""
            elif e.grant_id and getattr(e.grant, "donor", None):
                donor_guess = e.grant.donor.name
            recent_receipts.append({
                "reference": (e.reference or f"RV-{e.id:05d}").strip(),
                "date": e.entry_date,
                "amount_fmt": _fmt_money(amt),
                "source": donor_guess or (e.grant.code if e.grant_id else ""),
            })
    except Exception:
        pass

    incoming_kpis = [
        {"label": _("Total receipts this period"), "value": _fmt_money(total_receipts)},
        {"label": _("Total donor income"), "value": _fmt_money(donor_income)},
        {"label": _("Total grant income"), "value": _fmt_money(grant_income)},
        {"label": _("Outstanding receivables"), "value": _fmt_money(outstanding_ar)},
        {"label": _("Funds received this month"), "value": _fmt_money(received_month)},
        {"label": _("Pending receipts"), "value": f"{pending_n:,}"},
        {"label": _("Top donors"), "value": top_donors_display},
        {"label": _("Receipt entries this period"), "value": f"{posted_receipt_entries_period:,}"},
    ]

    return render(
        request,
        "tenant_portal/recv/incoming_fund_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "receivables",
            "active_item": "recv_center_home",
            "incoming_kpis": incoming_kpis,
            "recent_receipts": recent_receipts,
            "incoming_period_hint": f"{ps.isoformat()} — {pe.isoformat()}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_receipt_vouchers_view(request: HttpRequest) -> HttpResponse:
    """
    Dedicated Receipt Voucher screen.
    """
    from decimal import Decimal, InvalidOperation
    from django.utils import timezone
    from django.utils.dateparse import parse_date
    from django.db.models import Sum

    from tenant_finance.models import BankAccount, ChartAccount, JournalEntry, JournalLine, AuditLog
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    user = request.tenant_user

    if request.method == "POST":
        errors = []

        raw_date = (request.POST.get("voucher_date") or "").strip()
        project_id = request.POST.get("grant_id") or None
        fund_name = (request.POST.get("fund_name") or "").strip()
        reference_no = (request.POST.get("reference_no") or "").strip()
        received_from = (request.POST.get("received_from") or "").strip()
        receipt_method = (request.POST.get("receipt_method") or "").strip()
        deposit_account_id = request.POST.get("deposit_account_id") or ""
        income_account_id = request.POST.get("income_account_id") or ""
        description = (request.POST.get("description") or "").strip()

        voucher_date = None
        if not raw_date:
            errors.append("Voucher date is required.")
        else:
            voucher_date = parse_date(raw_date)
            if not voucher_date:
                errors.append("Voucher date is not a valid calendar date.")
            elif voucher_date > timezone.localdate():
                errors.append("Voucher date cannot be in the future.")

        grant = Grant.objects.using(tenant_db).filter(pk=project_id).select_related("donor", "bank_account").first() if project_id else None
        if not project_id or not grant:
            errors.append("Project must be selected.")
        else:
            # Block ended/inactive projects for both draft and post
            ended_by_date = bool(grant.end_date and grant.end_date < timezone.localdate())
            if grant.status != Grant.Status.ACTIVE or ended_by_date:
                errors.append("Receipts cannot be recorded for an ended or inactive project.")
        if not fund_name:
            errors.append("Fund / Donor is required.")
        if not received_from:
            errors.append("Received from / Donor / Project is required.")
        allowed_methods = {"bank", "cash", "transfer", "cheque", "mobile_money"}
        if receipt_method not in allowed_methods:
            errors.append("Receipt method must be Bank, Cash, Transfer, Cheque, or Mobile money.")

        try:
            amount = Decimal(str(request.POST.get("amount") or "0"))
        except (InvalidOperation, ValueError):
            amount = Decimal("0")
        if amount <= 0:
            errors.append("Amount must be greater than zero.")

        deposit_bank_account = (
            BankAccount.objects.using(tenant_db).select_related("account").filter(pk=deposit_account_id, is_active=True).first()
            if deposit_account_id
            else None
        )
        income_account = (
            ChartAccount.objects.using(tenant_db).filter(pk=income_account_id).first()
            if income_account_id
            else None
        )

        if not deposit_bank_account:
            errors.append("Deposit bank account must be selected.")
        if not income_account:
            errors.append("Income account must be selected.")

        action = (request.POST.get("action") or "").strip()

        # Additional validations only when posting (not just saving draft)
        if action == "post":
            # Required fields before posting
            if not received_from:
                errors.append("Received from / Donor / Project is required before posting.")
            if not receipt_method:
                errors.append("Receipt method is required before posting.")
            if not income_account:
                errors.append("Income account is required before posting.")
            if amount <= 0:
                errors.append("Amount must be greater than zero before posting.")

            # Deposit bank account required for non-cash methods
            if receipt_method in {"bank", "transfer", "cheque", "mobile_money"} and not deposit_bank_account:
                errors.append("Deposit bank account is required for bank, transfer, cheque, or mobile money receipts.")

            # Bank account must be linked to the selected project
            if grant and receipt_method in {"bank", "transfer", "cheque", "mobile_money"}:
                if not getattr(grant, "bank_account_id", None):
                    errors.append("The selected project does not have a linked bank account.")
                elif not deposit_bank_account or deposit_bank_account.id != grant.bank_account_id:
                    errors.append("Selected deposit bank account must match the bank account linked to the selected project.")

            # Accounting period must allow posting (Financial Setup)
            if voucher_date:
                try:
                    _finance_assert_open_period(voucher_date, tenant_db, getattr(request, "tenant_user_id", None))
                except ValueError as e:
                    errors.append(str(e))

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            status = (
                JournalEntry.Status.DRAFT
                if action == "save_draft"
                else JournalEntry.Status.POSTED
            )

            from tenant_finance.services.journal_posting import post_receipt_voucher

            entry = post_receipt_voucher(
                using=tenant_db,
                user=user,
                entry_date=voucher_date,
                memo=description or f"Receipt voucher from {received_from or 'N/A'}",
                grant=grant,
                deposit_chart_account=deposit_bank_account.account,
                income_chart_account=income_account,
                amount=amount,
                description=description,
                status=status,
            )

            try:
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.CREATE,
                    user_id=getattr(user, "id", None),
                    username=getattr(user, "full_name", "") or getattr(user, "email", ""),
                    old_data=None,
                    new_data={"status": entry.status, "reference": entry.reference},
                    summary=f"Created receipt voucher {entry.reference} ({entry.get_status_display()})",
                )
            except Exception:
                pass

            if status == JournalEntry.Status.DRAFT:
                messages.success(request, f"Receipt voucher {entry.reference} saved as draft.")
            else:
                messages.success(request, f"Receipt voucher {entry.reference} created and posted.")

            return redirect(reverse("tenant_portal:recv_receipt_vouchers"))

    from tenant_finance.models import ChartAccount as CA, BankAccount
    from tenant_grants.models import Donor

    # Projects: active only, and not ended by date (if end_date is set)
    grants = (
        Grant.objects.using(tenant_db)
        .filter(status=Grant.Status.ACTIVE)
        .select_related("donor", "bank_account", "bank_account__currency")
        .order_by("code")
    )
    today = timezone.localdate()
    grants = [g for g in grants if not (g.end_date and g.end_date < today)]

    # Deposit bank accounts: only those linked to active projects and active themselves
    linked_bank_account_ids = [g.bank_account_id for g in grants if getattr(g, "bank_account_id", None)]
    bank_accounts = (
        BankAccount.objects.using(tenant_db)
        .select_related("currency")
        .filter(is_active=True, id__in=linked_bank_account_ids)
        .order_by("bank_name", "account_name")
    )

    # Chart of Accounts (all active GL accounts) for income dropdown (per your request)
    income_accounts = CA.objects.using(tenant_db).filter(is_active=True).order_by("code")
    donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")

    from tenant_finance.models import JournalEntry as JE

    recent_qs = (
        JE.objects.using(tenant_db)
        .filter(reference__startswith="RV-")
        .select_related("grant")
        .order_by("-entry_date", "-id")[:50]
    )
    recent = []
    for je in recent_qs:
        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )
        recent.append(
            {
                "id": je.id,
                "reference": je.reference or f"RV-{je.id:05d}",
                "date": je.entry_date,
                "project": je.grant.title if je.grant else "",
                "amount": total,
                "memo": je.memo,
            }
        )

    return render(
        request,
        "tenant_portal/recv/receipt_vouchers.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "grants": grants,
            "donors": donors,
            "bank_accounts": bank_accounts,
            "income_accounts": income_accounts,
            "recent": recent,
            "active_submenu": "receivables",
            "active_item": "recv_receipt_vouchers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_donor_receipts_view(request: HttpRequest) -> HttpResponse:
    """
    Donor receipts tracking page: lists receipt vouchers related to donor/grant income.
    """
    from decimal import Decimal
    from django.db.models import Sum, Q
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    # Base queryset: all receipt vouchers (RV-...)
    qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="RV-")
        .select_related("grant", "grant__donor")
    )

    # Restrict to entries that have at least one income line we treat as grant/donor income.
    income_filter = Q(
        lines__account__type=ChartAccount.Type.INCOME,
        lines__credit__gt=0,
    ) & (
        Q(lines__account__name__icontains="grant")
        | Q(lines__account__name__icontains="donor")
        | Q(lines__account__category__name__icontains="grant")
        | Q(lines__account__category__name__icontains="donor")
    )
    qs = qs.filter(income_filter).distinct()

    # Filters
    voucher_no = (request.GET.get("ref") or "").strip()
    donor_name = (request.GET.get("donor") or "").strip()
    project_title = (request.GET.get("project") or "").strip()
    grant_code = (request.GET.get("grant") or "").strip()
    bank_name = (request.GET.get("bank") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()

    if voucher_no:
        qs = qs.filter(reference__icontains=voucher_no)
    if donor_name:
        qs = qs.filter(grant__donor__name__icontains=donor_name)
    if project_title:
        qs = qs.filter(grant__title__icontains=project_title)
    if grant_code:
        qs = qs.filter(grant__code__icontains=grant_code)

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)

    # Bank / cash account filter operates on the deposit (asset) line
    if bank_name:
        qs = qs.filter(
            lines__account__type=ChartAccount.Type.ASSET,
            lines__account__name__icontains=bank_name,
        )

    qs = qs.order_by("-entry_date", "-id")[:300]

    rows = []
    for je in qs:
        # Deposit (bank/cash) line: debit to asset
        deposit_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(
                entry=je,
                account__type=ChartAccount.Type.ASSET,
                debit__gt=0,
            )
            .first()
        )
        # Income line: credit to income account
        income_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(
                entry=je,
                account__type=ChartAccount.Type.INCOME,
                credit__gt=0,
            )
            .first()
        )

        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        rows.append(
            {
                "id": je.id,
                "reference": je.reference or f"RV-{je.id:05d}",
                "date": je.entry_date,
                "donor": getattr(getattr(je.grant, "donor", None), "name", "") if je.grant else "",
                "project": je.grant.title if je.grant else "",
                "grant_code": je.grant.code if je.grant else "",
                "user_reference": "",  # Reserved for future explicit reference field
                "bank_account": deposit_line.account if deposit_line else None,
                "amount": total,
                "status": je.get_status_display(),
            }
        )

    donors = (
        Grant.objects.using(tenant_db)
        .select_related("donor")
        .values_list("donor__name", flat=True)
        .distinct()
        .order_by("donor__name")
    )
    grants = Grant.objects.using(tenant_db).order_by("code")

    return render(
        request,
        "tenant_portal/recv/donor_receipts.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "donors": [d for d in donors if d],
            "grants": grants,
            "active_submenu": "receivables",
            "active_item": "recv_donor_receipts",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_bank_cash_receipts_view(request: HttpRequest) -> HttpResponse:
    """
    Bank & Cash Receipt Entries register: lists receipt vouchers affecting bank or cash accounts.
    """
    from decimal import Decimal
    from django.db.models import Sum, Q
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="RV-")
        .select_related("grant", "grant__donor", "approved_by")
    )

    # Must affect a bank/cash account (asset with debit)
    qs = qs.filter(lines__account__type=ChartAccount.Type.ASSET, lines__debit__gt=0).distinct()

    # Filters
    bank_acc = (request.GET.get("bank") or "").strip()
    receipt_method = (request.GET.get("method") or "").strip().lower()
    project_title = (request.GET.get("project") or "").strip()
    fund_name = (request.GET.get("fund") or "").strip()
    received_from = (request.GET.get("received_from") or "").strip()
    voucher_no = (request.GET.get("ref") or "").strip()
    ref_no = (request.GET.get("ref_no") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()
    status_code = (request.GET.get("status") or "").strip()

    # Default: show posted receipts
    if status_code:
        qs = qs.filter(status=status_code)
    else:
        qs = qs.filter(status=JournalEntry.Status.POSTED)

    if voucher_no:
        qs = qs.filter(reference__icontains=voucher_no)
    if project_title:
        qs = qs.filter(grant__title__icontains=project_title)
    if fund_name:
        qs = qs.filter(grant__donor__name__icontains=fund_name)
    if received_from:
        qs = qs.filter(memo__icontains=received_from)

    # Receipt method heuristic: based on bank/cash words in asset account name
    if receipt_method == "bank":
        qs = qs.filter(
            lines__account__type=ChartAccount.Type.ASSET,
            lines__account__name__icontains="bank",
        )
    elif receipt_method == "cash":
        cash_q = Q(lines__account__name__icontains="cash") | Q(
            lines__account__name__icontains="petty"
        )
        qs = qs.filter(cash_q, lines__account__type=ChartAccount.Type.ASSET)

    if bank_acc:
        qs = qs.filter(
            lines__account__type=ChartAccount.Type.ASSET,
            lines__account__name__icontains=bank_acc,
        )

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)

    # Amount range – applied after total calculation
    raw_min = (request.GET.get("amount_from") or "").strip()
    raw_max = (request.GET.get("amount_to") or "").strip()
    try:
        min_amount = Decimal(raw_min.replace(",", "")) if raw_min else None
    except Exception:
        min_amount = None
    try:
        max_amount = Decimal(raw_max.replace(",", "")) if raw_max else None
    except Exception:
        max_amount = None

    qs = qs.order_by("-entry_date", "-id")[:500]

    rows = []
    for je in qs:
        # Deposit (bank/cash) line
        deposit_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je, account__type=ChartAccount.Type.ASSET, debit__gt=0)
            .first()
        )
        # Income account line
        income_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je, account__type=ChartAccount.Type.INCOME, credit__gt=0)
            .first()
        )

        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        if min_amount is not None and total < min_amount:
            continue
        if max_amount is not None and total > max_amount:
            continue

        rows.append(
            {
                "id": je.id,
                "entry_date": je.entry_date,
                "voucher": je.reference or f"RV-{je.id:05d}",
                "receipt_date": je.entry_date,
                "receipt_method": "Bank" if deposit_line and "bank" in deposit_line.account.name.lower() else "Cash",
                "bank_account": deposit_line.account if deposit_line else None,
                "received_from": je.memo or "",
                "project": je.grant.title if je.grant else "",
                "fund": getattr(getattr(je.grant, "donor", None), "name", "") if je.grant else "",
                "income_account": income_line.account if income_line else None,
                "reference_no": ref_no or "",
                "amount": total,
                "status": je.get_status_display(),
                "posted_by": getattr(je.approved_by, "full_name", "") or getattr(je.approved_by, "email", ""),
            }
        )

    status_choices = JournalEntry.Status.choices

    return render(
        request,
        "tenant_portal/recv/bank_cash_receipts.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": status_choices,
            "active_submenu": "receivables",
            "active_item": "recv_bank_cash_receipts",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_income_register_view(request: HttpRequest) -> HttpResponse:
    """
    Income Register: list posted income transactions from receipt vouchers.
    """
    from decimal import Decimal
    from django.db.models import Sum, Q
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine

    tenant_db = request.tenant_db

    qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="RV-")
        .select_related("grant", "grant__donor")
    )
    # Must have at least one income line
    qs = qs.filter(lines__account__type=ChartAccount.Type.INCOME, lines__credit__gt=0).distinct()

    # Filters
    voucher_no = (request.GET.get("ref") or "").strip()
    income_account = (request.GET.get("income") or "").strip()
    received_from = (request.GET.get("received_from") or "").strip()
    project_title = (request.GET.get("project") or "").strip()
    donor_name = (request.GET.get("donor") or "").strip()
    bank_acc = (request.GET.get("bank") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()
    status_code = (request.GET.get("status") or "").strip()

    # Default: show posted
    if status_code:
        qs = qs.filter(status=status_code)
    else:
        qs = qs.filter(status=JournalEntry.Status.POSTED)

    if voucher_no:
        qs = qs.filter(reference__icontains=voucher_no)
    if project_title:
        qs = qs.filter(grant__title__icontains=project_title)
    if donor_name:
        qs = qs.filter(grant__donor__name__icontains=donor_name)
    if received_from:
        qs = qs.filter(memo__icontains=received_from)

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)

    # Income account filter
    if income_account:
        income_q = Q(lines__account__name__icontains=income_account) | Q(
            lines__account__code__icontains=income_account
        )
        qs = qs.filter(income_q, lines__account__type=ChartAccount.Type.INCOME)

    # Bank / cash filter via asset line
    if bank_acc:
        qs = qs.filter(
            lines__account__type=ChartAccount.Type.ASSET,
            lines__account__name__icontains=bank_acc,
        )

    # Amount range – after computing totals
    raw_min = (request.GET.get("amount_from") or "").strip()
    raw_max = (request.GET.get("amount_to") or "").strip()
    try:
        min_amount = Decimal(raw_min.replace(",", "")) if raw_min else None
    except Exception:
        min_amount = None
    try:
        max_amount = Decimal(raw_max.replace(",", "")) if raw_max else None
    except Exception:
        max_amount = None

    qs = qs.order_by("-entry_date", "-id")[:500]

    rows = []
    for je in qs:
        # Deposit line (bank/cash)
        deposit_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je, account__type=ChartAccount.Type.ASSET, debit__gt=0)
            .first()
        )
        # Income line
        income_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je, account__type=ChartAccount.Type.INCOME, credit__gt=0)
            .first()
        )

        amount = income_line.credit if income_line else Decimal("0")
        if min_amount is not None and amount < min_amount:
            continue
        if max_amount is not None and amount > max_amount:
            continue

        rows.append(
            {
                "id": je.id,
                "reference": je.reference or f"RV-{je.id:05d}",
                "date": je.entry_date,
                "income_account": income_line.account if income_line else None,
                "received_from": je.memo or "",
                "project": je.grant.title if je.grant else "",
                "fund": getattr(getattr(je.grant, "donor", None), "name", "") if je.grant else "",
                "receipt_method": "Bank"
                if deposit_line and "bank" in deposit_line.account.name.lower()
                else "Cash",
                "bank_account": deposit_line.account if deposit_line else None,
                "reference_no": "",
                "amount": amount,
                "status": je.get_status_display(),
            }
        )

    status_choices = JournalEntry.Status.choices

    return render(
        request,
        "tenant_portal/recv/income_register.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": status_choices,
            "active_submenu": "receivables",
            "active_item": "recv_income_register",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_incoming_fund_register_view(request: HttpRequest) -> HttpResponse:
    """
    Incoming Fund Register: summary per grant/donor of funding vs receipts.
    """
    from decimal import Decimal
    from django.db.models import Sum, Min, Max
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    donor_name = (request.GET.get("donor") or "").strip()
    project_title = (request.GET.get("project") or "").strip()
    grant_code = (request.GET.get("grant") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    grants_qs = Grant.objects.using(tenant_db).select_related("donor")
    if donor_name:
        grants_qs = grants_qs.filter(donor__name__icontains=donor_name)
    if project_title:
        grants_qs = grants_qs.filter(title__icontains=project_title)
    if grant_code:
        grants_qs = grants_qs.filter(code__icontains=grant_code)
    if status_filter:
        grants_qs = grants_qs.filter(status=status_filter)

    grants_qs = grants_qs.order_by("code")

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None

    # Collect receipt vouchers by grant
    receipts_qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="RV-", grant__isnull=False)
        .select_related("grant")
    )
    if from_date:
        receipts_qs = receipts_qs.filter(entry_date__gte=from_date)
    if to_date:
        receipts_qs = receipts_qs.filter(entry_date__lte=to_date)

    # Only income lines
    receipts_qs = receipts_qs.filter(
        lines__account__type=ChartAccount.Type.INCOME, lines__credit__gt=0
    ).distinct()

    stats = {}
    for je in receipts_qs:
        grant_id = je.grant_id
        if not grant_id:
            continue

        amount = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je, account__type=ChartAccount.Type.INCOME, credit__gt=0)
            .aggregate(t=Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        s = stats.setdefault(
            grant_id,
            {"total_received": Decimal("0"), "first": None, "latest": None},
        )
        s["total_received"] += amount
        if s["first"] is None or je.entry_date < s["first"]:
            s["first"] = je.entry_date
        if s["latest"] is None or je.entry_date > s["latest"]:
            s["latest"] = je.entry_date

    rows = []
    for g in grants_qs:
        s = stats.get(g.id, None)
        total_received = s["total_received"] if s else Decimal("0")
        remaining = (g.award_amount or Decimal("0")) - total_received
        rows.append(
            {
                "donor": g.donor.name,
                "project": g.title,
                "grant_code": g.code,
                "total_grant": g.award_amount,
                "total_received": total_received,
                "remaining": remaining,
                "first_date": s["first"] if s else None,
                "latest_date": s["latest"] if s else None,
                "status": g.get_status_display(),
            }
        )

    status_choices = Grant.Status.choices

    return render(
        request,
        "tenant_portal/recv/incoming_fund_register.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": status_choices,
            "active_submenu": "receivables",
            "active_item": "recv_incoming_fund_register",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_grant_income_tracking_view(request: HttpRequest) -> HttpResponse:
    """
    Grant Income Tracker: detailed grant/tranche receipt tracking.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    donor_name = (request.GET.get("donor") or "").strip()
    project_title = (request.GET.get("project") or "").strip()
    grant_code = (request.GET.get("grant") or "").strip()
    voucher_no = (request.GET.get("ref") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()
    status_code = (request.GET.get("status") or "").strip()

    qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="RV-", grant__isnull=False)
        .select_related("grant", "grant__donor")
    )

    if voucher_no:
        qs = qs.filter(reference__icontains=voucher_no)
    if donor_name:
        qs = qs.filter(grant__donor__name__icontains=donor_name)
    if project_title:
        qs = qs.filter(grant__title__icontains=project_title)
    if grant_code:
        qs = qs.filter(grant__code__icontains=grant_code)

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None
    if from_date:
        qs = qs.filter(entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry_date__lte=to_date)

    # Default status: posted
    if status_code:
        qs = qs.filter(status=status_code)
    else:
        qs = qs.filter(status=JournalEntry.Status.POSTED)

    qs = qs.order_by("grant__code", "entry_date", "id")

    # Precompute cumulative receipts per grant for balance remaining
    grant_totals = {}
    rows = []

    for je in qs:
        grant = je.grant
        if not grant:
            continue

        amount = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je, account__type=ChartAccount.Type.INCOME, credit__gt=0)
            .aggregate(t=Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        g_stats = grant_totals.setdefault(
            grant.id,
            {"award": grant.award_amount or Decimal("0"), "received_so_far": Decimal("0"), "count": 0},
        )
        g_stats["count"] += 1
        tranche_no = g_stats["count"]
        g_stats["received_so_far"] += amount
        balance_remaining = g_stats["award"] - g_stats["received_so_far"]

        # Deposit account
        deposit_line = (
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je, account__type=ChartAccount.Type.ASSET, debit__gt=0)
            .first()
        )

        rows.append(
            {
                "donor": grant.donor.name,
                "project": grant.title,
                "grant_code": grant.code,
                "tranche_no": tranche_no,
                "planned_amount": amount,  # placeholder until explicit plan is modeled
                "received_amount": amount,
                "receipt_date": je.entry_date,
                "voucher": je.reference or f"RV-{je.id:05d}",
                "bank_account": deposit_line.account if deposit_line else None,
                "reference_no": "",
                "balance_remaining": balance_remaining,
                "status": je.get_status_display(),
            }
        )

    status_choices = JournalEntry.Status.choices

    return render(
        request,
        "tenant_portal/recv/grant_income_tracking.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": status_choices,
            "active_submenu": "receivables",
            "active_item": "recv_grant_income_tracking",
        },
    )


def _receivable_accounts_q():
    """Q filter for receivable accounts: ASSET with 'receivable' in name or code."""
    from django.db.models import Q
    from tenant_finance.models import ChartAccount
    q = Q(type=ChartAccount.Type.ASSET) & (
        Q(name__icontains="receivable") | Q(code__icontains="receivable")
    )
    return q


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_receivable_ledger_view(request: HttpRequest) -> HttpResponse:
    """
    Receivable Ledger: full ledger of all receivable transactions (journal lines
    on receivable accounts) with running balance. Links to original transaction.
    """
    from decimal import Decimal
    from django.db.models import Q
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    receivable_q = _receivable_accounts_q()
    receivable_ids = list(
        ChartAccount.objects.using(tenant_db).filter(receivable_q).values_list("id", flat=True)
    )
    if not receivable_ids:
        return render(
            request,
            "tenant_portal/recv/receivable_ledger.html",
            {
                "tenant": request.tenant,
                "tenant_user": request.tenant_user,
                "rows": [],
                "status_choices": JournalEntry.Status.choices,
                "active_submenu": "receivables",
                "active_item": "recv_ledger",
            },
        )

    qs = (
        JournalLine.objects.using(tenant_db)
        .filter(account_id__in=receivable_ids)
        .select_related("entry", "account")
        .order_by("entry__entry_date", "entry__id", "id")
    )
    # Optional: only posted
    status_code = (request.GET.get("status") or "").strip()
    if status_code:
        qs = qs.filter(entry__status=status_code)

    ref = (request.GET.get("ref") or "").strip()
    if ref:
        qs = qs.filter(
            Q(entry__reference__icontains=ref) | Q(entry__reference__istartswith=ref)
        )
    debtor = (request.GET.get("debtor") or "").strip()
    if debtor:
        qs = qs.filter(
            Q(entry__memo__icontains=debtor)
            | Q(entry__grant__donor__name__icontains=debtor)
        )
    recv_type = (request.GET.get("type") or "").strip().lower()
    if recv_type == "invoice":
        qs = qs.filter(debit__gt=0)
    elif recv_type == "collection":
        qs = qs.filter(credit__gt=0)
    project = (request.GET.get("project") or "").strip()
    if project:
        qs = qs.filter(entry__grant__title__icontains=project)
    donor = (request.GET.get("donor") or "").strip()
    if donor:
        qs = qs.filter(entry__grant__donor__name__icontains=donor)
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()
    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None
    if from_date:
        qs = qs.filter(entry__entry_date__gte=from_date)
    if to_date:
        qs = qs.filter(entry__entry_date__lte=to_date)
    raw_due_from = (request.GET.get("due_from") or "").strip()
    raw_due_to = (request.GET.get("due_to") or "").strip()
    due_from = parse_date(raw_due_from) if raw_due_from else None
    due_to = parse_date(raw_due_to) if raw_due_to else None
    if due_from:
        qs = qs.filter(entry__entry_date__gte=due_from)
    if due_to:
        qs = qs.filter(entry__entry_date__lte=due_to)
    balance_min = request.GET.get("balance_from")
    balance_max = request.GET.get("balance_to")

    lines = list(qs)
    # Running balance per account
    balance_by_account = {}
    rows = []
    for line in lines:
        acc_id = line.account_id
        prev = balance_by_account.get(acc_id, Decimal("0"))
        bal = prev + (line.debit or Decimal("0")) - (line.credit or Decimal("0"))
        balance_by_account[acc_id] = bal

        entry = line.entry
        debtor_name = ""
        if entry.grant_id:
            g = getattr(entry, "grant", None)
            if g and getattr(g, "donor", None):
                debtor_name = g.donor.name
        if not debtor_name and entry.memo:
            debtor_name = entry.memo
        recv_type_label = "Collection" if (line.credit or 0) > 0 else "Invoice"
        project_title = entry.grant.title if entry.grant_id and getattr(entry, "grant", None) else ""
        fund_donor = ""
        if entry.grant_id and getattr(entry, "grant", None) and getattr(entry.grant, "donor", None):
            fund_donor = entry.grant.donor.name
        due_date = getattr(entry, "due_date", None) or entry.entry_date

        row = {
            "entry_date": entry.entry_date,
            "receivable_no": entry.reference or f"AR-{entry.id:05d}",
            "debtor_name": debtor_name,
            "receivable_type": recv_type_label,
            "description": line.description or entry.memo or "",
            "project": project_title,
            "fund_donor": fund_donor,
            "debit_amount": line.debit or Decimal("0"),
            "credit_amount": line.credit or Decimal("0"),
            "balance": bal,
            "due_date": due_date,
            "status": entry.get_status_display(),
            "reference_no": entry.reference or "",
            "entry_id": entry.id,
        }
        if balance_min is not None and balance_min != "":
            try:
                if bal < Decimal(str(balance_min).replace(",", "")):
                    continue
            except Exception:
                pass
        if balance_max is not None and balance_max != "":
            try:
                if bal > Decimal(str(balance_max).replace(",", "")):
                    continue
            except Exception:
                pass
        rows.append(row)

    transaction_url_name = "tenant_portal:recv_receipt_vouchers"
    return render(
        request,
        "tenant_portal/recv/receivable_ledger.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": JournalEntry.Status.choices,
            "transaction_url_name": transaction_url_name,
            "active_submenu": "receivables",
            "active_item": "recv_ledger",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def recv_outstanding_receivables_view(request: HttpRequest) -> HttpResponse:
    """
    Outstanding Receivables: only receivables with unpaid (positive) balance,
    grouped by account and grant. Links to original transaction.
    """
    from collections import defaultdict
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils import timezone
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    receivable_q = _receivable_accounts_q()
    receivable_accounts = list(
        ChartAccount.objects.using(tenant_db).filter(receivable_q).values_list("id", "code", "name")
    )
    if not receivable_accounts:
        return render(
            request,
            "tenant_portal/recv/outstanding_receivables.html",
            {
                "tenant": request.tenant,
                "tenant_user": request.tenant_user,
                "rows": [],
                "status_choices": JournalEntry.Status.choices,
                "active_submenu": "receivables",
                "active_item": "recv_outstanding",
            },
        )

    recv_ids = [a[0] for a in receivable_accounts]
    qs = (
        JournalLine.objects.using(tenant_db)
        .filter(account_id__in=recv_ids)
        .values("entry_id", "account_id")
        .annotate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit"),
        )
    )
    entry_ids = set()
    for row in qs:
        entry_ids.add(row["entry_id"])
    entries = {
        e.id: e
        for e in JournalEntry.objects.using(tenant_db)
        .filter(id__in=entry_ids)
        .select_related("grant")
    }
    # Group by (account_id, grant_id) for outstanding balance
    group = defaultdict(
        lambda: {
            "original": Decimal("0"),
            "collected": Decimal("0"),
            "first_ref": None,
            "first_entry_id": None,
            "due_date": None,
            "debtor_name": "",
            "description": "",
            "project": "",
            "fund_donor": "",
            "status": "",
        }
    )
    for row in qs:
        entry = entries.get(row["entry_id"])
        if not entry:
            continue
        key = (row["account_id"], entry.grant_id or 0)
        g = group[key]
        d = row["total_debit"] or Decimal("0")
        c = row["total_credit"] or Decimal("0")
        g["original"] += d
        g["collected"] += c
        if g["first_ref"] is None:
            ref = entry.reference or f"AR-{entry.id:05d}"
            g["first_ref"] = ref
            g["first_entry_id"] = entry.id
            g["due_date"] = getattr(entry, "due_date", None) or entry.entry_date
            g["debtor_name"] = ""
            if entry.grant and getattr(entry.grant, "donor", None):
                g["debtor_name"] = entry.grant.donor.name
            if not g["debtor_name"] and entry.memo:
                g["debtor_name"] = entry.memo
            g["description"] = entry.memo or ""
            g["project"] = entry.grant.title if entry.grant else ""
            g["fund_donor"] = entry.grant.donor.name if entry.grant and getattr(entry.grant, "donor", None) else ""
            g["status"] = entry.get_status_display()

    today = timezone.now().date()
    rows = []
    for (acc_id, grant_id), g in group.items():
        balance = g["original"] - g["collected"]
        if balance <= 0:
            continue
        due = g["due_date"] or today
        try:
            days_out = (today - due).days
        except Exception:
            days_out = 0

        # Filters
        debtor_filter = (request.GET.get("debtor") or "").strip()
        if debtor_filter and debtor_filter.lower() not in (g["debtor_name"] or "").lower():
            continue
        type_filter = (request.GET.get("type") or "").strip().lower()
        project_filter = (request.GET.get("project") or "").strip()
        if project_filter and project_filter.lower() not in (g["project"] or "").lower():
            continue
        donor_filter = (request.GET.get("donor") or "").strip()
        if donor_filter and donor_filter.lower() not in (g["fund_donor"] or "").lower():
            continue
        raw_due_from = (request.GET.get("due_from") or "").strip()
        raw_due_to = (request.GET.get("due_to") or "").strip()
        due_from = parse_date(raw_due_from) if raw_due_from else None
        due_to = parse_date(raw_due_to) if raw_due_to else None
        if due_from and due < due_from:
            continue
        if due_to and due > due_to:
            continue
        days_from = request.GET.get("days_from")
        days_to = request.GET.get("days_to")
        if days_from is not None and days_from != "":
            try:
                if days_out < int(days_from):
                    continue
            except ValueError:
                pass
        if days_to is not None and days_to != "":
            try:
                if days_out > int(days_to):
                    continue
            except ValueError:
                pass
        bal_from = request.GET.get("balance_from")
        bal_to = request.GET.get("balance_to")
        if bal_from is not None and bal_from != "":
            try:
                if balance < Decimal(str(bal_from).replace(",", "")):
                    continue
            except Exception:
                pass
        if bal_to is not None and bal_to != "":
            try:
                if balance > Decimal(str(bal_to).replace(",", "")):
                    continue
            except Exception:
                pass
        status_filter = (request.GET.get("status") or "").strip()
        if status_filter and g["status"] != status_filter:
            continue

        rows.append({
            "receivable_no": g["first_ref"],
            "entry_id": g["first_entry_id"],
            "debtor_name": g["debtor_name"],
            "receivable_type": "Receivable",
            "description": g["description"],
            "project": g["project"],
            "fund_donor": g["fund_donor"],
            "original_amount": g["original"],
            "amount_collected": g["collected"],
            "outstanding_balance": balance,
            "due_date": due,
            "days_outstanding": days_out,
            "status": g["status"],
        })

    rows.sort(key=lambda r: (r["due_date"], r["receivable_no"]))
    transaction_url_name = "tenant_portal:recv_receipt_vouchers"
    return render(
        request,
        "tenant_portal/recv/outstanding_receivables.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": JournalEntry.Status.choices,
            "transaction_url_name": transaction_url_name,
            "active_submenu": "receivables",
            "active_item": "recv_outstanding",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_financial_alerts_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from django.utils import timezone
    from datetime import timedelta
    from tenant_finance.models import ChartAccount, JournalLine
    from tenant_grants.models import BudgetLine, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    today = timezone.now().date()
    ninety_days = today + timedelta(days=90)

    alerts = []

    # Budget overruns
    budgets_by_grant = {
        r["grant_id"]: r["total"] or Decimal("0")
        for r in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
    }
    spend_by_grant = {
        r["entry__grant_id"]: r["spent"] or Decimal("0")
        for r in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(spent=Sum("debit"))
    }
    overrun_count = 0
    low_cash_count = 0
    expiring_count = 0
    display_alerts = []

    for g in Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE):
        budget = budgets_by_grant.get(g.id, Decimal("0"))
        spent = spend_by_grant.get(g.id, Decimal("0"))
        ceiling = budget if budget > 0 else Decimal(str(g.award_amount or 0))
        if ceiling > 0 and spent > ceiling:
            overrun_count += 1
            display_alerts.append({
                "type": "overrun",
                "type_label": "Budget overrun",
                "detail": f"{g.code}: spent {spent} exceeds ceiling {ceiling}",
                "amount": spent - ceiling,
            })

    cash_total = (
        JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.ASSET)
        .aggregate(t=Sum("debit") - Sum("credit"))
    ).get("t") or Decimal("0")
    if cash_total < Decimal("0"):
        low_cash_count += 1
        display_alerts.append({"type": "low_cash", "type_label": "Negative cash", "detail": "Total cash/bank balance is negative", "amount": cash_total})
    elif cash_total < Decimal("1000"):
        low_cash_count += 1
        display_alerts.append({"type": "low_cash", "type_label": "Low cash", "detail": "Total cash/bank balance below 1,000", "amount": cash_total})

    expiring = Grant.objects.using(tenant_db).filter(
        status=Grant.Status.ACTIVE, end_date__isnull=False, end_date__gte=today, end_date__lte=ninety_days
    ).order_by("end_date")
    for g in expiring:
        expiring_count += 1
        display_alerts.append({
            "type": "expiring",
            "type_label": "Expiring grant",
            "detail": f"{g.code} ends {g.end_date}",
            "amount": Decimal("0"),
        })

    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="financial_alerts.csv"'
        w = csv.writer(response)
        w.writerow(["Type", "Detail", "Amount/Value"])
        for a in display_alerts:
            w.writerow([a.get("type_label"), a.get("detail"), a.get("amount")])
        return response
    return render(
        request,
        "tenant_portal/finance/financial_alerts.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "dashboard",
            "active_item": "dashboard_alerts",
            "filters": f,
            "alerts": display_alerts,
            "overrun_count": overrun_count,
            "low_cash_count": low_cash_count,
            "expiring_count": expiring_count,
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_accounts_view(request: HttpRequest) -> HttpResponse:
    import csv
    from io import TextIOWrapper
    from collections import defaultdict

    from tenant_finance.models import ChartAccount, AccountCategory

    tenant_db = request.tenant_db
    # Export current Chart of Accounts as CSV
    if request.method == "GET" and request.GET.get("export") == "1":
        accounts = (
            ChartAccount.objects.using(tenant_db)
            .select_related("parent", "category")
            .order_by("code")
        )
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="chart_of_accounts.csv"'
        writer = csv.writer(response)
        writer.writerow(
            ["code", "name", "type", "parent_code", "category_code", "is_active"]
        )
        for a in accounts:
            writer.writerow(
                [
                    a.code,
                    a.name,
                    a.type,
                    a.parent.code if a.parent else "",
                    a.category.code if a.category else "",
                    "1" if a.is_active else "0",
                ]
            )
        return response

    if request.method == "POST":
        action = request.POST.get("action", "create")
        if action == "delete":
            account_id = request.POST.get("account_id")
            if account_id:
                acc = ChartAccount.objects.using(tenant_db).filter(pk=account_id).first()
                if acc:
                    if acc.is_used(using=tenant_db):
                        messages.error(
                            request,
                            "This account is used in transactions or setup and cannot be deleted. "
                            "Use Disable to deactivate it instead.",
                        )
                    else:
                        acc.delete(using=tenant_db)
                        messages.success(request, f"Account {acc.code} has been deleted.")
                else:
                    messages.error(request, "Account not found.")
            return redirect(reverse("tenant_portal:finance_accounts"))
        if action == "toggle_active":
            account_id = request.POST.get("account_id")
            if account_id:
                acc = ChartAccount.objects.using(tenant_db).filter(pk=account_id).first()
                if acc:
                    acc.is_active = not acc.is_active
                    acc.updated_by = request.tenant_user
                    acc.save(update_fields=["is_active", "updated_by"])
                    messages.success(request, f"Account {acc.code} is now {'active' if acc.is_active else 'inactive'}.")
            return redirect(reverse("tenant_portal:finance_accounts"))
        if action == "import":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Please choose a CSV file to import.")
                return redirect(reverse("tenant_portal:finance_accounts"))
            # Only CSV is supported for now. If an Excel file is uploaded, ask the user to save as CSV first.
            filename = (upload.name or "").lower()
            if not filename.endswith(".csv"):
                messages.error(
                    request,
                    "Only CSV files are supported. Please export or save your Chart of Accounts as CSV before importing.",
                )
                return redirect(reverse("tenant_portal:finance_accounts"))
            # Expect a CSV with header: code,name,type,parent_code,category_code,is_active
            try:
                wrapper = TextIOWrapper(upload, encoding="utf-8-sig", errors="ignore")
                reader = csv.DictReader(wrapper)
            except Exception:
                messages.error(
                    request,
                    "Unable to read the uploaded CSV file. Please check the encoding and try again.",
                )
                return redirect(reverse("tenant_portal:finance_accounts"))

            created_count = 0
            for row in reader:
                code = (row.get("code") or "").strip()
                name = (row.get("name") or "").strip()
                type_ = (row.get("type") or "").strip()
                if not code or not name or not type_:
                    continue
                if ChartAccount.objects.using(tenant_db).filter(code=code).exists():
                    continue
                parent = None
                parent_code = (row.get("parent_code") or "").strip()
                if parent_code:
                    parent = (
                        ChartAccount.objects.using(tenant_db)
                        .filter(code=parent_code)
                        .first()
                    )
                category = None
                category_code = (row.get("category_code") or "").strip()
                if category_code:
                    category = (
                        AccountCategory.objects.using(tenant_db)
                        .filter(code=category_code)
                        .first()
                    )
                is_active_raw = (row.get("is_active") or "").strip()
                is_active = is_active_raw not in ("0", "false", "False", "")
                st = ""
                if type_ in ("asset", "liability", "equity"):
                    st = ChartAccount.StatementType.BALANCE_SHEET
                elif type_ in ("income", "expense"):
                    st = ChartAccount.StatementType.INCOME_EXPENDITURE
                ChartAccount.objects.using(tenant_db).create(
                    code=code,
                    name=name,
                    type=type_,
                    statement_type=st or "",
                    parent=parent,
                    category=category,
                    is_active=is_active,
                    created_by=request.tenant_user,
                )
                created_count += 1

            messages.success(
                request,
                f"Imported {created_count} account(s) from the file.",
            )
            return redirect(reverse("tenant_portal:finance_accounts"))
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        type_ = (request.POST.get("type") or "").strip()
        parent_id = request.POST.get("parent_id") or None
        category_id = request.POST.get("category_id") or None
        status = (request.POST.get("status") or "active").strip().lower()
        if not code or not name or not type_:
            messages.error(request, "Please provide code, name, and type.")
        else:
            parent = ChartAccount.objects.using(tenant_db).filter(pk=parent_id).first() if parent_id else None
            category = AccountCategory.objects.using(tenant_db).filter(pk=category_id).first() if category_id else None
            is_active = status != "inactive"
            if ChartAccount.objects.using(tenant_db).filter(code=code).exists():
                messages.error(request, "An account with this code already exists.")
            else:
                acc = ChartAccount(
                    code=code,
                    name=name,
                    type=type_,
                    is_active=is_active,
                    parent=parent,
                    category=category,
                    created_by=request.tenant_user,
                )
                acc._state.db = tenant_db
                try:
                    acc.full_clean()
                except ValidationError as e:
                    for _field, errs in e.message_dict.items():
                        for msg in errs:
                            messages.error(request, msg)
                    return redirect(reverse("tenant_portal:finance_accounts"))
                acc.save(using=tenant_db)
                messages.success(request, "Account created.")
                return redirect(reverse("tenant_portal:finance_accounts"))

    accounts_qs = (
        ChartAccount.objects.using(tenant_db)
        .select_related("parent", "category")
        .order_by("code")
    )
    accounts = list(accounts_qs)

    # Build hierarchical tree structure (parent -> children) for UI
    children_map: dict[int | None, list[ChartAccount]] = defaultdict(list)
    for acc in accounts:
        children_map[acc.parent_id].append(acc)
    for key in children_map:
        children_map[key].sort(key=lambda a: a.code or "")

    tree_accounts: list[dict] = []

    def _walk(parent_id: int | None, depth: int) -> None:
        for acc in children_map.get(parent_id, []):
            has_children = acc.id in children_map
            tree_accounts.append(
                {
                    "account": acc,
                    "depth": depth,
                    "has_children": has_children,
                    "is_posting": not has_children,
                }
            )
            _walk(acc.id, depth + 1)

    _walk(None, 0)

    categories = AccountCategory.objects.using(tenant_db).order_by("display_order", "code")
    return render(
        request,
        "tenant_portal/finance/accounts.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "accounts": accounts,
            "tree_accounts": tree_accounts,
            "categories": categories,
            "types": ChartAccount.Type,
            "active_submenu": "core",
            "active_item": "core_chart",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_account_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit a chart account. Used accounts cannot change code, type, or statement type."""
    from tenant_finance.models import ChartAccount, AccountCategory

    tenant_db = request.tenant_db
    acc = get_object_or_404(ChartAccount.objects.using(tenant_db).select_related("parent", "category"), pk=pk)
    used = acc.is_used(using=tenant_db)
    categories = AccountCategory.objects.using(tenant_db).order_by("display_order", "code")
    accounts = list(
        ChartAccount.objects.using(tenant_db).exclude(pk=pk).select_related("parent", "category").order_by("code")
    )

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        type_ = (request.POST.get("type") or "").strip()
        code = (request.POST.get("code") or "").strip()
        statement_type = (request.POST.get("statement_type") or "").strip()
        parent_id = request.POST.get("parent_id") or None
        category_id = request.POST.get("category_id") or None
        status = (request.POST.get("status") or "active").strip().lower()

        if used:
            if code != acc.code or type_ != acc.type or statement_type != (acc.statement_type or ""):
                messages.error(
                    request,
                    "This account is used in transactions. Code, account type, and statement type cannot be changed.",
                )
                return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))

        if not name:
            messages.error(request, "Account name is required.")
            return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))
        if not code and not used:
            messages.error(request, "Account code is required.")
            return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))
        if not type_:
            messages.error(request, "Account type must be selected.")
            return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))

        if not used and code != acc.code and ChartAccount.objects.using(tenant_db).filter(code=code).exists():
            messages.error(request, "An account with this code already exists.")
            return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))

        parent = ChartAccount.objects.using(tenant_db).filter(pk=parent_id).first() if parent_id else None
        category = AccountCategory.objects.using(tenant_db).filter(pk=category_id).first() if category_id else None
        is_active = status != "inactive"

        if not used:
            acc.code = code
        acc.name = name
        acc.type = type_
        if not used:
            acc.statement_type = statement_type or ""
        acc.parent = parent
        acc.category = category
        acc.is_active = is_active
        acc.updated_by = request.tenant_user
        acc._state.db = tenant_db
        try:
            acc.full_clean()
        except ValidationError as e:
            for _field, errs in e.message_dict.items():
                for msg in errs:
                    messages.error(request, msg)
            return redirect(reverse("tenant_portal:finance_account_edit", args=[pk]))
        acc.save(using=tenant_db)
        messages.success(request, "Account updated.")
        return redirect(reverse("tenant_portal:finance_accounts"))

    statement_types = ChartAccount.StatementType.choices
    return render(
        request,
        "tenant_portal/finance/account_edit.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "account": acc,
            "used": used,
            "accounts": accounts,
            "categories": categories,
            "types": ChartAccount.Type,
            "statement_types": statement_types,
            "active_submenu": "core",
            "active_item": "core_chart",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_journals_view(request: HttpRequest) -> HttpResponse:
    """
    Core Accounting → Journal management → General journal (GL register / audit trail).
    """

    from decimal import Decimal

    from django.db.models import DecimalField, Q, Sum, Value
    from django.db.models.functions import Coalesce

    from tenant_finance.db_compat import journalentry_has_0034_schema
    from tenant_finance.models import ChartAccount, FiscalPeriod, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    user = request.tenant_user

    f = _parse_finance_filters(request)
    _apply_journal_quick_date_range(request, tenant_db, f)

    status = (request.GET.get("status") or "").strip()
    search = (request.GET.get("q") or "").strip()
    source_type_f = (request.GET.get("source_type") or "").strip()
    accounting_period_id = (
        request.GET.get("accounting_period_id") or request.GET.get("fiscal_period_id") or ""
    ).strip()
    quick_filter = (request.GET.get("quick") or "").strip()

    je_schema_0034 = journalentry_has_0034_schema(tenant_db)

    accounting_periods_all = list(
        FiscalPeriod.objects.using(tenant_db).select_related("fiscal_year").order_by("-start_date")
    )

    entries_qs = (
        JournalEntry.objects.using(tenant_db)
        .select_related("grant", "created_by", "currency")
        .filter(entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])
        .annotate(
            debit_total=Coalesce(
                Sum("lines__debit"),
                Value(0),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
            credit_total=Coalesce(
                Sum("lines__credit"),
                Value(0),
                output_field=DecimalField(max_digits=14, decimal_places=2),
            ),
        )
        .order_by("-entry_date", "-id")
    )
    if je_schema_0034:
        entries_qs = entries_qs.select_related("posted_by")
    else:
        entries_qs = entries_qs.defer(
            "posted_by",
            "is_system_generated",
            "source_document_no",
            "source_id",
            "source_type",
        )

    if accounting_period_id:
        fp = FiscalPeriod.objects.using(tenant_db).filter(pk=accounting_period_id).first()
        if fp:
            entries_qs = entries_qs.filter(
                entry_date__gte=fp.start_date,
                entry_date__lte=fp.end_date,
            )

    can_all_grants = user_has_permission(user, "finance:scope.all_grants", using=tenant_db)
    if not can_all_grants:
        try:
            allowed_grant_ids = list(
                user.assigned_grants.using(tenant_db).values_list("id", flat=True)
            )
        except Exception:
            allowed_grant_ids = []
        entries_qs = entries_qs.filter(grant_id__in=allowed_grant_ids) if allowed_grant_ids else entries_qs.none()

    if f["grant_id"]:
        if not can_all_grants:
            try:
                if not user.assigned_grants.using(tenant_db).filter(id=f["grant_id"]).exists():
                    entries_qs = entries_qs.none()
            except Exception:
                entries_qs = entries_qs.none()
        entries_qs = entries_qs.filter(grant_id=f["grant_id"])
    if f["donor_id"]:
        entries_qs = entries_qs.filter(grant__donor_id=f["donor_id"])
    if status:
        entries_qs = entries_qs.filter(status=status)
    if (
        je_schema_0034
        and source_type_f
        and source_type_f in dict(JournalEntry.SourceType.choices)
    ):
        entries_qs = entries_qs.filter(source_type=source_type_f)
    if search:
        q = (
            Q(reference__icontains=search)
            | Q(memo__icontains=search)
            | Q(lines__account__code__icontains=search)
            | Q(lines__account__name__icontains=search)
        )
        if je_schema_0034:
            q |= Q(source_document_no__icontains=search)
        entries_qs = entries_qs.filter(q).distinct()

    from django.db.models import Sum as SumAgg

    kpi_base = entries_qs.distinct()
    kpi_total_journals = kpi_base.count()
    kpi_posted = kpi_base.filter(status=JournalEntry.Status.POSTED).count()
    line_totals = (
        JournalLine.objects.using(tenant_db)
        .filter(entry_id__in=kpi_base.values("id"))
        .aggregate(td=SumAgg("debit"), tc=SumAgg("credit"))
    )
    kpi_total_debit = line_totals.get("td") or Decimal("0")
    kpi_total_credit = line_totals.get("tc") or Decimal("0")

    entries_slice = entries_qs[:200]

    entries = []
    for entry in entries_slice:
        journal_no = entry.reference or f"JV-{entry.entry_date.year}-{entry.id:04d}"
        jt = (entry.journal_type or "").strip()
        jt_label = jt.replace("_", " ").title() if jt else "—"
        if je_schema_0034:
            src_doc = (
                (entry.source_document_no or "").strip()
                or (entry.reference or "").strip()
                or "—"
            )
            st_label = entry.get_source_type_display() if entry.source_type else "—"
            pb = entry.posted_by
            pb_name = "—"
            if pb:
                pb_name = (
                    (getattr(pb, "get_full_name", lambda: "")() or "").strip()
                    or getattr(pb, "email", "")
                    or "—"
                )
            is_sys = bool(entry.is_system_generated)
            posted_at = entry.posted_at
            source_url = _journal_source_document_url(entry, True)
        else:
            src_doc = (entry.reference or "").strip() or "—"
            st_label = "—"
            pb_name = "—"
            is_sys = False
            posted_at = getattr(entry, "posted_at", None)
            source_url = None

        posting_date = (posted_at.date() if posted_at else None) or entry.entry_date
        acct_period_lbl = _accounting_period_label_for_date(accounting_periods_all, entry.entry_date)
        curr = getattr(entry, "currency", None)
        currency_code = (curr.code if curr else None) or "—"

        can_reverse_row = (
            entry.status == JournalEntry.Status.POSTED
            and not is_sys
            and user_has_permission(user, "finance:journals.reverse", using=tenant_db)
        )

        entries.append(
            {
                "id": entry.id,
                "journal_no": journal_no,
                "date": entry.entry_date,
                "posting_date": posting_date,
                "reference": entry.reference,
                "memo": entry.memo,
                "grant": entry.grant,
                "status": entry.status,
                "debit_total": entry.debit_total or Decimal("0"),
                "credit_total": entry.credit_total or Decimal("0"),
                "source_type": entry.source_type if je_schema_0034 else None,
                "source_type_label": st_label,
                "source_document_no": src_doc,
                "source_document_url": source_url,
                "journal_type": jt,
                "journal_type_label": jt_label,
                "is_system_generated": is_sys,
                "posted_by_name": pb_name,
                "posted_at": posted_at,
                "accounting_period_label": acct_period_lbl,
                "currency_code": currency_code,
                "audit_trail_url": _journal_audit_trail_url(entry.id),
                "can_reverse": can_reverse_row,
            }
        )

    from tenant_grants.models import Donor, Grant

    grants = Grant.objects.using(tenant_db).filter(status="active").order_by("code")
    if not can_all_grants:
        try:
            grants = grants.filter(id__in=user.assigned_grants.using(tenant_db).values_list("id", flat=True))
        except Exception:
            grants = grants.none()
    donors = Donor.objects.using(tenant_db).order_by("name")
    accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")
    accounting_periods = FiscalPeriod.objects.using(tenant_db).order_by("-start_date")[:48]

    can_reverse_journal = user_has_permission(user, "finance:journals.reverse", using=tenant_db)
    can_create_manual = user_has_permission(user, "finance.add_journalentry", using=tenant_db)

    from tenant_grants.models import ProjectBudgetLine, WorkplanActivity

    manual_journal_budget_lines = list(
        ProjectBudgetLine.objects.using(tenant_db)
        .select_related("project_budget__project")
        .order_by("project_budget__project__code", "id")[:500]
    )
    manual_journal_workplan_activities = list(
        WorkplanActivity.objects.using(tenant_db)
        .select_related("grant")
        .order_by("-id")[:500]
    )

    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="journal_entries.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "Journal No",
                "Journal type",
                "Source type",
                "Source document",
                "Entry date",
                "Posting date",
                "Accounting period",
                "Currency",
                "System JE",
                "Description",
                "Project / Grant",
                "Posted by",
                "Posted at",
                "Status",
                "Total Debit",
                "Total Credit",
            ]
        )
        for e in entries:
            writer.writerow(
                [
                    e["journal_no"],
                    e["journal_type_label"],
                    e["source_type_label"],
                    e["source_document_no"],
                    e["date"],
                    e["posting_date"],
                    e["accounting_period_label"],
                    e["currency_code"],
                    "Y" if e["is_system_generated"] else "N",
                    e["memo"] or "",
                    getattr(e["grant"], "code", ""),
                    e["posted_by_name"],
                    e["posted_at"] or "",
                    e["status"],
                    e["debit_total"],
                    e["credit_total"],
                ]
            )
        return response

    return render(
        request,
        "tenant_portal/finance/journals.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "entries": entries,
            "filters": f,
            "grants": grants,
            "donors": donors,
            "accounting_periods": accounting_periods,
            "filter_source_type": source_type_f,
            "filter_accounting_period_id": accounting_period_id,
            "quick_filter": quick_filter,
            "journal_preset_urls": _journal_list_preset_urls(request),
            "kpi_total_journals": kpi_total_journals,
            "kpi_posted_journals": kpi_posted,
            "kpi_total_debit": kpi_total_debit,
            "kpi_total_credit": kpi_total_credit,
            "export_csv_url": _finance_export_csv_url(request),
            "journal_statuses": JournalEntry.Status,
            "journal_source_types": JournalEntry.SourceType,
            "accounts": accounts,
            "can_reverse_journal": can_reverse_journal,
            "can_create_manual": can_create_manual,
            "open_manual_from_query": request.GET.get("open_manual") == "1",
            "journal_schema_0034": je_schema_0034,
            "manual_journal_budget_lines": manual_journal_budget_lines,
            "manual_journal_workplan_activities": manual_journal_workplan_activities,
            "active_submenu": "core",
            "active_item": "core_journals",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_journal_detail_view(request: HttpRequest, entry_id: int) -> HttpResponse:
    """Read-only journal register line detail (GL audit)."""
    from decimal import Decimal

    from django.shortcuts import get_object_or_404

    from tenant_finance.db_compat import journalentry_has_0034_schema, journalentry_has_0040_adjusting_schema
    from tenant_finance.models import FiscalPeriod, JournalEntry, JournalLine, OrganizationSettings

    tenant_db = request.tenant_db
    je_schema_0034 = journalentry_has_0034_schema(tenant_db)
    je_schema_0040 = journalentry_has_0040_adjusting_schema(tenant_db)
    je_qs = JournalEntry.objects.using(tenant_db).select_related(
        "grant", "donor", "created_by", "currency", "cost_center"
    )
    if je_schema_0034:
        je_qs = je_qs.select_related("posted_by")
    else:
        je_qs = je_qs.defer(
            "posted_by",
            "is_system_generated",
            "source_document_no",
            "source_id",
            "source_type",
        )
    entry = get_object_or_404(je_qs.prefetch_related("attachments"), pk=entry_id)

    if entry.grant_id and not user_has_permission(
        request.tenant_user, "finance:scope.all_grants", using=tenant_db
    ):
        try:
            if not request.tenant_user.assigned_grants.using(tenant_db).filter(
                id=entry.grant_id
            ).exists():
                return render(
                    request,
                    "tenant_portal/forbidden.html",
                    {
                        "tenant": request.tenant,
                        "tenant_user": request.tenant_user,
                        "reason": "You do not have access to this journal.",
                    },
                    status=403,
                )
        except Exception:
            return render(
                request,
                "tenant_portal/forbidden.html",
                {
                    "tenant": request.tenant,
                    "tenant_user": request.tenant_user,
                    "reason": "You do not have access to this journal.",
                },
                status=403,
            )

    lines = list(
        JournalLine.objects.using(tenant_db)
        .select_related(
            "account",
            "grant",
            "project_budget_line",
            "project_budget_line__project_budget__project",
            "workplan_activity",
            "workplan_activity__grant",
        )
        .filter(entry=entry)
        .order_by("id")
    )
    debit_total = sum((l.debit or Decimal("0")) for l in lines) if lines else Decimal("0")
    credit_total = sum((l.credit or Decimal("0")) for l in lines) if lines else Decimal("0")
    org_settings = OrganizationSettings.objects.using(tenant_db).first()
    can_reverse_journal = user_has_permission(
        request.tenant_user, "finance:journals.reverse", using=tenant_db
    )
    is_system_generated = bool(getattr(entry, "is_system_generated", False))
    can_reverse_effective = (
        can_reverse_journal
        and entry.status == JournalEntry.Status.POSTED
        and not is_system_generated
    )
    accounting_periods_all = list(
        FiscalPeriod.objects.using(tenant_db).select_related("fiscal_year").order_by("-start_date")
    )
    gl_for_period = (
        entry.posting_date
        if je_schema_0040 and getattr(entry, "posting_date", None)
        else entry.entry_date
    )
    accounting_period_label = _accounting_period_label_for_date(accounting_periods_all, gl_for_period)
    source_document_url = _journal_source_document_url(entry, je_schema_0034)
    posting_date = (
        entry.posting_date
        if je_schema_0040 and getattr(entry, "posting_date", None)
        else ((entry.posted_at.date() if entry.posted_at else None) or entry.entry_date)
    )
    currency_code = getattr(getattr(entry, "currency", None), "code", None) or "—"

    return render(
        request,
        "tenant_portal/finance/journal_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "entry": entry,
            "lines": lines,
            "debit_total": debit_total,
            "credit_total": credit_total,
            "org_settings": org_settings,
            "can_reverse_journal": can_reverse_effective,
            "journal_schema_0034": je_schema_0034,
            "journal_schema_0040": je_schema_0040,
            "is_system_generated": is_system_generated,
            "source_document_url": source_document_url,
            "accounting_period_label": accounting_period_label,
            "posting_date": posting_date,
            "currency_code": currency_code,
            "audit_trail_url": _journal_audit_trail_url(entry.id),
            "active_submenu": "core",
            "active_item": "core_journals",
        },
    )


@require_http_methods(["POST"])
@tenant_view(require_module="finance_grants", require_perm="finance.add_journalentry")
def finance_journal_create_view(request: HttpRequest) -> HttpResponse:
    """
    Create a manual journal entry from the list-page modal.
    """
    from decimal import Decimal

    from tenant_finance.models import AuditLog, JournalEntry, JournalLine

    tenant_db = request.tenant_db

    entry_date_str = request.POST.get("entry_date") or ""
    memo = (request.POST.get("memo") or "").strip()
    grant_id = request.POST.get("grant_id") or ""
    journal_type = (request.POST.get("journal_type") or "").strip() or "adjustment"

    from django.utils.dateparse import parse_date

    entry_date = parse_date(entry_date_str)
    if not entry_date:
        messages.error(request, "Please provide a valid journal date.")
        return redirect(reverse("tenant_portal:finance_journals"))

    try:
        _finance_assert_open_period(entry_date, tenant_db, request.tenant_user_id)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect(reverse("tenant_portal:finance_journals"))

    accounts = request.POST.getlist("line_account")
    debits = request.POST.getlist("line_debit")
    credits = request.POST.getlist("line_credit")
    descriptions = request.POST.getlist("line_description")
    line_pbl = request.POST.getlist("line_project_budget_line")
    line_wa = request.POST.getlist("line_workplan_activity")

    lines = []
    for idx in range(len(accounts)):
        if not accounts[idx]:
            continue
        pr = (line_pbl[idx] or "").strip() if idx < len(line_pbl) else ""
        wr = (line_wa[idx] or "").strip() if idx < len(line_wa) else ""
        lines.append(
            {
                "account_id": accounts[idx],
                "description": descriptions[idx] if idx < len(descriptions) else "",
                "debit": debits[idx] if idx < len(debits) else "0",
                "credit": credits[idx] if idx < len(credits) else "0",
                "project_budget_line_id": int(pr) if pr.isdigit() else None,
                "workplan_activity_id": int(wr) if wr.isdigit() else None,
            }
        )

    header = {
        "entry_date": entry_date,
        "memo": memo,
        "grant_id": grant_id,
        "journal_type": journal_type,
        "source": "manual",
    }

    try:
        _finance_validate_journal_payload(header, lines, tenant_db)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect(reverse("tenant_portal:finance_journals"))

    from django.core.exceptions import ValidationError as DjangoValidationError
    from django.db import transaction
    from tenant_grants.models import Grant

    with transaction.atomic(using=tenant_db):
        grant = None
        if grant_id:
            grant = Grant.objects.using(tenant_db).select_related("project").filter(pk=grant_id).first()
            if grant:
                tmp = JournalEntry(entry_date=entry_date, grant=grant)
                try:
                    tmp.full_clean()
                except DjangoValidationError as e:
                    err = " ".join(f"{k}: {v}" for k, v in (e.message_dict or {}).items()) or str(e)
                    messages.error(request, err)
                    return redirect(reverse("tenant_portal:finance_journals"))

        entry = JournalEntry.objects.using(tenant_db).create(
            entry_date=entry_date,
            memo=memo,
            grant=grant,
            status=JournalEntry.Status.DRAFT,
            created_by=request.tenant_user,
            journal_type=journal_type,
            source="manual",
            source_type=JournalEntry.SourceType.MANUAL,
            is_system_generated=False,
        )

        for line in lines:
            JournalLine.objects.using(tenant_db).create(
                entry=entry,
                account_id=line["account_id"],
                description=line["description"],
                debit=Decimal(line["debit"] or "0"),
                credit=Decimal(line["credit"] or "0"),
                grant_id=grant.pk if grant else None,
                project_budget_line_id=line.get("project_budget_line_id"),
                workplan_activity_id=line.get("workplan_activity_id"),
            )

        # Audit log
        AuditLog.objects.using(tenant_db).create(
            model_name="journalentry",
            object_id=entry.id,
            action=AuditLog.Action.CREATE,
            user_id=request.tenant_user.id if request.tenant_user else None,
            username=request.tenant_user.get_full_name()
            if getattr(request, "tenant_user", None)
            else "",
            summary=f"Created manual journal entry DRAFT on {entry.entry_date}",
        )

    messages.success(request, "Journal entry created in Draft status.")
    return redirect(reverse("tenant_portal:finance_journals"))


@require_http_methods(["POST"])
@tenant_view(require_module="finance_grants", require_perm="finance:journals.view")
def finance_journal_action_view(request: HttpRequest, entry_id: int) -> HttpResponse:
    """
    Handle status workflow actions: submit, approve, post, reverse.
    Per-action permissions are enforced below (create / approve / post / reverse).
    """
    from decimal import Decimal

    from rbac.models import user_has_permission
    from tenant_finance.models import (
        AuditLog,
        FiscalYear,
        JournalEntry,
        JournalLine,
        TransactionReversalRule,
    )
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    action = request.POST.get("action") or ""

    def _actor_name() -> str:
        u = getattr(request, "tenant_user", None)
        return (getattr(u, "full_name", "") or "").strip() or getattr(u, "email", "") or ""

    def _has_perm(code: str) -> bool:
        cached = getattr(request, "rbac_permission_codes", None)
        if isinstance(cached, set):
            return ("*" in cached) or (code in cached)
        return user_has_permission(request.tenant_user, code, using=tenant_db)

    required_action_perm = {
        "submit": "finance:journals.create",
        "approve": "finance:journals.approve",
        "reject": "finance:journals.approve",
        "post": "finance:journals.post",
        "reverse": "finance:journals.reverse",
    }.get(action)
    if required_action_perm and not _has_perm(required_action_perm):
        messages.error(request, "You do not have permission to perform this action.")
        return _finance_journal_action_redirect(request)

    entry = (
        JournalEntry.objects.using(tenant_db)
        .select_related("grant", "donor")
        .prefetch_related(
            "lines__account",
            "lines__grant",
            "lines__project_budget_line",
            "lines__workplan_activity",
        )
        .filter(pk=entry_id)
        .first()
    )
    if not entry:
        messages.error(request, "Journal entry not found.")
        return _finance_journal_action_redirect(request)

    # Data-level access: restrict grant-linked journals unless user has all-grants scope
    if not _has_perm("finance:scope.all_grants"):
        grant_ids = set()
        if entry.grant_id:
            grant_ids.add(entry.grant_id)
        for line in entry.lines.all():
            if getattr(line, "grant_id", None):
                grant_ids.add(line.grant_id)
        for gid in grant_ids:
            try:
                if not request.tenant_user.assigned_grants.using(tenant_db).filter(id=gid).exists():
                    messages.error(request, "You do not have access to this grant/project.")
                    return _finance_journal_action_redirect(request)
            except Exception:
                messages.error(request, "You do not have access to this grant/project.")
                return _finance_journal_action_redirect(request)

    from django.db import transaction

    try:
        with transaction.atomic(using=tenant_db):
            if action == "submit":
                if entry.status != JournalEntry.Status.DRAFT:
                    raise ValueError("Only draft journals can be submitted for approval.")
                # Validate rules before moving forward
                lines = [
                    {
                        "account_id": l.account_id,
                        "description": l.description,
                        "debit": l.debit,
                        "credit": l.credit,
                        "grant_id": getattr(l, "grant_id", None) or "",
                    }
                    for l in entry.lines.all()
                ]
                header = {
                    "entry_date": entry.entry_date,
                    "memo": entry.memo,
                    "grant_id": entry.grant_id or "",
                    "journal_type": entry.journal_type or "",
                    "source": entry.source or "",
                }
                gl_date = _finance_journal_gl_date(entry)
                _finance_assert_open_period(gl_date, tenant_db, request.tenant_user_id)
                _finance_validate_journal_payload(header, lines, tenant_db)
                _finance_validate_journal_line_grants(lines, tenant_db, gl_date)
                if entry.grant_id:
                    entry.refresh_from_db()
                    entry.grant = Grant.objects.using(tenant_db).select_related("project").get(pk=entry.grant_id)
                    try:
                        entry.full_clean()
                    except Exception as e:
                        if hasattr(e, "message_dict"):
                            err = " ".join(f"{k}: {v}" for k, v in e.message_dict.items())
                        else:
                            err = str(e)
                        raise ValueError(err)
                old_status = entry.status
                from django.utils import timezone as _dj_tz

                entry.status = JournalEntry.Status.PENDING_APPROVAL
                entry.submitted_by = request.tenant_user
                entry.submitted_at = _dj_tz.now()
                entry.save(update_fields=["status", "submitted_by", "submitted_at"])
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=_actor_name(),
                    summary=f"Journal status changed {old_status} → {entry.status}",
                )
                messages.success(request, "Journal submitted for approval.")

            elif action == "approve":
                if entry.status != JournalEntry.Status.PENDING_APPROVAL:
                    raise ValueError("Only journals pending approval can be approved.")
                if (
                    entry.created_by_id
                    and request.tenant_user
                    and entry.created_by_id == request.tenant_user.id
                    and not (_has_perm("finance:journals.override_maker_checker") or _has_perm("finance:maker_checker.override"))
                ):
                    raise ValueError(
                        "Maker-checker is enforced: you cannot approve a journal entry you created."
                    )
                old_status = entry.status
                from django.utils import timezone as _dj_tz

                entry.status = JournalEntry.Status.APPROVED
                entry.approved_by = request.tenant_user
                entry.approved_at = _dj_tz.now()
                entry.save(update_fields=["status", "approved_by", "approved_at"])
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=_actor_name(),
                    summary=f"Journal status changed {old_status} → {entry.status}",
                )
                messages.success(request, "Journal approved.")

            elif action == "reject":
                if entry.status != JournalEntry.Status.PENDING_APPROVAL:
                    raise ValueError("Only submitted journals (pending approval) can be returned to draft.")
                comment = (request.POST.get("correction_comment") or "").strip()
                old_status = entry.status
                entry.status = JournalEntry.Status.DRAFT
                entry.submitted_by_id = None
                entry.submitted_at = None
                entry.save(update_fields=["status", "submitted_by_id", "submitted_at"])
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=_actor_name(),
                    summary=(
                        f"Journal returned to draft (was {old_status})."
                        + (f" Comment: {comment}" if comment else "")
                    )[:255],
                )
                messages.success(request, "Journal returned to draft.")

            elif action == "post":
                if entry.status != JournalEntry.Status.APPROVED:
                    raise ValueError("Only approved journals can be posted.")
                # Backdated posting control (date before today)
                from django.utils import timezone as _tz

                today = _tz.localdate()
                gl_date = _finance_journal_gl_date(entry)
                if gl_date and gl_date < today and not _has_perm("finance:posting.backdated"):
                    raise ValueError("Backdated posting is restricted by policy.")
                if (
                    entry.created_by_id
                    and request.tenant_user
                    and entry.created_by_id == request.tenant_user.id
                    and not (_has_perm("finance:journals.override_maker_checker") or _has_perm("finance:maker_checker.override"))
                ):
                    raise ValueError(
                        "Maker-checker is enforced: you cannot post a journal entry you created."
                    )
                _finance_assert_open_period(gl_date, tenant_db, request.tenant_user_id)
                if entry.grant_id:
                    from decimal import Decimal
                    from django.db.models import Sum
                    from tenant_finance.models import ChartAccount, get_grant_posted_expense_total

                    existing_spent = get_grant_posted_expense_total(entry.grant_id, tenant_db)
                    entry_expense = (
                        entry.lines.using(tenant_db)
                        .filter(account__type=ChartAccount.Type.EXPENSE)
                        .aggregate(s=Sum("debit"))
                        .get("s") or Decimal("0")
                    )
                    grant = Grant.objects.using(tenant_db).get(pk=entry.grant_id)
                    grant_budget = getattr(grant, "award_amount", None)
                    if grant_budget is None:
                        grant_budget = getattr(grant, "amount_awarded", None)
                    grant_budget = Decimal(str(grant_budget or 0))
                    if grant_budget and existing_spent + entry_expense > grant_budget:
                        raise ValueError(
                            f"Grant budget would be exceeded. Budget: {grant_budget}, "
                            f"already posted: {existing_spent}, this entry expense: {entry_expense}."
                        )
                    from tenant_grants.services.project_budget_actuals import (
                        validate_journal_entry_expense_budget_dimensions,
                    )

                    dim_errs = validate_journal_entry_expense_budget_dimensions(entry, tenant_db)
                    if dim_errs:
                        raise ValueError(dim_errs[0])
                    from tenant_grants.restrictions import evaluate_journal_post_restrictions

                    override_dn = user_has_permission(
                        request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
                    )
                    viol = evaluate_journal_post_restrictions(
                        entry, tenant_db, has_override_permission=override_dn
                    )
                    for v in viol:
                        if not v.blocks_posting and v.compliance_level == "recommended":
                            messages.warning(request, v.message)
                    hard = [v for v in viol if v.blocks_posting]
                    if hard:
                        raise ValueError(hard[0].message)
                if not entry.reference:
                    jt = (entry.journal_type or "").lower()
                    if entry.source == "manual" and jt in ("adjustment", "adjusting", "adjusting_journal"):
                        entry.reference = _finance_generate_adjusting_journal_number(tenant_db, gl_date)
                    else:
                        entry.reference = _finance_generate_journal_number(tenant_db, gl_date)
                old_status = entry.status
                entry.status = JournalEntry.Status.POSTED
                from django.utils import timezone

                entry.posted_at = timezone.now()
                entry.posted_by = request.tenant_user
                entry.save(
                    update_fields=["status", "reference", "posted_at", "posted_by"]
                )
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=_actor_name(),
                    summary=f"Journal status changed {old_status} → {entry.status}",
                )
                messages.success(request, "Journal posted to the general ledger.")
                try:
                    from tenant_grants.services.project_budget_actuals import (
                        refresh_project_budget_and_activity_actuals,
                    )

                    refresh_project_budget_and_activity_actuals(tenant_db, entry=entry)
                except Exception:
                    pass

            elif action == "reverse":
                if entry.status != JournalEntry.Status.POSTED:
                    raise ValueError("Only posted journals can be reversed.")
                if getattr(entry, "is_system_generated", False):
                    raise ValueError(
                        "This journal was generated by the system from a source transaction. "
                        "Reverse or adjust it from that source document or contact Finance."
                    )

                # Load reversal rules (singleton per tenant)
                rule = (
                    TransactionReversalRule.objects.using(tenant_db)
                    .select_for_update()
                    .first()
                )
                if not rule or not rule.allow_reversal:
                    raise ValueError(
                        "Reversals are disabled by the finance configuration. "
                        "Contact your system administrator."
                    )

                # Enforce authorized roles for reversal (Finance Manager / Administrator equivalents)
                if rule.authorized_roles_for_reversal:
                    codes = [
                        c.strip()
                        for c in rule.authorized_roles_for_reversal.split(",")
                        if c.strip()
                    ]
                    if codes and not any(
                        user_has_permission(request.tenant_user, code, using=tenant_db)
                        for code in codes
                    ):
                        raise ValueError(
                            "You are not authorized to reverse posted vouchers. "
                            "Only designated finance managers or administrators may perform reversals."
                        )

                # Justification is always required for reversals
                reversal_reason = (request.POST.get("reversal_reason") or "").strip()
                if not reversal_reason:
                    raise ValueError(
                        "A justification comment is required when reversing a posted journal."
                    )

                reversal_date_str = request.POST.get("reversal_date") or ""
                from django.utils.dateparse import parse_date

                reversal_date = parse_date(reversal_date_str) or entry.entry_date
                # Always block reversals if the accounting period is closed
                _finance_assert_open_period(
                    reversal_date, tenant_db, request.tenant_user_id
                )

                # Fiscal year validation for cross-period reversals
                if rule.prevent_cross_period_reversal:
                    def _get_fiscal_year(d):
                        return (
                            FiscalYear.objects.using(tenant_db)
                            .filter(start_date__lte=d, end_date__gte=d)
                            .first()
                        )

                    orig_fy = _get_fiscal_year(entry.entry_date)
                    rev_fy = _get_fiscal_year(reversal_date)
                    if orig_fy and rev_fy and orig_fy.id != rev_fy.id:
                        allow_cross_period = False
                        codes = [
                            c.strip()
                            for c in (
                                rule.authorized_roles_for_cross_period_reversal or ""
                            ).split(",")
                            if c.strip()
                        ]
                        if codes:
                            allow_cross_period = any(
                                user_has_permission(
                                    request.tenant_user, code, using=tenant_db
                                )
                                for code in codes
                            )
                        if not allow_cross_period:
                            raise ValueError(
                                "Reversal across fiscal years is not allowed by policy. "
                                "Contact Finance for an authorized cross-year adjustment."
                            )

                reversal_ref = _finance_generate_journal_number(tenant_db, reversal_date)
                # If policy requires approval workflow for reversals, create the journal
                # in Pending Approval status so it follows the standard approve/post steps.
                reversal_status = (
                    JournalEntry.Status.PENDING_APPROVAL
                    if rule.require_reversal_approval
                    else JournalEntry.Status.POSTED
                )
                from django.utils import timezone as _dj_tz

                reversal_kwargs = {
                    "entry_date": reversal_date,
                    "memo": f"Reversal of {entry.reference or entry.id} — {reversal_reason}",
                    "grant": entry.grant,
                    "status": reversal_status,
                    "created_by": request.tenant_user,
                    "reference": reversal_ref,
                    "source": "reversal",
                    "journal_type": "reversal",
                    "source_type": JournalEntry.SourceType.REVERSAL,
                    "is_system_generated": True,
                    "source_document_no": reversal_ref,
                    "source_id": None,
                }
                if not rule.require_reversal_approval:
                    # Auto-approve when no approval workflow is required
                    reversal_kwargs["approved_by"] = request.tenant_user
                    reversal_kwargs["posted_at"] = _dj_tz.now()
                    reversal_kwargs["posted_by"] = request.tenant_user
                reversal = JournalEntry.objects.using(tenant_db).create(**reversal_kwargs)
                JournalEntry.objects.using(tenant_db).filter(pk=reversal.pk).update(
                    source_id=reversal.pk
                )
                for line in entry.lines.all():
                    JournalLine.objects.using(tenant_db).create(
                        entry=reversal,
                        account=line.account,
                        grant_id=getattr(line, "grant_id", None),
                        project_budget_line_id=getattr(line, "project_budget_line_id", None),
                        workplan_activity_id=getattr(line, "workplan_activity_id", None),
                        description=f"Reversal of line {line.id}",
                        debit=Decimal(line.credit or 0),
                        credit=Decimal(line.debit or 0),
                    )
                entry.reversed_by = reversal
                entry.status = JournalEntry.Status.REVERSED
                entry.save(update_fields=["reversed_by", "status"])

                # Audit trail: capture original and reversal voucher numbers, user, reason, and timestamp
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=request.tenant_user.get_full_name()
                    if getattr(request, "tenant_user", None)
                    else "",
                    summary=(
                        f"Journal {entry.reference or entry.id} reversed by {reversal.reference}. "
                        f"Reason: {reversal_reason}"
                    ),
                )
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=reversal.id,
                    action=AuditLog.Action.CREATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=request.tenant_user.get_full_name()
                    if getattr(request, "tenant_user", None)
                    else "",
                    summary=(
                        f"Reversal journal {reversal.reference} created for "
                        f"{entry.reference or entry.id}. Reason: {reversal_reason}"
                    ),
                )
                messages.success(request, f"Reversal journal {reversal.reference} created and posted.")

            else:
                raise ValueError("Unsupported journal action.")

    except ValueError as e:
        messages.error(request, str(e))

    return _finance_journal_action_redirect(request)


def _finance_assert_open_period(entry_date, tenant_db, request_user_id=None):
    """
    Ensure Financial Setup defines an accounting period for entry_date and posting is allowed
    (open, or soft-closed with role exception). Uses the same rules as model-level posting checks.
    """
    from tenant_finance.services.accounting_periods import assert_can_post
    from tenant_users.models import TenantUser

    user = None
    if request_user_id:
        user = TenantUser.objects.using(tenant_db).filter(pk=request_user_id).first()
    assert_can_post(using=tenant_db, dt=entry_date, user=user)


def _compute_bank_current_balance(bank_account, tenant_db):
    """
    Compute current balance for a bank account: opening balance + posted debits - posted credits.
    """
    from decimal import Decimal as _D
    from tenant_finance.models import JournalEntry, JournalLine

    bal = (
        JournalLine.objects.using(tenant_db)
        .filter(account_id=bank_account.account_id, entry__status=JournalEntry.Status.POSTED)
        .aggregate(b=Sum("debit") - Sum("credit"))
        .get("b")
        or _D("0")
    )
    return (bank_account.opening_balance or _D("0")) + bal


def _finance_generate_journal_number(tenant_db, entry_date):
    """Generate journal numbers using configured DocumentSeries, with fiscal-year support."""
    from tenant_finance.models import DocumentSeries, JournalEntry
    from tenant_finance.services.numbering import generate_document_number

    year = entry_date.year

    try:
        gen = generate_document_number(
            using=tenant_db,
            document_type=DocumentSeries.DocumentType.JOURNAL,
            entry_date=entry_date,
        )
        candidate = gen.value
        # Defensive duplicate check
        i = 1
        while JournalEntry.objects.using(tenant_db).filter(reference=candidate).exists():
            i += 1
            candidate = f"{gen.value}-{i}"
        return candidate
    except Exception:
        # Fallback: JV-YYYY-00001 style based on existing refs when no series configured.
        prefix = f"JV-{year}-"
    last = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith=prefix)
        .order_by("-reference")
        .first()
    )
    if last and last.reference and last.reference.startswith(prefix):
        try:
            last_seq = int(last.reference.split("-")[-1])
        except Exception:
            last_seq = 0
    else:
        last_seq = 0
    return f"{prefix}{last_seq + 1:05d}"


def _finance_generate_adjusting_journal_number(tenant_db, entry_date):
    """
    Generate AJV-YYYY-00001 style numbers specifically for adjusting journals.
    """
    from tenant_finance.models import JournalEntry

    year = entry_date.year
    prefix = f"AJV-{year}-"
    last = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith=prefix)
        .order_by("-reference")
        .first()
    )
    if last and last.reference and last.reference.startswith(prefix):
        try:
            last_seq = int(last.reference.split("-")[-1])
        except Exception:
            last_seq = 0
    else:
        last_seq = 0
    return f"{prefix}{last_seq + 1:05d}"


def _finance_validate_journal_payload(header, lines, tenant_db):
    """
    Validate journal business rules.

    header: dict with entry_date, memo, grant_id, journal_type, source
    lines: list of dicts with account_id, description, debit, credit
    """
    from decimal import Decimal

    from tenant_finance.models import ChartAccount

    if len(lines) < 2:
        raise ValueError("A journal must have at least two lines.")

    total_debit = Decimal("0")
    total_credit = Decimal("0")

    for idx, line in enumerate(lines, start=1):
        account_id = line.get("account_id")
        debit = Decimal(line.get("debit") or "0")
        credit = Decimal(line.get("credit") or "0")

        if debit and credit:
            raise ValueError(f"Line {idx}: a line cannot have both debit and credit.")
        if not debit and not credit:
            raise ValueError(f"Line {idx}: a line must have either debit or credit.")

        # Only active leaf (posting) accounts allowed; parent/summary accounts cannot be used in transactions
        account = (
            ChartAccount.objects.using(tenant_db)
            .filter(pk=account_id, is_active=True)
            .first()
        )
        if not account:
            raise ValueError(f"Line {idx}: account is not an active posting account.")
        if not account.is_leaf(using=tenant_db):
            raise ValueError(
                f"Line {idx}: only leaf (posting) accounts can be used in transactions. "
                f"Account {account.code} is a summary account and cannot receive entries."
            )

        total_debit += debit
        total_credit += credit

    if total_debit != total_credit:
        raise ValueError("Total debit must equal total credit.")

    return total_debit, total_credit


def _finance_journal_gl_date(entry) -> object:
    """Effective GL date for period controls (posting date when set, else journal date)."""
    pd = getattr(entry, "posting_date", None)
    return pd or entry.entry_date


def _finance_validate_journal_line_grants(lines: list, tenant_db: str, gl_date) -> None:
    """Ensure optional line-level grants are active and valid for gl_date."""
    from django.core.exceptions import ValidationError as DjangoValidationError

    from tenant_finance.models import JournalEntry
    from tenant_grants.models import Grant

    for idx, line in enumerate(lines, start=1):
        gid = line.get("grant_id")
        if not gid:
            continue
        grant = Grant.objects.using(tenant_db).select_related("project").filter(pk=gid).first()
        if not grant or grant.status != "active":
            raise ValueError(f"Line {idx}: project/grant must be active.")
        tmp = JournalEntry(entry_date=gl_date, grant=grant)
        try:
            tmp.full_clean()
        except DjangoValidationError as e:
            err = " ".join(f"{k}: {v}" for k, v in (e.message_dict or {}).items()) or str(e)
            raise ValueError(f"Line {idx}: {err}") from e


def _finance_journal_action_redirect(request: HttpRequest):
    """Redirect after journal workflow action (adjusting vs general register)."""
    target = (request.POST.get("next") or "").strip()
    if target == "adjusting":
        return redirect(reverse("tenant_portal:finance_adjusting_journals"))
    if target == "approval":
        return redirect(reverse("tenant_portal:finance_journal_approval"))
    return redirect(reverse("tenant_portal:finance_journals"))


def _journal_enterprise_category(entry) -> str:
    """Bucket for journal approval UI: general | recurring | adjusting | system."""
    if getattr(entry, "is_system_generated", False):
        return "system"
    jt = (entry.journal_type or "").lower()
    if jt in ("adjusting_journal", "adjustment"):
        return "adjusting"
    if "recurring" in jt:
        return "recurring"
    return "general"


def _journal_number_display(entry) -> str:
    ref = (getattr(entry, "reference", None) or "").strip()
    if ref:
        return ref
    return f"JE-{entry.entry_date.year}-{entry.id:05d}"


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_account_categories_view(request: HttpRequest) -> HttpResponse:
    import csv
    from io import TextIOWrapper

    from django.db.models import Count

    from django.conf import settings

    from tenant_finance.models import AccountCategory
    from tenant_portal.migration_checks import ensure_account_category_schema

    tenant_db = request.tenant_db

    if not ensure_account_category_schema(
        tenant_db, auto_migrate=getattr(settings, "TENANT_AUTO_MIGRATE", False)
    ):
        return render(
            request,
            "tenant_portal/finance/tenant_migration_required.html",
            {
                "tenant": request.tenant,
                "tenant_user": request.tenant_user,
                "tenant_db": tenant_db,
                "migration_label": "tenant_finance.0033_accountcategory_enterprise_fields",
                "active_submenu": "core",
                "active_item": "core_categories",
            },
            status=503,
        )

    def _filtered_qs():
        qs = (
            AccountCategory.objects.using(tenant_db)
            .select_related("parent_category")
            .annotate(usage_count=Count("accounts"))
        )
        ft = (request.GET.get("type") or "").strip()
        fs = (request.GET.get("statement") or "").strip()
        fst = (request.GET.get("status") or "").strip()
        if ft in dict(AccountCategory.CategoryType.choices):
            qs = qs.filter(category_type=ft)
        if fs in dict(AccountCategory.StatementType.choices):
            qs = qs.filter(statement_type=fs)
        if fst in dict(AccountCategory.Status.choices):
            qs = qs.filter(status=fst)
        return qs.order_by("display_order", "code")

    # Export categories (respects current filters; omit export flag from filters)
    if request.method == "GET" and request.GET.get("export") == "1":
        categories = _filtered_qs()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="account_categories.csv"'
        writer = csv.writer(response)
        writer.writerow(
            [
                "code",
                "name",
                "statement_type",
                "category_type",
                "normal_balance",
                "parent_code",
                "is_system",
                "status",
                "description",
                "display_order",
            ]
        )
        for c in categories:
            writer.writerow(
                [
                    c.code,
                    c.name,
                    c.statement_type or "",
                    c.category_type,
                    c.normal_balance,
                    c.parent_category.code if c.parent_category_id else "",
                    "1" if c.is_system else "0",
                    c.status or AccountCategory.Status.ACTIVE,
                    (c.description or "").replace("\n", " "),
                    c.display_order,
                ]
            )
        return response

    if request.method == "POST":
        action = request.POST.get("action", "create")

        if action == "delete":
            category_id = request.POST.get("category_id")
            cat = (
                AccountCategory.objects.using(tenant_db).filter(pk=category_id).first()
                if category_id
                else None
            )
            if cat:
                try:
                    cat.delete(using=tenant_db)
                    messages.success(request, _("Category deleted."))
                except ValidationError as exc:
                    messages.error(request, " ".join(exc.messages))
            return redirect(reverse("tenant_portal:finance_account_categories"))

        if action == "update":
            category_id = request.POST.get("category_id")
            cat = (
                AccountCategory.objects.using(tenant_db)
                .annotate(_in_use=Count("accounts"))
                .filter(pk=category_id)
                .first()
                if category_id
                else None
            )
            if not cat:
                messages.error(request, _("Category not found."))
                return redirect(reverse("tenant_portal:finance_account_categories"))
            in_use = getattr(cat, "_in_use", 0) > 0
            name = (request.POST.get("name") or "").strip()
            description = (request.POST.get("description") or "").strip()
            display_order = request.POST.get("display_order")
            status = (request.POST.get("status") or cat.status or AccountCategory.Status.ACTIVE).strip()
            if status not in dict(AccountCategory.Status.choices):
                status = AccountCategory.Status.ACTIVE
            cat.name = name or cat.name
            cat.description = description
            if display_order is not None and str(display_order).strip().isdigit():
                cat.display_order = int(display_order)
            cat.status = status
            # Non-system + linked accounts: only name, description, order, status (enterprise lock)
            if not cat.is_system and not in_use:
                code = (request.POST.get("code") or "").strip()
                statement_type = (request.POST.get("statement_type") or "").strip()
                category_type = (request.POST.get("category_type") or "").strip()
                normal_balance = (request.POST.get("normal_balance") or "").strip()
                parent_id = (request.POST.get("parent_category_id") or "").strip()
                if code:
                    cat.code = code
                if statement_type in dict(AccountCategory.StatementType.choices):
                    cat.statement_type = statement_type
                if category_type in dict(AccountCategory.CategoryType.choices):
                    cat.category_type = category_type
                if normal_balance in dict(AccountCategory.NormalBalance.choices):
                    cat.normal_balance = normal_balance
                if parent_id:
                    parent = AccountCategory.objects.using(tenant_db).filter(pk=parent_id).first()
                    cat.parent_category = parent
                else:
                    cat.parent_category = None
            try:
                cat.save(using=tenant_db)
                messages.success(request, _("Category updated."))
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))
            rq = (request.POST.get("return_qs") or "").strip()
            if rq:
                return redirect(f"{reverse('tenant_portal:finance_account_categories')}?{rq}")
            return redirect(reverse("tenant_portal:finance_account_categories"))

        if action == "toggle_status":
            category_id = request.POST.get("category_id")
            if category_id:
                cat = AccountCategory.objects.using(tenant_db).filter(pk=category_id).first()
                if cat:
                    cat.status = (
                        AccountCategory.Status.INACTIVE
                        if cat.status != AccountCategory.Status.INACTIVE
                        else AccountCategory.Status.ACTIVE
                    )
                    try:
                        cat.save(using=tenant_db, update_fields=["status"])
                        messages.success(
                            request,
                            f"Category {cat.code} is now {'active' if cat.status == AccountCategory.Status.ACTIVE else 'inactive'}.",
                        )
                    except ValidationError as exc:
                        messages.error(request, " ".join(exc.messages))
            return redirect(reverse("tenant_portal:finance_account_categories"))

        if action == "import":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Please choose a CSV file to import.")
                return redirect(reverse("tenant_portal:finance_account_categories"))
            filename = (upload.name or "").lower()
            if not filename.endswith(".csv"):
                messages.error(
                    request,
                    "Only CSV files are supported. Please export or save your Account Categories as CSV before importing.",
                )
                return redirect(reverse("tenant_portal:finance_account_categories"))
            try:
                wrapper = TextIOWrapper(upload, encoding="utf-8-sig", errors="ignore")
                reader = csv.DictReader(wrapper)
            except Exception:
                messages.error(
                    request,
                    "Unable to read the uploaded CSV file. Please check the encoding and try again.",
                )
                return redirect(reverse("tenant_portal:finance_account_categories"))

            created_count = 0
            for row in reader:
                code = (row.get("code") or "").strip()
                name = (row.get("name") or "").strip()
                statement_type = (row.get("statement_type") or "").strip()
                category_type = (row.get("category_type") or "").strip()
                normal_balance = (row.get("normal_balance") or "").strip()
                status = (row.get("status") or "active").strip().lower() or "active"
                description = (row.get("description") or "").strip()
                parent_code = (row.get("parent_code") or "").strip()
                if not code or not name or not statement_type:
                    continue
                if AccountCategory.objects.using(tenant_db).filter(code=code).exists():
                    continue
                if statement_type not in dict(AccountCategory.StatementType.choices):
                    continue
                if category_type not in dict(AccountCategory.CategoryType.choices):
                    continue
                if normal_balance not in dict(AccountCategory.NormalBalance.choices):
                    normal_balance = (
                        AccountCategory.NormalBalance.DEBIT
                        if category_type in (AccountCategory.CategoryType.ASSET, AccountCategory.CategoryType.EXPENSE)
                        else AccountCategory.NormalBalance.CREDIT
                    )
                if status not in dict(AccountCategory.Status.choices):
                    status = AccountCategory.Status.ACTIVE
                parent = None
                if parent_code:
                    parent = AccountCategory.objects.using(tenant_db).filter(code=parent_code).first()
                obj = AccountCategory(
                    code=code,
                    name=name,
                    statement_type=statement_type,
                    category_type=category_type,
                    normal_balance=normal_balance,
                    status=status,
                    description=description,
                    parent_category=parent,
                    is_system=False,
                )
                try:
                    obj.save(using=tenant_db)
                    created_count += 1
                except ValidationError:
                    continue
            if created_count > 0:
                messages.success(
                    request,
                    f"Imported {created_count} category(ies) from the file.",
                )
            return redirect(reverse("tenant_portal:finance_account_categories"))

        # create
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        statement_type = (request.POST.get("statement_type") or "").strip()
        category_type = (request.POST.get("category_type") or "").strip()
        normal_balance = (request.POST.get("normal_balance") or "").strip()
        status = (request.POST.get("status") or "active").strip().lower() or "active"
        description = (request.POST.get("description") or "").strip()
        parent_id = (request.POST.get("parent_category_id") or "").strip()
        display_order = request.POST.get("display_order")
        if not code or not name or not statement_type or not category_type:
            messages.error(
                request,
                _("Please provide code, name, statement type, and category class."),
            )
        elif AccountCategory.objects.using(tenant_db).filter(code=code).exists():
            messages.error(request, _("Code must be unique."))
        elif statement_type not in dict(AccountCategory.StatementType.choices):
            messages.error(request, _("Please select a valid statement type."))
        elif category_type not in dict(AccountCategory.CategoryType.choices):
            messages.error(request, _("Please select a valid category class."))
        else:
            if status not in dict(AccountCategory.Status.choices):
                status = AccountCategory.Status.ACTIVE
            if normal_balance not in dict(AccountCategory.NormalBalance.choices):
                normal_balance = (
                    AccountCategory.NormalBalance.DEBIT
                    if category_type
                    in (AccountCategory.CategoryType.ASSET, AccountCategory.CategoryType.EXPENSE)
                    else AccountCategory.NormalBalance.CREDIT
                )
            parent = None
            if parent_id:
                parent = AccountCategory.objects.using(tenant_db).filter(pk=parent_id).first()
            do = 0
            if display_order is not None and str(display_order).strip().isdigit():
                do = int(display_order)
            obj = AccountCategory(
                code=code,
                name=name,
                statement_type=statement_type,
                category_type=category_type,
                normal_balance=normal_balance,
                status=status,
                description=description,
                parent_category=parent,
                display_order=do,
                is_system=False,
            )
            try:
                obj.save(using=tenant_db)
                messages.success(request, _("Category created."))
                return redirect(reverse("tenant_portal:finance_account_categories"))
            except ValidationError as exc:
                messages.error(request, " ".join(exc.messages))

    categories = _filtered_qs()
    parent_options = AccountCategory.objects.using(tenant_db).order_by("display_order", "code")

    edit_category = None
    edit_id = (request.GET.get("edit") or "").strip()
    if edit_id:
        edit_category = (
            AccountCategory.objects.using(tenant_db)
            .filter(pk=edit_id)
            .select_related("parent_category")
            .annotate(usage_count=Count("accounts"))
            .first()
        )

    filter_q = request.GET.copy()
    filter_q.pop("edit", None)
    filter_querystring = filter_q.urlencode()
    has_filters = bool(
        (request.GET.get("type") or "").strip()
        or (request.GET.get("statement") or "").strip()
        or (request.GET.get("status") or "").strip()
    )

    return render(
        request,
        "tenant_portal/finance/account_categories.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "categories": categories,
            "filter_type": (request.GET.get("type") or "").strip(),
            "filter_statement": (request.GET.get("statement") or "").strip(),
            "filter_status": (request.GET.get("status") or "").strip(),
            "filter_querystring": filter_querystring,
            "has_filters": has_filters,
            "statement_types": AccountCategory.StatementType,
            "category_types": AccountCategory.CategoryType,
            "normal_balances": AccountCategory.NormalBalance,
            "statuses": AccountCategory.Status,
            "parent_options": parent_options,
            "edit_category": edit_category,
            "active_submenu": "core",
            "active_item": "core_categories",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_general_ledger_view(request: HttpRequest) -> HttpResponse:
    import csv
    from decimal import Decimal

    from django.contrib import messages
    from django.db.models import Q

    from tenant_finance.models import BankAccount, ChartAccount, JournalEntry, OrganizationSettings
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
        user_sees_all_grants,
    )
    from tenant_grants.models import BudgetLine, Donor, Grant, Project

    tenant_db = request.tenant_db
    f = _parse_gl_filters(request)
    account_id = (request.GET.get("account_id") or "").strip()

    gid = (f.get("grant_id") or "").strip()
    if gid.isdigit() and not assert_grant_filter_allowed(request.tenant_user, tenant_db, int(gid)):
        messages.error(request, "You do not have access to the selected grant.")
        f = {**f, "grant_id": ""}
        gid = ""

    lines_qs = (
        posted_journal_lines(tenant_db)
        .filter(
            gl_date__gte=f["period_start"],
            gl_date__lte=f["period_end"],
        )
        .select_related(
            "account",
            "grant",
            "grant__project",
            "grant__donor",
            "entry",
            "entry__grant",
            "entry__grant__project",
            "entry__grant__donor",
            "entry__donor",
            "entry__posted_by",
        )
        .order_by("account__code", "gl_date", "entry_id", "id")
    )
    lines_qs = restrict_journal_lines_by_grant_scope(lines_qs, request.tenant_user, tenant_db)

    if account_id.isdigit():
        lines_qs = lines_qs.filter(account_id=int(account_id))
    if gid.isdigit():
        g = int(gid)
        lines_qs = lines_qs.filter(Q(grant_id=g) | Q(entry__grant_id=g))
    did = (f.get("donor_id") or "").strip()
    if did.isdigit():
        d = int(did)
        lines_qs = lines_qs.filter(
            Q(entry__donor_id=d)
            | Q(entry__grant__donor_id=d)
            | Q(grant__donor_id=d)
        )
    pid = (f.get("project_id") or "").strip()
    if pid.isdigit():
        p = int(pid)
        lines_qs = lines_qs.filter(
            Q(entry__grant__project_id=p) | Q(grant__project_id=p)
        )
    st = (f.get("source_type") or "").strip()
    if st:
        lines_qs = lines_qs.filter(entry__source_type=st)
    jn = (f.get("journal_no") or "").strip()
    if jn:
        lines_qs = lines_qs.filter(
            Q(entry__reference__icontains=jn) | Q(entry__source_document_no__icontains=jn)
        )

    bl_txt = (f.get("budget_line") or "").strip()
    if bl_txt:
        bl_pairs = list(
            BudgetLine.objects.using(tenant_db)
            .filter(Q(category__icontains=bl_txt) | Q(description__icontains=bl_txt))
            .exclude(grant_id__isnull=True)
            .exclude(account_id__isnull=True)
            .values_list("grant_id", "account_id")
        )
        if not bl_pairs:
            lines_qs = lines_qs.none()
        else:
            q_bl = Q()
            for bg, ba in bl_pairs:
                q_bl |= (Q(grant_id=bg) | Q(entry__grant_id=bg)) & Q(account_id=ba)
            lines_qs = lines_qs.filter(q_bl)

    all_lines = list(lines_qs)

    grant_ids: set[int] = set()
    account_ids: set[int] = set()
    for ln in all_lines:
        eg = ln.grant_id or (ln.entry.grant_id if ln.entry else None)
        if eg:
            grant_ids.add(eg)
        account_ids.add(ln.account_id)
    budget_map: dict[tuple[int | None, int], str] = {}
    if grant_ids and account_ids:
        for bl in (
            BudgetLine.objects.using(tenant_db)
            .filter(grant_id__in=grant_ids, account_id__in=account_ids)
            .only("grant_id", "account_id", "category", "description")
        ):
            label = (bl.category or "").strip()
            if not label:
                label = (bl.description or "").strip()
            budget_map[(bl.grant_id, bl.account_id)] = label or "—"

    def _eff_grant_id(ln):
        return ln.grant_id or (ln.entry.grant_id if ln.entry else None)

    def _grant_label(ln, ent):
        g = ln.grant or ent.grant
        if g:
            return f"{g.code} — {g.title}"
        return "—"

    def _donor_label(ln, ent):
        if ent.donor_id and getattr(ent, "donor", None):
            return ent.donor.name
        g = ln.grant or ent.grant
        if g and getattr(g, "donor", None):
            return g.donor.name
        return "—"

    rows: list[dict] = []
    running_balance_by_account: dict[int, Decimal] = {}
    current_acc_id: int | None = None

    for line in all_lines:
        ent = line.entry
        acc_id = line.account_id
        if current_acc_id != acc_id:
            current_acc_id = acc_id
            running_balance_by_account[acc_id] = Decimal("0")

        dr = line.debit or Decimal("0")
        cr = line.credit or Decimal("0")
        running_balance_by_account[acc_id] += dr - cr

        eg_id = _eff_grant_id(line)
        bkey = (eg_id, acc_id)
        activity = budget_map.get(bkey)
        if activity is None and ent.grant_id:
            activity = budget_map.get((ent.grant_id, acc_id))
        if activity is None:
            activity = "—"

        gl_date = getattr(line, "gl_date", None) or ent.posting_date or ent.entry_date
        posted_by = "—"
        if ent.posted_by_id and getattr(ent, "posted_by", None):
            posted_by = ent.posted_by.get_full_name()

        rows.append(
            {
                "gl_date": gl_date,
                "entry_id": ent.pk,
                "journal_number": _gl_journal_number(ent),
                "account_label": f"{line.account.code} — {line.account.name}",
                "grant": _grant_label(line, ent),
                "donor": _donor_label(line, ent),
                "activity": activity,
                "memo": (line.description or ent.memo or "").strip() or "—",
                "debit": line.debit,
                "credit": line.credit,
                "running_balance": running_balance_by_account[acc_id],
                "account_code": line.account.code,
                "account_name": line.account.name,
                "account_id": acc_id,
                "posted_by": posted_by,
                "entry_source_type": (getattr(ent, "source_type", None) or "").strip(),
                "entry_journal_type": (getattr(ent, "journal_type", None) or "").strip().lower(),
            }
        )

    liquidity_account_ids = set(
        BankAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .values_list("account_id", flat=True)
    )
    gl_recon_receipts = Decimal("0")
    gl_recon_payments = Decimal("0")
    gl_recon_rv_debits = Decimal("0")
    gl_recon_other_debits = Decimal("0")
    gl_recon_pv_credits = Decimal("0")
    gl_recon_other_credits = Decimal("0")
    gl_recon_closing_last: dict[int, Decimal] = {}
    rv_st = JournalEntry.SourceType.RECEIPT_VOUCHER
    pv_st = JournalEntry.SourceType.PAYMENT_VOUCHER
    for r in rows:
        if r["account_id"] not in liquidity_account_ids:
            continue
        dr = r["debit"] or Decimal("0")
        cr = r["credit"] or Decimal("0")
        gl_recon_receipts += dr
        gl_recon_payments += cr
        gl_recon_closing_last[r["account_id"]] = r["running_balance"]
        st = r.get("entry_source_type") or ""
        jt = r.get("entry_journal_type") or ""
        if dr > 0:
            if st == rv_st or jt == "receipt_voucher":
                gl_recon_rv_debits += dr
            else:
                gl_recon_other_debits += dr
        if cr > 0:
            if st == pv_st or jt == "payment_voucher":
                gl_recon_pv_credits += cr
            else:
                gl_recon_other_credits += cr
    gl_recon_net = gl_recon_receipts - gl_recon_payments
    gl_recon_closing = sum(gl_recon_closing_last.values(), Decimal("0"))

    grand_total_debit = Decimal("0")
    grand_total_credit = Decimal("0")
    for r in rows:
        grand_total_debit += r["debit"] or Decimal("0")
        grand_total_credit += r["credit"] or Decimal("0")
    closing_balance = grand_total_credit - grand_total_debit

    account_sections: list[dict] = []
    section_key: int | None = None
    for r in rows:
        aid = r["account_id"]
        if aid != section_key:
            section_key = aid
            account_sections.append(
                {
                    "type": "header",
                    "code": r["account_code"],
                    "name": r["account_name"],
                }
            )
        account_sections.append({"type": "row", **r})

    accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")
    grants_qs = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    grants = grants_qs
    if user_sees_all_grants(request.tenant_user, tenant_db):
        donors = Donor.objects.using(tenant_db).order_by("name")
    else:
        d_ids = [x for x in grants.values_list("donor_id", flat=True).distinct() if x]
        donors = (
            Donor.objects.using(tenant_db).filter(pk__in=d_ids).order_by("name")
            if d_ids
            else Donor.objects.using(tenant_db).none()
        )
    projects = Project.objects.using(tenant_db).filter(status=Project.Status.ACTIVE).order_by("code")

    selected_grant = None
    if (f.get("grant_id") or "").strip().isdigit():
        selected_grant = (
            grants_qs.filter(pk=int(f["grant_id"])).select_related("donor").first()
        )
    selected_donor = None
    if (f.get("donor_id") or "").strip().isdigit():
        selected_donor = donors.filter(pk=int(f["donor_id"])).first()
    selected_project = None
    if (f.get("project_id") or "").strip().isdigit():
        selected_project = projects.filter(pk=int(f["project_id"])).first()

    org_settings = OrganizationSettings.objects.using(tenant_db).first()
    source_type_choices = list(JournalEntry.SourceType.choices)

    if request.GET.get("format") == "csv":
        from django.utils import timezone as dj_tz

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="general_ledger_export.csv"'
        response.write("\ufeff")
        w = csv.writer(response)
        legal_name = ""
        if org_settings and (org_settings.organization_name or "").strip():
            legal_name = org_settings.organization_name.strip()
        if not legal_name:
            legal_name = getattr(request.tenant, "display_name", None) or getattr(
                request.tenant, "name", ""
            )
        period_label = f'{f["period_start"]} – {f["period_end"]}'
        generated = dj_tz.localtime(dj_tz.now()).strftime("%Y-%m-%d %H:%M")
        w.writerow([legal_name])
        w.writerow(["Report", "General Ledger"])
        w.writerow(["Report period", period_label])
        w.writerow(["Generated", generated])
        w.writerow([])
        w.writerow(
            [
                "Date",
                "Journal",
                "Account",
                "Grant",
                "Donor",
                "Activity",
                "Description",
                "Debit",
                "Credit",
                "Balance",
            ]
        )
        for r in rows:
            w.writerow(
                [
                    r["gl_date"],
                    r["journal_number"],
                    r["account_label"],
                    r["grant"],
                    r["donor"],
                    r["activity"],
                    r["memo"],
                    r["debit"],
                    r["credit"],
                    r["running_balance"],
                ]
            )
        return response
    return render(
        request,
        "tenant_portal/finance/general_ledger.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_gl",
            "filters": f,
            "account_sections": account_sections,
            "row_count": len(rows),
            "grand_total_debit": grand_total_debit,
            "grand_total_credit": grand_total_credit,
            "closing_balance": closing_balance,
            "gl_recon_receipts": gl_recon_receipts,
            "gl_recon_payments": gl_recon_payments,
            "gl_recon_rv_debits": gl_recon_rv_debits,
            "gl_recon_other_debits": gl_recon_other_debits,
            "gl_recon_pv_credits": gl_recon_pv_credits,
            "gl_recon_other_credits": gl_recon_other_credits,
            "gl_recon_net": gl_recon_net,
            "gl_recon_closing": gl_recon_closing,
            "grants": grants,
            "donors": donors,
            "projects": projects,
            "accounts": accounts,
            "selected_account_id": account_id,
            "selected_grant": selected_grant,
            "selected_donor": selected_donor,
            "selected_project": selected_project,
            "org_settings": org_settings,
            "export_csv_url": _finance_export_csv_url(request),
            "source_type_choices": source_type_choices,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_account_ledger_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal

    from django.db.models import Q
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
    )

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    account_id = request.GET.get("account_id")
    gid = (f.get("grant_id") or "").strip()
    if gid.isdigit() and not assert_grant_filter_allowed(request.tenant_user, tenant_db, int(gid)):
        gid = ""
        f = {**f, "grant_id": ""}
    lines_qs = (
        posted_journal_lines(tenant_db)
        .filter(gl_date__gte=f["period_start"], gl_date__lte=f["period_end"])
        .select_related("entry", "account")
        .order_by("gl_date", "id")
    )
    lines_qs = restrict_journal_lines_by_grant_scope(lines_qs, request.tenant_user, tenant_db)
    if account_id:
        lines_qs = lines_qs.filter(account_id=account_id)
    if gid.isdigit():
        g = int(gid)
        lines_qs = lines_qs.filter(Q(entry__grant_id=g) | Q(grant_id=g))
    rows = [
        {
            "date": getattr(l, "gl_date", None) or l.entry.entry_date,
            "ref": l.entry.reference or f"JE-{l.entry.id}",
            "memo": l.entry.memo or l.description,
            "debit": l.debit,
            "credit": l.credit,
        }
        for l in lines_qs[:500]
    ]
    accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")
    from tenant_grants.models import Grant

    grants = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="account_ledger.csv"'
        w = csv.writer(response)
        w.writerow(["Date", "Reference", "Memo", "Debit", "Credit"])
        for row in rows:
            w.writerow([row.get("date"), row.get("ref"), row.get("memo"), row.get("debit"), row.get("credit")])
        return response
    return render(
        request,
        "tenant_portal/finance/account_ledger.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_account_ledger",
            "filters": f,
            "rows": rows,
            "accounts": accounts,
            "grants": grants,
            "selected_account_id": account_id,
            "export_csv_url": _finance_export_csv_url(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_trial_balance_view(request: HttpRequest) -> HttpResponse:
    import csv
    from decimal import Decimal

    from django.contrib import messages
    from django.db.models import Q, Sum
    from django.utils import timezone as dj_tz
    from django.utils.translation import gettext as _

    from tenant_finance.models import ChartAccount, OrganizationSettings
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        is_balanced_debit_credit,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
        user_sees_all_grants,
    )
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    # All TB filters must narrow `base_qs` here so HTML print, PDF, and CSV stay aligned.
    # When adding GET params: filter base_qs, add a row in the CSV header block, and a line in trial_balance.html print header.

    gid = (f.get("grant_id") or "").strip()
    if gid.isdigit() and not assert_grant_filter_allowed(request.tenant_user, tenant_db, int(gid)):
        messages.error(request, "You do not have access to the selected grant.")
        f = {**f, "grant_id": ""}
        gid = ""

    base_qs = posted_journal_lines(tenant_db).filter(
        gl_date__gte=f["period_start"],
        gl_date__lte=f["period_end"],
    )
    base_qs = restrict_journal_lines_by_grant_scope(base_qs, request.tenant_user, tenant_db)
    if gid.isdigit():
        g = int(gid)
        base_qs = base_qs.filter(Q(grant_id=g) | Q(entry__grant_id=g))
    did = (f.get("donor_id") or "").strip()
    if did.isdigit():
        d_int = int(did)
        base_qs = base_qs.filter(Q(donor_id=d_int) | Q(entry__donor_id=d_int))

    agg_qs = (
        base_qs.values("account_id", "account__code", "account__name", "account__type")
        .annotate(debit_total=Sum("debit"), credit_total=Sum("credit"))
        .order_by("account__code")
    )

    type_labels = dict(ChartAccount.Type.choices)
    rows: list[dict] = []
    for r in agg_qs:
        aid = r["account_id"]
        dr = r.get("debit_total") or Decimal("0")
        cr = r.get("credit_total") or Decimal("0")
        bal = dr - cr
        if bal == 0:
            continue
        rows.append(
            {
                "account_id": aid,
                "code": r.get("account__code") or "",
                "name": r.get("account__name") or "",
                "type": r.get("account__type") or "",
                "type_label": type_labels.get(r.get("account__type"), r.get("account__type") or ""),
                "debit": dr,
                "credit": cr,
                "balance": bal,
            }
        )

    total_dr = sum(x["debit"] for x in rows)
    total_cr = sum(x["credit"] for x in rows)
    total_balance = sum(x["balance"] for x in rows)
    tb_balanced = is_balanced_debit_credit(total_dr, total_cr)

    grants_qs = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    grants = grants_qs
    if user_sees_all_grants(request.tenant_user, tenant_db):
        donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    else:
        d_ids = [x for x in grants.values_list("donor_id", flat=True).distinct() if x]
        donors = (
            Donor.objects.using(tenant_db)
            .filter(pk__in=d_ids, status=Donor.Status.ACTIVE)
            .order_by("name")
            if d_ids
            else Donor.objects.using(tenant_db).none()
        )
    selected_grant = None
    if gid.isdigit():
        selected_grant = grants_qs.filter(pk=int(gid)).first()
    selected_donor = None
    if did.isdigit():
        selected_donor = donors.filter(pk=int(did)).first()
    org_settings = OrganizationSettings.objects.using(tenant_db).first()

    if request.GET.get("format") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="trial_balance.csv"'
        response.write("\ufeff")
        w = csv.writer(response)
        legal_name = ""
        if org_settings and (org_settings.organization_name or "").strip():
            legal_name = org_settings.organization_name.strip()
        if not legal_name:
            legal_name = getattr(request.tenant, "display_name", None) or getattr(
                request.tenant, "name", ""
            )
        period_display = (
            f'{f["period_start"].strftime("%d/%m/%Y")} – {f["period_end"].strftime("%d/%m/%Y")}'
        )
        generated = dj_tz.localtime(dj_tz.now()).strftime("%Y-%m-%d %H:%M")
        grant_export = str(_("All"))
        if selected_grant:
            grant_export = f"{selected_grant.code} — {selected_grant.title}"
        donor_export = str(_("All"))
        if selected_donor:
            donor_export = selected_donor.name
        w.writerow([legal_name])
        w.writerow([str(_("Report")), str(_("Trial balance"))])
        w.writerow([str(_("Report period")), period_display])
        w.writerow([str(_("Grant")), grant_export])
        w.writerow([str(_("Donor")), donor_export])
        w.writerow([str(_("Generated")), generated])
        w.writerow([])
        w.writerow(
            [
                str(_("Code")),
                str(_("Account name")),
                str(_("Account type")),
                str(_("Debit")),
                str(_("Credit")),
                str(_("Balance")),
            ]
        )
        for row in rows:
            w.writerow(
                [
                    row["code"],
                    row["name"],
                    row["type_label"],
                    row["debit"],
                    row["credit"],
                    row["balance"],
                ]
            )
        w.writerow([])
        w.writerow([str(_("Totals")), "", "", total_dr, total_cr, total_balance])
        return response
    return render(
        request,
        "tenant_portal/finance/trial_balance.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_tb",
            "filters": f,
            "rows": rows,
            "total_debit": total_dr,
            "total_credit": total_cr,
            "total_balance": total_balance,
            "grants": grants,
            "donors": donors,
            "selected_grant": selected_grant,
            "selected_donor": selected_donor,
            "org_settings": org_settings,
            "export_csv_url": _finance_export_csv_url(request),
            "report_validation": {"tb_balanced": tb_balanced, "tb_total_debit": total_dr, "tb_total_credit": total_cr},
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_recurring_journals_view(request: HttpRequest) -> HttpResponse:
    import json

    from django.contrib import messages
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, RecurringJournal
    from tenant_finance.services.recurring_journal import (
        parse_line_rows,
        save_recurring_template,
    )
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_manage = user_has_permission(user, "module:finance.manage", using=tenant_db)

    if request.method == "POST" and can_manage:
        action = (request.POST.get("action") or "").strip().lower()
        if action == "delete":
            rid = request.POST.get("id")
            if str(rid or "").isdigit():
                RecurringJournal.objects.using(tenant_db).filter(pk=int(rid)).delete()
                messages.success(request, "Recurring journal template deleted.")
            return redirect(reverse("tenant_portal:finance_recurring_journals"))
        if action == "set_status":
            rid = request.POST.get("id")
            st = (request.POST.get("status") or "").strip().lower()
            if str(rid or "").isdigit() and st in {
                c[0] for c in RecurringJournal.Status.choices
            }:
                RecurringJournal.objects.using(tenant_db).filter(pk=int(rid)).update(status=st)
                messages.success(request, "Status updated.")
            return redirect(reverse("tenant_portal:finance_recurring_journals"))
        if action == "run_now":
            messages.info(
                request,
                "Run now posts this template to the general ledger when the scheduler is connected. "
                "Next run date is shown for planning.",
            )
            return redirect(reverse("tenant_portal:finance_recurring_journals"))
        if action in ("create", "update"):
            name = (request.POST.get("name") or "").strip()
            reference_prefix = (request.POST.get("reference_prefix") or "").strip()
            description = (request.POST.get("description") or "").strip()
            frequency = (request.POST.get("frequency") or "").strip()
            status = (request.POST.get("status") or RecurringJournal.Status.ACTIVE).strip()
            start_d = parse_date(request.POST.get("start_date") or "")
            end_d = parse_date(request.POST.get("end_date") or "") or None
            raw_lines = parse_line_rows(request.POST)
            edit_id = request.POST.get("edit_id")
            eid = int(edit_id) if str(edit_id or "").strip().isdigit() else None
            if action == "update" and not eid:
                messages.error(request, "Missing template to update.")
            elif not start_d:
                messages.error(request, "Start date is required.")
            else:
                rj, err = save_recurring_template(
                    tenant_db=tenant_db,
                    journal_id=eid if action == "update" else None,
                    name=name,
                    reference_prefix=reference_prefix,
                    frequency=frequency,
                    start_date=start_d,
                    end_date=end_d,
                    description=description,
                    status=status,
                    raw_lines=raw_lines,
                )
                if err:
                    messages.error(request, err)
                else:
                    messages.success(
                        request,
                        f"Saved “{rj.name}”. Next run: {rj.next_run_date or '—'}.",
                    )
            return redirect(reverse("tenant_portal:finance_recurring_journals"))

    list_qs = (
        RecurringJournal.objects.using(tenant_db)
        .prefetch_related(
            "lines",
            "lines__account",
            "lines__grant",
            "lines__grant__project",
        )
        .order_by("name")
    )
    accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")
    grants = list(
        Grant.objects.using(tenant_db)
        .filter(status=Grant.Status.ACTIVE)
        .select_related("project")
        .order_by("code")[:500]
    )

    recurring_payload = []
    for rj in list_qs:
        recurring_payload.append(
            {
                "id": rj.id,
                "name": rj.name,
                "reference_prefix": rj.reference_prefix or "",
                "description": rj.description or "",
                "frequency": rj.frequency,
                "start_date": rj.start_date.isoformat() if rj.start_date else "",
                "end_date": rj.end_date.isoformat() if rj.end_date else "",
                "status": rj.status,
                "next_run_date": rj.next_run_date.isoformat() if rj.next_run_date else "",
                "lines": [
                    {
                        "account_id": str(l.account_id),
                        "grant_id": str(l.grant_id) if l.grant_id else "",
                        "description": l.description or "",
                        "debit": str(l.debit) if l.debit else "",
                        "credit": str(l.credit) if l.credit else "",
                    }
                    for l in sorted(rj.lines.all(), key=lambda x: (x.display_order, x.id))
                ],
            }
        )

    return render(
        request,
        "tenant_portal/finance/recurring_journals.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "recurring_list": list_qs,
            "accounts": accounts,
            "grants": grants,
            "recurring_json": json.dumps(recurring_payload),
            "frequency_choices": RecurringJournal.Frequency.choices,
            "status_choices": RecurringJournal.Status.choices,
            "can_manage": can_manage,
            "active_submenu": "core",
            "active_item": "core_recurring",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_fiscal_periods_view(request: HttpRequest) -> HttpResponse:
    """Legacy URL: period calendar is maintained under Financial Setup → Fiscal structure."""
    return redirect(reverse("tenant_portal:setup_fiscal_years_list"))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_period_closing_view(request: HttpRequest) -> HttpResponse:
    """Legacy URL: use Financial Setup → Accounting periods (close / reopen with audit)."""
    return redirect(reverse("tenant_portal:setup_accounting_periods_list"))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_reopen_period_view(request: HttpRequest) -> HttpResponse:
    """Legacy URL: reopen closed periods from Financial Setup → Accounting periods."""
    return redirect(reverse("tenant_portal:setup_accounting_periods_list"))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_statement_financial_position_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal

    from django.contrib import messages
    from django.db.models import Q, Sum
    from django.http import HttpResponse
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
        statement_equation_delta,
        user_sees_all_grants,
    )
    from tenant_finance.services.sfp_buckets import partition_assets, partition_liabilities, sum_amounts
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_sfp_filters(request)
    as_of = f["as_at"]
    grant_id = f["grant_id"]
    donor_id = f["donor_id"]

    gid_int = int(grant_id) if str(grant_id).isdigit() else None
    if gid_int and not assert_grant_filter_allowed(request.tenant_user, tenant_db, gid_int):
        messages.error(request, "You do not have access to the selected grant.")
        grant_id = ""
        gid_int = None
        f = {**f, "grant_id": ""}

    selected_grant = None
    if gid_int:
        selected_grant = (
            Grant.objects.using(tenant_db).select_related("donor").filter(pk=gid_int).first()
        )
        if not selected_grant:
            grant_id = ""
            gid_int = None
            f = {**f, "grant_id": ""}

    if grant_id and donor_id and selected_grant and str(selected_grant.donor_id) != str(donor_id):
        messages.warning(
            request,
            "The selected donor does not match this grant; the donor filter was ignored.",
        )
        donor_id = ""
        f = {**f, "donor_id": ""}

    if donor_id and not Donor.objects.using(tenant_db).filter(pk=donor_id).exists():
        donor_id = ""
        f = {**f, "donor_id": ""}

    line_qs = posted_journal_lines(tenant_db).filter(
        gl_date__lte=as_of,
        account__type__in=[
            ChartAccount.Type.ASSET,
            ChartAccount.Type.LIABILITY,
            ChartAccount.Type.EQUITY,
        ],
    )
    line_qs = restrict_journal_lines_by_grant_scope(line_qs, request.tenant_user, tenant_db)
    if gid_int:
        line_qs = line_qs.filter(Q(entry__grant_id=gid_int) | Q(grant_id=gid_int))
    if donor_id:
        did = int(donor_id)
        line_qs = line_qs.filter(Q(entry__donor_id=did) | Q(donor_id=did))

    qs = (
        line_qs.values(
            "account__type",
            "account__code",
            "account__name",
            "account__category__code",
        )
        .annotate(bal=Sum("debit") - Sum("credit"))
        .order_by("account__code")
    )

    rows = list(qs)
    assets_raw = [r for r in rows if r["account__type"] == "asset"]
    liabilities_raw = [r for r in rows if r["account__type"] == "liability"]
    equity_raw = [r for r in rows if r["account__type"] == "equity"]

    assets_current, assets_non_current = partition_assets(assets_raw)
    liabilities_current, liabilities_non_current = partition_liabilities(liabilities_raw)
    fund_balance_rows = [(r["account__code"], r["account__name"], r["bal"]) for r in equity_raw]

    total_current_assets = sum_amounts(assets_current)
    total_non_current_assets = sum_amounts(assets_non_current)
    total_assets = total_current_assets + total_non_current_assets

    total_current_liabilities = sum_amounts(liabilities_current)
    total_non_current_liabilities = sum_amounts(liabilities_non_current)
    total_fund_balance = sum_amounts(fund_balance_rows)
    total_liabilities_and_fund_balance = (
        total_current_liabilities + total_non_current_liabilities + total_fund_balance
    )

    sfp_balance_delta = statement_equation_delta(total_assets, total_liabilities_and_fund_balance)
    report_validation = {
        "sfp_equation_ok": abs(sfp_balance_delta) <= Decimal("0.02"),
        "sfp_balance_delta": sfp_balance_delta,
    }

    grants = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db)
        .select_related("donor")
        .filter(status__in=[Grant.Status.ACTIVE, Grant.Status.CLOSED])
        .order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if user_sees_all_grants(request.tenant_user, tenant_db):
        donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    else:
        d_ids = [x for x in grants.values_list("donor_id", flat=True).distinct() if x]
        donors = (
            Donor.objects.using(tenant_db)
            .filter(pk__in=d_ids, status=Donor.Status.ACTIVE)
            .order_by("name")
            if d_ids
            else Donor.objects.using(tenant_db).none()
        )
    grant_donor_map = {str(g.pk): str(g.donor_id) for g in grants if g.donor_id}

    selected_donor = None
    if donor_id:
        selected_donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
    elif selected_grant and selected_grant.donor_id:
        selected_donor = selected_grant.donor

    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="statement_of_financial_position.csv"'
        w = csv.writer(response)
        _grant_txt = (
            f"{selected_grant.title} ({selected_grant.code})" if selected_grant else "All grants"
        )
        _donor_txt = selected_donor.name if selected_donor else "—"
        _official_csv_preamble(
            w,
            request,
            "Statement of financial position (balance sheet)",
            [
                ("As at", str(f["as_at"])),
                ("Grant", _grant_txt),
                ("Donor", _donor_txt),
            ],
        )
        w.writerow(["Section", "Account Code", "Account Name", "Amount"])
        for code, name, bal in assets_current:
            w.writerow(["Current assets", code or "", name or "", str(bal or "0")])
        w.writerow(["Current assets", "", "Total current assets", str(total_current_assets)])
        for code, name, bal in assets_non_current:
            w.writerow(["Non-current assets", code or "", name or "", str(bal or "0")])
        w.writerow(["Non-current assets", "", "Total non-current assets", str(total_non_current_assets)])
        w.writerow(["Assets", "", "Total assets", str(total_assets)])
        for code, name, bal in liabilities_current:
            w.writerow(["Current liabilities", code or "", name or "", str(bal or "0")])
        w.writerow(["Current liabilities", "", "Total current liabilities", str(total_current_liabilities)])
        for code, name, bal in liabilities_non_current:
            w.writerow(["Non-current liabilities", code or "", name or "", str(bal or "0")])
        w.writerow(
            ["Non-current liabilities", "", "Total non-current liabilities", str(total_non_current_liabilities)]
        )
        for code, name, bal in fund_balance_rows:
            w.writerow(["Fund balance", code or "", name or "", str(bal or "0")])
        w.writerow(["Fund balance", "", "Total fund balance", str(total_fund_balance)])
        w.writerow(
            [
                "Liabilities and fund balance",
                "",
                "Total liabilities and fund balance",
                str(total_liabilities_and_fund_balance),
            ]
        )
        return response

    return render(
        request,
        "tenant_portal/finance/statement_financial_position.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_sfp",
            "filters": f,
            "assets_current": assets_current,
            "assets_non_current": assets_non_current,
            "liabilities_current": liabilities_current,
            "liabilities_non_current": liabilities_non_current,
            "fund_balance_rows": fund_balance_rows,
            "total_current_assets": total_current_assets,
            "total_non_current_assets": total_non_current_assets,
            "total_assets": total_assets,
            "total_current_liabilities": total_current_liabilities,
            "total_non_current_liabilities": total_non_current_liabilities,
            "total_fund_balance": total_fund_balance,
            "total_liabilities_and_fund_balance": total_liabilities_and_fund_balance,
            "grants": grants,
            "donors": donors,
            "selected_grant": selected_grant,
            "selected_donor": selected_donor,
            "grant_donor_map": grant_donor_map,
            "export_csv_url": _finance_export_csv_url(request),
            "report_validation": report_validation,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_statement_of_activities_view(request: HttpRequest) -> HttpResponse:
    import csv

    from django.contrib import messages
    from django.db.models import Q, Sum
    from django.http import HttpResponse
    from tenant_finance.models import ChartAccount
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
        user_sees_all_grants,
    )
    from tenant_finance.services.soa_layout import group_income_by_category, partition_expenses_by_bucket
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    grant_id = (f.get("grant_id") or "").strip()
    donor_id = (f.get("donor_id") or "").strip()

    gid_int = int(grant_id) if str(grant_id).isdigit() else None
    if gid_int and not assert_grant_filter_allowed(request.tenant_user, tenant_db, gid_int):
        messages.error(request, "You do not have access to the selected grant.")
        grant_id = ""
        gid_int = None
        f = {**f, "grant_id": ""}

    selected_grant = None
    if gid_int:
        selected_grant = (
            Grant.objects.using(tenant_db).select_related("donor").filter(pk=gid_int).first()
        )
        if not selected_grant:
            grant_id = ""
            gid_int = None
            f = {**f, "grant_id": ""}

    if grant_id and donor_id and selected_grant and str(selected_grant.donor_id) != str(donor_id):
        messages.warning(
            request,
            "The selected donor does not match this grant; the donor filter was ignored.",
        )
        donor_id = ""
        f = {**f, "donor_id": ""}

    if donor_id and not Donor.objects.using(tenant_db).filter(pk=donor_id).exists():
        donor_id = ""
        f = {**f, "donor_id": ""}

    line_qs = posted_journal_lines(tenant_db).filter(
        gl_date__gte=f["period_start"],
        gl_date__lte=f["period_end"],
        account__type__in=[ChartAccount.Type.INCOME, ChartAccount.Type.EXPENSE],
    )
    line_qs = restrict_journal_lines_by_grant_scope(line_qs, request.tenant_user, tenant_db)
    if gid_int:
        line_qs = line_qs.filter(Q(entry__grant_id=gid_int) | Q(grant_id=gid_int))
    if donor_id:
        did = int(donor_id)
        line_qs = line_qs.filter(Q(entry__donor_id=did) | Q(donor_id=did))

    qs = (
        line_qs.values(
            "account__type",
            "account__code",
            "account__name",
            "account__category__code",
            "account__category__name",
            "account__category__display_order",
        )
        .annotate(bal=Sum("debit") - Sum("credit"))
        .order_by("account__code")
    )

    rows = list(qs)
    income_raw = [r for r in rows if r["account__type"] == "income"]
    expense_raw = [r for r in rows if r["account__type"] == "expense"]

    income_sections = group_income_by_category(income_raw)
    total_income = sum(s["subtotal"] for s in income_sections)

    program_lines, administrative_lines, support_lines = partition_expenses_by_bucket(expense_raw)
    total_program = sum(x[2] for x in program_lines)
    total_administrative = sum(x[2] for x in administrative_lines)
    total_support = sum(x[2] for x in support_lines)
    total_expense = total_program + total_administrative + total_support

    expense_sections = [
        {"title": "Program expenses", "lines": program_lines, "subtotal": total_program},
        {"title": "Administrative expenses", "lines": administrative_lines, "subtotal": total_administrative},
        {"title": "Support costs", "lines": support_lines, "subtotal": total_support},
    ]

    surplus = total_income - total_expense
    surplus_abs = abs(surplus)

    report_validation = {
        "soa_result_ok": True,
        "soa_income_expense_delta": surplus,
    }

    grants = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db)
        .select_related("donor")
        .filter(status__in=[Grant.Status.ACTIVE, Grant.Status.CLOSED])
        .order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if user_sees_all_grants(request.tenant_user, tenant_db):
        donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    else:
        d_ids = [x for x in grants.values_list("donor_id", flat=True).distinct() if x]
        donors = (
            Donor.objects.using(tenant_db)
            .filter(pk__in=d_ids, status=Donor.Status.ACTIVE)
            .order_by("name")
            if d_ids
            else Donor.objects.using(tenant_db).none()
        )
    grant_donor_map = {str(g.pk): str(g.donor_id) for g in grants if g.donor_id}

    selected_donor = None
    if donor_id:
        selected_donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
    elif selected_grant and selected_grant.donor_id:
        selected_donor = selected_grant.donor

    if request.GET.get("format") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="statement_of_activities.csv"'
        w = csv.writer(response)
        _gtxt = f"{selected_grant.title} ({selected_grant.code})" if selected_grant else "All grants"
        _dtxt = selected_donor.name if selected_donor else "—"
        _official_csv_preamble(
            w,
            request,
            "Statement of activities (income and expenditure)",
            [
                ("Period", f"{f['period_start']} to {f['period_end']}"),
                ("Grant", _gtxt),
                ("Donor", _dtxt),
            ],
        )
        w.writerow(["Section", "Category", "Account Code", "Account Name", "Amount"])
        for sec in income_sections:
            for code, name, bal in sec["lines"]:
                w.writerow(["Income", sec["category_label"], code or "", name or "", str(bal or "0")])
            w.writerow(["Income", sec["category_label"], "", "Subtotal", str(sec["subtotal"])])
        w.writerow(["Income", "", "", "Total income", str(total_income)])
        for es in expense_sections:
            for code, name, bal in es["lines"]:
                w.writerow(["Expenditure", es["title"], code or "", name or "", str(bal or "0")])
            w.writerow(["Expenditure", es["title"], "", "Subtotal", str(es["subtotal"])])
        w.writerow(["Expenditure", "", "", "Total expenditure", str(total_expense)])
        w.writerow(["Result", "", "", "Surplus / (deficit)", str(surplus)])
        return response

    return render(
        request,
        "tenant_portal/finance/statement_of_activities.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_soa",
            "filters": f,
            "income_sections": income_sections,
            "expense_sections": expense_sections,
            "total_income": total_income,
            "total_expense": total_expense,
            "surplus": surplus,
            "surplus_abs": surplus_abs,
            "grants": grants,
            "donors": donors,
            "selected_grant": selected_grant,
            "selected_donor": selected_donor,
            "grant_donor_map": grant_donor_map,
            "export_csv_url": _finance_export_csv_url(request),
            "report_validation": report_validation,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_cash_flow_statement_view(request: HttpRequest) -> HttpResponse:
    import csv

    from decimal import Decimal

    from django.contrib import messages
    from django.db.models import Q, Sum
    from django.http import HttpResponse
    from tenant_finance.models import Currency, FinancialDimension
    from tenant_finance.services.cash_flow_statement import (
        cash_and_bank_chart_account_ids_extended,
        cash_roll_forward,
        compute_cash_flow_buckets,
    )
    from tenant_finance.services.financial_reporting import (
        assert_grant_filter_allowed,
        filter_grants_for_report_dropdown,
        posted_journal_lines,
        restrict_journal_lines_by_grant_scope,
        user_sees_all_grants,
    )
    from tenant_grants.models import Donor, Grant, Project

    tenant_db = request.tenant_db
    f = _parse_cfs_filters(request)
    grant_id = (f.get("grant_id") or "").strip()
    donor_id = (f.get("donor_id") or "").strip()
    currency_id = (f.get("currency_id") or "").strip()
    dimension_id = (f.get("dimension_id") or "").strip()
    project_id = (f.get("project_id") or "").strip()

    gid_int = int(grant_id) if str(grant_id).isdigit() else None
    if gid_int and not assert_grant_filter_allowed(request.tenant_user, tenant_db, gid_int):
        messages.error(request, "You do not have access to the selected grant.")
        grant_id = ""
        gid_int = None
        f = {**f, "grant_id": ""}

    selected_grant = None
    if gid_int:
        selected_grant = (
            Grant.objects.using(tenant_db).select_related("donor", "project").filter(pk=gid_int).first()
        )
        if not selected_grant:
            grant_id = ""
            gid_int = None
            f = {**f, "grant_id": ""}

    if grant_id and donor_id and selected_grant and str(selected_grant.donor_id) != str(donor_id):
        messages.warning(
            request,
            "The selected donor does not match this grant; the donor filter was ignored.",
        )
        donor_id = ""
        f = {**f, "donor_id": ""}

    if donor_id and not Donor.objects.using(tenant_db).filter(pk=donor_id).exists():
        donor_id = ""
        f = {**f, "donor_id": ""}

    cid_int = int(currency_id) if str(currency_id).isdigit() else None
    if cid_int and not Currency.objects.using(tenant_db).filter(pk=cid_int).exists():
        cid_int = None
        currency_id = ""
        f = {**f, "currency_id": ""}

    dim_int = int(dimension_id) if str(dimension_id).isdigit() else None
    if dim_int and not FinancialDimension.objects.using(tenant_db).filter(pk=dim_int).exists():
        dim_int = None
        dimension_id = ""
        f = {**f, "dimension_id": ""}

    proj_int = int(project_id) if str(project_id).isdigit() else None
    if proj_int and not Project.objects.using(tenant_db).filter(pk=proj_int).exists():
        proj_int = None
        project_id = ""
        f = {**f, "project_id": ""}

    cash_ids = cash_and_bank_chart_account_ids_extended(tenant_db)
    cash_scope_ok = bool(cash_ids)

    base = posted_journal_lines(tenant_db)
    base = restrict_journal_lines_by_grant_scope(base, request.tenant_user, tenant_db)
    if gid_int:
        base = base.filter(Q(entry__grant_id=gid_int) | Q(grant_id=gid_int))
    if donor_id:
        did = int(donor_id)
        base = base.filter(Q(entry__donor_id=did) | Q(donor_id=did))
    if cid_int:
        base = base.filter(entry__currency_id=cid_int)
    if dim_int:
        base = base.filter(entry__dimension_id=dim_int)
    if proj_int:
        base = base.filter(Q(entry__grant__project_id=proj_int) | Q(grant__project_id=proj_int))

    opening_cash, closing_cash, period_cash_movement = cash_roll_forward(
        base_journal_lines=base,
        cash_account_ids=cash_ids,
        period_start=f["period_start"],
        period_end=f["period_end"],
    )

    bucket_totals, bucket_details, net_from_buckets = compute_cash_flow_buckets(
        tenant_db=tenant_db,
        period_start=f["period_start"],
        period_end=f["period_end"],
        cash_account_ids=cash_ids,
        base_journal_lines=base,
    )

    cfs_net_operating = bucket_totals["operating"]
    cfs_net_investing = bucket_totals["investing"]
    cfs_net_financing = bucket_totals["financing"]
    net_cash = net_from_buckets

    cfs_roll_forward_delta = (opening_cash + period_cash_movement) - closing_cash
    cfs_bucket_delta = period_cash_movement - net_from_buckets
    cfs_recon_delta = (opening_cash + net_from_buckets) - closing_cash
    eps = Decimal("0.02")
    report_validation = {
        "cfs_cash_scope_ok": cash_scope_ok,
        "cfs_roll_forward_ok": (not cash_scope_ok) or (abs(cfs_roll_forward_delta) <= eps),
        "cfs_bucket_ok": (not cash_scope_ok) or (abs(cfs_bucket_delta) <= eps),
        "cfs_recon_ok": (not cash_scope_ok) or (abs(cfs_recon_delta) <= eps),
        "cfs_recon_delta": cfs_recon_delta,
        "cfs_roll_forward_delta": cfs_roll_forward_delta,
        "cfs_bucket_delta": cfs_bucket_delta,
        "opening_cash": opening_cash,
        "closing_cash": closing_cash,
        "period_cash_movement": period_cash_movement,
    }

    grants = filter_grants_for_report_dropdown(
        Grant.objects.using(tenant_db)
        .select_related("donor", "project")
        .filter(status__in=[Grant.Status.ACTIVE, Grant.Status.CLOSED])
        .order_by("code"),
        request.tenant_user,
        tenant_db,
    )
    if user_sees_all_grants(request.tenant_user, tenant_db):
        donors = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    else:
        d_ids = [x for x in grants.values_list("donor_id", flat=True).distinct() if x]
        donors = (
            Donor.objects.using(tenant_db)
            .filter(pk__in=d_ids, status=Donor.Status.ACTIVE)
            .order_by("name")
            if d_ids
            else Donor.objects.using(tenant_db).none()
        )
    grant_donor_map = {str(g.pk): str(g.donor_id) for g in grants if g.donor_id}

    selected_donor = None
    if donor_id:
        selected_donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
    elif selected_grant and selected_grant.donor_id:
        selected_donor = selected_grant.donor

    currencies = Currency.objects.using(tenant_db).filter(status=Currency.Status.ACTIVE).order_by("code")
    dimensions = (
        FinancialDimension.objects.using(tenant_db)
        .filter(status=FinancialDimension.Status.ACTIVE)
        .order_by("dimension_code")[:200]
    )
    project_ids = [x for x in grants.values_list("project_id", flat=True).distinct() if x]
    projects = (
        Project.objects.using(tenant_db).filter(pk__in=project_ids).order_by("code")
        if project_ids
        else Project.objects.using(tenant_db).none()
    )
    selected_project = (
        Project.objects.using(tenant_db).filter(pk=proj_int).first() if proj_int else None
    )
    selected_currency = (
        Currency.objects.using(tenant_db).filter(pk=cid_int).first() if cid_int else None
    )
    selected_dimension = (
        FinancialDimension.objects.using(tenant_db).filter(pk=dim_int).first()
        if dim_int
        else None
    )

    export_fmt = (request.GET.get("format") or "").lower()
    if export_fmt in ("csv", "xlsx", "pdf"):
        _gtxt = f"{selected_grant.title} ({selected_grant.code})" if selected_grant else "All grants"
        _dtxt = selected_donor.name if selected_donor else "—"
        _ctxt = selected_currency.code if selected_currency else "All"
        _prj = f"{selected_project.code} — {selected_project.name}" if selected_project else "All"
        _dim = (
            f"{selected_dimension.dimension_code} — {selected_dimension.dimension_name}"
            if selected_dimension
            else "All"
        )
        meta = [
            ("Period", f"{f['period_start']} to {f['period_end']}"),
            ("Grant", _gtxt),
            ("Donor", _dtxt),
            ("Project", _prj),
            ("Currency", _ctxt),
            ("Fund / analysis (dimension)", _dim),
        ]
        headers = ["Section", "Date", "Reference", "Description", "Amount"]
        rows: list[list] = []

        def _add_section(title: str, subrows: list[dict], subtotal: Decimal):
            for r in subrows:
                gd = r.get("gl_date")
                ds = gd.isoformat() if gd else ""
                rows.append(
                    [
                        title,
                        ds,
                        r.get("label") or "",
                        (r.get("memo") or "")[:200],
                        str(r.get("amount") or "0"),
                    ]
                )
            rows.append([title, "", "", f"Net {title.lower()} cash flows", str(subtotal)])

        _add_section("Operating activities", bucket_details["operating"], cfs_net_operating)
        _add_section("Investing activities", bucket_details["investing"], cfs_net_investing)
        _add_section("Financing activities", bucket_details["financing"], cfs_net_financing)
        rows.append(["Summary", "", "", "Opening cash and cash equivalents", str(opening_cash)])
        rows.append(["Summary", "", "", "Net increase/(decrease) in cash (period)", str(period_cash_movement)])
        rows.append(["Summary", "", "", "Closing cash and cash equivalents", str(closing_cash)])
        rows.append(
            [
                "Check",
                "",
                "",
                "Opening + net change − closing (should be 0)",
                str(cfs_recon_delta),
            ]
        )

        if export_fmt == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="cash_flow_statement.csv"'
            w = csv.writer(response)
            _official_csv_preamble(w, request, "Statement of cash flows (IFRS-style)", meta)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
            return response

        return _export_table_response(
            export_format=export_fmt,
            filename_base="cash_flow_statement",
            title="Statement of cash flows",
            headers=headers,
            rows=rows,
            request=request,
            include_official_header=True,
        )

    export_urls = _cfs_export_urls(request)

    return render(
        request,
        "tenant_portal/finance/cash_flow_statement.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_cfs",
            "filters": f,
            "cfs_operating_entries": bucket_details["operating"],
            "cfs_investing_entries": bucket_details["investing"],
            "cfs_financing_entries": bucket_details["financing"],
            "cfs_net_operating": cfs_net_operating,
            "cfs_net_investing": cfs_net_investing,
            "cfs_net_financing": cfs_net_financing,
            "net_cash": net_cash,
            "period_cash_movement": period_cash_movement,
            "grants": grants,
            "donors": donors,
            "currencies": currencies,
            "dimensions": dimensions,
            "projects": projects,
            "selected_grant": selected_grant,
            "selected_donor": selected_donor,
            "selected_project": selected_project,
            "selected_currency": selected_currency,
            "selected_dimension": selected_dimension,
            "grant_donor_map": grant_donor_map,
            "export_csv_url": export_urls.get("csv", ""),
            "export_xlsx_url": export_urls.get("xlsx", ""),
            "export_pdf_url": export_urls.get("pdf", ""),
            "report_validation": report_validation,
            "official_report_period_line": f"Period: {f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_opening_balances_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal, InvalidOperation

    from django.contrib import messages
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.utils.dateparse import parse_date
    from tenant_finance.models import ChartAccount, OpeningBalance
    from tenant_finance.services.financial_reporting import (
        fiscal_year_containing,
        is_balanced_debit_credit,
        opening_balances_totals,
    )

    tenant_db = request.tenant_db
    if request.method == "POST":
        as_of = parse_date(request.POST.get("as_of_date") or "")
        account_id = request.POST.get("account_id")
        debit = request.POST.get("debit") or "0"
        credit = request.POST.get("credit") or "0"
        if as_of and account_id:
            try:
                dr = Decimal(debit)
                cr = Decimal(credit)
            except (InvalidOperation, ValueError):
                messages.error(request, "Invalid amounts.")
            else:
                if dr < 0 or cr < 0:
                    messages.error(request, "Debit and credit amounts cannot be negative.")
                elif dr > 0 and cr > 0:
                    messages.error(
                        request,
                        "Enter either a debit or a credit for opening balance, not both.",
                    )
                else:
                    fy = fiscal_year_containing(tenant_db, as_of)
                    if fy:
                        OpeningBalance.objects.using(tenant_db).filter(
                            account_id=account_id,
                            as_of_date__gte=fy.start_date,
                            as_of_date__lte=fy.end_date,
                        ).delete()
                        OpeningBalance.objects.using(tenant_db).create(
                            account_id=account_id,
                            as_of_date=as_of,
                            debit=dr,
                            credit=cr,
                        )
                    else:
                        OpeningBalance.objects.using(tenant_db).update_or_create(
                            account_id=account_id,
                            as_of_date=as_of,
                            defaults={"debit": dr, "credit": cr},
                        )
                    td, tc = opening_balances_totals(tenant_db)
                    if not is_balanced_debit_credit(td, tc):
                        messages.warning(
                            request,
                            f"Opening balances are not in balance (total debits {td} vs credits {tc}). "
                            "Adjust other accounts so total debits equal total credits.",
                        )
                    else:
                        messages.success(request, "Opening balance saved.")
        return redirect(reverse("tenant_portal:finance_opening_balances"))
    balances = OpeningBalance.objects.using(tenant_db).select_related("account").order_by(
        "as_of_date", "account__code"
    )
    accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code")
    td, tc = opening_balances_totals(tenant_db)
    return render(
        request,
        "tenant_portal/finance/opening_balances.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "balances": balances,
            "accounts": accounts,
            "active_submenu": "core",
            "active_item": "core_fs",
            "opening_balance_total_debit": td,
            "opening_balance_total_credit": tc,
            "opening_balances_balanced": is_balanced_debit_credit(td, tc),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_audit_trail_view(request: HttpRequest) -> HttpResponse:
    from tenant_finance.models import AuditLog

    tenant_db = request.tenant_db
    from rbac.models import user_has_permission as _uhp
    cached = getattr(request, "rbac_permission_codes", None)
    if not (isinstance(cached, set) and ("*" in cached or "finance:audit.view" in cached)) and not _uhp(
        request.tenant_user, "finance:audit.view", using=tenant_db
    ):
        return render(
            request,
            "tenant_portal/forbidden.html",
            {"tenant": request.tenant, "tenant_user": request.tenant_user, "reason": "You do not have permission to view audit logs."},
            status=403,
        )
    f = _parse_finance_filters(request)
    logs = AuditLog.objects.using(tenant_db).order_by("-changed_at")
    mn = (request.GET.get("model_name") or "").strip().lower()
    oid = request.GET.get("object_id") or ""
    if mn and oid.isdigit():
        logs = logs.filter(model_name=mn, object_id=int(oid))
    logs = logs[:200]
    if request.GET.get("format") == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="audit_trail.csv"'
        w = csv.writer(response)
        w.writerow(["When", "Action", "Model", "Object ID", "User", "Summary"])
        for log in logs:
            w.writerow([log.changed_at, log.get_action_display(), log.model_name, log.object_id, log.username or "", log.summary or ""])
        return response
    filter_model = mn or ""
    filter_object_id = int(oid) if oid.isdigit() else None
    return render(
        request,
        "tenant_portal/finance/audit_trail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_audit",
            "filters": f,
            "logs": logs,
            "export_csv_url": _finance_export_csv_url(request),
            "audit_filter_model": filter_model,
            "audit_filter_object_id": filter_object_id,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="finance:journals.view")
def finance_journal_approval_view(request: HttpRequest) -> HttpResponse:
    """
    Enterprise journal approval queue: draft → submitted (pending) → approved → posted.
    Actions POST to finance_journal_action (approve / reject / post) with next=approval.
    """
    from decimal import Decimal, InvalidOperation

    from django.db.models import Q, Sum

    from rbac.models import user_has_permission as _user_has_permission
    from tenant_finance.models import FiscalPeriod, JournalEntry
    from tenant_grants.models import Grant
    from tenant_users.models import TenantUser

    tenant_db = request.tenant_db
    cached = getattr(request, "rbac_permission_codes", None)

    def _has(code: str) -> bool:
        if isinstance(cached, set):
            return ("*" in cached) or (code in cached)
        return _user_has_permission(request.tenant_user, code, using=tenant_db)

    if not (_has("finance:journals.approve") or _has("finance:journals.post")):
        return render(
            request,
            "tenant_portal/forbidden.html",
            {
                "tenant": request.tenant,
                "tenant_user": request.tenant_user,
                "reason": "You need journal approval or posting permission to use this page.",
            },
            status=403,
        )

    can_approve = _has("finance:journals.approve")
    can_post = _has("finance:journals.post")
    can_override_maker = _has("finance:journals.override_maker_checker") or _has("finance:maker_checker.override")

    f = _parse_finance_filters(request)
    accounting_periods_all = list(
        FiscalPeriod.objects.using(tenant_db).select_related("fiscal_year").order_by("-start_date")
    )

    qs = (
        JournalEntry.objects.using(tenant_db)
        .exclude(reference__startswith="PV-")
        .select_related("grant", "created_by", "submitted_by", "approved_by")
    )

    # Status filter (default: submitted / pending approval only)
    if "status" not in request.GET:
        qs = qs.filter(status=JournalEntry.Status.PENDING_APPROVAL)
    else:
        st = (request.GET.get("status") or "").strip()
        if st == "all":
            pass
        elif st:
            qs = qs.filter(status=st)
        else:
            qs = qs.filter(status=JournalEntry.Status.PENDING_APPROVAL)

    # Journal category (general / recurring / adjusting / system)
    cat = (request.GET.get("journal_category") or "").strip()
    if cat == "system":
        qs = qs.filter(is_system_generated=True)
    elif cat == "adjusting":
        qs = qs.filter(is_system_generated=False).filter(
            Q(journal_type__iexact="adjusting_journal") | Q(journal_type__iexact="adjustment")
        )
    elif cat == "recurring":
        qs = qs.filter(is_system_generated=False).filter(journal_type__icontains="recurring")
    elif cat == "general":
        qs = qs.filter(is_system_generated=False).exclude(
            Q(journal_type__iexact="adjusting_journal") | Q(journal_type__iexact="adjustment")
        ).exclude(journal_type__icontains="recurring")

    qs = qs.filter(entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])

    if f["grant_id"]:
        qs = qs.filter(grant_id=f["grant_id"])

    created_by_id = (request.GET.get("created_by_id") or "").strip()
    if created_by_id.isdigit():
        qs = qs.filter(created_by_id=int(created_by_id))

    qs = qs.annotate(total_amount=Sum("lines__debit")).distinct()

    amt_min_s = (request.GET.get("amount_min") or "").strip()
    amt_max_s = (request.GET.get("amount_max") or "").strip()
    try:
        if amt_min_s:
            qs = qs.filter(total_amount__gte=Decimal(amt_min_s))
    except (InvalidOperation, ValueError, TypeError):
        pass
    try:
        if amt_max_s:
            qs = qs.filter(total_amount__lte=Decimal(amt_max_s))
    except (InvalidOperation, ValueError, TypeError):
        pass

    qs = qs.order_by("-entry_date", "-id")[:350]

    rows = []
    for entry in qs:
        gl_date = _finance_journal_gl_date(entry)
        posted_at = getattr(entry, "posted_at", None)
        posting_display = gl_date
        if getattr(entry, "posting_date", None):
            posting_display = entry.posting_date
        elif posted_at:
            posting_display = posted_at.date()
        acct_lbl = _accounting_period_label_for_date(accounting_periods_all, posting_display)
        cat_code = _journal_enterprise_category(entry)
        cat_labels = {
            "general": "General",
            "recurring": "Recurring",
            "adjusting": "Adjusting",
            "system": "System",
        }
        total_amt = getattr(entry, "total_amount", None) or Decimal("0")
        rows.append(
            {
                "entry": entry,
                "journal_no": _journal_number_display(entry),
                "category": cat_code,
                "category_label": cat_labels.get(cat_code, cat_code.title()),
                "posting_date": posting_display,
                "accounting_period_label": acct_lbl,
                "total_amount": total_amt,
                "can_approve_row": can_approve
                and entry.status == JournalEntry.Status.PENDING_APPROVAL
                and (
                    entry.created_by_id != getattr(request.tenant_user, "id", None) or can_override_maker
                ),
                "can_reject_row": can_approve and entry.status == JournalEntry.Status.PENDING_APPROVAL,
                "can_post_row": can_post and entry.status == JournalEntry.Status.APPROVED
                and (
                    entry.created_by_id != getattr(request.tenant_user, "id", None) or can_override_maker
                ),
            }
        )

    grants = Grant.objects.using(tenant_db).filter(status="active").order_by("code")
    created_by_users = (
        TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("full_name", "email")
    )

    if request.GET.get("format") == "csv":
        import csv

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="journal_approval.csv"'
        w = csv.writer(response)
        w.writerow(
            [
                "Journal no",
                "Type",
                "Journal date",
                "Posting date",
                "Accounting period",
                "Reference",
                "Description",
                "Project / grant",
                "Total amount",
                "Created by",
                "Submitted at",
                "Submitted by",
                "Status",
                "Approved by",
                "Approved at",
            ]
        )
        for r in rows:
            e = r["entry"]
            w.writerow(
                [
                    r["journal_no"],
                    r["category_label"],
                    e.entry_date,
                    r["posting_date"],
                    r["accounting_period_label"],
                    e.reference or "",
                    e.memo or "",
                    getattr(e.grant, "code", "") if e.grant_id else "",
                    r["total_amount"],
                    e.created_by.get_full_name() if e.created_by else "",
                    getattr(e, "submitted_at", None) or "",
                    e.submitted_by.get_full_name() if getattr(e, "submitted_by_id", None) else "",
                    e.get_status_display(),
                    e.approved_by.get_full_name() if e.approved_by else "",
                    getattr(e, "approved_at", None) or "",
                ]
            )
        return response

    return render(
        request,
        "tenant_portal/finance/journal_approval.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "filters": f,
            "journal_category": (request.GET.get("journal_category") or "").strip(),
            "status_filter": (request.GET.get("status") or "").strip(),
            "amount_min": (request.GET.get("amount_min") or "").strip(),
            "amount_max": (request.GET.get("amount_max") or "").strip(),
            "created_by_id": (request.GET.get("created_by_id") or "").strip(),
            "grants": grants,
            "created_by_users": created_by_users,
            "journal_statuses": JournalEntry.Status,
            "can_approve": can_approve,
            "can_post": can_post,
            "export_csv_url": _finance_export_csv_url(request),
            "active_submenu": "core",
            "active_item": "core_approval",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_home_view(request: HttpRequest) -> HttpResponse:
    """Funds & Donors module home: dashboard with KPIs and navigation to grant workflows."""
    from datetime import date, timedelta
    from decimal import Decimal

    from django.db.models import Sum
    from django.utils.translation import gettext as _

    tenant_db = request.tenant_db
    total_donors = 0
    active_grants_n = 0
    active_projects_n = 0
    total_budget_dec = Decimal("0")
    utilized_dec = Decimal("0")
    remaining_dec = Decimal("0")
    expiring_soon_n = 0

    def _fmt_money(d: Decimal) -> str:
        d = d.quantize(Decimal("0.01"))
        neg = d < 0
        d = abs(d)
        whole_s, _, frac = f"{d:.2f}".partition(".")
        whole_s = "{:,}".format(int(whole_s))
        return ("-" if neg else "") + whole_s + "." + frac

    try:
        from tenant_finance.models import ChartAccount, JournalLine
        from tenant_grants.models import BudgetLine, Donor, Grant, Project

        total_donors = Donor.objects.using(tenant_db).count()
        active_grants_qs = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE)
        active_grants_n = active_grants_qs.count()
        active_projects_n = Project.objects.using(tenant_db).filter(status=Project.Status.ACTIVE).count()

        budgets_by_grant = {
            r["grant_id"]: r["total"] or Decimal("0")
            for r in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
        }
        spend_by_grant = {
            r["entry__grant_id"]: r["spent"] or Decimal("0")
            for r in JournalLine.objects.using(tenant_db)
            .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
            .values("entry__grant_id")
            .annotate(spent=Sum("debit"))
        }

        total_ceiling = Decimal("0")
        utilized_dec = Decimal("0")
        for g in active_grants_qs.only("id", "award_amount"):
            b = budgets_by_grant.get(g.id, Decimal("0"))
            ceiling = b if b > 0 else Decimal(str(g.award_amount or 0))
            total_ceiling += ceiling
            utilized_dec += spend_by_grant.get(g.id, Decimal("0"))

        total_budget_dec = total_ceiling
        remaining_dec = total_ceiling - utilized_dec

        today = date.today()
        horizon = today + timedelta(days=90)
        for g in active_grants_qs.only("revised_end_date", "original_end_date", "end_date"):
            ed = g.effective_end_date()
            if ed and today <= ed <= horizon:
                expiring_soon_n += 1
    except Exception:
        pass

    funds_kpis = [
        {"label": _("Total donors"), "value": f"{total_donors:,}"},
        {"label": _("Active grants"), "value": f"{active_grants_n:,}"},
        {"label": _("Active projects"), "value": f"{active_projects_n:,}"},
        {"label": _("Total grant budget"), "value": _fmt_money(total_budget_dec)},
        {"label": _("Utilized budget"), "value": _fmt_money(utilized_dec)},
        {"label": _("Remaining budget"), "value": _fmt_money(remaining_dec)},
        {"label": _("Expiring soon (90d)"), "value": f"{expiring_soon_n:,}"},
    ]

    return render(
        request,
        "tenant_portal/grants/funds_donors_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_center_home",
            "funds_kpis": funds_kpis,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_donors_view(request: HttpRequest) -> HttpResponse:
    import re
    from tenant_grants.models import Donor

    tenant_db = request.tenant_db
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        if not name:
            messages.error(request, "Donor name is required.")
        else:
            base = re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-")[:40].upper() or "DONOR"
            code = base
            n = 0
            while Donor.objects.using(tenant_db).filter(code__iexact=code).exists():
                n += 1
                code = f"{base}-{n}"
            Donor.objects.using(tenant_db).create(
                code=code, name=name, email=email, status=Donor.Status.ACTIVE
            )
            messages.success(request, "Donor created.")
            redir = reverse("tenant_portal:grants_donors")
            if (request.POST.get("report_nav") or "").strip() == "1":
                redir = f"{redir}?report_nav=1"
            return redirect(redir)

    donors = list(_active_donors_queryset(tenant_db))
    from_reporting = (request.GET.get("report_nav") or "").strip() == "1"
    ctx = {
        "tenant": request.tenant,
        "tenant_user": request.tenant_user,
        "donors": donors,
        "active_submenu": "reports" if from_reporting else "funds",
    }
    if from_reporting:
        ctx["active_item"] = "report_donor_summary"
    return render(request, "tenant_portal/grants/donors.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grants_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date

    from tenant_finance.models import BankAccount
    from tenant_grants.models import Donor, Grant

    tenant_db = request.tenant_db

    # Program/Project manage permission (create/edit); others are read-only.
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)

    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to create or edit grants.")
            return redirect(reverse("tenant_portal:grants_grants"))

        code = (request.POST.get("code") or "").strip()
        title = (request.POST.get("title") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        bank_account_id = (request.POST.get("bank_account_id") or "").strip()
        raw_award_amount = (request.POST.get("award_amount") or "").replace(",", "").strip()
        start_date = parse_date(request.POST.get("start_date") or "")
        end_date = parse_date(request.POST.get("end_date") or "")

        errors: list[str] = []

        if not code:
            errors.append("Grant code is required.")
        if not title:
            errors.append("Grant title is required.")
        if not donor_id:
            errors.append("Donor is required.")
        if not bank_account_id:
            errors.append("Bank account is required.")

        # Uniqueness of grant code within tenant DB
        if code:
            if Grant.objects.using(tenant_db).filter(code__iexact=code).exists():
                errors.append("Grant code must be unique.")

        # Validate donor existence
        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if not donor:
            errors.append("Selected donor does not exist.")

        # Validate bank account (active only)
        bank_account = (
            BankAccount.objects.using(tenant_db)
            .filter(pk=bank_account_id, is_active=True)
            .first()
            if bank_account_id
            else None
        )
        if not bank_account:
            errors.append("Selected bank account is invalid or inactive.")

        # Validate award amount > 0
        award_amount = Decimal("0")
        if raw_award_amount:
            try:
                award_amount = Decimal(raw_award_amount)
            except (InvalidOperation, ValueError):
                errors.append("Grant amount must be a valid number.")
        if award_amount <= 0:
            errors.append("Grant amount must be greater than zero.")

        # Validate dates
        if not start_date:
            errors.append("Start date is required.")
        if end_date and start_date and end_date <= start_date:
            errors.append("End date must be later than start date.")

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            signed_date = parse_date(request.POST.get("signed_date") or "")
            reporting_rules = (request.POST.get("reporting_rules") or "").strip()
            donor_restrictions = (request.POST.get("donor_restrictions") or "").strip()
            grant = Grant.objects.using(tenant_db).create(
                code=code,
                title=title,
                donor=donor,
                bank_account=bank_account,
                status=Grant.Status.ACTIVE,
                award_amount=award_amount,
                start_date=start_date,
                end_date=end_date,
                signed_date=signed_date or None,
                reporting_rules=reporting_rules,
                donor_restrictions=donor_restrictions,
            )
            if request.FILES.get("signed_contract_document"):
                grant.signed_contract_document = request.FILES["signed_contract_document"]
                grant.save(using=tenant_db)
            messages.success(request, "Grant agreement created.")
            return redirect(reverse("tenant_portal:grants_grants"))

    f = _parse_grants_filters(request)
    qs = Grant.objects.using(tenant_db).select_related("donor").order_by("-created_at")
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        qs = qs.filter(pk=f["grant_id"])
    if f.get("q"):
        from django.db.models import Q
        qs = qs.filter(
            Q(code__icontains=f["q"]) | Q(title__icontains=f["q"]) | Q(project_name__icontains=f["q"])
        )
    grants = list(qs.select_related("bank_account")[:200])
    donors = list(_active_donors_queryset(tenant_db))
    bank_accounts = BankAccount.objects.using(tenant_db).filter(is_active=True).order_by("bank_name", "account_name")
    export_format = request.GET.get("format") or ""
    if export_format:
        rows = [
            [
                g.code,
                g.title,
                g.donor.name if g.donor else "",
                g.bank_account.account_name if g.bank_account else "",
                g.award_amount,
                g.start_date or "",
                g.end_date or "",
                g.status,
            ]
            for g in grants
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="grant_agreements",
            title="Grant Agreements",
            headers=["Code", "Title", "Donor", "Amount", "Start date", "End date", "Status"],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/grants.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "grants": grants,
            "donors": donors,
            "bank_accounts": bank_accounts,
            "filters": f,
            "active_submenu": "funds",
            "active_item": "funds_grant_agreements",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
            "can_manage": can_manage,
        },
    )


def _active_donors_queryset(tenant_db):
    """Donors with status=active for grant/agreement/tracking dropdowns; inactive/archived remain for history only."""
    from tenant_grants.models import Donor
    return Donor.active.using(tenant_db).order_by("name")


def _parse_grants_filters(request):
    """Parse donor_id, grant_id, project_id, deadline_status, period for Funds & Donors pages."""
    from django.utils.dateparse import parse_date
    from django.utils import timezone
    today = timezone.now().date()
    period_start = parse_date(request.GET.get("period_start") or "")
    period_end = parse_date(request.GET.get("period_end") or "")
    if not period_start:
        period_start = today.replace(day=1)
    if not period_end:
        period_end = today
    return {
        "period_start": period_start,
        "period_end": period_end,
        "donor_id": request.GET.get("donor_id") or "",
        "grant_id": request.GET.get("grant_id") or "",
        "project_id": request.GET.get("project_id") or "",
        "deadline_status": (request.GET.get("deadline_status") or "").strip().lower(),
        "q": (request.GET.get("q") or "").strip(),
    }


def _grants_export_url(request):
    from urllib.parse import urlencode
    q = request.GET.copy()
    q["format"] = "csv"
    return request.path + "?" + q.urlencode()


def _grants_export_urls(request):
    q = request.GET.copy()
    q["format"] = "csv"
    csv_url = request.path + "?" + q.urlencode()
    q["format"] = "xlsx"
    xlsx_url = request.path + "?" + q.urlencode()
    q["format"] = "pdf"
    pdf_url = request.path + "?" + q.urlencode()
    return {"csv": csv_url, "xlsx": xlsx_url, "pdf": pdf_url}


def _export_table_response(
    *,
    export_format: str,
    filename_base: str,
    title: str,
    headers: list[str],
    rows: list[list],
    request: HttpRequest | None = None,
    include_official_header: bool = False,
):
    """
    Export a simple tabular report to CSV/XLSX/PDF.
    rows: list of row lists aligned to headers.
    When include_official_header and request are set, prepend tenant legal name (and logo on PDF)
    for official donor/financial-style exports.
    """
    export_format = (export_format or "").lower()
    legal = _tenant_legal_name_for_official_reports(request) if (include_official_header and request) else ""

    if export_format == "csv":
        import csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        w = csv.writer(response)
        if legal:
            _official_csv_preamble(w, request, title)
        w.writerow(headers)
        for r in rows:
            w.writerow([("" if v is None else str(v)) for v in r])
        return response

    if export_format == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        if legal:
            ws.append([legal])
            ws.append([title])
            ws.append([])
        ws.append(headers)
        for r in rows:
            ws.append([("" if v is None else v) for v in r])
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    if export_format == "pdf":
        import html
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
        styles = getSampleStyleSheet()
        story = []
        if include_official_header and request:
            logo = _official_report_pdf_logo_flowable(request)
            if logo:
                story.append(logo)
                story.append(Spacer(1, 10))
            if legal:
                story.append(Paragraph(html.escape(legal), styles["Title"]))
                story.append(Spacer(1, 6))
        story.append(Paragraph(html.escape(title), styles["Heading2"] if legal else styles["Title"]))
        story.append(Spacer(1, 12))
        data = [headers] + [[("" if v is None else str(v)) for v in r] for r in rows]
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5F0FF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
                ]
            )
        )
        story.append(tbl)
        doc.build(story)
        pdf = buf.getvalue()
        buf.close()
        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.pdf"'
        return response

    return None


def _parse_donor_register_filters(request):
    """Parse q, donor_type, donor_category, status for Donor Register."""
    return {
        "q": (request.GET.get("q") or "").strip(),
        "donor_type": request.GET.get("donor_type") or "",
        "donor_category": request.GET.get("donor_category") or "",
        "status": request.GET.get("status") or "",
    }


def _donor_register_export_urls(request):
    """Export URLs for Donor Register (preserve filter params)."""
    q = request.GET.copy()
    q["format"] = "csv"
    csv_url = request.path + "?" + q.urlencode()
    q["format"] = "xlsx"
    xlsx_url = request.path + "?" + q.urlencode()
    q["format"] = "pdf"
    pdf_url = request.path + "?" + q.urlencode()
    return {"csv": csv_url, "xlsx": xlsx_url, "pdf": pdf_url}


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_register_view(request: HttpRequest) -> HttpResponse:
    """Donor Register: master donor setup with search, filters, export, and create modal."""
    from django.db.models import Count, Q
    from tenant_grants.models import Donor

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)

    if request.method == "POST" and can_manage:
        code = (request.POST.get("code") or "").strip().upper()
        name = (request.POST.get("name") or "").strip()
        short_name = (request.POST.get("short_name") or "").strip()
        donor_type = (request.POST.get("donor_type") or "").strip() or None
        donor_category = (request.POST.get("donor_category") or "").strip() or None
        contact_person = (request.POST.get("contact_person") or "").strip()
        email = (request.POST.get("email") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        address = (request.POST.get("address") or "").strip()
        country = (request.POST.get("country") or "").strip()
        website = (request.POST.get("website") or "").strip()
        preferred_currency = (request.POST.get("preferred_currency") or "").strip()
        default_restriction_type = (request.POST.get("default_restriction_type") or "").strip() or None
        default_reporting_frequency = (request.POST.get("default_reporting_frequency") or "").strip() or None
        status = (request.POST.get("status") or Donor.Status.ACTIVE).strip()
        notes = (request.POST.get("notes") or "").strip()

        errors = []
        if not code:
            errors.append("Donor code is required.")
        if not name:
            errors.append("Donor name is required.")
        if code and Donor.objects.using(tenant_db).filter(code__iexact=code).exists():
            errors.append("Donor code already exists.")
        if name and Donor.objects.using(tenant_db).filter(name__iexact=name).exists():
            errors.append("Donor name already exists.")
        if status and status not in (Donor.Status.ACTIVE, Donor.Status.INACTIVE, Donor.Status.ARCHIVED):
            status = Donor.Status.ACTIVE

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            Donor.objects.using(tenant_db).create(
                code=code,
                name=name,
                short_name=short_name,
                donor_type=donor_type,
                donor_category=donor_category,
                contact_person=contact_person,
                email=email,
                phone=phone,
                address=address,
                country=country,
                website=website or "",
                preferred_currency=preferred_currency,
                default_restriction_type=default_restriction_type,
                default_reporting_frequency=default_reporting_frequency,
                status=status,
                notes=notes,
            )
            messages.success(request, "Donor created.")
            return redirect(reverse("tenant_portal:grants_donor_register"))

    f = _parse_donor_register_filters(request)
    qs = Donor.objects.using(tenant_db).annotate(grants_count=Count("grants")).order_by("name")
    if f["q"]:
        qs = qs.filter(
            Q(code__icontains=f["q"])
            | Q(name__icontains=f["q"])
            | Q(short_name__icontains=f["q"])
            | Q(contact_person__icontains=f["q"])
            | Q(email__icontains=f["q"])
            | Q(country__icontains=f["q"])
        )
    if f["donor_type"]:
        qs = qs.filter(donor_type=f["donor_type"])
    if f["donor_category"]:
        qs = qs.filter(donor_category=f["donor_category"])
    if f["status"]:
        qs = qs.filter(status=f["status"])
    donors = list(qs[:500])
    donor_register_kpi = {
        "total_shown": len(donors),
        "active_shown": sum(1 for d in donors if d.status == Donor.Status.ACTIVE),
        "grants_linked": sum(int(getattr(d, "grants_count", 0) or 0) for d in donors),
    }
    donors_truncated = len(donors) >= 500

    export_format = request.GET.get("format") or ""
    if export_format:
        rows = [
            [
                d.code,
                d.name,
                d.short_name or "",
                d.get_donor_type_display() if d.donor_type else "",
                d.get_donor_category_display() if d.donor_category else "",
                d.contact_person or "",
                d.email or "",
                d.phone or "",
                d.country or "",
                d.preferred_currency or "",
                d.get_default_restriction_type_display() if d.default_restriction_type else "",
                d.get_default_reporting_frequency_display() if d.default_reporting_frequency else "",
                d.get_status_display(),
                getattr(d, "grants_count", 0),
            ]
            for d in donors
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="donor_register",
            title="Donor Register",
            headers=[
                "Code", "Name", "Short Name", "Donor Type", "Donor Category",
                "Contact Person", "Email", "Phone", "Country", "Preferred Currency",
                "Restriction Type", "Reporting Frequency", "Status", "Grants count",
            ],
            rows=rows,
        )
        if resp:
            return resp

    return render(
        request,
        "tenant_portal/grants/donor_register.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "donors": donors,
            "filters": f,
            "active_submenu": "funds",
            "active_item": "funds_donor_register",
            "donor_types": Donor.DonorType,
            "donor_categories": Donor.DonorCategory,
            "donor_statuses": Donor.Status,
            "restriction_types": Donor.DefaultRestrictionType,
            "reporting_frequencies": Donor.DefaultReportingFrequency,
            "export_csv_url": _donor_register_export_urls(request)["csv"],
            "export_xlsx_url": _donor_register_export_urls(request)["xlsx"],
            "export_pdf_url": _donor_register_export_urls(request)["pdf"],
            "can_manage": can_manage,
            "donor_register_kpi": donor_register_kpi,
            "donors_truncated": donors_truncated,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_funding_sources_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import FundingSource, Donor

    tenant_db = request.tenant_db
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        funding_type = request.POST.get("funding_type") or "grant"
        donor_id = request.POST.get("donor_id") or None
        description = (request.POST.get("description") or "").strip()
        if not name:
            messages.error(request, "Name is required.")
        else:
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
            FundingSource.objects.using(tenant_db).create(
                name=name, funding_type=funding_type, donor=donor, description=description,
            )
            messages.success(request, "Funding source created.")
            return redirect(reverse("tenant_portal:grants_funding_sources"))
    f = _parse_grants_filters(request)
    funding_type_sel = (request.GET.get("funding_type") or "").strip()
    filters_display = {**f, "funding_type": funding_type_sel}

    sources_qs = FundingSource.objects.using(tenant_db).select_related("donor").order_by("name")
    donors = list(_active_donors_queryset(tenant_db))
    if f["donor_id"]:
        sources_qs = sources_qs.filter(donor_id=f["donor_id"])
    if f.get("q"):
        sources_qs = sources_qs.filter(name__icontains=f["q"])
    if funding_type_sel:
        sources_qs = sources_qs.filter(funding_type=funding_type_sel)

    sources_list = list(sources_qs)
    funding_kpi = {
        "total_shown": len(sources_list),
        "with_donor": sum(1 for s in sources_list if s.donor_id),
        "active": sum(1 for s in sources_list if s.is_active),
    }

    if request.GET.get("format"):
        rows = [
            [
                s.name,
                (s.get_funding_type_display() if getattr(s, "funding_type", None) else ""),
                (s.donor.name if s.donor else ""),
                (s.description or ""),
                (_("Active") if s.is_active else _("Inactive")),
            ]
            for s in sources_list
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="funding_sources",
            title="Funding Sources",
            headers=["Name", "Type", "Donor", "Description", "Status"],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/funding_sources.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "sources": sources_list,
            "donors": donors,
            "filters": filters_display,
            "funding_types": FundingSource.FundingType,
            "funding_kpi": funding_kpi,
            "active_submenu": "funds",
            "active_item": "funds_funding_sources",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
            "official_report_period_line": f"{f['period_start']} — {f['period_end']}",
        },
    )


def _parse_donor_agreement_filters(request: HttpRequest) -> dict:
    from django.utils.dateparse import parse_date

    base = _parse_grants_filters(request)
    return {
        **base,
        "q": (request.GET.get("q") or "").strip(),
        "signed_from": parse_date(request.GET.get("signed_from") or "") or None,
        "signed_to": parse_date(request.GET.get("signed_to") or "") or None,
        "ag_start_from": parse_date(request.GET.get("ag_start_from") or "") or None,
        "ag_start_to": parse_date(request.GET.get("ag_start_to") or "") or None,
        "ag_end_from": parse_date(request.GET.get("ag_end_from") or "") or None,
        "ag_end_to": parse_date(request.GET.get("ag_end_to") or "") or None,
        "currency_id": (request.GET.get("currency_id") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
        "agreement_type": (request.GET.get("agreement_type") or "").strip(),
    }


def _sync_donor_agreement_expiry(tenant_db: str) -> None:
    from django.utils import timezone
    from tenant_grants.models import DonorAgreement

    today = timezone.now().date()
    DonorAgreement.objects.using(tenant_db).filter(end_date__isnull=False, end_date__lt=today).exclude(
        status__in=[
            DonorAgreement.Status.DRAFT,
            DonorAgreement.Status.CLOSED,
            DonorAgreement.Status.EXPIRED,
        ],
    ).update(status=DonorAgreement.Status.EXPIRED)


def _generate_donor_agreement_code(tenant_db: str) -> str:
    from datetime import date

    from tenant_grants.models import DonorAgreement

    y = date.today().year
    prefix = f"DAG-{y}-"
    n = DonorAgreement.objects.using(tenant_db).filter(agreement_code__startswith=prefix).count() + 1
    code = f"{prefix}{n:04d}"
    while DonorAgreement.objects.using(tenant_db).filter(agreement_code=code).exists():
        n += 1
        code = f"{prefix}{n:04d}"
    return code


def _donor_agreement_filter_queryset(qs, f: dict):
    from django.db.models import Q

    if f.get("q"):
        qv = f["q"]
        qs = qs.filter(
            Q(title__icontains=qv)
            | Q(agreement_code__icontains=qv)
            | Q(donor__name__icontains=qv)
            | Q(reference_number__icontains=qv)
        )
    if f.get("donor_id"):
        qs = qs.filter(donor_id=f["donor_id"])
    if f.get("currency_id"):
        qs = qs.filter(currency_id=f["currency_id"])
    if f.get("status"):
        qs = qs.filter(status=f["status"])
    if f.get("agreement_type"):
        qs = qs.filter(agreement_type=f["agreement_type"])
    if f.get("signed_from"):
        qs = qs.filter(signed_date__gte=f["signed_from"])
    if f.get("signed_to"):
        qs = qs.filter(signed_date__lte=f["signed_to"])
    if f.get("ag_start_from"):
        qs = qs.filter(start_date__gte=f["ag_start_from"])
    if f.get("ag_start_to"):
        qs = qs.filter(start_date__lte=f["ag_start_to"])
    if f.get("ag_end_from"):
        qs = qs.filter(end_date__gte=f["ag_end_from"])
    if f.get("ag_end_to"):
        qs = qs.filter(end_date__lte=f["ag_end_to"])
    return qs


def _parse_donor_condition_filters(request: HttpRequest) -> dict:
    from django.utils.dateparse import parse_date

    base = _parse_grants_filters(request)
    return {
        **base,
        "q": (request.GET.get("q") or "").strip(),
        "funding_source_id": (request.GET.get("funding_source_id") or "").strip(),
        "category": (request.GET.get("category") or "").strip(),
        "restriction_type": (request.GET.get("restriction_type") or "").strip(),
        "compliance_level": (request.GET.get("compliance_level") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
        "eff_start_from": parse_date(request.GET.get("eff_start_from") or "") or None,
        "eff_start_to": parse_date(request.GET.get("eff_start_to") or "") or None,
        "eff_end_from": parse_date(request.GET.get("eff_end_from") or "") or None,
        "eff_end_to": parse_date(request.GET.get("eff_end_to") or "") or None,
    }


def _donor_condition_filter_queryset(qs, f: dict):
    from django.db.models import Q

    if f.get("q"):
        qv = f["q"]
        qs = qs.filter(
            Q(description__icontains=qv)
            | Q(conditions__icontains=qv)
            | Q(internal_notes__icontains=qv)
            | Q(donor__name__icontains=qv)
            | Q(grant__code__icontains=qv)
            | Q(restriction_code__icontains=qv)
            | Q(project__code__icontains=qv)
        )
    if f.get("donor_id"):
        qs = qs.filter(donor_id=f["donor_id"])
    if f.get("grant_id"):
        qs = qs.filter(grant_id=f["grant_id"])
    if f.get("funding_source_id"):
        qs = qs.filter(funding_source_id=f["funding_source_id"])
    if f.get("category"):
        qs = qs.filter(category=f["category"])
    if f.get("restriction_type"):
        qs = qs.filter(restriction_type=f["restriction_type"])
    if f.get("compliance_level"):
        qs = qs.filter(compliance_level=f["compliance_level"])
    if f.get("status"):
        qs = qs.filter(status=f["status"])
    if f.get("eff_start_from"):
        qs = qs.filter(
            Q(effective_start__isnull=True) | Q(effective_start__gte=f["eff_start_from"])
        )
    if f.get("eff_start_to"):
        qs = qs.filter(
            Q(effective_start__isnull=True) | Q(effective_start__lte=f["eff_start_to"])
        )
    if f.get("eff_end_from"):
        qs = qs.filter(Q(effective_end__isnull=True) | Q(effective_end__gte=f["eff_end_from"]))
    if f.get("eff_end_to"):
        qs = qs.filter(Q(effective_end__isnull=True) | Q(effective_end__lte=f["eff_end_to"]))
    return qs


def _donor_restrictions_scope_for_officer(qs, tenant_user, tenant_db: str):
    from django.db.models import Q

    from tenant_grants.models import Grant, GrantAssignment

    if getattr(tenant_user, "is_tenant_admin", False):
        return qs
    if user_has_permission(tenant_user, "module:grants.manage", using=tenant_db):
        return qs
    if user_has_permission(tenant_user, "grants:donor_restrictions.manage", using=tenant_db):
        return qs
    assigned = list(
        GrantAssignment.objects.using(tenant_db)
        .filter(officer=tenant_user, is_active=True)
        .values_list("grant_id", flat=True)
    )
    if not assigned:
        return qs.none()
    proj_ids = [
        p
        for p in Grant.objects.using(tenant_db)
        .filter(pk__in=assigned)
        .values_list("project_id", flat=True)
        if p
    ]
    return qs.filter(Q(grant_id__in=assigned) | Q(project_id__in=proj_ids))


def _parse_optional_funding_limit(raw: str | None):
    from decimal import Decimal, InvalidOperation

    s = (raw or "").strip().replace(",", "")
    if not s:
        return True, None
    try:
        d = Decimal(s)
        if d <= 0:
            return False, None
        return True, d
    except InvalidOperation:
        return False, None


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_agreements_view(request: HttpRequest) -> HttpResponse:
    from datetime import timedelta
    from decimal import Decimal

    from django.db.models import Count
    from django.urls import reverse
    from django.utils import timezone
    from django.utils.dateparse import parse_date

    from tenant_finance.models import Currency
    from tenant_grants.models import (
        Donor,
        DonorAgreement,
        DonorAgreementAttachment,
        DonorAgreementGrant,
        DonorAgreementProject,
        FundingSource,
        Grant,
        Project,
    )

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    _sync_donor_agreement_expiry(tenant_db)

    if request.method == "POST" and can_manage:
        donor_id = (request.POST.get("donor_id") or "").strip()
        title = (request.POST.get("title") or "").strip()
        agreement_code_in = (request.POST.get("agreement_code") or "").strip()
        agreement_type = (request.POST.get("agreement_type") or DonorAgreement.AgreementType.GRANT).strip()
        reference_number = (request.POST.get("reference_number") or "").strip()
        status = (request.POST.get("status") or DonorAgreement.Status.DRAFT).strip()
        funding_source_id = (request.POST.get("funding_source_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        ok_lim, funding_limit = _parse_optional_funding_limit(request.POST.get("funding_limit"))
        payment_terms_summary = (request.POST.get("payment_terms_summary") or "").strip()
        installment_notes = (request.POST.get("installment_notes") or "").strip()
        signed_date = parse_date(request.POST.get("signed_date") or "")
        start_date = parse_date(request.POST.get("start_date") or "")
        end_date = parse_date(request.POST.get("end_date") or "")
        reporting_frequency = (request.POST.get("reporting_frequency") or "").strip()
        terms_summary = (request.POST.get("terms_summary") or "").strip()
        internal_notes = (request.POST.get("internal_notes") or "").strip()
        restriction_summary = (request.POST.get("restriction_summary") or "").strip()
        upload = request.FILES.get("agreement_file")
        grant_ids = [int(x) for x in request.POST.getlist("grant_ids") if str(x).strip().isdigit()]
        project_ids = [int(x) for x in request.POST.getlist("project_ids") if str(x).strip().isdigit()]

        errors: list[str] = []
        if not donor_id:
            errors.append("Donor is required.")
        if not title:
            errors.append("Agreement title is required.")
        if not ok_lim:
            errors.append("Funding limit must be a positive number when provided.")
        if start_date and end_date and end_date < start_date:
            errors.append("End date cannot be earlier than start date.")
        if status not in {c for c, _ in DonorAgreement.Status.choices}:
            status = DonorAgreement.Status.DRAFT
        if agreement_type not in {c for c, _ in DonorAgreement.AgreementType.choices}:
            agreement_type = DonorAgreement.AgreementType.GRANT

        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if donor_id and not donor:
            errors.append("Invalid donor.")

        allow_mg = request.POST.get("allow_multiple_grants") == "1"
        allow_mp = request.POST.get("allow_multiple_projects") == "1"
        if not allow_mg and len(grant_ids) > 1:
            errors.append("Only one grant may be linked when multiple grants are not allowed.")
        if not allow_mp and len(project_ids) > 1:
            errors.append("Only one project may be linked when multiple projects are not allowed.")

        for gid in grant_ids:
            g = Grant.objects.using(tenant_db).filter(pk=gid).first()
            if not g:
                errors.append("Invalid grant selection.")
                break
            if donor and g.donor_id != donor.pk:
                errors.append(f"Grant {g.code} does not belong to the selected donor.")
                break

        code = agreement_code_in or _generate_donor_agreement_code(tenant_db)
        if DonorAgreement.objects.using(tenant_db).filter(agreement_code__iexact=code).exists():
            errors.append("Agreement code already exists.")

        fs = None
        if funding_source_id.isdigit():
            fs = FundingSource.objects.using(tenant_db).filter(pk=int(funding_source_id)).first()
        cur = None
        if currency_id.isdigit():
            cur = Currency.objects.using(tenant_db).filter(pk=int(currency_id)).first()

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            assert donor is not None
            ag = DonorAgreement(
                agreement_code=code,
                donor=donor,
                title=title,
                agreement_type=agreement_type,
                reference_number=reference_number,
                status=status,
                funding_source=fs,
                currency=cur,
                funding_limit=funding_limit,
                payment_terms_summary=payment_terms_summary,
                installment_notes=installment_notes,
                signed_date=signed_date,
                start_date=start_date,
                end_date=end_date,
                reporting_frequency=reporting_frequency or "",
                compliance_financial_reporting=request.POST.get("compliance_financial_reporting") == "1",
                compliance_narrative_reporting=request.POST.get("compliance_narrative_reporting") == "1",
                compliance_audit_required=request.POST.get("compliance_audit_required") == "1",
                compliance_special_conditions=request.POST.get("compliance_special_conditions") == "1",
                restricted_funding=request.POST.get("restricted_funding") == "1",
                restriction_summary=restriction_summary,
                allow_multiple_grants=allow_mg,
                allow_multiple_projects=allow_mp,
                terms_summary=terms_summary,
                internal_notes=internal_notes,
            )
            try:
                ag.full_clean()
            except ValidationError as ve:
                for _f, msgs in ve.error_dict.items():
                    for m in msgs:
                        messages.error(request, str(m))
                return redirect(reverse("tenant_portal:grants_donor_agreements"))
            ag.save(using=tenant_db)
            if upload:
                ag.file = upload
                ag.original_filename = getattr(upload, "name", "") or ""
                ag.save(using=tenant_db, update_fields=["file", "original_filename", "updated_at"])
            for gid in grant_ids:
                DonorAgreementGrant.objects.using(tenant_db).create(agreement=ag, grant_id=gid)
            for pid in project_ids:
                DonorAgreementProject.objects.using(tenant_db).create(agreement=ag, project_id=pid)
            for fobj in request.FILES.getlist("amendment_files"):
                DonorAgreementAttachment.objects.using(tenant_db).create(
                    agreement=ag,
                    kind=DonorAgreementAttachment.Kind.AMENDMENT,
                    file=fobj,
                    original_filename=getattr(fobj, "name", "") or "",
                )
            for fobj in request.FILES.getlist("supporting_files"):
                DonorAgreementAttachment.objects.using(tenant_db).create(
                    agreement=ag,
                    kind=DonorAgreementAttachment.Kind.SUPPORTING,
                    file=fobj,
                    original_filename=getattr(fobj, "name", "") or "",
                )
            messages.success(request, "Agreement created.")
            return redirect(reverse("tenant_portal:grants_donor_agreement_detail", args=[ag.pk]))

    f = _parse_donor_agreement_filters(request)
    qs = (
        DonorAgreement.objects.using(tenant_db)
        .select_related("donor", "currency", "funding_source")
        .prefetch_related("grant_links__grant")
        .annotate(linked_grants_count=Count("grant_links", distinct=True))
        .order_by("-signed_date", "-created_at")
    )
    qs = _donor_agreement_filter_queryset(qs, f)
    agreements = list(qs[:500])

    today = timezone.now().date()
    soon = today + timedelta(days=90)
    expiring_soon = sum(
        1
        for a in agreements
        if a.status == DonorAgreement.Status.ACTIVE
        and a.end_date
        and today < a.end_date <= soon
    )
    expired_n = sum(1 for a in agreements if a.status == DonorAgreement.Status.EXPIRED)
    missing_file_n = sum(1 for a in agreements if not a.file)
    over_limit_ids: list[int] = []
    for a in agreements:
        if a.funding_limit is None:
            continue
        total_award = Decimal("0")
        for link in a.grant_links.all():
            total_award += link.grant.award_amount or Decimal("0")
        if total_award > (a.funding_limit or Decimal("0")):
            over_limit_ids.append(a.pk)

    agreement_kpi = {
        "total": len(agreements),
        "active": sum(1 for a in agreements if a.status == DonorAgreement.Status.ACTIVE),
        "draft": sum(1 for a in agreements if a.status == DonorAgreement.Status.DRAFT),
        "expired": sum(1 for a in agreements if a.status == DonorAgreement.Status.EXPIRED),
        "closed": sum(1 for a in agreements if a.status == DonorAgreement.Status.CLOSED),
    }

    donors = list(_active_donors_queryset(tenant_db))
    currencies = list(Currency.objects.using(tenant_db).filter(status=Currency.Status.ACTIVE).order_by("code"))
    funding_sources = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:500])
    projects = list(Project.objects.using(tenant_db).order_by("code")[:500])

    if request.GET.get("format"):
        rows = []
        for a in agreements:
            rows.append(
                [
                    a.agreement_code,
                    a.donor.name if a.donor else "",
                    a.title,
                    a.get_agreement_type_display(),
                    a.signed_date or "",
                    a.start_date or "",
                    a.end_date or "",
                    a.funding_limit or "",
                    a.currency.code if a.currency else "",
                    getattr(a, "linked_grants_count", 0) or 0,
                    a.get_status_display(),
                    (_("Yes") if a.file else _("No")),
                ]
            )
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="donor_agreements",
            title="Donor Agreements",
            headers=[
                "Agreement code",
                "Donor",
                "Title",
                "Agreement type",
                "Signed date",
                "Start date",
                "End date",
                "Funding limit",
                "Currency",
                "Linked grants",
                "Status",
                "Signed file",
            ],
            rows=rows,
        )
        if resp:
            return resp

    return render(
        request,
        "tenant_portal/grants/donor_agreements.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "agreements": agreements,
            "donors": donors,
            "currencies": currencies,
            "funding_sources": funding_sources,
            "grants": grants,
            "projects": projects,
            "filters": f,
            "can_manage": can_manage,
            "agreement_kpi": agreement_kpi,
            "alert_expiring_soon": expiring_soon,
            "alert_expired": expired_n,
            "alert_missing_file": missing_file_n,
            "over_limit_ids": set(over_limit_ids),
            "active_submenu": "funds",
            "active_item": "funds_donor_agreements",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
            "official_report_period_line": f"{f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_agreement_detail_view(request: HttpRequest, agreement_id: int) -> HttpResponse:
    from decimal import Decimal

    from datetime import timedelta

    from django.http import Http404
    from django.utils import timezone

    from tenant_grants.models import DonorAgreement

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    ag = (
        DonorAgreement.objects.using(tenant_db)
        .select_related("donor", "currency", "funding_source")
        .prefetch_related("grant_links__grant", "project_links__project", "attachments")
        .filter(pk=agreement_id)
        .first()
    )
    if not ag:
        raise Http404("Agreement not found")

    total_award = Decimal("0")
    for link in ag.grant_links.all():
        total_award += link.grant.award_amount or Decimal("0")
    over_limit = ag.funding_limit is not None and total_award > ag.funding_limit

    today = timezone.now().date()
    expiring_soon = (
        ag.status == DonorAgreement.Status.ACTIVE
        and ag.end_date
        and today < ag.end_date <= today + timedelta(days=90)
    )
    _fn = ((ag.original_filename or "") + (ag.file.name if ag.file else "")).lower()
    show_pdf_preview = bool(ag.file and _fn.endswith(".pdf"))

    return render(
        request,
        "tenant_portal/grants/donor_agreement_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "agreement": ag,
            "can_manage": can_manage,
            "total_award": total_award,
            "over_limit": over_limit,
            "expiring_soon": expiring_soon,
            "missing_signed_file": not bool(ag.file),
            "show_pdf_preview": show_pdf_preview,
            "active_submenu": "funds",
            "active_item": "funds_donor_agreements",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_donor_agreement_edit_view(request: HttpRequest, agreement_id: int) -> HttpResponse:
    from decimal import Decimal

    from django.http import Http404
    from django.urls import reverse
    from django.utils.dateparse import parse_date

    from tenant_finance.models import Currency
    from tenant_grants.models import (
        Donor,
        DonorAgreement,
        DonorAgreementGrant,
        DonorAgreementProject,
        FundingSource,
        Grant,
        Project,
    )

    tenant_db = request.tenant_db
    ag = DonorAgreement.objects.using(tenant_db).filter(pk=agreement_id).first()
    if not ag:
        raise Http404("Agreement not found")

    if ag.status == DonorAgreement.Status.CLOSED:
        messages.error(request, "Closed agreements cannot be edited.")
        return redirect(reverse("tenant_portal:grants_donor_agreement_detail", args=[ag.pk]))

    if request.method == "POST":
        donor_id = (request.POST.get("donor_id") or "").strip()
        title = (request.POST.get("title") or "").strip()
        agreement_type = (request.POST.get("agreement_type") or ag.agreement_type).strip()
        reference_number = (request.POST.get("reference_number") or "").strip()
        status = (request.POST.get("status") or ag.status).strip()
        funding_source_id = (request.POST.get("funding_source_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        ok_lim, funding_limit = _parse_optional_funding_limit(request.POST.get("funding_limit"))
        payment_terms_summary = (request.POST.get("payment_terms_summary") or "").strip()
        installment_notes = (request.POST.get("installment_notes") or "").strip()
        signed_date = parse_date(request.POST.get("signed_date") or "")
        start_date = parse_date(request.POST.get("start_date") or "")
        end_date = parse_date(request.POST.get("end_date") or "")
        reporting_frequency = (request.POST.get("reporting_frequency") or "").strip()
        terms_summary = (request.POST.get("terms_summary") or "").strip()
        internal_notes = (request.POST.get("internal_notes") or "").strip()
        restriction_summary = (request.POST.get("restriction_summary") or "").strip()
        grant_ids = [int(x) for x in request.POST.getlist("grant_ids") if str(x).strip().isdigit()]
        project_ids = [int(x) for x in request.POST.getlist("project_ids") if str(x).strip().isdigit()]

        errors: list[str] = []
        if not donor_id:
            errors.append("Donor is required.")
        if not title:
            errors.append("Agreement title is required.")
        if not ok_lim:
            errors.append("Funding limit must be a positive number when provided.")
        if start_date and end_date and end_date < start_date:
            errors.append("End date cannot be earlier than start date.")

        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if donor_id and not donor:
            errors.append("Invalid donor.")

        allow_mg = request.POST.get("allow_multiple_grants") == "1"
        allow_mp = request.POST.get("allow_multiple_projects") == "1"
        if not allow_mg and len(grant_ids) > 1:
            errors.append("Only one grant may be linked when multiple grants are not allowed.")
        if not allow_mp and len(project_ids) > 1:
            errors.append("Only one project may be linked when multiple projects are not allowed.")

        for gid in grant_ids:
            g = Grant.objects.using(tenant_db).filter(pk=gid).first()
            if not g:
                errors.append("Invalid grant selection.")
                break
            if donor and g.donor_id != donor.pk:
                errors.append(f"Grant {g.code} does not belong to the selected donor.")
                break

        fs = None
        if funding_source_id.isdigit():
            fs = FundingSource.objects.using(tenant_db).filter(pk=int(funding_source_id)).first()
        cur = None
        if currency_id.isdigit():
            cur = Currency.objects.using(tenant_db).filter(pk=int(currency_id)).first()

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            assert donor is not None
            ag.donor = donor
            ag.title = title
            ag.agreement_type = agreement_type
            ag.reference_number = reference_number
            ag.status = status
            ag.funding_source = fs
            ag.currency = cur
            ag.funding_limit = funding_limit
            ag.payment_terms_summary = payment_terms_summary
            ag.installment_notes = installment_notes
            ag.signed_date = signed_date
            ag.start_date = start_date
            ag.end_date = end_date
            ag.reporting_frequency = reporting_frequency or ""
            ag.compliance_financial_reporting = request.POST.get("compliance_financial_reporting") == "1"
            ag.compliance_narrative_reporting = request.POST.get("compliance_narrative_reporting") == "1"
            ag.compliance_audit_required = request.POST.get("compliance_audit_required") == "1"
            ag.compliance_special_conditions = request.POST.get("compliance_special_conditions") == "1"
            ag.restricted_funding = request.POST.get("restricted_funding") == "1"
            ag.restriction_summary = restriction_summary
            ag.allow_multiple_grants = allow_mg
            ag.allow_multiple_projects = allow_mp
            ag.terms_summary = terms_summary
            ag.internal_notes = internal_notes
            upload = request.FILES.get("agreement_file")
            if upload:
                ag.file = upload
                ag.original_filename = getattr(upload, "name", "") or ""
            try:
                ag.full_clean()
            except ValidationError as ve:
                for _f, msgs in ve.error_dict.items():
                    for m in msgs:
                        messages.error(request, str(m))
                return redirect(reverse("tenant_portal:grants_donor_agreement_edit", args=[ag.pk]))
            ag.save(using=tenant_db)
            DonorAgreementGrant.objects.using(tenant_db).filter(agreement=ag).delete()
            DonorAgreementProject.objects.using(tenant_db).filter(agreement=ag).delete()
            for gid in grant_ids:
                DonorAgreementGrant.objects.using(tenant_db).create(agreement=ag, grant_id=gid)
            for pid in project_ids:
                DonorAgreementProject.objects.using(tenant_db).create(agreement=ag, project_id=pid)
            messages.success(request, "Agreement updated.")
            return redirect(reverse("tenant_portal:grants_donor_agreement_detail", args=[ag.pk]))

    donors = list(_active_donors_queryset(tenant_db))
    currencies = list(Currency.objects.using(tenant_db).filter(status=Currency.Status.ACTIVE).order_by("code"))
    funding_sources = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:500])
    projects = list(Project.objects.using(tenant_db).order_by("code")[:500])
    selected_grant_ids = list(ag.grant_links.values_list("grant_id", flat=True))
    selected_project_ids = list(ag.project_links.values_list("project_id", flat=True))

    return render(
        request,
        "tenant_portal/grants/donor_agreement_edit.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "agreement": ag,
            "donors": donors,
            "currencies": currencies,
            "funding_sources": funding_sources,
            "grants": grants,
            "projects": projects,
            "selected_grant_ids": set(selected_grant_ids),
            "selected_project_ids": set(selected_project_ids),
            "active_submenu": "funds",
            "active_item": "funds_donor_agreements",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_agreement_file_view(request: HttpRequest, agreement_id: int) -> HttpResponse:
    from django.http import FileResponse, Http404

    from tenant_grants.models import DonorAgreement

    tenant_db = request.tenant_db
    ag = DonorAgreement.objects.using(tenant_db).filter(pk=agreement_id).first()
    if not ag or not ag.file:
        raise Http404("File not found")
    resp = FileResponse(
        ag.file.open("rb"),
        as_attachment=True,
        filename=ag.original_filename or f"{ag.agreement_code}.pdf",
    )
    return resp


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_agreement_attachment_file_view(request: HttpRequest, agreement_id: int, attachment_id: int) -> HttpResponse:
    from django.http import FileResponse, Http404

    from tenant_grants.models import DonorAgreementAttachment

    tenant_db = request.tenant_db
    att = (
        DonorAgreementAttachment.objects.using(tenant_db)
        .filter(pk=attachment_id, agreement_id=agreement_id)
        .first()
    )
    if not att or not att.file:
        raise Http404("Attachment not found")
    return FileResponse(
        att.file.open("rb"),
        as_attachment=True,
        filename=att.original_filename or f"attachment-{attachment_id}",
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_donor_agreement_close_view(request: HttpRequest, agreement_id: int) -> HttpResponse:
    from django.http import Http404, HttpResponseRedirect
    from django.urls import reverse

    from tenant_grants.models import DonorAgreement

    tenant_db = request.tenant_db
    if request.method != "POST":
        return HttpResponseRedirect(reverse("tenant_portal:grants_donor_agreement_detail", args=[agreement_id]))
    ag = DonorAgreement.objects.using(tenant_db).filter(pk=agreement_id).first()
    if not ag:
        raise Http404("Agreement not found")
    ag.status = DonorAgreement.Status.CLOSED
    ag.save(using=tenant_db, update_fields=["status", "updated_at"])
    messages.success(request, "Agreement closed.")
    return redirect(reverse("tenant_portal:grants_donor_agreement_detail", args=[ag.pk]))


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_conditions_view(request: HttpRequest) -> HttpResponse:
    from datetime import timedelta
    from decimal import Decimal, InvalidOperation

    from django.core.exceptions import ValidationError
    from django.utils import timezone
    from django.utils.dateparse import parse_date

    from tenant_finance.models import AccountCategory
    from tenant_grants.models import (
        BudgetLine,
        Donor,
        DonorRestriction,
        FundingSource,
        Grant,
        Project,
    )
    from tenant_grants.restrictions import sync_donor_restriction_expiry

    tenant_db = request.tenant_db
    sync_donor_restriction_expiry(tenant_db)
    can_manage_restrictions = user_has_permission(
        request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
    )

    if request.method == "POST" and can_manage_restrictions and request.POST.get("action") == "create":
        donor_id = (request.POST.get("donor_id") or "").strip()
        description = (request.POST.get("description") or "").strip()
        if not donor_id or not description:
            messages.error(request, "Donor and description are required.")
        else:
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
            if not donor:
                messages.error(request, "Invalid donor.")
            else:
                fs_id = (request.POST.get("funding_source_id") or "").strip()
                g_id = (request.POST.get("grant_id") or "").strip()
                p_id = (request.POST.get("project_id") or "").strip()
                bl_id = (request.POST.get("budget_line_id") or "").strip()
                ac_id = (request.POST.get("account_category_id") or "").strip()
                fs = FundingSource.objects.using(tenant_db).filter(pk=fs_id).first() if fs_id else None
                grant = Grant.objects.using(tenant_db).filter(pk=g_id).first() if g_id else None
                project = Project.objects.using(tenant_db).filter(pk=p_id).first() if p_id else None
                budget_line = BudgetLine.objects.using(tenant_db).filter(pk=bl_id).first() if bl_id else None
                acct_cat = AccountCategory.objects.using(tenant_db).filter(pk=ac_id).first() if ac_id else None
                code_in = (request.POST.get("restriction_code") or "").strip()
                cat = (request.POST.get("category") or DonorRestriction.Category.OTHER).strip()
                rtype = (request.POST.get("restriction_type") or DonorRestriction.RestrictionType.OTHER).strip()
                clevel = (request.POST.get("compliance_level") or DonorRestriction.ComplianceLevel.MANDATORY).strip()
                status = (request.POST.get("status") or DonorRestriction.Status.DRAFT).strip()
                applies_scope = (request.POST.get("applies_scope") or DonorRestriction.AppliesScope.DONOR_WIDE).strip()
                eff_s = parse_date(request.POST.get("effective_start") or "")
                eff_e = parse_date(request.POST.get("effective_end") or "")

                def _dec(raw):
                    s = (raw or "").strip().replace(",", "")
                    if not s:
                        return None
                    try:
                        return Decimal(s)
                    except InvalidOperation:
                        return None

                dr = DonorRestriction(
                    restriction_code=code_in or "",
                    donor=donor,
                    funding_source=fs,
                    grant=grant,
                    project=project,
                    budget_line=budget_line,
                    account_category=acct_cat,
                    category=cat
                    if cat in {c[0] for c in DonorRestriction.Category.choices}
                    else DonorRestriction.Category.OTHER,
                    restriction_type=(
                        rtype
                        if rtype in {c[0] for c in DonorRestriction.RestrictionType.choices}
                        else DonorRestriction.RestrictionType.OTHER
                    ),
                    compliance_level=(
                        clevel
                        if clevel in {c[0] for c in DonorRestriction.ComplianceLevel.choices}
                        else DonorRestriction.ComplianceLevel.MANDATORY
                    ),
                    status=status
                    if status in {c[0] for c in DonorRestriction.Status.choices}
                    else DonorRestriction.Status.DRAFT,
                    applies_scope=(
                        applies_scope
                        if applies_scope in {c[0] for c in DonorRestriction.AppliesScope.choices}
                        else DonorRestriction.AppliesScope.DONOR_WIDE
                    ),
                    description=description,
                    conditions=(request.POST.get("conditions") or "").strip(),
                    internal_notes=(request.POST.get("internal_notes") or "").strip(),
                    effective_start=eff_s,
                    effective_end=eff_e,
                    enforce_budget_validation=request.POST.get("enforce_budget_validation") == "1",
                    enforce_procurement_validation=request.POST.get("enforce_procurement_validation") == "1",
                    enforce_expense_eligibility=request.POST.get("enforce_expense_eligibility") == "1",
                    require_supporting_documents=request.POST.get("require_supporting_documents") == "1",
                    require_approval_override=request.POST.get("require_approval_override") == "1",
                    max_budget_percentage=_dec(request.POST.get("max_budget_percentage")),
                    max_expense_per_transaction=_dec(request.POST.get("max_expense_per_transaction")),
                    max_procurement_threshold=_dec(request.POST.get("max_procurement_threshold")),
                )
                try:
                    dr.full_clean()
                except ValidationError as ve:
                    for _field, err_list in ve.error_dict.items():
                        for err in err_list:
                            messages.error(request, str(err))
                else:
                    dr.save(using=tenant_db)
                    messages.success(request, "Restriction created.")
                    return redirect(
                        reverse("tenant_portal:grants_donor_condition_detail", args=[dr.pk])
                    )

    f = _parse_donor_condition_filters(request)
    qs = (
        DonorRestriction.objects.using(tenant_db)
        .select_related(
            "donor",
            "grant",
            "project",
            "funding_source",
        )
        .order_by("-created_at")
    )
    qs = _donor_condition_filter_queryset(qs, f)
    qs = _donor_restrictions_scope_for_officer(qs, request.tenant_user, tenant_db)
    restrictions = list(qs[:500])
    today = timezone.now().date()
    soon = today + timedelta(days=90)
    alert_expiring = sum(
        1
        for r in restrictions
        if r.status == DonorRestriction.Status.ACTIVE
        and r.effective_end
        and today < r.effective_end <= soon
    )
    mandatory_active = sum(
        1
        for r in restrictions
        if r.status == DonorRestriction.Status.ACTIVE
        and r.compliance_level == DonorRestriction.ComplianceLevel.MANDATORY
    )

    donors = list(_active_donors_queryset(tenant_db))
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:500])
    projects = list(Project.objects.using(tenant_db).order_by("code")[:500])
    funding_sources = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    budget_lines = list(
        BudgetLine.objects.using(tenant_db).select_related("grant").order_by("grant__code", "id")[:500]
    )
    account_categories = list(
        AccountCategory.objects.using(tenant_db).filter(status=AccountCategory.Status.ACTIVE).order_by("code")[:300]
    )

    if request.GET.get("format"):
        rows = [
            [
                r.restriction_code or "",
                r.donor.name if r.donor else "",
                r.funding_source.name if r.funding_source else "",
                r.grant.code if r.grant else "",
                r.get_restriction_type_display(),
                r.get_category_display(),
                r.effective_start or "",
                r.effective_end or "",
                r.get_compliance_level_display(),
                r.get_status_display(),
                (r.description or "")[:200],
            ]
            for r in restrictions
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="donor_restrictions",
            title="Donor Conditions & Restrictions",
            headers=[
                "Restriction code",
                "Donor",
                "Funding source",
                "Grant",
                "Restriction type",
                "Category",
                "Effective start",
                "Effective end",
                "Compliance level",
                "Status",
                "Description summary",
            ],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/donor_conditions.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "restrictions": restrictions,
            "donors": donors,
            "grants": grants,
            "projects": projects,
            "funding_sources": funding_sources,
            "budget_lines": budget_lines,
            "account_categories": account_categories,
            "filters": f,
            "restriction_types": DonorRestriction.RestrictionType,
            "restriction_categories": DonorRestriction.Category,
            "compliance_levels": DonorRestriction.ComplianceLevel,
            "restriction_statuses": DonorRestriction.Status,
            "applies_scopes": DonorRestriction.AppliesScope,
            "can_manage_restrictions": can_manage_restrictions,
            "alert_expiring": alert_expiring,
            "mandatory_active": mandatory_active,
            "active_submenu": "funds",
            "active_item": "funds_donor_conditions",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


def _get_donor_restriction_or_404(restriction_id: int, tenant_db: str, tenant_user):
    from django.http import Http404
    from tenant_grants.models import DonorRestriction

    r = (
        DonorRestriction.objects.using(tenant_db)
        .select_related("donor", "grant", "project", "funding_source", "budget_line", "account_category")
        .filter(pk=restriction_id)
        .first()
    )
    if not r:
        raise Http404("Restriction not found")
    qs = DonorRestriction.objects.using(tenant_db).filter(pk=restriction_id)
    scoped = _donor_restrictions_scope_for_officer(qs, tenant_user, tenant_db)
    if not scoped.exists():
        raise Http404("Restriction not found")
    return r


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_condition_detail_view(request: HttpRequest, restriction_id: int) -> HttpResponse:
    from tenant_grants.restrictions import sync_donor_restriction_expiry

    tenant_db = request.tenant_db
    sync_donor_restriction_expiry(tenant_db)
    r = _get_donor_restriction_or_404(restriction_id, tenant_db, request.tenant_user)
    can_manage_restrictions = user_has_permission(
        request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
    )
    return render(
        request,
        "tenant_portal/grants/donor_condition_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "restriction": r,
            "can_manage_restrictions": can_manage_restrictions,
            "active_submenu": "funds",
            "active_item": "funds_donor_conditions",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_condition_edit_view(request: HttpRequest, restriction_id: int) -> HttpResponse:
    from decimal import Decimal, InvalidOperation

    from django.core.exceptions import ValidationError
    from django.utils.dateparse import parse_date

    from tenant_finance.models import AccountCategory
    from tenant_grants.models import (
        BudgetLine,
        Donor,
        DonorRestriction,
        FundingSource,
        Grant,
        Project,
    )
    from tenant_grants.restrictions import sync_donor_restriction_expiry

    tenant_db = request.tenant_db
    if not user_has_permission(
        request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
    ):
        from django.http import HttpResponseForbidden

        return HttpResponseForbidden("You cannot edit donor restrictions.")
    sync_donor_restriction_expiry(tenant_db)
    r = _get_donor_restriction_or_404(restriction_id, tenant_db, request.tenant_user)

    if request.method == "POST":
        donor_id = (request.POST.get("donor_id") or "").strip()
        description = (request.POST.get("description") or "").strip()
        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if not donor or not description:
            messages.error(request, "Donor and description are required.")
        else:
            fs_id = (request.POST.get("funding_source_id") or "").strip()
            g_id = (request.POST.get("grant_id") or "").strip()
            p_id = (request.POST.get("project_id") or "").strip()
            bl_id = (request.POST.get("budget_line_id") or "").strip()
            ac_id = (request.POST.get("account_category_id") or "").strip()
            fs = FundingSource.objects.using(tenant_db).filter(pk=fs_id).first() if fs_id else None
            grant = Grant.objects.using(tenant_db).filter(pk=g_id).first() if g_id else None
            project = Project.objects.using(tenant_db).filter(pk=p_id).first() if p_id else None
            budget_line = BudgetLine.objects.using(tenant_db).filter(pk=bl_id).first() if bl_id else None
            acct_cat = AccountCategory.objects.using(tenant_db).filter(pk=ac_id).first() if ac_id else None
            code_in = (request.POST.get("restriction_code") or "").strip()
            cat = (request.POST.get("category") or DonorRestriction.Category.OTHER).strip()
            rtype = (request.POST.get("restriction_type") or DonorRestriction.RestrictionType.OTHER).strip()
            clevel = (request.POST.get("compliance_level") or DonorRestriction.ComplianceLevel.MANDATORY).strip()
            status = (request.POST.get("status") or DonorRestriction.Status.DRAFT).strip()
            applies_scope = (request.POST.get("applies_scope") or DonorRestriction.AppliesScope.DONOR_WIDE).strip()
            eff_s = parse_date(request.POST.get("effective_start") or "")
            eff_e = parse_date(request.POST.get("effective_end") or "")

            def _dec(raw):
                s = (raw or "").strip().replace(",", "")
                if not s:
                    return None
                try:
                    return Decimal(s)
                except InvalidOperation:
                    return None

            r.donor = donor
            r.funding_source = fs
            r.grant = grant
            r.project = project
            r.budget_line = budget_line
            r.account_category = acct_cat
            if code_in:
                r.restriction_code = code_in
            r.category = (
                cat if cat in {c[0] for c in DonorRestriction.Category.choices} else DonorRestriction.Category.OTHER
            )
            r.restriction_type = (
                rtype
                if rtype in {c[0] for c in DonorRestriction.RestrictionType.choices}
                else DonorRestriction.RestrictionType.OTHER
            )
            r.compliance_level = (
                clevel
                if clevel in {c[0] for c in DonorRestriction.ComplianceLevel.choices}
                else DonorRestriction.ComplianceLevel.MANDATORY
            )
            r.status = (
                status
                if status in {c[0] for c in DonorRestriction.Status.choices}
                else DonorRestriction.Status.DRAFT
            )
            r.applies_scope = (
                applies_scope
                if applies_scope in {c[0] for c in DonorRestriction.AppliesScope.choices}
                else DonorRestriction.AppliesScope.DONOR_WIDE
            )
            r.description = description
            r.conditions = (request.POST.get("conditions") or "").strip()
            r.internal_notes = (request.POST.get("internal_notes") or "").strip()
            r.effective_start = eff_s
            r.effective_end = eff_e
            r.enforce_budget_validation = request.POST.get("enforce_budget_validation") == "1"
            r.enforce_procurement_validation = request.POST.get("enforce_procurement_validation") == "1"
            r.enforce_expense_eligibility = request.POST.get("enforce_expense_eligibility") == "1"
            r.require_supporting_documents = request.POST.get("require_supporting_documents") == "1"
            r.require_approval_override = request.POST.get("require_approval_override") == "1"
            r.max_budget_percentage = _dec(request.POST.get("max_budget_percentage"))
            r.max_expense_per_transaction = _dec(request.POST.get("max_expense_per_transaction"))
            r.max_procurement_threshold = _dec(request.POST.get("max_procurement_threshold"))
            try:
                r.full_clean()
            except ValidationError as ve:
                for _field, err_list in ve.error_dict.items():
                    for err in err_list:
                        messages.error(request, str(err))
            else:
                r.save(using=tenant_db)
                messages.success(request, "Restriction updated.")
                return redirect(reverse("tenant_portal:grants_donor_condition_detail", args=[r.pk]))

    donors = list(_active_donors_queryset(tenant_db))
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:500])
    projects = list(Project.objects.using(tenant_db).order_by("code")[:500])
    funding_sources = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    budget_lines = list(
        BudgetLine.objects.using(tenant_db).select_related("grant").order_by("grant__code", "id")[:500]
    )
    account_categories = list(
        AccountCategory.objects.using(tenant_db).filter(status=AccountCategory.Status.ACTIVE).order_by("code")[:300]
    )
    return render(
        request,
        "tenant_portal/grants/donor_condition_edit.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "restriction": r,
            "donors": donors,
            "grants": grants,
            "projects": projects,
            "funding_sources": funding_sources,
            "budget_lines": budget_lines,
            "account_categories": account_categories,
            "restriction_types": DonorRestriction.RestrictionType,
            "restriction_categories": DonorRestriction.Category,
            "compliance_levels": DonorRestriction.ComplianceLevel,
            "restriction_statuses": DonorRestriction.Status,
            "applies_scopes": DonorRestriction.AppliesScope,
            "active_submenu": "funds",
            "active_item": "funds_donor_conditions",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_condition_deactivate_view(request: HttpRequest, restriction_id: int) -> HttpResponse:
    from tenant_grants.models import DonorRestriction

    tenant_db = request.tenant_db
    if not user_has_permission(
        request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
    ):
        from django.http import HttpResponseForbidden

        return HttpResponseForbidden("You cannot deactivate donor restrictions.")
    if request.method != "POST":
        from django.http import HttpResponseNotAllowed

        return HttpResponseNotAllowed(["POST"])
    r = _get_donor_restriction_or_404(restriction_id, tenant_db, request.tenant_user)
    r.status = DonorRestriction.Status.INACTIVE
    r.save(using=tenant_db, update_fields=["status", "updated_at"])
    messages.success(request, "Restriction deactivated.")
    return redirect(reverse("tenant_portal:grants_donor_conditions"))


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_allocation_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import GrantAllocation, Grant, Donor
    from decimal import Decimal

    tenant_db = request.tenant_db
    if request.method == "POST":
        grant_id = request.POST.get("grant_id")
        donor_id = request.POST.get("donor_id")
        amount = request.POST.get("amount") or None
        percentage = request.POST.get("percentage") or None
        if not grant_id or not donor_id:
            messages.error(request, "Grant and donor are required.")
        else:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
            if grant and donor:
                amt = Decimal(amount) if amount else None
                pct = Decimal(percentage) if percentage else None
                GrantAllocation.objects.using(tenant_db).update_or_create(
                    grant=grant, donor=donor, defaults={"amount": amt, "percentage": pct}
                )
                messages.success(request, "Allocation saved.")
                return redirect(reverse("tenant_portal:grants_grant_allocation"))
    f = _parse_grants_filters(request)
    qs = GrantAllocation.objects.using(tenant_db).select_related("grant", "donor").order_by("grant__code", "donor__name")
    if f["grant_id"]:
        qs = qs.filter(grant_id=f["grant_id"])
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    allocations = list(qs)
    grants = Grant.objects.using(tenant_db).select_related("donor").order_by("code")
    donors = list(_active_donors_queryset(tenant_db))
    if request.GET.get("format"):
        rows = [
            [
                f"{a.grant.code} — {a.grant.title}" if a.grant else "",
                a.donor.name if a.donor else "",
                a.amount or "",
                a.percentage or "",
            ]
            for a in allocations
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="grant_allocation",
            title="Grant Allocation (Multi-donor)",
            headers=["Grant", "Donor", "Amount", "Percentage"],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/grant_allocation.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "allocations": allocations,
            "grants": grants,
            "donors": donors,
            "filters": f,
            "active_submenu": "funds",
            "active_item": "funds_grant_allocation",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


def _parse_tracking_filters(request: HttpRequest) -> dict:
    return {
        "donor_id": request.GET.get("donor_id") or None,
        "tracking_id": request.GET.get("tracking_id") or None,
        "pipeline_stage": request.GET.get("pipeline_stage") or None,
        "period_start": request.GET.get("period_start") or None,
        "period_end": request.GET.get("period_end") or None,
    }


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_tracking_view(request: HttpRequest) -> HttpResponse:
    """Pre-award pipeline: list GrantTracking records; Convert to Agreement when approved."""
    from decimal import Decimal
    from tenant_grants.models import GrantTracking

    tenant_db = request.tenant_db
    f = _parse_tracking_filters(request)

    qs = GrantTracking.objects.using(tenant_db).select_related(
        "donor", "grant_manager", "project"
    ).order_by("-updated_at", "-created_at")
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    if f["tracking_id"]:
        qs = qs.filter(pk=f["tracking_id"])
    if f["pipeline_stage"]:
        qs = qs.filter(pipeline_stage=f["pipeline_stage"])
    if f.get("period_start"):
        qs = qs.filter(submission_deadline__gte=f["period_start"])
    if f.get("period_end"):
        qs = qs.filter(submission_deadline__lte=f["period_end"])

    rows = []
    for t in qs:
        grant_manager_display = ""
        if t.grant_manager_id:
            grant_manager_display = getattr(t.grant_manager, "full_name", None) or getattr(t.grant_manager, "email", "") or ""
        if not grant_manager_display and getattr(t, "grant_owner", None):
            grant_manager_display = t.grant_owner
        project_display = (t.project.name if t.project_id else None) or t.project_name or ""
        amount_awarded = t.amount_awarded
        try:
            ag = t.grant_agreement
            if amount_awarded is None:
                amount_awarded = ag.award_amount
            r = {
                "tracking": t,
                "pipeline_stage": t.get_pipeline_stage_display(),
                "grant_type": t.get_grant_type_display() if t.grant_type else "",
                "priority": t.get_priority_display() if t.priority else "",
                "submission_deadline": t.submission_deadline,
                "date_submitted": t.date_submitted,
                "project_name": project_display,
                "amount_requested": t.amount_requested or Decimal("0"),
                "grant_manager_display": grant_manager_display,
                "amount_awarded": amount_awarded,
                "upcoming_deadline": None,
                "notes": t.notes or "",
            }
            r["amount_rewarded"] = ag.award_amount
            r["agreement_code"] = ag.code
            r["can_convert"] = False
        except Exception:
            r = {
                "tracking": t,
                "pipeline_stage": t.get_pipeline_stage_display(),
                "grant_type": t.get_grant_type_display() if t.grant_type else "",
                "priority": t.get_priority_display() if t.priority else "",
                "submission_deadline": t.submission_deadline,
                "date_submitted": t.date_submitted,
                "project_name": project_display,
                "amount_requested": t.amount_requested or Decimal("0"),
                "grant_manager_display": grant_manager_display,
                "amount_awarded": amount_awarded,
                "upcoming_deadline": None,
                "notes": t.notes or "",
            }
            r["amount_rewarded"] = None
            r["agreement_code"] = None
            r["can_convert"] = t.can_convert_to_agreement()
        rows.append(r)

    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")
    trackings = list(qs[:500])

    export_format = request.GET.get("format") or ""
    if export_format:
        export_rows = [
            [
                row["pipeline_stage"], row["tracking"].code, row["tracking"].title,
                row["tracking"].donor.name if row["tracking"].donor else "",
                row["grant_type"], row["priority"], row.get("grant_manager_display") or "",
                row["submission_deadline"] or "", row["date_submitted"] or "",
                row["project_name"], row["amount_requested"], row.get("amount_awarded") or row.get("amount_rewarded") or "",
                row["notes"],
            ]
            for row in rows
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="grant_tracking",
            title="Grant Tracking (Pre-Award Pipeline)",
            headers=[
                "Pipeline Stage", "ID", "Grant Name", "Donor", "Grant Type", "Priority",
                "Grant Manager", "Submission Deadline", "Date Submitted",
                "Project Name", "Amount Requested", "Amount Awarded", "Notes",
            ],
            rows=export_rows,
        )
        if resp:
            return resp
    from tenant_grants.models import GrantTracking as GT
    return render(
        request,
        "tenant_portal/grants/grant_tracking.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "trackings": trackings,
            "donors": donors,
            "filters": f,
            "pipeline_stages": GT.PipelineStage,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


def _validate_tracking_stage(stage: str, data: dict) -> list:
    """Stage-based validation for Grant Tracking. Returns list of error messages."""
    from tenant_grants.models import GrantTracking
    errors = []
    stages_require_date_submitted = (
        GrantTracking.PipelineStage.PROPOSAL_SUBMITTED,
        GrantTracking.PipelineStage.UNDER_REVIEW,
        GrantTracking.PipelineStage.CLARIFICATION_REQUESTED,
        GrantTracking.PipelineStage.APPROVED,
    )
    if stage in (s.value for s in stages_require_date_submitted) and not data.get("date_submitted"):
        errors.append("Date submitted is required when pipeline stage is Proposal Submitted or later.")
    if stage == GrantTracking.PipelineStage.APPROVED:
        if not data.get("project_id") and not (data.get("project_name") or "").strip():
            errors.append("Project is required when stage is Approved.")
        if data.get("amount_requested") is None or (data.get("amount_requested") or 0) <= 0:
            errors.append("Amount requested must be greater than zero when stage is Approved.")
    if stage in (GrantTracking.PipelineStage.REJECTED, GrantTracking.PipelineStage.CANCELLED):
        if not (data.get("notes") or "").strip():
            errors.append("Notes are required when stage is Rejected or Cancelled.")
    return errors


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_tracking_create_view(request: HttpRequest) -> HttpResponse:
    """
    Pre-award pipeline: create a GrantTracking record (opportunity → proposal → submitted → approved/rejected).
    No bank account or award amount here; those are set when converting to Grant Agreement.
    """
    from decimal import Decimal, InvalidOperation
    from django.utils import timezone
    from django.utils.dateparse import parse_date

    from tenant_grants.models import Donor, GrantTracking, GrantTrackingDocument

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)

    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to create tracking records.")
            return redirect(reverse("tenant_portal:grants_grant_tracking_create"))

        action = (request.POST.get("action") or "").strip()
        if action == "cancel":
            return redirect(reverse("tenant_portal:grants_grant_tracking"))

        title = (request.POST.get("grant_name") or request.POST.get("title") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        pipeline_stage = (request.POST.get("pipeline_stage") or "").strip() or GrantTracking.PipelineStage.OPPORTUNITY
        grant_type = (request.POST.get("grant_type") or "").strip() or GrantTracking.GrantType.OTHER
        priority = (request.POST.get("priority") or "").strip() or GrantTracking.Priority.MEDIUM
        grant_manager_id = (request.POST.get("grant_manager_id") or "").strip()
        submission_deadline = parse_date(request.POST.get("submission_deadline") or "")
        date_submitted = parse_date(request.POST.get("date_submitted") or "")
        project_id = (request.POST.get("project_id") or "").strip()
        project_name = (request.POST.get("project_name") or "").strip()
        raw_requested = (request.POST.get("amount_requested") or "").replace(",", "").strip()
        raw_awarded = (request.POST.get("amount_awarded") or "").replace(",", "").strip()
        notes = (request.POST.get("notes") or "").strip()

        errors = []
        if not title:
            errors.append("Grant name is required.")
        elif len(title) < 3:
            errors.append("Grant name must be at least 3 characters.")
        if not donor_id:
            errors.append("Donor is required.")
        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if not donor:
            errors.append("Selected donor does not exist.")
        if not grant_type:
            errors.append("Grant type is required.")
        if not priority:
            errors.append("Priority is required.")
        if not grant_manager_id:
            errors.append("Grant manager is required.")
        if not submission_deadline:
            errors.append("Submission deadline is required.")
        amount_requested = None
        if raw_requested:
            try:
                amount_requested = Decimal(raw_requested)
                if amount_requested <= 0:
                    errors.append("Amount requested must be greater than zero.")
            except (InvalidOperation, ValueError):
                errors.append("Amount requested must be a valid number.")
        else:
            errors.append("Amount requested is required.")
        amount_awarded = None
        if raw_awarded:
            try:
                amount_awarded = Decimal(raw_awarded)
                if amount_awarded < 0:
                    errors.append("Amount awarded cannot be negative.")
                if amount_requested and amount_awarded and amount_awarded > amount_requested:
                    can_override = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
                    if not can_override:
                        errors.append("Amount awarded cannot exceed amount requested.")
            except (InvalidOperation, ValueError):
                errors.append("Amount awarded must be a valid number.")
        errors.extend(_validate_tracking_stage(pipeline_stage, {
            "date_submitted": date_submitted, "project_id": project_id, "project_name": project_name,
            "amount_requested": amount_requested, "notes": notes,
        }))

        if action == "submit_grant" and not errors:
            pipeline_stage = GrantTracking.PipelineStage.SUBMITTED
            if not date_submitted:
                date_submitted = timezone.now().date()

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            code = (request.POST.get("code") or "").strip()
            if not code:
                y = str(timezone.now().year)
                n = GrantTracking.objects.using(tenant_db).count() + 1
                code = f"TRK-{y}-{n:04d}"
            if GrantTracking.objects.using(tenant_db).filter(code__iexact=code).exists():
                messages.error(request, "Tracking code already exists. Choose a unique code.")
            else:
                grant_manager = None
                if grant_manager_id:
                    grant_manager = __import__("tenant_users.models", fromlist=["TenantUser"]).TenantUser.objects.using(tenant_db).filter(pk=grant_manager_id).first()
                project = None
                if project_id:
                    project = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(pk=project_id).first()
                tracking = GrantTracking.objects.using(tenant_db).create(
                    code=code,
                    title=title,
                    donor=donor,
                    pipeline_stage=pipeline_stage,
                    grant_type=grant_type or GrantTracking.GrantType.OTHER,
                    priority=priority or GrantTracking.Priority.MEDIUM,
                    grant_manager=grant_manager,
                    submission_deadline=submission_deadline or None,
                    date_submitted=date_submitted or None,
                    project=project,
                    project_name=project_name or (project.name if project else ""),
                    amount_requested=amount_requested,
                    amount_awarded=amount_awarded,
                    notes=notes,
                )
                doc_file = request.FILES.get("document")
                if doc_file:
                    fn = getattr(doc_file, "name", "") or ""
                    if fn.lower().endswith(".zip"):
                        GrantTrackingDocument.objects.using(tenant_db).create(
                            tracking=tracking,
                            file=doc_file,
                            original_filename=fn,
                        )
                    else:
                        messages.warning(request, "Only ZIP files are accepted. Upload skipped.")
                if action == "submit_grant":
                    messages.success(request, "Tracking submitted successfully.")
                else:
                    messages.success(request, "Tracking saved as draft.")
                return redirect(reverse("tenant_portal:grants_grant_tracking"))

    donors = list(_active_donors_queryset(tenant_db))
    projects = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(is_active=True).order_by("name")
    tenant_users = __import__("tenant_users.models", fromlist=["TenantUser"]).TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("email")
    return render(
        request,
        "tenant_portal/grants/grant_tracking_create.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "donors": donors,
            "projects": projects,
            "tenant_users": tenant_users,
            "grant_types": GrantTracking.GrantType,
            "pipeline_stages": GrantTracking.PipelineStage,
            "priority_choices": GrantTracking.Priority,
            "can_manage": can_manage,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_tracking_edit_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """Edit a GrantTracking record; enforce stage-based validation on save."""
    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date

    from tenant_grants.models import Donor, GrantTracking, GrantTrackingDocument

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).select_related("donor").first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))

    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to edit tracking.")
            return redirect(reverse("tenant_portal:grants_grant_tracking_edit", args=[tracking_id]))

        action = (request.POST.get("action") or "").strip()
        if action == "cancel":
            return redirect(reverse("tenant_portal:grants_grant_tracking"))

        title = (request.POST.get("grant_name") or request.POST.get("title") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        pipeline_stage = (request.POST.get("pipeline_stage") or "").strip() or tracking.pipeline_stage
        grant_type = (request.POST.get("grant_type") or "").strip() or GrantTracking.GrantType.OTHER
        priority = (request.POST.get("priority") or "").strip() or GrantTracking.Priority.MEDIUM
        grant_manager_id = (request.POST.get("grant_manager_id") or "").strip()
        submission_deadline = parse_date(request.POST.get("submission_deadline") or "")
        date_submitted = parse_date(request.POST.get("date_submitted") or "")
        project_id = (request.POST.get("project_id") or "").strip()
        project_name = (request.POST.get("project_name") or "").strip()
        raw_requested = (request.POST.get("amount_requested") or "").replace(",", "").strip()
        raw_awarded = (request.POST.get("amount_awarded") or "").replace(",", "").strip()
        notes = (request.POST.get("notes") or "").strip()

        errors = []
        if not title:
            errors.append("Grant name is required.")
        elif len(title) < 3:
            errors.append("Grant name must be at least 3 characters.")
        if not donor_id:
            errors.append("Donor is required.")
        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        if not donor:
            errors.append("Selected donor does not exist.")
        if not grant_manager_id:
            errors.append("Grant manager is required.")
        if not submission_deadline:
            errors.append("Submission deadline is required.")
        amount_requested = tracking.amount_requested
        if raw_requested:
            try:
                amount_requested = Decimal(raw_requested)
                if amount_requested <= 0:
                    errors.append("Amount requested must be greater than zero.")
            except (InvalidOperation, ValueError):
                errors.append("Amount requested must be a valid number.")
        amount_awarded = getattr(tracking, "amount_awarded", None)
        if raw_awarded:
            try:
                amount_awarded = Decimal(raw_awarded)
                if amount_awarded < 0:
                    errors.append("Amount awarded cannot be negative.")
                if amount_requested and amount_awarded and amount_awarded > amount_requested:
                    if not user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db):
                        errors.append("Amount awarded cannot exceed amount requested.")
            except (InvalidOperation, ValueError):
                errors.append("Amount awarded must be a valid number.")
        errors.extend(_validate_tracking_stage(pipeline_stage, {
            "date_submitted": date_submitted, "project_id": project_id, "project_name": project_name,
            "amount_requested": amount_requested, "notes": notes,
        }))

        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            grant_manager = __import__("tenant_users.models", fromlist=["TenantUser"]).TenantUser.objects.using(tenant_db).filter(pk=grant_manager_id).first() if grant_manager_id else None
            project = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(pk=project_id).first() if project_id else None
            tracking.title = title
            tracking.donor = donor
            tracking.pipeline_stage = pipeline_stage
            tracking.grant_type = grant_type or GrantTracking.GrantType.OTHER
            tracking.priority = priority or GrantTracking.Priority.MEDIUM
            tracking.grant_manager = grant_manager
            tracking.submission_deadline = submission_deadline or None
            tracking.date_submitted = date_submitted or None
            tracking.project = project
            tracking.project_name = project_name or (project.name if project else "")
            tracking.amount_requested = amount_requested
            if amount_awarded is not None:
                tracking.amount_awarded = amount_awarded
            tracking.notes = notes
            tracking.save(using=tenant_db)
            doc_file = request.FILES.get("document")
            if doc_file:
                fn = getattr(doc_file, "name", "") or ""
                if fn.lower().endswith(".zip"):
                    GrantTrackingDocument.objects.using(tenant_db).create(
                        tracking=tracking,
                        file=doc_file,
                        original_filename=fn,
                    )
            messages.success(request, "Tracking updated.")
            return redirect(reverse("tenant_portal:grants_grant_tracking"))

    donors = list(_active_donors_queryset(tenant_db))
    projects = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(is_active=True).order_by("name")
    tenant_users = __import__("tenant_users.models", fromlist=["TenantUser"]).TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("email")
    return render(
        request,
        "tenant_portal/grants/grant_tracking_edit.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "tracking": tracking,
            "donors": donors,
            "projects": projects,
            "tenant_users": tenant_users,
            "grant_types": GrantTracking.GrantType,
            "pipeline_stages": GrantTracking.PipelineStage,
            "priority_choices": GrantTracking.Priority,
            "can_manage": can_manage,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_tracking_detail_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """View a single Grant Tracking record (read-only)."""
    from tenant_grants.models import GrantTracking

    tenant_db = request.tenant_db
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).select_related(
        "donor", "grant_manager", "project"
    ).first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    return render(
        request,
        "tenant_portal/grants/grant_tracking_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "tracking": tracking,
            "can_manage": can_manage,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_grant_tracking_update_stage_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """Update pipeline stage only; notes required when Rejected or Cancelled."""
    from django.utils.dateparse import parse_date
    from tenant_grants.models import GrantTracking

    tenant_db = request.tenant_db
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).select_related("donor").first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))

    if request.method == "POST":
        new_stage = (request.POST.get("pipeline_stage") or "").strip()
        notes = (request.POST.get("notes") or "").strip()
        date_submitted = parse_date(request.POST.get("date_submitted") or "")
        project_id = request.POST.get("project_id") or None
        errors = _validate_tracking_stage(new_stage, {
            "date_submitted": date_submitted or tracking.date_submitted,
            "project_id": project_id or tracking.project_id,
            "project_name": tracking.project_name,
            "amount_requested": tracking.amount_requested,
            "notes": notes or tracking.notes,
        })
        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            tracking.pipeline_stage = new_stage
            if notes:
                tracking.notes = notes
            if date_submitted and not tracking.date_submitted:
                tracking.date_submitted = date_submitted
            if project_id and not tracking.project_id:
                tracking.project = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(pk=project_id).first()
            tracking.save(using=tenant_db)
            messages.success(request, "Pipeline stage updated.")
            return redirect(reverse("tenant_portal:grants_grant_tracking_detail", args=[tracking_id]))

    projects = __import__("tenant_grants.models", fromlist=["Project"]).Project.objects.using(tenant_db).filter(is_active=True).order_by("name")
    return render(
        request,
        "tenant_portal/grants/grant_tracking_update_stage.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "tracking": tracking,
            "pipeline_stages": GrantTracking.PipelineStage,
            "projects": projects,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_grant_tracking_delete_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """Delete a Grant Tracking record only if still in Opportunity stage."""
    from tenant_grants.models import GrantTracking

    tenant_db = request.tenant_db
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    if not tracking.can_delete():
        messages.error(request, "Only records in Opportunity stage can be deleted.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    if request.method == "POST":
        tracking.delete(using=tenant_db)
        messages.success(request, "Tracking record deleted.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    return render(
        request,
        "tenant_portal/grants/grant_tracking_delete_confirm.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "tracking": tracking,
            "active_submenu": "funds",
            "active_item": "funds_grant_tracking",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_grant_tracking_convert_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """Create a Grant Agreement from an approved GrantTracking record. GET = confirm; POST = create and redirect."""
    from django.utils import timezone
    from tenant_grants.models import Grant, GrantTracking

    tenant_db = request.tenant_db
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).select_related("donor").first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    if not tracking.can_convert_to_agreement():
        messages.error(request, "Only approved tracking records without an existing agreement can be converted.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))

    if request.method != "POST":
        return render(
            request,
            "tenant_portal/grants/grant_tracking_convert_confirm.html",
            {
                "tenant": request.tenant,
                "tenant_user": request.tenant_user,
                "tracking": tracking,
                "active_submenu": "funds",
                "active_item": "funds_grant_tracking",
            },
        )

    code = (request.POST.get("code") or "").strip() or f"GR-{tracking.code}"
    if Grant.objects.using(tenant_db).filter(code__iexact=code).exists():
        code = f"GR-{tracking.code}-{timezone.now().strftime('%Y%m%d')}"
    grant = Grant.objects.using(tenant_db).create(
        code=code,
        title=tracking.title,
        donor=tracking.donor,
        status=Grant.Status.DRAFT,
        source_tracking=tracking,
        grant_type=tracking.grant_type or Grant.GrantType.OTHER,
        priority=tracking.priority or Grant.Priority.MEDIUM,
        project_name=tracking.project_name or "",
        amount_requested=tracking.amount_requested,
        award_amount=tracking.amount_requested or 0,
        notes=tracking.notes or "",
    )
    messages.success(request, "Grant agreement created. Complete bank account, signed date, and activate below.")
    return redirect(reverse("tenant_portal:grants_agreement_create_from_tracking", args=[tracking_id]))


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_agreement_create_from_tracking_view(request: HttpRequest, tracking_id: int) -> HttpResponse:
    """Show agreement form prefilled from tracking; used after convert. Agreement already exists, redirect to edit or show form with grant_id."""
    from tenant_grants.models import Grant, GrantTracking

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    tracking = GrantTracking.objects.using(tenant_db).filter(pk=tracking_id).select_related("donor").first()
    if not tracking:
        messages.error(request, "Tracking record not found.")
        return redirect(reverse("tenant_portal:grants_grants"))
    try:
        grant = tracking.grant_agreement
    except Exception:
        grant = None
    if not grant:
        messages.error(request, "No agreement linked to this tracking. Use Convert to Grant Agreement from the tracking list.")
        return redirect(reverse("tenant_portal:grants_grant_tracking"))
    from tenant_finance.models import BankAccount
    donors = list(_active_donors_queryset(tenant_db))
    bank_accounts = BankAccount.objects.using(tenant_db).filter(is_active=True).order_by("bank_name", "account_name")
    from django.utils.dateparse import parse_date
    from decimal import Decimal, InvalidOperation

    if request.method == "POST" and can_manage:
        bank_account_id = (request.POST.get("bank_account_id") or "").strip()
        start_date = parse_date(request.POST.get("start_date") or "")
        end_date = parse_date(request.POST.get("end_date") or "")
        signed_date = parse_date(request.POST.get("signed_date") or "")
        raw_award = (request.POST.get("award_amount") or "").replace(",", "").strip()
        reporting_rules = (request.POST.get("reporting_rules") or "").strip()
        donor_restrictions = (request.POST.get("donor_restrictions") or "").strip()
        errors = []
        if not bank_account_id:
            errors.append("Bank account is required.")
        if not start_date:
            errors.append("Start date is required.")
        if end_date and start_date and end_date <= start_date:
            errors.append("End date must be later than start date.")
        award_amount = grant.award_amount or 0
        if raw_award:
            try:
                award_amount = Decimal(raw_award)
            except (InvalidOperation, ValueError):
                errors.append("Award amount must be a valid number.")
        if award_amount <= 0:
            errors.append("Award amount must be greater than zero.")
        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            bank_account = BankAccount.objects.using(tenant_db).filter(pk=bank_account_id, is_active=True).first() if bank_account_id else None
            if not bank_account:
                messages.error(request, "Selected bank account is invalid or inactive.")
            else:
                grant.bank_account = bank_account
                grant.start_date = start_date
                grant.end_date = end_date or None
                grant.signed_date = signed_date or None
                grant.award_amount = award_amount
                grant.reporting_rules = reporting_rules
                grant.donor_restrictions = donor_restrictions
                if request.POST.get("activate") == "1":
                    grant.status = Grant.Status.ACTIVE
                if request.FILES.get("signed_contract_document"):
                    grant.signed_contract_document = request.FILES["signed_contract_document"]
                grant.save(using=tenant_db)
                messages.success(request, "Grant agreement updated. It is now the official source for funds and reporting when active.")
                return redirect(reverse("tenant_portal:grants_grants"))
    return render(
        request,
        "tenant_portal/grants/agreement_from_tracking.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "tracking": tracking,
            "grant": grant,
            "donors": donors,
            "bank_accounts": bank_accounts,
            "can_manage": can_manage,
            "active_submenu": "funds",
            "active_item": "funds_grant_agreements",
        },
    )


def _parse_workplan_filters(request):
    """Parse filters for Grant Workplan list."""
    from django.utils.dateparse import parse_date
    return {
        "grant_id": request.GET.get("grant_id") or "",
        "donor_id": request.GET.get("donor_id") or "",
        "workplan_status": request.GET.get("workplan_status") or "",
        "responsible_department": (request.GET.get("responsible_department") or "").strip(),
        "activity_status": request.GET.get("activity_status") or "",
        "start_date": request.GET.get("start_date") or "",
        "end_date": request.GET.get("end_date") or "",
    }


def _grant_workplan_export_urls(request):
    """Export URLs for Grant Workplan (Excel, PDF); preserves filter params."""
    from urllib.parse import urlencode
    q = request.GET.copy()
    q["format"] = "xlsx"
    xlsx_url = request.path + "?" + q.urlencode()
    q["format"] = "pdf"
    pdf_url = request.path + "?" + q.urlencode()
    return {"xlsx": xlsx_url, "pdf": pdf_url}


def _validate_workplan_activity(
    tenant_db,
    grant,
    start_date,
    end_date,
    budget_amount,
    exclude_activity_id=None,
    project_budget_line_id=None,
):
    """Validate workplan activity: dates within grant period; total activity budgets <= grant award_amount."""
    from decimal import Decimal
    from django.db.models import Sum

    from tenant_grants.models import WorkplanActivity
    from tenant_grants.services.project_budget_actuals import project_has_budget_lines

    errors = []
    if not grant:
        return errors
    grant_start = grant.start_date
    grant_end = grant.end_date
    if start_date and grant_start and start_date < grant_start:
        errors.append("Activity start date must be on or after grant start date.")
    if start_date and grant_end and start_date > grant_end:
        errors.append("Activity start date must be on or before grant end date.")
    if end_date and grant_start and end_date < grant_start:
        errors.append("Activity end date must be on or after grant start date.")
    if end_date and grant_end and end_date > grant_end:
        errors.append("Activity end date must be on or before grant end date.")
    if start_date and end_date and end_date < start_date:
        errors.append("Activity end date must be after start date.")
    grant_total = grant.award_amount or Decimal("0")
    qs = WorkplanActivity.objects.using(tenant_db).filter(grant=grant)
    if exclude_activity_id:
        qs = qs.exclude(pk=exclude_activity_id)
    existing_sum = qs.aggregate(s=Sum("budget_amount")).get("s") or Decimal("0")
    new_budget = budget_amount or Decimal("0")
    if grant_total > 0 and (existing_sum + new_budget > grant_total):
        errors.append(
            f"Total workplan budget would exceed grant allocation ({grant_total}). "
            f"Current activities total {existing_sum}; this activity would add {new_budget}."
        )
    if grant.project_id and project_has_budget_lines(tenant_db, grant.project_id):
        if not project_budget_line_id:
            errors.append(
                "Select a project budget line for this activity (project has a budget structure)."
            )
    return errors


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_workplan_view(request: HttpRequest) -> HttpResponse:
    """Grant workplan: filter panel, workplan table, Create workplan and Raise PR forms."""
    from django.core.exceptions import ValidationError
    from django.utils.dateparse import parse_date
    from decimal import Decimal
    from tenant_grants.models import Donor, Grant, ProjectBudgetLine, WorkplanActivity

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        if form_type == "create_workplan" and can_manage:
            grant_id = request.POST.get("grant_id")
            title = (request.POST.get("workplan_title") or "").strip()
            donor_id = request.POST.get("donor_id") or None
            component = (request.POST.get("workplan_component") or "").strip()
            budget_line = (request.POST.get("workplan_budget_line") or "").strip()
            pbl_raw = (request.POST.get("workplan_project_budget_line") or "").strip()
            procurement_req = (request.POST.get("workplan_procurement") or "").strip()
            dept = (request.POST.get("workplan_department") or "").strip()
            staff = (request.POST.get("workplan_staff") or "").strip()
            start_date = parse_date(request.POST.get("workplan_start_date") or "")
            end_date = parse_date(request.POST.get("workplan_end_date") or "")
            budget = request.POST.get("workplan_budget") or None
            notes = (request.POST.get("workplan_description") or "").strip()
            if grant_id and title:
                grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
                if grant:
                    budget_val = Decimal("0")
                    if budget:
                        try:
                            budget_val = Decimal(str(budget).replace(",", ""))
                        except Exception:
                            pass
                    pbl = (
                        ProjectBudgetLine.objects.using(tenant_db)
                        .select_related("project_budget")
                        .filter(pk=int(pbl_raw))
                        .first()
                        if pbl_raw.isdigit()
                        else None
                    )
                    errs = _validate_workplan_activity(
                        tenant_db,
                        grant,
                        start_date,
                        end_date,
                        budget_val,
                        exclude_activity_id=None,
                        project_budget_line_id=pbl.pk if pbl else None,
                    )
                    if errs:
                        for e in errs:
                            messages.error(request, e)
                    else:
                        donor = None
                        if donor_id:
                            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
                        if not donor:
                            donor = grant.donor
                        act = WorkplanActivity(
                            grant=grant,
                            donor=donor,
                            activity=title,
                            component_output=component,
                            budget_line=budget_line,
                            procurement_requirement=procurement_req,
                            responsible_department=dept,
                            responsible_staff=staff,
                            start_date=start_date or None,
                            end_date=end_date or None,
                            budget_amount=budget_val,
                            notes=notes,
                            description=notes,
                            project_budget_line=pbl,
                        )
                        try:
                            act.full_clean()
                        except ValidationError as ve:
                            for _f, msg_list in ve.error_dict.items():
                                for msg in msg_list:
                                    messages.error(request, str(msg))
                        else:
                            act.save(using=tenant_db)
                            messages.success(request, "Workplan activity created.")
                else:
                    messages.error(request, "Grant not found.")
            else:
                messages.error(request, "Grant and activity name are required.")
        elif form_type == "approve_for_pr" and can_manage:
            activity_id = request.POST.get("activity_id")
            if activity_id:
                act = WorkplanActivity.objects.using(tenant_db).filter(pk=activity_id).first()
                if act:
                    act.approved_for_pr = True
                    act.save(using=tenant_db, update_fields=["approved_for_pr"])
                    messages.success(request, f"Activity {act.workplan_code} approved for PR.")
                else:
                    messages.error(request, "Activity not found.")
            else:
                messages.error(request, "Activity is required.")
        elif form_type == "raise_pr":
            activity_id = request.POST.get("pr_activity_id") or request.POST.get("pr_grant_id")
            if activity_id:
                return redirect(
                    reverse("tenant_portal:grants_pr_create") + "?activity_id=" + str(activity_id)
                )
            messages.error(request, "Select a workplan activity to raise a PR.")
        elif form_type == "import_workplan" and can_manage:
            from tenant_grants.models import Donor, ProjectBudgetLine
            from django.db.models import Sum
            import csv
            import io
            upload = request.FILES.get("workplan_import_file")
            if not upload:
                messages.error(request, "Please select a file to import.")
            else:
                def _cell(row, idx):
                    v = (row + (None,) * (idx + 1))[idx] if row else None
                    return (str(v).strip() if v is not None and str(v).strip() else "") or None

                def _parse_date(s):
                    if not s:
                        return None
                    return parse_date(s) or parse_date(s.replace(".", "-").replace("/", "-"))

                rows_data = []
                filename = (getattr(upload, "name", "") or "").lower()
                try:
                    if filename.endswith(".csv"):
                        content = upload.read()
                        if isinstance(content, bytes):
                            content = content.decode("utf-8-sig", errors="replace")
                        reader = csv.reader(io.StringIO(content))
                        raw_rows = list(reader)
                        if not raw_rows:
                            raise ValueError("CSV is empty")
                        headers = [h.strip().lower().replace(" ", "_").replace("/", "_") for h in raw_rows[0]]
                        for r in raw_rows[1:]:
                            row = {}
                            for i, h in enumerate(headers):
                                if i < len(r):
                                    row[h] = (r[i] or "").strip()
                                else:
                                    row[h] = ""
                            rows_data.append(row)
                    else:
                        import openpyxl
                        wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
                        ws = wb.active
                        raw_rows = list(ws.iter_rows(values_only=True))
                        wb.close()
                        if not raw_rows:
                            raise ValueError("Excel sheet is empty")
                        headers = [str(h).strip().lower().replace(" ", "_").replace("/", "_") if h else "" for h in raw_rows[0]]
                        for r in raw_rows[1:]:
                            row = {}
                            for i, h in enumerate(headers):
                                if h:
                                    row[h] = (str(r[i]).strip() if r and i < len(r) and r[i] is not None else "") or ""
                                else:
                                    row[h] = ""
                            rows_data.append(row)
                except Exception as e:
                    messages.error(request, f"Could not read file: {e}")
                    return redirect(reverse("tenant_portal:grants_grant_workplan") + "?" + request.GET.urlencode())

                def _get(row, *keys, default=""):
                    for k in keys:
                        v = row.get(k)
                        if v is not None and str(v).strip():
                            return str(v).strip()
                    return default

                created = 0
                batch_budget_by_grant = {}
                for idx, row in enumerate(rows_data):
                    grant_code = _get(row, "grant", "grant_code")
                    activity_title = _get(row, "activity", "activity_name", "title")
                    if not grant_code or not activity_title:
                        continue
                    grant = Grant.objects.using(tenant_db).filter(code=grant_code).first()
                    if not grant:
                        messages.warning(request, f"Row {idx + 2}: Grant '{grant_code}' not found; skipped.")
                        continue
                    donor = None
                    donor_val = _get(row, "donor")
                    if donor_val:
                        donor = Donor.objects.using(tenant_db).filter(name__iexact=donor_val).first() or Donor.objects.using(tenant_db).filter(code__iexact=donor_val).first()
                    if not donor:
                        donor = grant.donor
                    start_date = _parse_date(_get(row, "start_date", "start"))
                    end_date = _parse_date(_get(row, "end_date", "end"))
                    budget_val = Decimal("0")
                    try:
                        b = _get(row, "budget_amount", "budget", "amount")
                        if b:
                            budget_val = Decimal(str(b).replace(",", ""))
                    except Exception:
                        pass
                    pbl_imp = _get(row, "project_budget_line_id", "project_budget_line")
                    pbl_imp_obj = (
                        ProjectBudgetLine.objects.using(tenant_db)
                        .filter(pk=int(pbl_imp))
                        .first()
                        if str(pbl_imp).isdigit()
                        else None
                    )
                    errs = _validate_workplan_activity(
                        tenant_db,
                        grant,
                        start_date,
                        end_date,
                        budget_val,
                        exclude_activity_id=None,
                        project_budget_line_id=pbl_imp_obj.pk if pbl_imp_obj else None,
                    )
                    existing_sum = batch_budget_by_grant.get(grant.id)
                    if existing_sum is None:
                        existing_sum = WorkplanActivity.objects.using(tenant_db).filter(grant=grant).aggregate(s=Sum("budget_amount")).get("s") or Decimal("0")
                    if grant.award_amount and (existing_sum + budget_val > grant.award_amount):
                        errs = errs or []
                        errs.append(f"Total workplan budget would exceed grant allocation ({grant.award_amount}).")
                    if errs:
                        for e in errs:
                            messages.warning(request, f"Row {idx + 2}: {e}")
                        continue
                    activity_status_val = _get(row, "activity_status", "status").lower() or "planned"
                    if activity_status_val not in ("planned", "in_progress", "completed"):
                        activity_status_val = "planned"
                    imp_act = WorkplanActivity(
                        grant=grant,
                        donor=donor,
                        activity=activity_title,
                        component_output=_get(row, "component_output", "component", "output"),
                        budget_line=_get(row, "budget_line", "budget_line"),
                        procurement_requirement=_get(row, "procurement", "procurement_requirement"),
                        responsible_department=_get(row, "responsible_department", "department", "dept"),
                        responsible_staff=_get(row, "responsible_staff", "staff"),
                        start_date=start_date or None,
                        end_date=end_date or None,
                        budget_amount=budget_val,
                        activity_status=activity_status_val,
                        notes="",
                        project_budget_line=pbl_imp_obj,
                    )
                    try:
                        imp_act.full_clean()
                    except ValidationError as ve:
                        for _f, msg_list in ve.error_dict.items():
                            for msg in msg_list:
                                messages.warning(request, f"Row {idx + 2}: {msg}")
                        continue
                    imp_act.save(using=tenant_db)
                    batch_budget_by_grant[grant.id] = existing_sum + budget_val
                    created += 1
                if created:
                    messages.success(request, f"Imported {created} workplan activity(ies).")
                elif not rows_data:
                    messages.warning(request, "No valid rows to import. Check file format and column names.")
                return redirect(reverse("tenant_portal:grants_grant_workplan") + "?" + request.GET.urlencode())
        return redirect(reverse("tenant_portal:grants_grant_workplan") + "?" + request.GET.urlencode())

    # Template download (GET)
    if request.method == "GET" and request.GET.get("download_template"):
        from django.http import HttpResponse
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workplan import"
        headers = [
            "Grant",
            "Donor",
            "Activity",
            "Component/Output",
            "Budget line",
            "Project_budget_line_id",
            "Procurement",
            "Responsible department",
            "Responsible staff",
            "Start date",
            "End date",
            "Budget amount",
            "Activity status",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        sample = [
            "GRANT-001",
            "Donor name",
            "Q1 field activities",
            "Output 1",
            "Travel",
            "",
            "Vehicle hire",
            "Programme",
            "John Doe",
            "2025-01-01",
            "2025-03-31",
            "5000.00",
            "planned",
        ]
        for col, v in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=v)
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="workplan_import_template.xlsx"'
        wb.save(resp)
        return resp

    f = _parse_workplan_filters(request)
    qs = (
        WorkplanActivity.objects.using(tenant_db)
        .select_related("grant", "donor", "project_budget_line", "project_budget_line__project_budget__project")
        .order_by("-created_at")
    )
    if f["grant_id"]:
        qs = qs.filter(grant_id=f["grant_id"])
    if f["donor_id"]:
        from django.db.models import Q
        qs = qs.filter(Q(donor_id=f["donor_id"]) | Q(grant__donor_id=f["donor_id"]))
    if f["workplan_status"]:
        qs = qs.filter(workplan_status=f["workplan_status"])
    if f["responsible_department"]:
        qs = qs.filter(responsible_department=f["responsible_department"])
    if f["activity_status"]:
        qs = qs.filter(activity_status=f["activity_status"])
    if f["start_date"]:
        qs = qs.filter(start_date__gte=f["start_date"])
    if f["end_date"]:
        qs = qs.filter(end_date__lte=f["end_date"])
    workplans = list(qs[:500])

    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:200])
    donors = list(_active_donors_queryset(tenant_db))
    departments = list(
        WorkplanActivity.objects.using(tenant_db)
        .values_list("responsible_department", flat=True)
        .distinct()
    )
    departments = [d for d in departments if d]

    export_format = request.GET.get("format") or ""
    if export_format in ("xlsx", "pdf"):
        rows = [
            [
                w.workplan_code,
                w.grant.code if w.grant else "",
                w.donor_display(),
                w.activity,
                w.component_output,
                (w.project_budget_line.category if w.project_budget_line_id else "") or (w.budget_line or ""),
                w.procurement_requirement[:80] if w.procurement_requirement else "",
                w.responsible_department,
                w.responsible_staff,
                w.start_date or "",
                w.end_date or "",
                w.budget_amount or "",
                w.actual_cost or 0,
                "Yes" if w.approved_for_pr else "No",
                w.get_activity_status_display(),
            ]
            for w in workplans
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="grant_workplan",
            title="Grant Workplan",
            headers=[
                "Workplan ID",
                "Grant",
                "Donor",
                "Activity",
                "Component/Output",
                "Budget line (project / text)",
                "Procurement",
                "Responsible Department",
                "Responsible Staff",
                "Start Date",
                "End Date",
                "Planned amount",
                "Actual cost",
                "Approved for PR",
                "Activity Status",
            ],
            rows=rows,
        )
        if resp:
            return resp

    export_urls = _grant_workplan_export_urls(request)
    workplan_import_template_url = reverse("tenant_portal:grants_grant_workplan") + "?download_template=1"
    project_budget_lines = list(
        ProjectBudgetLine.objects.using(tenant_db)
        .select_related("project_budget__project")
        .order_by("project_budget__project__code", "id")[:500]
    )
    return render(
        request,
        "tenant_portal/grants/grant_workplan.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_grant_workplan",
            "workplans": workplans,
            "grants": grants,
            "donors": donors,
            "departments": departments,
            "filters": f,
            "export_xlsx_url": export_urls["xlsx"],
            "workplan_import_template_url": workplan_import_template_url,
            "workplan_status_choices": WorkplanActivity.WorkplanStatus.choices,
            "activity_status_choices": WorkplanActivity.ActivityStatus.choices,
            "can_manage": can_manage,
            "project_budget_lines": project_budget_lines,
            "approved_activities_for_pr": [
                a for a in workplans if getattr(a, "approved_for_pr", False)
            ],
        },
    )


def _parse_pr_filters(request):
    from django.utils.dateparse import parse_date
    return {
        "grant_id": request.GET.get("grant_id") or "",
        "donor_id": request.GET.get("donor_id") or "",
        "status": request.GET.get("status") or "",
        "activity_id": request.GET.get("activity_id") or "",
        "date_from": request.GET.get("date_from") or "",
        "date_to": request.GET.get("date_to") or "",
    }


def _validate_pr_submission(pr, tenant_db):
    """
    Validate PR for submit: active grant, approved workplan activity, valid budget line,
    sufficient available allocation, PR date within grant period. Returns list of error strings.
    """
    from decimal import Decimal
    from tenant_grants.models import Grant
    errors = []
    if not pr.grant_id:
        errors.append("PR must be linked to a grant.")
        return errors
    grant = pr.grant if hasattr(pr, "_grant") else Grant.objects.using(tenant_db).filter(pk=pr.grant_id).first()
    if grant and getattr(grant, "status", None) != Grant.Status.ACTIVE:
        errors.append("Grant must be active.")
    if not pr.workplan_activity_id:
        errors.append("PR must be linked to a workplan activity.")
    else:
        act = pr.workplan_activity if getattr(pr, "_activity", None) else None
        if not act:
            from tenant_grants.models import WorkplanActivity
            act = WorkplanActivity.objects.using(tenant_db).filter(pk=pr.workplan_activity_id).first()
        if act and not getattr(act, "approved_for_pr", False):
            errors.append("Workplan activity must be approved for PR.")
        if act:
            remaining = act.remaining_budget_for_pr(using=tenant_db)
            this_total = getattr(pr, "effective_total", lambda: pr.estimated_total_cost or Decimal("0"))()
            if this_total > remaining:
                errors.append(
                    f"PR total ({this_total}) exceeds activity remaining budget ({remaining})."
                )
    if not (pr.budget_line or "").strip():
        errors.append("Budget line is required.")
    if grant and pr.pr_date:
        if grant.start_date and pr.pr_date < grant.start_date:
            errors.append("PR date must be on or after grant start date.")
        if grant.end_date and pr.pr_date > grant.end_date:
            errors.append("PR date must be on or before grant end date.")
    # Grant allocation: total PR value for this grant (excluding rejected/cancelled) must not exceed award_amount
    if grant and (grant.award_amount or 0) > 0:
        from django.db.models import Sum
        from tenant_grants.models import PurchaseRequisition as PRModel
        other_pr_total = (
            PRModel.objects.using(tenant_db)
            .filter(grant_id=grant.id)
            .exclude(status__in=(PRModel.Status.REJECTED, PRModel.Status.CANCELLED))
            .exclude(pk=pr.pk)
            .aggregate(s=Sum("estimated_total_cost"))
            .get("s")
        ) or Decimal("0")
        this_total = getattr(pr, "effective_total", lambda: pr.estimated_total_cost or Decimal("0"))()
        if other_pr_total + this_total > grant.award_amount:
            errors.append(
                f"Total PR value for this grant would exceed grant allocation ({grant.award_amount}). "
                f"Existing PRs total {other_pr_total}; this PR adds {this_total}."
            )
    return errors


def _log_pr_status(pr, from_status, to_status, performed_by, comment, using):
    from tenant_grants.models import PurchaseRequisitionStatusLog
    PurchaseRequisitionStatusLog.objects.using(using).create(
        pr=pr,
        from_status=from_status or "",
        to_status=to_status,
        performed_by=performed_by,
        comment=comment or "",
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_pr_list_view(request: HttpRequest) -> HttpResponse:
    """Purchase Requisition list: filters, status badges, export, action menu."""
    from tenant_grants.models import PurchaseRequisition, Grant
    from django.db.models import Q

    tenant_db = request.tenant_db
    f = _parse_pr_filters(request)
    qs = (
        PurchaseRequisition.objects.using(tenant_db)
        .select_related("grant", "donor", "workplan_activity")
        .order_by("-pr_date", "-created_at")
    )
    if f["grant_id"]:
        qs = qs.filter(grant_id=f["grant_id"])
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    if f["status"]:
        qs = qs.filter(status=f["status"])
    if f["activity_id"]:
        qs = qs.filter(workplan_activity_id=f["activity_id"])
    if f["date_from"]:
        qs = qs.filter(pr_date__gte=f["date_from"])
    if f["date_to"]:
        qs = qs.filter(pr_date__lte=f["date_to"])
    prs = list(qs[:300])
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:200])
    donors = list(_active_donors_queryset(tenant_db))
    export_format = request.GET.get("format") or ""
    if export_format in ("xlsx", "pdf"):
        rows = [
            [
                pr.pr_number,
                pr.pr_date,
                pr.grant.code if pr.grant else "",
                pr.donor.name if pr.donor else "",
                pr.workplan_activity.workplan_code if pr.workplan_activity else "",
                pr.budget_line,
                pr.item_description[:80] if pr.item_description else "",
                pr.quantity,
                pr.estimated_unit_cost,
                pr.estimated_total_cost,
                pr.get_procurement_method_display() if pr.procurement_method else "",
                pr.get_priority_display(),
                pr.get_status_display(),
            ]
            for pr in prs
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="purchase_requisitions",
            title="Purchase Requisitions",
            headers=[
                "PR Number", "PR Date", "Grant", "Donor", "Workplan", "Budget Line",
                "Item", "Qty", "Unit Cost", "Total", "Procurement", "Priority", "Status",
            ],
            rows=rows,
        )
        if resp:
            return resp
    q = request.GET.copy()
    q["format"] = "xlsx"
    export_xlsx_url = request.path + "?" + q.urlencode()
    q["format"] = "pdf"
    export_pdf_url = request.path + "?" + q.urlencode()
    return render(
        request,
        "tenant_portal/grants/purchase_requisitions.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_pr",
            "prs": prs,
            "grants": grants,
            "donors": donors,
            "filters": f,
            "export_xlsx_url": export_xlsx_url,
            "export_pdf_url": export_pdf_url,
            "pr_status_choices": PurchaseRequisition.Status.choices,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_pr_create_view(request: HttpRequest) -> HttpResponse:
    """Create PR from an approved workplan activity. Validates: activity approved, PR total <= activity remaining budget."""
    from django.utils.dateparse import parse_date
    from django.utils import timezone
    from decimal import Decimal
    from tenant_grants.models import (
        PurchaseRequisition,
        WorkplanActivity,
        Grant,
    )

    tenant_db = request.tenant_db
    activity_id = request.GET.get("activity_id") or request.POST.get("workplan_activity_id")
    activity = None
    if activity_id:
        activity = (
            WorkplanActivity.objects.using(tenant_db)
            .select_related("grant", "donor")
            .filter(pk=activity_id)
            .first()
        )
    grants = list(Grant.objects.using(tenant_db).select_related("donor").order_by("code")[:200])
    donors = list(_active_donors_queryset(tenant_db))
    approved_activities = list(
        WorkplanActivity.objects.using(tenant_db)
        .filter(approved_for_pr=True)
        .select_related("grant", "donor")
        .order_by("-created_at")[:200]
    )
    for a in approved_activities:
        a.remaining_budget_display = a.remaining_budget_for_pr(using=tenant_db)

    if request.method == "POST":
        wid = request.POST.get("workplan_activity_id")
        act = None
        if wid:
            act = WorkplanActivity.objects.using(tenant_db).select_related("grant").filter(pk=wid).first()
        if not act:
            messages.error(request, "Select an approved workplan activity.")
            return redirect(reverse("tenant_portal:grants_pr_create"))
        if not act.approved_for_pr:
            messages.error(request, "Only activities approved for PR can be used. Approve the activity from Grant Workplan first.")
            return redirect(reverse("tenant_portal:grants_pr_create"))
        grant = act.grant
        if grant and getattr(grant, "status", None) != Grant.Status.ACTIVE:
            messages.error(request, "Grant must be active to raise a PR.")
            return redirect(reverse("tenant_portal:grants_pr_create"))
        pr_date = parse_date(request.POST.get("pr_date") or "")
        if not pr_date:
            pr_date = timezone.now().date()
        if grant:
            if grant.start_date and pr_date < grant.start_date:
                messages.error(request, "PR date must be on or after grant start date.")
                return redirect(reverse("tenant_portal:grants_pr_create"))
            if grant.end_date and pr_date > grant.end_date:
                messages.error(request, "PR date must be on or before grant end date.")
                return redirect(reverse("tenant_portal:grants_pr_create"))
        # Parse line items: line_0_item, line_0_qty, line_0_unit_cost, line_0_budget_line, line_1_*, ...
        from tenant_grants.models import PurchaseRequisitionLine
        lines_data = []
        for i in range(200):
            item = (request.POST.get(f"line_{i}_item") or "").strip()
            if not item and i == 0:
                # Legacy single-line from main form
                item = (request.POST.get("item_description") or "").strip()
                if not item:
                    break
                qty = Decimal("1")
                try:
                    qty = Decimal(str(request.POST.get("quantity") or "1").replace(",", ""))
                except Exception:
                    pass
                unit = Decimal("0")
                try:
                    unit = Decimal(str(request.POST.get("estimated_unit_cost") or "0").replace(",", ""))
                except Exception:
                    pass
                bl = (request.POST.get("budget_line") or act.budget_line or "").strip()
                lines_data.append({"item": item, "quantity": qty, "unit_cost": unit, "budget_line": bl})
                break
            if not item:
                continue
            qty = Decimal("1")
            try:
                qty = Decimal(str(request.POST.get(f"line_{i}_qty") or "1").replace(",", ""))
            except Exception:
                pass
            unit = Decimal("0")
            try:
                unit = Decimal(str(request.POST.get(f"line_{i}_unit_cost") or "0").replace(",", ""))
            except Exception:
                pass
            bl = (request.POST.get(f"line_{i}_budget_line") or "").strip()
            lines_data.append({"item": item, "quantity": qty, "unit_cost": unit, "budget_line": bl})
        if not lines_data:
            messages.error(request, "Add at least one line with an item description.")
            return redirect(reverse("tenant_portal:grants_pr_create"))
        budget_line_val = lines_data[0].get("budget_line") or (request.POST.get("budget_line") or act.budget_line or "").strip()
        if not budget_line_val:
            budget_line_val = (act.budget_line or "").strip()
        total = sum(d["quantity"] * d["unit_cost"] for d in lines_data)
        remaining = act.remaining_budget_for_pr(using=tenant_db)
        _do_pr_create = False
        if total > remaining:
            messages.error(
                request,
                f"PR total ({total}) exceeds activity remaining budget ({remaining}). "
                f"Activity budget: {act.budget_amount or 0}; already raised PRs total: {act.total_pr_value(using=tenant_db)}.",
            )
        elif grant and (grant.award_amount or 0) > 0:
            from django.db.models import Sum
            existing_pr_total = (
                PurchaseRequisition.objects.using(tenant_db)
                .filter(grant=grant)
                .exclude(status__in=(PurchaseRequisition.Status.REJECTED, PurchaseRequisition.Status.CANCELLED))
                .aggregate(s=Sum("estimated_total_cost"))
                .get("s")
            ) or Decimal("0")
            # Use effective totals: for existing PRs with lines we'd need to sum their lines; here we use header sum which may undercount. For simplicity we use aggregate on estimated_total_cost; you could add a DB function later.
            if existing_pr_total + total > grant.award_amount:
                messages.error(
                    request,
                    f"PR total ({total}) would exceed grant allocation ({grant.award_amount}). "
                    f"Existing PRs for this grant total {existing_pr_total}.",
                )
            else:
                _do_pr_create = True
        else:
            _do_pr_create = True
        if _do_pr_create:
            from tenant_grants.restrictions import evaluate_procurement_restrictions

            donor_chk = act.donor or (act.grant.donor if act.grant else None)
            if donor_chk and grant:
                override_dn = user_has_permission(
                    request.tenant_user, "grants:donor_restrictions.manage", using=tenant_db
                )
                pv = evaluate_procurement_restrictions(
                    using=tenant_db,
                    donor_id=donor_chk.pk,
                    grant_id=grant.pk,
                    estimated_amount=total,
                    project_id=grant.project_id,
                    funding_source_id=None,
                    has_override_permission=override_dn,
                )
                for v in pv:
                    if not v.blocks_posting and v.compliance_level == "recommended":
                        messages.warning(request, v.message)
                hard = [v for v in pv if v.blocks_posting]
                if hard:
                    messages.error(request, hard[0].message)
                    return redirect(reverse("tenant_portal:grants_pr_create"))
            last = (
                PurchaseRequisition.objects.using(tenant_db)
                .order_by("-id")
                .values_list("id", flat=True)
                .first()
            )
            next_num = (last or 0) + 1
            pr_number = f"PR-{next_num:05d}"
            while PurchaseRequisition.objects.using(tenant_db).filter(pr_number=pr_number).exists():
                next_num += 1
                pr_number = f"PR-{next_num:05d}"
            donor = act.donor or (act.grant.donor if act.grant else None)
            delivery_date_val = parse_date(request.POST.get("delivery_date") or "") or None
            pr_obj = PurchaseRequisition.objects.using(tenant_db).create(
                pr_number=pr_number,
                pr_date=pr_date,
                grant=act.grant,
                donor=donor,
                workplan_activity=act,
                budget_line=budget_line_val,
                item_description=lines_data[0]["item"][:255] if lines_data else "",
                quantity=Decimal("0"),
                estimated_unit_cost=Decimal("0"),
                estimated_total_cost=total,
                procurement_method=request.POST.get("procurement_method") or PurchaseRequisition.ProcurementMethod.OTHER,
                priority=request.POST.get("priority") or PurchaseRequisition.Priority.MEDIUM,
                delivery_date=delivery_date_val,
                justification=(request.POST.get("justification") or "").strip(),
                status=PurchaseRequisition.Status.DRAFT,
                requested_by=request.tenant_user,
            )
            for idx, d in enumerate(lines_data):
                PurchaseRequisitionLine.objects.using(tenant_db).create(
                    pr=pr_obj,
                    line_number=idx + 1,
                    item_description=d["item"],
                    quantity=d["quantity"],
                    estimated_unit_cost=d["unit_cost"],
                    budget_line=d.get("budget_line") or "",
                )
            from tenant_grants.models import PurchaseRequisitionAttachment
            for f in request.FILES.getlist("attachment"):
                if f:
                    PurchaseRequisitionAttachment.objects.using(tenant_db).create(
                        pr=pr_obj,
                        file=f,
                        original_filename=getattr(f, "name", "") or "",
                        uploaded_by=request.tenant_user,
                    )
            messages.success(request, f"Purchase Requisition {pr_number} created with {len(lines_data)} line(s).")
            return redirect(reverse("tenant_portal:grants_pr_list"))
    return render(
        request,
        "tenant_portal/grants/purchase_requisition_create.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_pr",
            "activity": activity,
            "grants": grants,
            "donors": donors,
            "approved_activities": approved_activities,
            "procurement_methods": PurchaseRequisition.ProcurementMethod.choices,
            "pr_priorities": PurchaseRequisition.Priority.choices,
            "today": timezone.now().date(),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_pr_detail_view(request: HttpRequest, pr_id: int) -> HttpResponse:
    """PR detail: view, audit trail, and workflow actions (submit, approve, reject, return, procurement, cancel)."""
    from django.utils import timezone
    from tenant_grants.models import (
        PurchaseRequisition,
        PurchaseRequisitionStatusLog,
        Grant,
    )

    tenant_db = request.tenant_db
    from tenant_grants.models import PurchaseRequisitionLine
    pr = (
        PurchaseRequisition.objects.using(tenant_db)
        .select_related("grant", "donor", "workplan_activity", "requested_by", "line_manager_approved_by", "procurement_officer", "cancelled_by")
        .prefetch_related("lines", "attachments")
        .filter(pk=pr_id)
        .first()
    )
    if not pr:
        from django.http import Http404
        raise Http404("Purchase requisition not found.")

    can_manage = user_has_permission(request.tenant_user, "module:grants.manage", using=tenant_db)
    can_line_manager = user_has_permission(request.tenant_user, "module:grants.pr_line_manager_approve", using=tenant_db)
    can_procurement = user_has_permission(request.tenant_user, "module:grants.pr_procurement_process", using=tenant_db)
    is_requester = bool(
        request.tenant_user and pr.requested_by_id and request.tenant_user.id == pr.requested_by_id
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "submit":
            if pr.status != PurchaseRequisition.Status.DRAFT:
                messages.error(request, "Only draft PRs can be submitted.")
            elif not (can_manage or is_requester):
                messages.error(request, "You do not have permission to submit this PR.")
            else:
                errs = _validate_pr_submission(pr, tenant_db)
                if errs:
                    for e in errs:
                        messages.error(request, e)
                else:
                    old = pr.status
                    pr.status = PurchaseRequisition.Status.PENDING_LINE_MANAGER_APPROVAL
                    pr.submitted_at = timezone.now()
                    pr.save(using=tenant_db, update_fields=["status", "submitted_at"])
                    _log_pr_status(pr, old, pr.status, request.tenant_user, "Submitted for line manager approval.", tenant_db)
                    messages.success(request, "PR submitted for line manager approval.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "approve":
            if pr.status != PurchaseRequisition.Status.PENDING_LINE_MANAGER_APPROVAL:
                messages.error(request, "Only PRs pending line manager approval can be approved.")
            elif not can_line_manager:
                messages.error(request, "You do not have permission to approve PRs (Line Manager).")
            else:
                old = pr.status
                pr.status = PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER
                pr.line_manager_approved_at = timezone.now()
                pr.line_manager_approved_by = request.tenant_user
                pr.line_manager_rejection_comment = ""
                pr.line_manager_return_comment = ""
                pr.save(using=tenant_db, update_fields=["status", "line_manager_approved_at", "line_manager_approved_by_id", "line_manager_rejection_comment", "line_manager_return_comment"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, request.POST.get("comment") or "Approved by line manager.", tenant_db)
                messages.success(request, "PR approved by line manager.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "reject":
            if pr.status != PurchaseRequisition.Status.PENDING_LINE_MANAGER_APPROVAL:
                messages.error(request, "Only PRs pending line manager approval can be rejected.")
            elif not can_line_manager:
                messages.error(request, "You do not have permission to reject PRs (Line Manager).")
            else:
                comment = (request.POST.get("comment") or "").strip()
                if not comment:
                    messages.error(request, "Rejection reason is required.")
                else:
                    old = pr.status
                    pr.status = PurchaseRequisition.Status.REJECTED
                    pr.line_manager_rejection_comment = comment
                    pr.save(using=tenant_db, update_fields=["status", "line_manager_rejection_comment"])
                    _log_pr_status(pr, old, pr.status, request.tenant_user, comment, tenant_db)
                    messages.success(request, "PR rejected.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "return":
            if pr.status != PurchaseRequisition.Status.PENDING_LINE_MANAGER_APPROVAL:
                messages.error(request, "Only PRs pending line manager approval can be returned.")
            elif not can_line_manager:
                messages.error(request, "You do not have permission to return PRs (Line Manager).")
            else:
                comment = (request.POST.get("comment") or "").strip()
                if not comment:
                    messages.error(request, "Return reason is required.")
                else:
                    old = pr.status
                    pr.status = PurchaseRequisition.Status.DRAFT
                    pr.line_manager_return_comment = comment
                    pr.save(using=tenant_db, update_fields=["status", "line_manager_return_comment"])
                    _log_pr_status(pr, old, pr.status, request.tenant_user, comment, tenant_db)
                    messages.success(request, "PR returned for correction.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "assign_to_procurement":
            if pr.status != PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER:
                messages.error(request, "Only PRs approved by line manager can be assigned to procurement.")
            elif not can_procurement:
                messages.error(request, "You do not have permission to process PRs (Procurement Officer).")
            else:
                old = pr.status
                pr.status = PurchaseRequisition.Status.ASSIGNED_TO_PROCUREMENT
                pr.assigned_to_procurement_at = timezone.now()
                pr.procurement_officer = request.tenant_user
                pr.save(using=tenant_db, update_fields=["status", "assigned_to_procurement_at", "procurement_officer_id"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, "Assigned to procurement.", tenant_db)
                messages.success(request, "PR assigned to procurement.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "under_processing":
            if pr.status not in (PurchaseRequisition.Status.ASSIGNED_TO_PROCUREMENT, PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER):
                messages.error(request, "PR must be assigned to procurement first.")
            elif not can_procurement:
                messages.error(request, "You do not have permission to process PRs.")
            else:
                old = pr.status
                pr.status = PurchaseRequisition.Status.UNDER_PROCUREMENT_PROCESSING
                if not pr.assigned_to_procurement_at:
                    pr.assigned_to_procurement_at = timezone.now()
                    pr.procurement_officer = request.tenant_user
                pr.save(using=tenant_db, update_fields=["status", "assigned_to_procurement_at", "procurement_officer_id"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, "Under procurement processing.", tenant_db)
                messages.success(request, "PR marked under procurement processing.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "po_issued":
            if pr.status != PurchaseRequisition.Status.UNDER_PROCUREMENT_PROCESSING:
                messages.error(request, "PR must be under procurement processing before marking PO issued.")
            elif not can_procurement:
                messages.error(request, "You do not have permission to process PRs.")
            else:
                old = pr.status
                pr.status = PurchaseRequisition.Status.PO_ISSUED
                pr.po_issued_at = timezone.now()
                pr.save(using=tenant_db, update_fields=["status", "po_issued_at"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, request.POST.get("comment") or "PO issued.", tenant_db)
                messages.success(request, "PO issued.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "fulfilled":
            if pr.status != PurchaseRequisition.Status.PO_ISSUED:
                messages.error(request, "PR must have PO issued before marking fulfilled.")
            elif not can_procurement:
                messages.error(request, "You do not have permission to process PRs.")
            else:
                old = pr.status
                pr.status = PurchaseRequisition.Status.FULFILLED
                pr.fulfilled_at = timezone.now()
                pr.save(using=tenant_db, update_fields=["status", "fulfilled_at"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, request.POST.get("comment") or "Fulfilled.", tenant_db)
                messages.success(request, "PR fulfilled.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "cancel":
            if pr.is_terminal():
                messages.error(request, "This PR is already in a terminal status.")
            elif pr.status != PurchaseRequisition.Status.DRAFT and not can_manage:
                messages.error(request, "Only managers can cancel a submitted PR.")
            elif pr.status == PurchaseRequisition.Status.DRAFT and not (can_manage or is_requester):
                messages.error(request, "You do not have permission to cancel this PR.")
            else:
                comment = (request.POST.get("comment") or "").strip()
                old = pr.status
                pr.status = PurchaseRequisition.Status.CANCELLED
                pr.cancelled_at = timezone.now()
                pr.cancelled_by = request.tenant_user
                pr.cancellation_comment = comment
                pr.save(using=tenant_db, update_fields=["status", "cancelled_at", "cancelled_by_id", "cancellation_comment"])
                _log_pr_status(pr, old, pr.status, request.tenant_user, comment or "Cancelled.", tenant_db)
                messages.success(request, "PR cancelled.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "add_line" and pr.can_edit_lines() and (can_manage or is_requester):
            from decimal import Decimal
            item = (request.POST.get("new_line_item") or "").strip()
            if not item:
                messages.error(request, "Item description is required.")
            else:
                try:
                    qty = Decimal(str(request.POST.get("new_line_qty") or "1").replace(",", ""))
                except Exception:
                    qty = Decimal("1")
                try:
                    unit = Decimal(str(request.POST.get("new_line_unit_cost") or "0").replace(",", ""))
                except Exception:
                    unit = Decimal("0")
                bl = (request.POST.get("new_line_budget_line") or "").strip()
                new_line_total = qty * unit
                new_pr_total = pr.effective_total() + new_line_total
                act = pr.workplan_activity
                remaining_activity = act.remaining_budget_for_pr(using=tenant_db)
                if new_line_total > remaining_activity:
                    messages.error(request, f"Adding this line would exceed activity remaining budget ({remaining_activity}).")
                else:
                    from django.db.models import Max
                    next_num = (pr.lines.using(tenant_db).aggregate(m=Max("line_number"))["m"] or 0) + 1
                    PurchaseRequisitionLine.objects.using(tenant_db).create(
                        pr=pr,
                        line_number=next_num,
                        item_description=item,
                        quantity=qty,
                        estimated_unit_cost=unit,
                        budget_line=bl,
                    )
                    from django.db.models import Sum
                    pr.estimated_total_cost = (
                        PurchaseRequisitionLine.objects.using(tenant_db)
                        .filter(pr=pr)
                        .aggregate(s=Sum("estimated_total_cost"))
                        .get("s") or Decimal("0")
                    )
                    pr.save(using=tenant_db, update_fields=["estimated_total_cost"])
                    messages.success(request, "Line added.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "add_attachment" and pr.can_edit_lines() and (can_manage or is_requester):
            from tenant_grants.models import PurchaseRequisitionAttachment
            f = request.FILES.get("attachment_file")
            if f:
                PurchaseRequisitionAttachment.objects.using(tenant_db).create(
                    pr=pr,
                    file=f,
                    original_filename=getattr(f, "name", "") or "",
                    uploaded_by=request.tenant_user,
                )
                messages.success(request, "Attachment added.")
            else:
                messages.error(request, "Select a file to upload.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "remove_attachment" and pr.can_edit_lines() and (can_manage or is_requester):
            from tenant_grants.models import PurchaseRequisitionAttachment
            att_id = request.POST.get("attachment_id")
            att = PurchaseRequisitionAttachment.objects.using(tenant_db).filter(pr=pr, pk=att_id).first()
            if att:
                att.delete(using=tenant_db)
                messages.success(request, "Attachment removed.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

        if action == "remove_line" and pr.can_edit_lines() and (can_manage or is_requester):
            line_id = request.POST.get("line_id")
            line = PurchaseRequisitionLine.objects.using(tenant_db).filter(pr=pr, pk=line_id).first()
            if line and pr.lines.using(tenant_db).count() > 1:
                line.delete(using=tenant_db)
                from django.db.models import Sum
                from decimal import Decimal
                pr.estimated_total_cost = (
                    PurchaseRequisitionLine.objects.using(tenant_db)
                    .filter(pr=pr)
                    .aggregate(s=Sum("estimated_total_cost"))
                    .get("s") or Decimal("0")
                )
                pr.save(using=tenant_db, update_fields=["estimated_total_cost"])
                messages.success(request, "Line removed.")
            elif line and pr.lines.using(tenant_db).count() <= 1:
                messages.error(request, "At least one line is required.")
            else:
                messages.error(request, "Line not found.")
            return redirect(reverse("tenant_portal:grants_pr_detail", args=[pr_id]))

    status_logs = list(
        PurchaseRequisitionStatusLog.objects.using(tenant_db)
        .filter(pr=pr)
        .select_related("performed_by")
        .order_by("-performed_at")[:50]
    )
    can_submit = (pr.status == PurchaseRequisition.Status.DRAFT) and (can_manage or is_requester)
    can_approve_reject_return = pr.can_line_manager_act() and can_line_manager
    can_do_procurement = pr.can_procurement_act() and can_procurement
    can_cancel = not pr.is_terminal() and (can_manage or (pr.status == PurchaseRequisition.Status.DRAFT and is_requester))
    can_edit_lines = pr.can_edit_lines() and (can_manage or is_requester)
    pr_effective_total = pr.effective_total()
    return render(
        request,
        "tenant_portal/grants/purchase_requisition_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "funds_pr",
            "pr": pr,
            "pr_effective_total": pr_effective_total,
            "status_logs": status_logs,
            "can_submit": can_submit,
            "can_approve_reject_return": can_approve_reject_return,
            "can_edit_lines": can_edit_lines,
            "can_do_procurement": can_do_procurement,
            "can_cancel": can_cancel,
            "pr_status_choices": PurchaseRequisition.Status.choices,
        },
    )


# ----- Procurement Management -----

def _procurement_approved_pr_queryset(tenant_db):
    """PRs that procurement can create POs from: approved by LM, not cancelled/rejected/fulfilled."""
    from tenant_grants.models import PurchaseRequisition
    return (
        PurchaseRequisition.objects.using(tenant_db)
        .filter(
            status__in=(
                PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER,
                PurchaseRequisition.Status.ASSIGNED_TO_PROCUREMENT,
                PurchaseRequisition.Status.UNDER_PROCUREMENT_PROCESSING,
                PurchaseRequisition.Status.PO_ISSUED,
            )
        )
        .select_related("grant", "donor", "workplan_activity")
        .prefetch_related("lines")
        .order_by("-pr_date")
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def procurement_po_list_view(request: HttpRequest) -> HttpResponse:
    """Purchase Order list with filters."""
    from tenant_grants.models import PurchaseOrder, Supplier, PurchaseRequisition

    tenant_db = request.tenant_db
    status = request.GET.get("status") or ""
    supplier_id = request.GET.get("supplier_id") or ""
    pr_id = request.GET.get("pr_id") or ""
    qs = (
        PurchaseOrder.objects.using(tenant_db)
        .select_related("pr", "supplier", "pr__grant", "pr__workplan_activity")
        .order_by("-order_date", "-created_at")
    )
    if status:
        qs = qs.filter(status=status)
    if supplier_id:
        qs = qs.filter(supplier_id=supplier_id)
    if pr_id:
        qs = qs.filter(pr_id=pr_id)
    pos = list(qs[:200])
    suppliers = list(Supplier.objects.using(tenant_db).filter(is_active=True).order_by("name")[:100])
    return render(
        request,
        "tenant_portal/grants/procurement_po_list.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "pos": pos,
            "suppliers": suppliers,
            "filters": {"status": status, "supplier_id": supplier_id, "pr_id": pr_id},
            "po_status_choices": PurchaseOrder.Status.choices,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.pr_procurement_process")
def procurement_po_create_view(request: HttpRequest) -> HttpResponse:
    """Create PO from an approved PR. Enforces thresholds and budget."""
    from django.utils import timezone
    from django.utils.dateparse import parse_date
    from decimal import Decimal
    from tenant_grants.models import (
        PurchaseRequisition,
        PurchaseOrder,
        PurchaseOrderLine,
        PurchaseRequisitionLine,
        Supplier,
        ProcurementThreshold,
    )

    tenant_db = request.tenant_db
    pr_id = request.GET.get("pr_id") or request.POST.get("pr_id")
    pr = None
    if pr_id:
        pr = (
            PurchaseRequisition.objects.using(tenant_db)
            .select_related("grant", "workplan_activity")
            .prefetch_related("lines")
            .filter(
                status__in=(
                    PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER,
                    PurchaseRequisition.Status.ASSIGNED_TO_PROCUREMENT,
                    PurchaseRequisition.Status.UNDER_PROCUREMENT_PROCESSING,
                    PurchaseRequisition.Status.PO_ISSUED,
                )
            )
            .filter(pk=pr_id)
            .first()
        )
    approved_prs = list(_procurement_approved_pr_queryset(tenant_db)[:100])
    suppliers = list(Supplier.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    thresholds = list(ProcurementThreshold.objects.using(tenant_db).order_by("amount_min"))

    if request.method == "POST":
        pr_id_post = request.POST.get("pr_id")
        if pr_id_post and not pr:
            pr = (
                PurchaseRequisition.objects.using(tenant_db)
                .select_related("grant", "workplan_activity")
                .prefetch_related("lines")
                .filter(
                    status__in=(
                        PurchaseRequisition.Status.APPROVED_BY_LINE_MANAGER,
                        PurchaseRequisition.Status.ASSIGNED_TO_PROCUREMENT,
                        PurchaseRequisition.Status.UNDER_PROCUREMENT_PROCESSING,
                        PurchaseRequisition.Status.PO_ISSUED,
                    )
                )
                .filter(pk=pr_id_post)
                .first()
            )
        if not pr:
            messages.error(request, "Select an approved PR.")
        else:
            supplier_id = request.POST.get("supplier_id")
            order_date = parse_date(request.POST.get("order_date") or "") or timezone.now().date()
            expected_delivery = parse_date(request.POST.get("expected_delivery_date") or "") or None
            notes = (request.POST.get("notes") or "").strip()
            supplier = Supplier.objects.using(tenant_db).filter(pk=supplier_id).first() if supplier_id else None
            if not supplier:
                messages.error(request, "Select a supplier.")
            else:
                total = pr.effective_total()
                threshold = ProcurementThreshold.for_amount(total, using=tenant_db)
                method = (request.POST.get("procurement_method") or "").strip() or (threshold.method if threshold else "direct_purchase")
                if threshold and not method:
                    method = threshold.method
                requires_approval = threshold.requires_po_approval if threshold else False
                approval_limit = threshold.po_approval_limit if threshold else None
                if requires_approval and approval_limit is not None and total >= approval_limit:
                    initial_status = PurchaseOrder.Status.PENDING_APPROVAL
                else:
                    initial_status = PurchaseOrder.Status.DRAFT
                last = PurchaseOrder.objects.using(tenant_db).order_by("-id").values_list("id", flat=True).first()
                next_num = (last or 0) + 1
                po_number = f"PO-{next_num:05d}"
                while PurchaseOrder.objects.using(tenant_db).filter(po_number=po_number).exists():
                    next_num += 1
                    po_number = f"PO-{next_num:05d}"
                po = PurchaseOrder.objects.using(tenant_db).create(
                    pr=pr,
                    po_number=po_number,
                    supplier=supplier,
                    procurement_method=method,
                    order_date=order_date,
                    expected_delivery_date=expected_delivery,
                    total_amount=total,
                    status=initial_status,
                    notes=notes,
                )
                for pr_line in pr.lines.all():
                    PurchaseOrderLine.objects.using(tenant_db).create(
                        po=po,
                        pr_line=pr_line,
                        item_description=pr_line.item_description,
                        quantity=pr_line.quantity,
                        unit_price=pr_line.estimated_unit_cost or Decimal("0"),
                        amount=(pr_line.quantity or Decimal("0")) * (pr_line.estimated_unit_cost or Decimal("0")),
                    )
                po.total_amount = sum((l.amount for l in po.lines.all()), Decimal("0"))
                po.save(using=tenant_db, update_fields=["total_amount"])
                messages.success(request, f"Purchase Order {po_number} created.")
                return redirect(reverse("tenant_portal:procurement_po_detail", args=[po.pk]))
    return render(
        request,
        "tenant_portal/grants/procurement_po_create.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "pr": pr,
            "approved_prs": approved_prs,
            "suppliers": suppliers,
            "thresholds": thresholds,
            "today": timezone.now().date(),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def procurement_po_detail_view(request: HttpRequest, po_id: int) -> HttpResponse:
    """PO detail: lines, GRs, Invoices; actions Approve, Send, Record GR, Add Invoice."""
    from django.utils import timezone
    from tenant_grants.models import (
        PurchaseOrder,
        PurchaseOrderLine,
        GoodsReceipt,
        SupplierInvoice,
    )

    tenant_db = request.tenant_db
    po = (
        PurchaseOrder.objects.using(tenant_db)
        .select_related("pr", "supplier", "pr__grant", "pr__workplan_activity")
        .prefetch_related("lines", "lines__pr_line", "goods_receipts", "invoices")
        .filter(pk=po_id)
        .first()
    )
    if not po:
        from django.http import Http404
        raise Http404("Purchase order not found.")
    can_process = user_has_permission(request.tenant_user, "module:grants.pr_procurement_process", using=tenant_db)

    if request.method == "POST" and can_process:
        action = (request.POST.get("action") or "").strip()
        if action == "approve_po":
            if po.status == PurchaseOrder.Status.PENDING_APPROVAL:
                po.status = PurchaseOrder.Status.APPROVED
                po.approved_at = timezone.now()
                po.approved_by = request.tenant_user
                po.save(using=tenant_db, update_fields=["status", "approved_at", "approved_by_id"])
                messages.success(request, "PO approved.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
        if action == "send_po":
            if po.status in (PurchaseOrder.Status.DRAFT, PurchaseOrder.Status.APPROVED):
                po.status = PurchaseOrder.Status.SENT
                po.save(using=tenant_db, update_fields=["status"])
                messages.success(request, "PO marked as sent to supplier.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
        if action == "close_po":
            if po.status == PurchaseOrder.Status.RECEIVED:
                po.status = PurchaseOrder.Status.CLOSED
                po.save(using=tenant_db, update_fields=["status"])
                messages.success(request, "PO closed.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
        if action == "submit_invoice_for_approval":
            inv_id = request.POST.get("invoice_id")
            inv = SupplierInvoice.objects.using(tenant_db).filter(po=po, pk=inv_id).first()
            if inv and inv.status == SupplierInvoice.Status.DRAFT:
                inv.status = SupplierInvoice.Status.PENDING_APPROVAL
                inv.save(using=tenant_db, update_fields=["status"])
                messages.success(request, "Invoice submitted for approval.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
        if action == "approve_invoice":
            inv_id = request.POST.get("invoice_id")
            inv = SupplierInvoice.objects.using(tenant_db).filter(po=po, pk=inv_id).first()
            if inv and inv.status == SupplierInvoice.Status.PENDING_APPROVAL:
                inv.status = SupplierInvoice.Status.APPROVED
                inv.approved_at = timezone.now()
                inv.approved_by = request.tenant_user
                inv.save(using=tenant_db, update_fields=["status", "approved_at", "approved_by_id"])
                messages.success(request, "Invoice approved.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
        if action == "mark_invoice_paid":
            inv_id = request.POST.get("invoice_id")
            inv = SupplierInvoice.objects.using(tenant_db).filter(po=po, pk=inv_id).first()
            if inv:
                from django.utils.dateparse import parse_date
                inv.status = SupplierInvoice.Status.PAID
                inv.payment_reference = (request.POST.get("payment_reference") or "").strip()
                inv.payment_date = parse_date(request.POST.get("payment_date") or "") or timezone.now().date()
                inv.save(using=tenant_db, update_fields=["status", "payment_reference", "payment_date"])
                messages.success(request, "Invoice marked as paid.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))

    return render(
        request,
        "tenant_portal/grants/procurement_po_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "po": po,
            "can_process": can_process,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.pr_procurement_process")
def procurement_gr_create_view(request: HttpRequest, po_id: int) -> HttpResponse:
    """Record Goods Receipt against a PO."""
    from django.utils import timezone
    from django.utils.dateparse import parse_date
    from decimal import Decimal
    from tenant_grants.models import PurchaseOrder, GoodsReceipt, GoodsReceiptLine, PurchaseOrderLine

    tenant_db = request.tenant_db
    po = PurchaseOrder.objects.using(tenant_db).prefetch_related("lines").filter(pk=po_id).first()
    if not po:
        from django.http import Http404
        raise Http404("Purchase order not found.")
    if request.method == "POST":
        receipt_date = parse_date(request.POST.get("receipt_date") or "") or timezone.now().date()
        notes = (request.POST.get("notes") or "").strip()
        existing = po.goods_receipts.using(tenant_db).count()
        gr_number = f"GR-{po.po_number}-{existing + 1}"
        gr = GoodsReceipt.objects.using(tenant_db).create(
            po=po,
            gr_number=gr_number,
            receipt_date=receipt_date,
            received_by=request.tenant_user,
            notes=notes,
        )
        for line in po.lines.all():
            qty_key = f"qty_line_{line.id}"
            try:
                qty = Decimal(str(request.POST.get(qty_key) or "0").replace(",", ""))
            except Exception:
                qty = Decimal("0")
            if qty > 0:
                GoodsReceiptLine.objects.using(tenant_db).create(
                    gr=gr,
                    po_line=line,
                    quantity_received=qty,
                )
                line.received_quantity = (line.received_quantity or Decimal("0")) + qty
                line.save(using=tenant_db, update_fields=["received_quantity"])
        total_received = sum((l.received_quantity or Decimal("0")) for l in po.lines.all())
        total_ordered = sum((l.quantity or Decimal("0")) for l in po.lines.all())
        if total_ordered and total_received >= total_ordered:
            po.status = PurchaseOrder.Status.RECEIVED
        else:
            po.status = PurchaseOrder.Status.PARTIALLY_RECEIVED
        po.save(using=tenant_db, update_fields=["status"])
        messages.success(request, f"Goods Receipt {gr.gr_number} recorded.")
        return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
    return render(
        request,
        "tenant_portal/grants/procurement_gr_create.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "po": po,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.pr_procurement_process")
def procurement_invoice_create_view(request: HttpRequest, po_id: int) -> HttpResponse:
    """Record Supplier Invoice against a PO."""
    from django.utils.dateparse import parse_date
    from decimal import Decimal
    from tenant_grants.models import PurchaseOrder, SupplierInvoice

    tenant_db = request.tenant_db
    po = PurchaseOrder.objects.using(tenant_db).select_related("supplier").filter(pk=po_id).first()
    if not po:
        from django.http import Http404
        raise Http404("Purchase order not found.")
    if request.method == "POST":
        invoice_number = (request.POST.get("invoice_number") or "").strip()
        if not invoice_number:
            messages.error(request, "Invoice number is required.")
        else:
            invoice_date = parse_date(request.POST.get("invoice_date") or "") or None
            due_date = parse_date(request.POST.get("due_date") or "") or None
            try:
                total_amount = Decimal(str(request.POST.get("total_amount") or "0").replace(",", ""))
            except Exception:
                total_amount = po.total_amount
            SupplierInvoice.objects.using(tenant_db).create(
                po=po,
                invoice_number=invoice_number,
                supplier=po.supplier,
                invoice_date=invoice_date or po.order_date,
                due_date=due_date,
                total_amount=total_amount,
                status=SupplierInvoice.Status.DRAFT,
            )
            messages.success(request, "Invoice recorded.")
            return redirect(reverse("tenant_portal:procurement_po_detail", args=[po_id]))
    return render(
        request,
        "tenant_portal/grants/procurement_invoice_create.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "po": po,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def procurement_suppliers_view(request: HttpRequest) -> HttpResponse:
    """Supplier list and simple add form (manage permission to add)."""
    from tenant_grants.models import Supplier

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.pr_procurement_process", using=tenant_db)
    if request.method == "POST" and can_manage:
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        if code and name:
            if Supplier.objects.using(tenant_db).filter(code=code).exists():
                messages.error(request, f"Supplier code '{code}' already exists.")
            else:
                Supplier.objects.using(tenant_db).create(
                    code=code,
                    name=name,
                    contact_person=(request.POST.get("contact_person") or "").strip(),
                    email=(request.POST.get("email") or "").strip(),
                    phone=(request.POST.get("phone") or "").strip(),
                    address=(request.POST.get("address") or "").strip(),
                )
                messages.success(request, "Supplier added.")
        return redirect(reverse("tenant_portal:procurement_suppliers"))
    suppliers = list(Supplier.objects.using(tenant_db).order_by("name")[:200])
    return render(
        request,
        "tenant_portal/grants/procurement_suppliers.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "suppliers": suppliers,
            "can_manage": can_manage,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def procurement_thresholds_view(request: HttpRequest) -> HttpResponse:
    """Procurement thresholds list and add (manage to add/edit)."""
    from tenant_grants.models import ProcurementThreshold

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:grants.pr_procurement_process", using=tenant_db)
    if request.method == "POST" and can_manage:
        from decimal import Decimal
        try:
            amt_min = Decimal(str(request.POST.get("amount_min") or "0").replace(",", ""))
        except Exception:
            amt_min = Decimal("0")
        amt_max = None
        if request.POST.get("amount_max"):
            try:
                amt_max = Decimal(str(request.POST.get("amount_max") or "").replace(",", ""))
            except Exception:
                pass
        method = (request.POST.get("method") or "").strip() or "direct_purchase"
        requires = request.POST.get("requires_po_approval") == "on"
        limit = None
        if request.POST.get("po_approval_limit"):
            try:
                limit = Decimal(str(request.POST.get("po_approval_limit") or "").replace(",", ""))
            except Exception:
                pass
        ProcurementThreshold.objects.using(tenant_db).create(
            amount_min=amt_min,
            amount_max=amt_max,
            method=method,
            requires_po_approval=requires,
            po_approval_limit=limit,
        )
        messages.success(request, "Threshold added.")
        return redirect(reverse("tenant_portal:procurement_thresholds"))
    thresholds = list(ProcurementThreshold.objects.using(tenant_db).order_by("amount_min"))
    return render(
        request,
        "tenant_portal/grants/procurement_thresholds.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "funds",
            "active_item": "procurement",
            "thresholds": thresholds,
            "can_manage": can_manage,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_utilization_view(request: HttpRequest) -> HttpResponse:
    """Redirect to finance grant utilization or render same data in grants context."""
    return redirect(reverse("tenant_portal:finance_grant_utilization") + "?" + request.GET.urlencode())


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_timeline_view(request: HttpRequest) -> HttpResponse:
    from django.utils import timezone
    from datetime import timedelta
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    today = timezone.now().date()
    ninety = today + timedelta(days=90)
    f = _parse_grants_filters(request)
    qs = Grant.objects.using(tenant_db).select_related("donor").filter(
        status=Grant.Status.ACTIVE
    ).order_by("start_date", "end_date")
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        qs = qs.filter(pk=f["grant_id"])
    expiring = [g for g in qs if g.end_date and today <= g.end_date <= ninety]
    all_grants = list(qs[:100])
    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")
    if request.GET.get("format"):
        rows = [
            [
                f"{g.code} — {g.title}",
                g.donor.name if g.donor else "",
                g.start_date or "",
                g.end_date or "",
                g.status,
                ("Yes" if g in expiring else "No"),
            ]
            for g in all_grants
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="grant_timeline",
            title="Grant Start & End Periods",
            headers=["Grant", "Donor", "Start date", "End date", "Status", "Expiring soon"],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/grant_timeline.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "grants": all_grants,
            "expiring": expiring,
            "filters": f,
            "donors": donors,
            "active_submenu": "funds",
            "active_item": "funds_grant_timeline",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_reporting_requirements_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import ReportingRequirement, Donor

    tenant_db = request.tenant_db
    if request.method == "POST":
        donor_id = request.POST.get("donor_id")
        name = (request.POST.get("name") or "").strip()
        format_desc = (request.POST.get("format_description") or "").strip()
        frequency = request.POST.get("frequency") or ""
        if not donor_id or not name:
            messages.error(request, "Donor and name are required.")
        else:
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
            if donor:
                ReportingRequirement.objects.using(tenant_db).create(
                    donor=donor, name=name, format_description=format_desc, frequency=frequency or None,
                )
                messages.success(request, "Reporting requirement added.")
                return redirect(reverse("tenant_portal:grants_reporting_requirements"))
    f = _parse_grants_filters(request)
    qs = ReportingRequirement.objects.using(tenant_db).select_related("donor").filter(is_active=True).order_by("donor__name", "name")
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    requirements = list(qs)
    donors = list(_active_donors_queryset(tenant_db))
    if request.GET.get("format"):
        rows = [
            [
                r.donor.name if r.donor else "",
                r.name,
                (r.get_frequency_display() if getattr(r, "frequency", None) else ""),
                (r.format_description or ""),
            ]
            for r in requirements
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="reporting_requirements",
            title="Donor Reporting Requirements",
            headers=["Donor", "Requirement", "Frequency", "Format"],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/reporting_requirements.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "requirements": requirements,
            "donors": donors,
            "filters": f,
            "frequencies": ReportingRequirement.Frequency,
            "active_submenu": "funds",
            "active_item": "funds_reporting_reqs",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_grant_financial_reports_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_grants.models import Grant
    from tenant_finance.models import ChartAccount, JournalLine

    tenant_db = request.tenant_db
    f = _parse_grants_filters(request)
    grants_qs = Grant.objects.using(tenant_db).select_related("donor").order_by("code")
    if f["donor_id"]:
        grants_qs = grants_qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        grants_qs = grants_qs.filter(pk=f["grant_id"])
    budget_by_grant = {}
    for g in grants_qs:
        budget_total = g.budget_lines.using(tenant_db).aggregate(t=Sum("amount")).get("t") or Decimal("0")
        budget_by_grant[g.id] = budget_total
    spend_by_grant = {
        r["entry__grant_id"]: r["total"] or Decimal("0")
        for r in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(total=Sum("debit"))
    }
    rows = []
    for g in grants_qs:
        budget = budget_by_grant.get(g.id, Decimal("0"))
        spent = spend_by_grant.get(g.id, Decimal("0"))
        rows.append({"grant": g, "budget": budget, "spent": spent, "remaining": budget - spent})
    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")
    if request.GET.get("format"):
        export_rows = [
            [
                f"{row['grant'].code} — {row['grant'].title}",
                row["grant"].donor.name if row["grant"].donor else "",
                row["budget"],
                row["spent"],
                row["remaining"],
            ]
            for row in rows
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="grant_financial_reports",
            title="Grant Financial Reports",
            headers=["Grant", "Donor", "Budget", "Spent", "Remaining"],
            rows=export_rows,
            request=request,
            include_official_header=True,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/grant_financial_reports.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "grants": grants_qs,
            "donors": donors,
            "filters": f,
            "active_submenu": "funds",
            "active_item": "funds_grant_reports",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
            "official_report_period_line": f"Period: {f['period_start']} — {f['period_end']}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_donor_compliance_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import DonorRestriction, Grant
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalLine

    tenant_db = request.tenant_db
    f = _parse_grants_filters(request)
    restrictions = (
        DonorRestriction.objects.using(tenant_db)
        .select_related("donor", "grant", "funding_source", "project")
        .order_by("donor__name", "restriction_code")
    )
    if f["donor_id"]:
        restrictions = restrictions.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        restrictions = restrictions.filter(grant_id=f["grant_id"])
    restrictions = list(restrictions[:200])
    spend_by_grant = {
        r["entry__grant_id"]: r["total"] or Decimal("0")
        for r in JournalLine.objects.using(tenant_db)
        .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
        .values("entry__grant_id")
        .annotate(total=Sum("debit"))
    }
    compliance_rows = [
        {
            "restriction": rest,
            "grant_spend": spend_by_grant.get(rest.grant_id) if rest.grant_id else None,
        }
        for rest in restrictions
    ]
    donors = list(_active_donors_queryset(tenant_db))
    grants = Grant.objects.using(tenant_db).order_by("code")
    if request.GET.get("format"):
        rows = [
            [
                row["restriction"].restriction_code or "",
                row["restriction"].donor.name if row["restriction"].donor else "",
                (row["restriction"].grant.code if row["restriction"].grant else ""),
                (
                    row["restriction"].funding_source.name
                    if row["restriction"].funding_source_id
                    else ""
                ),
                row["restriction"].get_category_display(),
                row["restriction"].get_restriction_type_display(),
                row["restriction"].get_compliance_level_display(),
                row["restriction"].get_status_display(),
                (
                    row["restriction"].effective_start.isoformat()
                    if row["restriction"].effective_start
                    else ""
                ),
                (
                    row["restriction"].effective_end.isoformat()
                    if row["restriction"].effective_end
                    else ""
                ),
                row["restriction"].description,
            ]
            for row in compliance_rows
        ]
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="donor_compliance",
            title="Donor Compliance Monitoring",
            headers=[
                "Code",
                "Donor",
                "Grant",
                "Funding source",
                "Category",
                "Restriction type",
                "Compliance",
                "Status",
                "Effective start",
                "Effective end",
                "Description",
            ],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/donor_compliance.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "compliance_rows": compliance_rows,
            "spend_by_grant": spend_by_grant,
            "donors": donors,
            "grants": grants,
            "filters": f,
            "active_submenu": "funds",
            "active_item": "funds_donor_compliance",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_reporting_deadlines_view(request: HttpRequest) -> HttpResponse:
    from django.db.models import Q
    from django.utils import timezone
    from django.utils.dateparse import parse_date
    from tenant_grants.models import ReportingDeadline, ReportingRequirement, Donor, Grant, Project
    from tenant_users.models import TenantUser
    from datetime import timedelta

    tenant_db = request.tenant_db
    today = timezone.now().date()

    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "mark_submitted":
            dl_id = request.POST.get("deadline_id")
            if dl_id:
                d = (
                    ReportingDeadline.objects.using(tenant_db)
                    .filter(pk=dl_id)
                    .select_related("donor", "grant", "project")
                    .first()
                )
                if d and not d.is_submitted_record():
                    sub_date = parse_date(request.POST.get("submitted_date") or "") or today
                    d.status = ReportingDeadline.Status.SUBMITTED
                    d.submitted_at = timezone.now()
                    d.submitted_date = sub_date
                    d.save(update_fields=["status", "submitted_at", "submitted_date"])
                    messages.success(request, "Marked as submitted.")
                    return redirect(reverse("tenant_portal:grants_reporting_deadlines"))
        if action == "create_deadline":
            donor_id = request.POST.get("donor_id") or ""
            grant_id = request.POST.get("grant_id") or ""
            project_id = request.POST.get("project_id") or ""
            requirement_id = request.POST.get("requirement_id") or ""
            title = (request.POST.get("title") or "").strip()
            deadline_date = parse_date(request.POST.get("deadline_date") or "")
            reporting_period_from = parse_date(request.POST.get("reporting_period_from") or "")
            reporting_period_to = parse_date(request.POST.get("reporting_period_to") or "")
            submitted_date = parse_date(request.POST.get("submitted_date") or "")
            notes = (request.POST.get("notes") or "").strip()
            priority = (request.POST.get("priority") or ReportingDeadline.Priority.NORMAL).strip()
            if priority not in dict(ReportingDeadline.Priority.choices):
                priority = ReportingDeadline.Priority.NORMAL
            try:
                reminder_days_before = int(request.POST.get("reminder_days_before") or "7")
            except ValueError:
                reminder_days_before = 7
            reminder_days_before = max(1, min(reminder_days_before, 365))
            responsible_id = request.POST.get("responsible_user_id") or ""
            reviewer_id = request.POST.get("reviewer_user_id") or ""
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first() if grant_id else None
            project = Project.objects.using(tenant_db).filter(pk=project_id).first() if project_id else None
            req = (
                ReportingRequirement.objects.using(tenant_db).filter(pk=requirement_id).first()
                if requirement_id
                else None
            )
            responsible = (
                TenantUser.objects.using(tenant_db).filter(pk=responsible_id).first()
                if responsible_id
                else None
            )
            reviewer = (
                TenantUser.objects.using(tenant_db).filter(pk=reviewer_id).first() if reviewer_id else None
            )
            if not donor or not title or not deadline_date:
                messages.error(request, "Please provide donor, title, and deadline date.")
            elif not grant and not project:
                messages.error(request, "Select a grant or a project (or both).")
            else:
                initial_status = (
                    ReportingDeadline.Status.SUBMITTED
                    if submitted_date
                    else ReportingDeadline.Status.OPEN
                )
                submitted_at = timezone.now() if submitted_date else None
                ReportingDeadline.objects.using(tenant_db).create(
                    donor=donor,
                    grant=grant,
                    project=project,
                    requirement=req,
                    title=title,
                    reporting_period_from=reporting_period_from,
                    reporting_period_to=reporting_period_to,
                    deadline_date=deadline_date,
                    submitted_date=submitted_date or None,
                    submitted_at=submitted_at,
                    status=initial_status,
                    notes=notes,
                    priority=priority,
                    reminder_days_before=reminder_days_before,
                    responsible_user=responsible,
                    reviewer_user=reviewer,
                )
                messages.success(request, "Reporting deadline created.")
                return redirect(reverse("tenant_portal:grants_reporting_deadlines"))
    f = _parse_grants_filters(request)
    if not (request.GET.get("period_start") or "").strip():
        f["period_start"] = today - timedelta(days=180)
    if not (request.GET.get("period_end") or "").strip():
        f["period_end"] = today + timedelta(days=730)
    qs = (
        ReportingDeadline.objects.using(tenant_db)
        .select_related(
            "donor",
            "grant",
            "grant__project",
            "project",
            "requirement",
            "responsible_user",
            "reviewer_user",
        )
        .filter(
            deadline_date__gte=f["period_start"],
            deadline_date__lte=f["period_end"],
        )
    )
    if f["donor_id"]:
        qs = qs.filter(donor_id=f["donor_id"])
    if f["grant_id"]:
        qs = qs.filter(grant_id=f["grant_id"])
    if f["project_id"]:
        qs = qs.filter(Q(project_id=f["project_id"]) | Q(grant__project_id=f["project_id"]))
    deadlines = list(qs.order_by("deadline_date", "id")[:500])
    _allowed_disp = {k for k, _ in ReportingDeadline.DisplayStatus.choices}
    if f["deadline_status"] in _allowed_disp:
        deadlines = [d for d in deadlines if d.display_status() == f["deadline_status"]]
    overdue = [d for d in deadlines if d.is_overdue()]
    reminder_alerts = []
    for d in deadlines:
        hit = d.reminder_milestone_hit(today)
        if hit is not None:
            reminder_alerts.append({"deadline": d, "milestone": hit})
    reminder_alerts.sort(key=lambda x: (-x["milestone"], x["deadline"].deadline_date))
    donors = list(_active_donors_queryset(tenant_db))
    grants = Grant.objects.using(tenant_db).select_related("project").order_by("code")
    projects = Project.objects.using(tenant_db).order_by("code")
    requirements = (
        ReportingRequirement.objects.using(tenant_db)
        .select_related("donor")
        .filter(is_active=True)
        .order_by("donor__name", "name")
    )
    tenant_users = TenantUser.objects.using(tenant_db).filter(is_active=True).order_by(
        "full_name", "email"
    )
    if request.GET.get("format"):
        rows = []
        for d in deadlines:
            ep = d.effective_project()
            period_s = ""
            if d.reporting_period_from or d.reporting_period_to:
                period_s = f"{d.reporting_period_from or '—'} → {d.reporting_period_to or '—'}"
            dr = d.days_remaining()
            ru = d.responsible_user
            rv = d.reviewer_user
            res_lbl = ((ru.full_name or "").strip() or (ru.email if ru else "")) if ru else ""
            rev_lbl = ((rv.full_name or "").strip() or (rv.email if rv else "")) if rv else ""
            rows.append(
                [
                    d.title,
                    d.donor.name if d.donor else "",
                    d.grant.code if d.grant else "",
                    ep.code if ep else "",
                    period_s,
                    d.deadline_date,
                    "" if dr is None else dr,
                    d.display_status_label(),
                    "Yes" if d.is_overdue() else "No",
                    d.submitted_date or (d.submitted_at.date() if d.submitted_at else ""),
                    d.get_priority_display(),
                    d.reminder_days_before,
                    res_lbl,
                    rev_lbl,
                    d.notes or "",
                ]
            )
        resp = _export_table_response(
            export_format=request.GET.get("format") or "",
            filename_base="reporting_deadlines",
            title="Donor Reporting Deadlines",
            headers=[
                "Title",
                "Donor",
                "Grant",
                "Project",
                "Period",
                "Deadline",
                "Days remaining",
                "Status",
                "Overdue",
                "Submitted date",
                "Priority",
                "Reminder days before",
                "Responsible",
                "Reviewer",
                "Notes",
            ],
            rows=rows,
        )
        if resp:
            return resp
    return render(
        request,
        "tenant_portal/grants/reporting_deadlines.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "today": today,
            "deadlines": deadlines,
            "overdue": overdue,
            "reminder_alerts": reminder_alerts,
            "donors": donors,
            "grants": grants,
            "projects": projects,
            "requirements": requirements,
            "tenant_users": tenant_users,
            "filters": f,
            "deadline_status_choices": ReportingDeadline.DisplayStatus.choices,
            "priority_choices": ReportingDeadline.Priority.choices,
            "reminder_presets": ReportingDeadline.REMINDER_MILESTONE_DAYS,
            "active_submenu": "funds",
            "active_item": "funds_reporting_deadlines",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.manage")
def grants_budgets_view(request: HttpRequest) -> HttpResponse:
    from tenant_grants.models import Grant, BudgetLine
    from tenant_finance.models import ChartAccount

    tenant_db = request.tenant_db
    if request.method == "POST":
        grant_id = request.POST.get("grant_id")
        account_id = request.POST.get("account_id") or None
        category = (request.POST.get("category") or "").strip()
        description = (request.POST.get("description") or "").strip()
        amount = request.POST.get("amount") or "0"
        notes = (request.POST.get("notes") or "").strip()
        if not grant_id or not category:
            messages.error(request, "Please select a project/grant and provide a category.")
        else:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            account = (
                ChartAccount.objects.using(tenant_db).filter(pk=account_id).first()
                if account_id
                else None
            )
            BudgetLine.objects.using(tenant_db).create(
                grant=grant,
                account=account,
                category=category,
                description=description,
                amount=amount or 0,
                notes=notes,
            )
            messages.success(request, "Budget line created.")
            return redirect(reverse("tenant_portal:grants_budgets"))

    grants = Grant.objects.using(tenant_db).order_by("-created_at")[:100]
    accounts = ChartAccount.objects.using(tenant_db).order_by("code")
    budget_lines = (
        BudgetLine.objects.using(tenant_db)
        .select_related("grant", "account")
        .order_by("-id")[:200]
    )
    return render(
        request,
        "tenant_portal/grants/budgets.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "grants": grants,
            "budget_lines": budget_lines,
            "accounts": accounts,
            "active_submenu": "budget",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budgeting_center_view(request: HttpRequest) -> HttpResponse:
    """Budgeting module home: KPIs and navigation to budget setup, approval, monitoring, and forecasting."""
    from decimal import Decimal

    from django.db.models import Sum
    from django.utils.translation import gettext as _

    tenant_db = request.tenant_db
    total_budgets = 0
    active_budgets = 0
    draft_budgets = 0
    approved_budgets = 0
    pending_approvals = 0
    util_pct_str = "0%"
    over_budget_n = 0
    variance_alerts_n = 0

    try:
        from tenant_finance.models import ChartAccount, JournalLine
        from tenant_grants.models import BudgetLine, Grant, GrantApproval

        grant_ids = list(
            BudgetLine.objects.using(tenant_db).values_list("grant_id", flat=True).distinct()
        )
        total_budgets = len(grant_ids)
        if grant_ids:
            active_budgets = Grant.objects.using(tenant_db).filter(
                pk__in=grant_ids, status=Grant.Status.ACTIVE
            ).count()
            draft_budgets = Grant.objects.using(tenant_db).filter(
                pk__in=grant_ids, status=Grant.Status.DRAFT
            ).count()
            approved_budgets = (
                Grant.objects.using(tenant_db)
                .filter(pk__in=grant_ids, approvals__status=GrantApproval.Status.APPROVED)
                .distinct()
                .count()
            )

        pending_approvals = GrantApproval.objects.using(tenant_db).filter(
            status=GrantApproval.Status.PENDING
        ).count()

        budgets_by_grant = {
            r["grant_id"]: r["total"] or Decimal("0")
            for r in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
        }
        spend_by_grant = {
            r["entry__grant_id"]: r["spent"] or Decimal("0")
            for r in JournalLine.objects.using(tenant_db)
            .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
            .values("entry__grant_id")
            .annotate(spent=Sum("debit"))
        }

        active_with_lines = Grant.objects.using(tenant_db).filter(
            status=Grant.Status.ACTIVE, pk__in=grant_ids or []
        ).only("id", "award_amount")

        total_ceiling = Decimal("0")
        total_spent = Decimal("0")
        for g in active_with_lines:
            b = budgets_by_grant.get(g.id, Decimal("0"))
            ceiling = b if b > 0 else Decimal(str(g.award_amount or 0))
            spent = spend_by_grant.get(g.id, Decimal("0"))
            if ceiling > 0:
                total_ceiling += ceiling
                total_spent += spent
                if spent > ceiling:
                    over_budget_n += 1
                elif spent >= ceiling * Decimal("0.9"):
                    variance_alerts_n += 1

        if total_ceiling > 0:
            pct = (total_spent / total_ceiling) * Decimal("100")
            util_pct_str = f"{pct.quantize(Decimal('0.1'))}%"
    except Exception:
        pass

    budget_kpis = [
        {"label": _("Total budgets"), "value": f"{total_budgets:,}"},
        {"label": _("Active budgets"), "value": f"{active_budgets:,}"},
        {"label": _("Draft budgets"), "value": f"{draft_budgets:,}"},
        {"label": _("Approved budgets"), "value": f"{approved_budgets:,}"},
        {"label": _("Pending approvals"), "value": f"{pending_approvals:,}"},
        {"label": _("Budget utilization %"), "value": util_pct_str},
        {"label": _("Over-budget projects"), "value": f"{over_budget_n:,}"},
        {"label": _("Budget variance alerts"), "value": f"{variance_alerts_n:,}"},
    ]

    return render(
        request,
        "tenant_portal/budget/budgeting_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "budget",
            "active_item": "budget_center_home",
            "budget_kpis": budget_kpis,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_creation_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Creation: wrapper around grants_budgets_view using full-width layout, filters, and exports.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import JournalLine, ChartAccount
    from tenant_grants.models import Grant, BudgetLine, Donor
    from django.utils.dateparse import parse_date
    import openpyxl

    tenant_db = request.tenant_db

    # Handle Excel import for budget lines
    if request.method == "POST" and request.POST.get("action") == "import":
        upload = request.FILES.get("budget_file")
        if not upload:
            messages.error(request, "Please choose an Excel file to import.")
        else:
            try:
                wb = openpyxl.load_workbook(upload)
                ws = wb.active
                created = 0
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    code, category, amount, description = (row + (None,) * 4)[:4]
                    if not code or not category:
                        continue
                    grant = Grant.objects.using(tenant_db).filter(code=str(code).strip()).first()
                    if not grant:
                        continue
                    try:
                        amt = Decimal(str(amount or "0"))
                    except Exception:
                        amt = Decimal("0")
                    BudgetLine.objects.using(tenant_db).create(
                        grant=grant,
                        category=str(category).strip(),
                        description=(description or "") if description is not None else "",
                        amount=amt,
                    )
                    created += 1
                messages.success(request, f"Imported {created} budget lines from Excel.")
                return redirect(reverse("tenant_portal:budget_creation"))
            except Exception:
                messages.error(request, "Could not read the Excel file. Please check the format.")

    # Manual budget line creation mirrors grants_budgets_view
    if request.method == "POST" and request.POST.get("action") != "import":
        grant_id = request.POST.get("grant_id")
        account_id = request.POST.get("account_id") or None
        category = (request.POST.get("category") or "").strip()
        description = (request.POST.get("description") or "").strip()
        amount = request.POST.get("amount") or "0"
        notes = (request.POST.get("notes") or "").strip()
        if not grant_id or not category:
            messages.error(request, "Please select a project/grant and provide a category.")
        else:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            account = (
                ChartAccount.objects.using(tenant_db).filter(pk=account_id).first()
                if account_id
                else None
            )
            BudgetLine.objects.using(tenant_db).create(
                grant=grant,
                account=account,
                category=category,
                description=description,
                amount=amount or 0,
                notes=notes,
            )
            messages.success(request, "Budget line created.")
            return redirect(reverse("tenant_portal:budget_creation"))

    # Filters
    from django.utils import timezone
    f = _parse_finance_filters(request)
    # Adapt: allow donor filter too
    donor_id = request.GET.get("donor_id") or ""
    if donor_id:
        f["donor_id"] = donor_id

    # Budget lines filtered by grant/donor
    grants_qs = Grant.objects.using(tenant_db).select_related("donor").order_by("code")
    if f.get("donor_id"):
        grants_qs = grants_qs.filter(donor_id=f["donor_id"])
    if f.get("grant_id"):
        grants_qs = grants_qs.filter(pk=f["grant_id"])
    grants = list(grants_qs)
    grant_ids = [g.id for g in grants]

    budget_lines = (
        BudgetLine.objects.using(tenant_db)
        .select_related("grant", "account")
        .filter(grant_id__in=grant_ids or BudgetLine.objects.using(tenant_db).values("grant_id"))
        .order_by("grant__code", "id")
    )

    # Totals per grant
    totals_by_grant = (
        budget_lines.values("grant_id")
        .annotate(total=Sum("amount"))
    )
    totals_map = {row["grant_id"]: row["total"] for row in totals_by_grant}

    # Actual spend per grant for quick summary and links to transactions
    spend_by_grant = {
        r["entry__grant_id"]: r["spent"]
        for r in JournalLine.objects.using(tenant_db)
        .filter(
            account__type=ChartAccount.Type.EXPENSE,
            entry__grant_id__isnull=False,
            entry__entry_date__gte=f["period_start"],
            entry__entry_date__lte=f["period_end"],
        )
        .values("entry__grant_id")
        .annotate(spent=Sum("debit"))
    }

    export_format = request.GET.get("format") or ""
    if export_format:
        rows = []
        for b in budget_lines:
            g = b.grant
            rows.append(
                [
                    g.code if g else "",
                    g.title if g else "",
                    g.donor.name if getattr(g, "donor", None) else "",
                    b.category,
                    b.amount,
                    b.notes or "",
                ]
            )
        resp = _export_table_response(
            export_format=export_format,
            filename_base="budget_creation",
            title="Budget Creation",
            headers=["Grant code", "Grant title", "Donor", "Category", "Amount", "Notes"],
            rows=rows,
        )
        if resp:
            return resp

    donors = __import__("tenant_grants.models", fromlist=["Donor"]).Donor.objects.using(tenant_db).order_by("name")
    accounts = ChartAccount.objects.using(tenant_db).order_by("code")
    return render(
        request,
        "tenant_portal/budget/budget_creation.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "filters": f,
            "grants": grants,
            "donors": donors,
            "budget_lines": budget_lines,
            "totals_map": totals_map,
            "spend_by_grant": spend_by_grant,
            "accounts": accounts,
            "active_submenu": "budget",
            "active_item": "budget_creation",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_templates_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Templates: currently reuses the Budget Creation screen.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:budget_creation") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_versions_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Versions: currently reuses the Budget Creation screen (one version per project).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:budget_creation") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_structures_view(request: HttpRequest) -> HttpResponse:
    """
    Project & Donor Budget Structure: reuse Budget Creation (budgets by project/donor).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:budget_creation") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_approval_workflow_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Approval Workflow: reuse grant approvals screen (same approval engine).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:grants_approvals") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_approvals_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Approvals: reuse grant approvals screen.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:grants_approvals") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_revisions_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Revisions: reuse grant approvals screen (approvals for revised budgets).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:grants_approvals") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_monitoring_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Monitoring: reuse project financial status dashboard.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_project_financial_status") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_bva_view(request: HttpRequest) -> HttpResponse:
    """
    Budget vs Actual Analysis: reuse finance budget vs actual report.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_budget_vs_actual") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_variance_view(request: HttpRequest) -> HttpResponse:
    """
    Variance Analysis: reuse finance budget vs actual report (shows variance).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_budget_vs_actual") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_control_warnings_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Control Warnings: reuse finance financial alerts dashboard.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_financial_alerts") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_forecasting_view(request: HttpRequest) -> HttpResponse:
    """
    Budget forecasting: estimation of budget linked to a project,
    reusing the monthly / quarterly / yearly expense trend analysis.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_expense_trend") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def budget_adjustments_view(request: HttpRequest) -> HttpResponse:
    """
    Budget Adjustments: reuse fund balances screen (reallocation context).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_fund_balances") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:accounts.view")
def cash_dashboard_view(request: HttpRequest) -> HttpResponse:
    """
    Cash & Bank module dashboard: entry page when user selects Cash & Bank in top navigation.
    Summary cards, quick actions, and recent activity.
    """
    from decimal import Decimal

    from django.db.models import Sum

    from tenant_finance.models import BankAccount, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_manage = user_has_permission(user, "module:finance.manage", using=tenant_db)

    # All bank/cash account IDs (organizational)
    bank_accounts = list(
        BankAccount.objects.using(tenant_db).select_related("currency", "account").order_by("bank_name")
    )
    cash_account_ids = [ba.account_id for ba in bank_accounts if ba.account_id]

    # Total Bank Balance & Total Cash Balance (same source: posted balances on these accounts)
    total_bank_balance = Decimal("0")
    total_cash_balance = Decimal("0")
    if cash_account_ids:
        bal_rows = (
            JournalLine.objects.using(tenant_db)
            .filter(account_id__in=cash_account_ids, entry__status=JournalEntry.Status.POSTED)
            .values("account_id")
            .annotate(bal=Sum("debit") - Sum("credit"))
        )
        total_bank_balance = sum((r.get("bal") or Decimal("0")) for r in bal_rows)
        total_cash_balance = total_bank_balance  # same pool for this module

    # Petty Cash Accounts: count where account name suggests petty
    petty_count = sum(
        1 for ba in bank_accounts if ba.account_name and "petty" in ba.account_name.lower()
    )

    # Pending Reconciliation: placeholder (e.g. count of active accounts as "to reconcile")
    pending_recon = sum(1 for ba in bank_accounts if getattr(ba, "is_active", True))

    # Recent activity: last 10 posted entries (no grant) that touch cash/bank accounts
    recent_activity = []
    if cash_account_ids:
        from django.db.models import Q

        entry_ids = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account_id__in=cash_account_ids,
                entry__status=JournalEntry.Status.POSTED,
                entry__grant__isnull=True,
            )
            .values_list("entry_id", flat=True)
            .distinct()
        )
        entries = (
            JournalEntry.objects.using(tenant_db)
            .filter(id__in=entry_ids)
            .select_related("created_by")
            .order_by("-entry_date", "-id")[:10]
        )
        for entry in entries:
            lines = list(
                JournalLine.objects.using(tenant_db)
                .filter(entry=entry, account_id__in=cash_account_ids)
                .values_list("debit", "credit")
            )
            amount = sum((d or Decimal("0")) - (c or Decimal("0")) for d, c in lines)
            ref = entry.reference or f"JE-{entry.id:05d}"
            recent_activity.append({
                "date": entry.entry_date,
                "reference": ref,
                "description": (entry.memo or "—")[:80],
                "amount": amount,
            })

    return render(
        request,
        "tenant_portal/cash/dashboard.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "cash",
            "active_item": "cash_dashboard",
            "total_bank_balance": total_bank_balance,
            "total_cash_balance": total_cash_balance,
            "petty_count": petty_count,
            "pending_recon": pending_recon,
            "recent_activity": recent_activity,
            "can_manage": can_manage,
            "global_indicators": _get_global_financial_indicators(request),
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def cost_home_view(request: HttpRequest) -> HttpResponse:
    """Multi-donor sharing module home: allocation KPIs, distribution summary, and navigation."""
    from decimal import Decimal

    from django.db.models import Sum
    from django.utils.translation import gettext as _

    from tenant_finance.models import BudgetControlRule
    from tenant_grants.models import BudgetTemplate, Grant, GrantAllocation

    tenant_db = request.tenant_db

    def _fmt_money(d: Decimal) -> str:
        d = d.quantize(Decimal("0.01"))
        neg = d < 0
        d = abs(d)
        whole_s, _, frac = f"{d:.2f}".partition(".")
        whole_s = "{:,}".format(int(whole_s))
        return ("-" if neg else "") + whole_s + "." + frac

    total_alloc = 0
    active_rules_n = 0
    rules_enforcing_post_n = 0
    grants_sharing_n = 0
    templates_n = 0
    pct_lines_n = 0
    donors_in_pool_n = 0
    total_allocated_amt = Decimal("0")
    pct_distribution: list[dict] = []
    recent_allocations: list[dict] = []

    try:
        total_alloc = GrantAllocation.objects.using(tenant_db).count()
        active_rules_n = BudgetControlRule.objects.using(tenant_db).filter(is_active=True).count()
        rules_enforcing_post_n = BudgetControlRule.objects.using(tenant_db).filter(
            is_active=True, check_before_posting=True
        ).count()
        grants_sharing_n = (
            Grant.objects.using(tenant_db).filter(allocations__isnull=False).distinct().count()
        )
        templates_n = BudgetTemplate.objects.using(tenant_db).count()
        pct_lines_n = GrantAllocation.objects.using(tenant_db).filter(percentage__isnull=False).count()
        donors_in_pool_n = GrantAllocation.objects.using(tenant_db).values("donor_id").distinct().count()
        agg = (
            GrantAllocation.objects.using(tenant_db)
            .filter(amount__isnull=False)
            .aggregate(t=Sum("amount"))
            .get("t")
        )
        total_allocated_amt = agg or Decimal("0")

        pct_qs = (
            GrantAllocation.objects.using(tenant_db)
            .filter(percentage__isnull=False)
            .select_related("grant", "donor")
            .order_by("-percentage")[:12]
        )
        for row in pct_qs:
            p = row.percentage
            if p is None:
                continue
            try:
                pv = float(p)
            except (TypeError, ValueError):
                continue
            pv = max(0.0, min(100.0, pv))
            label = f"{row.grant.code} — {row.donor.name}"
            pct_distribution.append({"label": label, "pct": pv})

        recent_qs = (
            GrantAllocation.objects.using(tenant_db)
            .select_related("grant", "donor")
            .order_by("-id")[:10]
        )
        for row in recent_qs:
            amt = row.amount
            pct = row.percentage
            amt_s = _fmt_money(Decimal(str(amt))) if amt is not None else "—"
            pct_s = f"{pct}%" if pct is not None else "—"
            recent_allocations.append({
                "grant": row.grant.code,
                "donor": row.donor.name,
                "amount_fmt": amt_s,
                "pct_fmt": pct_s,
            })
    except Exception:
        pass

    cost_kpis = [
        {"label": _("Total shared cost allocations"), "value": f"{total_alloc:,}"},
        {"label": _("Active allocation rules"), "value": f"{active_rules_n:,}"},
        {"label": _("Rules enforced at posting"), "value": f"{rules_enforcing_post_n:,}"},
        {"label": _("Projects using cost sharing"), "value": f"{grants_sharing_n:,}"},
        {"label": _("Allocation templates"), "value": f"{templates_n:,}"},
        {"label": _("Lines with % split"), "value": f"{pct_lines_n:,}"},
        {"label": _("Donors in allocation pool"), "value": f"{donors_in_pool_n:,}"},
        {"label": _("Total allocated amounts"), "value": _fmt_money(total_allocated_amt)},
    ]

    return render(
        request,
        "tenant_portal/cost/multi_donor_sharing_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "cost",
            "active_item": "cost_center_home",
            "cost_kpis": cost_kpis,
            "pct_distribution": pct_distribution,
            "recent_allocations": recent_allocations,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def controls_home_view(request: HttpRequest) -> HttpResponse:
    """
    Governance Center: module home with KPIs, grouped navigation cards, and recent activity.
    """
    from datetime import timedelta

    from django.db.models import Q
    from django.urls import reverse
    from django.utils import timezone
    from django.utils.formats import date_format
    from django.utils.translation import gettext as _

    tenant_db = request.tenant_db
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    since_30d = now - timedelta(days=30)

    pending_total = 0
    docs_month = 0
    audit_30d = 0
    workflows_active = 0
    compliance_open = 0
    recent_rows: list[dict] = []

    try:
        from tenant_grants.models import GrantApproval
        from tenant_finance.models import (
            ApprovalWorkflow,
            AuditLog,
            JournalEntry,
            JournalEntryAttachment,
        )

        pending_total = GrantApproval.objects.using(tenant_db).filter(
            status=GrantApproval.Status.PENDING
        ).count() + JournalEntry.objects.using(tenant_db).filter(
            status=JournalEntry.Status.PENDING_APPROVAL
        ).count()

        audit_30d = AuditLog.objects.using(tenant_db).filter(changed_at__gte=since_30d).count()

        workflows_active = ApprovalWorkflow.objects.using(tenant_db).filter(
            status=ApprovalWorkflow.Status.ACTIVE
        ).count()

        docs_month = JournalEntryAttachment.objects.using(tenant_db).filter(
            uploaded_at__gte=month_start
        ).count()
    except Exception:
        pass

    try:
        from tenant_audit_risk.models import InvestigationAttachment, RiskAlert

        docs_month += InvestigationAttachment.objects.using(tenant_db).filter(
            uploaded_at__gte=month_start
        ).count()

        compliance_open = RiskAlert.objects.using(tenant_db).filter(
            status=RiskAlert.Status.OPEN
        ).count()
    except Exception:
        pass

    try:
        from tenant_finance.models import AuditLog

        log_qs = (
            AuditLog.objects.using(tenant_db)
            .filter(
                Q(summary__icontains="approv")
                | Q(summary__icontains="pending")
                | Q(summary__icontains="submit")
            )
            .order_by("-changed_at")[:8]
        )
        for log in log_qs:
            recent_rows.append(
                {
                    "when_display": date_format(log.changed_at, "SHORT_DATETIME_FORMAT"),
                    "user": log.username or "—",
                    "summary": (log.summary or log.get_action_display())[:200],
                }
            )
    except Exception:
        pass

    gov_kpis = [
        {
            "label": _("Pending approvals"),
            "value": f"{pending_total:,}",
            "href": reverse("tenant_portal:finance_journal_approval"),
        },
        {
            "label": _("Documents (this month)"),
            "value": f"{docs_month:,}",
            "href": reverse("tenant_portal:audit_risk_evidence"),
        },
        {
            "label": _("Audit entries (30 days)"),
            "value": f"{audit_30d:,}",
            "href": reverse("tenant_portal:finance_audit_trail"),
        },
        {
            "label": _("Active workflows"),
            "value": f"{workflows_active:,}",
            "href": reverse("tenant_portal:setup_approval_workflows_list"),
        },
        {
            "label": _("Open compliance alerts"),
            "value": f"{compliance_open:,}",
            "href": reverse("tenant_portal:audit_risk_compliance"),
        },
    ]

    return render(
        request,
        "tenant_portal/governance/governance_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "controls",
            "active_item": "controls_home",
            "gov_kpis": gov_kpis,
            "gov_recent_activity": recent_rows,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="cashbank:accounts.view")
def cash_bank_accounts_view(request: HttpRequest) -> HttpResponse:
    """
    Bank Accounts master data page.

    - Master setup for org bank accounts (receipts, payments, transfers, reconciliation).
    - Each bank account must be linked to a valid asset GL account; selection restricted to asset-type.
    - Unique account numbers enforced (DB + create validation).
    - Current balance calculated from posted transactions only.
    - Deletion prevented when any journal lines exist for the linked GL account; Deactivate used instead.
    - Finance Manager / Officer (module:finance.manage) can create, activate/deactivate, and delete (when no transactions).
    - Auditor (module:finance.view only) has read-only access.
    """
    from decimal import Decimal, InvalidOperation

    from django.db.models import Q, Sum, Count
    from django.utils.dateparse import parse_date

    from tenant_finance.models import BankAccount, ChartAccount, Currency, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    user = request.tenant_user

    # Permissions: manage = create/edit/activate/deactivate, view = read-only.
    can_manage = user_has_permission(user, "module:finance.manage", using=tenant_db)

    # Handle create / status change (only for users with manage permission).
    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to modify bank accounts.")
            return redirect(reverse("tenant_portal:cash_bank_accounts"))

        action = (request.POST.get("action") or "").strip()

        if action == "toggle_active":
            ba_id = request.POST.get("bank_account_id") or ""
            if ba_id:
                ba = BankAccount.objects.using(tenant_db).filter(pk=ba_id).first()
                if ba:
                    ba.is_active = not ba.is_active
                    ba.save(update_fields=["is_active"])
                    messages.success(
                        request,
                        f"Bank account {ba.account_number} is now {'active' if ba.is_active else 'inactive'}.",
                    )
            return redirect(reverse("tenant_portal:cash_bank_accounts"))

        if action == "delete":
            ba_id = request.POST.get("bank_account_id") or ""
            if ba_id:
                ba = BankAccount.objects.using(tenant_db).filter(pk=ba_id).first()
                if ba:
                    has_lines = (
                        JournalLine.objects.using(tenant_db)
                        .filter(account_id=ba.account_id)
                        .exists()
                    )
                    if has_lines:
                        messages.error(
                            request,
                            "Cannot delete this bank account because it has existing transactions. Deactivate it instead.",
                        )
                    else:
                        ba.delete()
                        messages.success(request, "Bank account deleted.")
                else:
                    messages.error(request, "Bank account not found.")
            return redirect(reverse("tenant_portal:cash_bank_accounts"))

        # Create new bank account
        bank_name = (request.POST.get("bank_name") or "").strip()
        account_name = (request.POST.get("account_name") or "").strip()
        account_number = (request.POST.get("account_number") or "").strip()
        branch = (request.POST.get("branch") or "").strip()
        office = (request.POST.get("office") or "").strip()
        description = (request.POST.get("description") or "").strip()
        currency_id = request.POST.get("currency_id") or ""
        account_id = request.POST.get("account_id") or ""
        raw_opening_balance = (request.POST.get("opening_balance") or "").replace(",", "").strip()
        raw_opening_date = (request.POST.get("opening_balance_date") or "").strip()
        status = (request.POST.get("status") or "active").strip().lower()

        errors = []

        if not bank_name:
            errors.append("Bank name is required.")
        if not account_name:
            errors.append("Account name is required.")
        if not account_number:
            errors.append("Account number is required.")
        if not currency_id:
            errors.append("Currency is required.")
        if not account_id:
            errors.append("Linked chart of account is required.")

        # Validate uniqueness of account number within tenant DB.
        if account_number:
            exists = BankAccount.objects.using(tenant_db).filter(account_number__iexact=account_number).exists()
            if exists:
                errors.append("Account number must be unique. Another bank account already uses this number.")

        # Validate opening balance
        opening_balance = Decimal("0")
        if raw_opening_balance:
            try:
                opening_balance = Decimal(raw_opening_balance)
            except (InvalidOperation, ValueError):
                errors.append("Opening balance must be a numeric amount.")

        # Validate opening balance date
        opening_balance_date = None
        if raw_opening_date:
            opening_balance_date = parse_date(raw_opening_date)
            if not opening_balance_date:
                errors.append("Opening balance date is not a valid date.")
            else:
                from django.utils import timezone

                if opening_balance_date > timezone.localdate():
                    errors.append("Opening balance date cannot be in the future.")

        # Resolve FK objects
        currency = Currency.objects.using(tenant_db).filter(pk=currency_id).first() if currency_id else None
        if not currency:
            errors.append("Selected currency does not exist.")

        account = (
            ChartAccount.objects.using(tenant_db)
            .filter(pk=account_id, is_active=True)
            .first()
            if account_id
            else None
        )
        if not account:
            errors.append("Selected chart of account is invalid or inactive.")

        if errors:
            for msg in errors:
                messages.error(request, msg)
            # Re-render list with modal open and form data for same-page correction
            request._bank_account_create_errors = True
            request._bank_account_create_post = {
                "bank_name": bank_name,
                "account_name": account_name,
                "account_number": account_number,
                "branch": branch,
                "office": office,
                "description": description,
                "currency_id": currency_id or "",
                "account_id": account_id or "",
                "opening_balance": raw_opening_balance or "",
                "opening_balance_date": raw_opening_date or "",
                "status": status,
            }
        else:
            is_active = status != "inactive"
            BankAccount.objects.using(tenant_db).create(
                bank_name=bank_name,
                account_name=account_name,
                account_number=account_number,
                branch=branch,
                office=office,
                description=description,
                currency=currency,
                account=account,
                opening_balance=opening_balance,
                opening_balance_date=opening_balance_date,
                is_active=is_active,
            )
            messages.success(request, "Bank account created.")
            return redirect(reverse("tenant_portal:cash_bank_accounts"))

    # Filters (GET)
    bank_name_f = (request.GET.get("bank_name") or "").strip()
    status_f = (request.GET.get("status") or "").strip().lower()
    currency_f = (request.GET.get("currency") or "").strip()
    coa_f = (request.GET.get("account_id") or "").strip()
    q = (request.GET.get("q") or "").strip()
    view_mode = (request.GET.get("view") or "").strip().lower()

    qs = (
        BankAccount.objects.using(tenant_db)
        .select_related("currency", "account")
        .order_by("bank_name", "account_name")
    )

    if bank_name_f:
        qs = qs.filter(bank_name__icontains=bank_name_f)
    if currency_f:
        qs = qs.filter(currency__code__iexact=currency_f)
    if status_f == "active":
        qs = qs.filter(is_active=True)
    elif status_f == "inactive":
        qs = qs.filter(is_active=False)
    if coa_f:
        qs = qs.filter(account_id=coa_f)
    if q:
        qs = qs.filter(Q(account_name__icontains=q) | Q(account_number__icontains=q))

    bank_accounts = list(qs)

    # Current balance per linked GL account:
    # opening balance (from BankAccount) + posted journal movements.
    balances = {}
    transaction_exists = {}  # account_id -> True if any journal lines exist (for delete guard)
    if bank_accounts:
        account_ids = [ba.account_id for ba in bank_accounts if ba.account_id]
        if account_ids:
            bal_rows = (
                JournalLine.objects.using(tenant_db)
                .filter(account_id__in=account_ids, entry__status=JournalEntry.Status.POSTED)
                .values("account_id")
                .annotate(balance=Sum("debit") - Sum("credit"))
            )
            for row in bal_rows:
                balances[row["account_id"]] = row.get("balance") or Decimal("0")
            line_counts = (
                JournalLine.objects.using(tenant_db)
                .filter(account_id__in=account_ids)
                .values("account_id")
                .annotate(cnt=Count("id"))
            )
            for row in line_counts:
                transaction_exists[row["account_id"]] = (row.get("cnt") or 0) > 0

        # Add opening balances on top of journal movements
        for ba in bank_accounts:
            if not ba.account_id:
                continue
            base = balances.get(ba.account_id, Decimal("0"))
            balances[ba.account_id] = base + (ba.opening_balance or Decimal("0"))

    rows = []
    for ba in bank_accounts:
        rows.append(
            {
                "obj": ba,
                "current_balance": balances.get(ba.account_id, Decimal("0")),
                "has_transactions": transaction_exists.get(ba.account_id, False),
            }
        )

    currencies = Currency.objects.using(tenant_db).order_by("code")

    # Prefer active asset-type accounts for linking bank accounts. If none exist
    # (e.g. tenant just seeded the chart of accounts with different types),
    # fall back to all active accounts so the dropdown is never empty.
    coa_qs_base = ChartAccount.objects.using(tenant_db).filter(is_active=True)
    coa_asset = coa_qs_base.filter(type=ChartAccount.Type.ASSET)
    if coa_asset.exists():
        coa_accounts = coa_asset.order_by("code")
    else:
        coa_accounts = coa_qs_base.order_by("code")

    # KPI stats (all bank accounts, unfiltered)
    all_bank = list(
        BankAccount.objects.using(tenant_db).values_list(
            "id", "account_id", "is_active", "opening_balance"
        )
    )
    kpi_total_count = len(all_bank)
    kpi_active_count = sum(1 for _ in all_bank if _[2])
    kpi_inactive_count = kpi_total_count - kpi_active_count
    all_account_ids = [x[1] for x in all_bank if x[1]]
    kpi_total_balance = Decimal("0")
    if all_account_ids:
        kpi_bal_rows = (
            JournalLine.objects.using(tenant_db)
            .filter(account_id__in=all_account_ids, entry__status=JournalEntry.Status.POSTED)
            .values("account_id")
            .annotate(bal=Sum("debit") - Sum("credit"))
        )
        journal_balances = {r["account_id"]: (r.get("bal") or Decimal("0")) for r in kpi_bal_rows}
        # Add opening balances from each bank account
        for _id, account_id, is_active, opening_balance in all_bank:
            if not account_id:
                continue
            base = journal_balances.get(account_id, Decimal("0"))
            journal_balances[account_id] = base + (opening_balance or Decimal("0"))
        kpi_total_balance = sum(journal_balances.values())

    return render(
        request,
        "tenant_portal/finance/bank_accounts.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "currencies": currencies,
            "coa_accounts": coa_accounts,
            "filters": {
                "bank_name": bank_name_f,
                "status": status_f,
                "currency": currency_f,
                "account_id": coa_f,
                "q": q,
            },
            "can_manage": can_manage,
            "show_create_only": view_mode == "create",
            "show_create_modal": getattr(request, "_bank_account_create_errors", False),
            "create_form_data": getattr(request, "_bank_account_create_post", {}),
            "kpi_total_count": kpi_total_count,
            "kpi_total_balance": kpi_total_balance,
            "kpi_active_count": kpi_active_count,
            "kpi_inactive_count": kpi_inactive_count,
            "active_submenu": "cash",
            "active_item": "fund_bank_accounts",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def cash_cash_accounts_view(request: HttpRequest) -> HttpResponse:
    """
    Cash Accounts: reuse finance accounts page (cash/bank accounts).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_accounts") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def cash_petty_accounts_view(request: HttpRequest) -> HttpResponse:
    """
    Petty Cash Accounts: reuse finance accounts page to manage petty cash accounts.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_accounts") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbook:entries.view")
def cash_main_cashbook_view(request: HttpRequest) -> HttpResponse:
    """
    Main Cashbook: organizational cash transactions not linked to donor-funded projects.
    Shows lines that hit cash/bank GL accounts (from BankAccount), grant_id is null,
    with running balance. Filters: date, office, transaction type.
    """
    from decimal import Decimal

    from django.db.models import Q

    from tenant_finance.models import BankAccount, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    office_filter = (request.GET.get("office") or "").strip()
    txn_type_filter = (request.GET.get("transaction_type") or "").strip().lower()

    # GL account IDs that are linked to bank/cash accounts (organizational)
    cash_account_ids = list(
        BankAccount.objects.using(tenant_db).values_list("account_id", flat=True).distinct()
    )
    if not cash_account_ids:
        rows = []
        offices = []
    else:
        qs = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account_id__in=cash_account_ids,
                entry__status=JournalEntry.Status.POSTED,
                entry__grant__isnull=True,
            )
            .filter(
                entry__entry_date__gte=f["period_start"],
                entry__entry_date__lte=f["period_end"],
            )
            .select_related("entry", "entry__created_by", "account")
            .order_by("entry__entry_date", "entry_id", "id")
        )
        if office_filter:
            # Filter by bank account office (exact match from dropdown)
            bank_account_ids_office = list(
                BankAccount.objects.using(tenant_db)
                .filter(office=office_filter)
                .values_list("account_id", flat=True)
            )
            if bank_account_ids_office:
                qs = qs.filter(account_id__in=bank_account_ids_office)
            else:
                qs = qs.none()
        if txn_type_filter and txn_type_filter != "all":
            type_prefix = {"payment": "PV", "receipt": "RV", "journal": "JV", "other": ""}.get(
                txn_type_filter
            )
            if type_prefix == "":
                # Other: reference not starting with PV, RV, JV
                qs = qs.exclude(
                    Q(entry__reference__istartswith="PV")
                    | Q(entry__reference__istartswith="RV")
                    | Q(entry__reference__istartswith="JV")
                )
            else:
                qs = qs.filter(entry__reference__istartswith=type_prefix)

        # Build account_id -> office lookup from BankAccount
        bank_by_account = {
            ba.account_id: ba
            for ba in BankAccount.objects.using(tenant_db).select_related("account")
        }
        offices = sorted(
            {ba.office for ba in bank_by_account.values() if ba.office},
            key=lambda x: (x or "").lower(),
        )

        running_balance = Decimal("0")
        rows = []
        for line in qs[:500]:
            entry = line.entry
            debit = line.debit or Decimal("0")
            credit = line.credit or Decimal("0")
            running_balance += debit - credit
            ba = bank_by_account.get(line.account_id)
            office = (ba.office or "—") if ba else "—"
            ref = entry.reference or f"JE-{entry.id:05d}"
            desc = (line.description or entry.memo or "—")[:120]
            entered_by = ""
            if entry.created_by_id:
                u = getattr(entry, "created_by", None)
                entered_by = getattr(u, "username", None) or getattr(u, "email", None) or str(entry.created_by_id)
            rows.append({
                "date": entry.entry_date,
                "reference": ref,
                "description": desc,
                "debit": debit,
                "credit": credit,
                "balance": running_balance,
                "office": office,
                "entered_by": entered_by or "—",
                "status": entry.get_status_display() if hasattr(entry, "get_status_display") else entry.status,
            })

    return render(
        request,
        "tenant_portal/cash/main_cashbook.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "cash",
            "active_item": "cash_main_cashbook",
            "rows": rows,
            "filters": {
                "period_start": f["period_start"],
                "period_end": f["period_end"],
                "office": office_filter,
                "transaction_type": txn_type_filter,
            },
            "offices": offices,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="cashbank:accounts.view")
def cash_bank_transfers_view(request: HttpRequest) -> HttpResponse:
    """
    Bank-to-bank GL transfers: Dr destination bank GL, Cr source bank GL.
    """
    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date
    from django.contrib import messages

    from tenant_finance.models import BankAccount, ChartAccount
    from tenant_grants.models import Grant
    from tenant_finance.services.journal_posting import post_bank_transfer

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_post = user_has_permission(user, "module:finance.manage", using=tenant_db)

    bank_rows = (
        BankAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .select_related("account", "currency")
        .order_by("bank_name", "account_name")
    )

    error = None
    if request.method == "POST" and can_post:
        entry_date = parse_date(request.POST.get("transfer_date") or "")
        grant_id = (request.POST.get("grant_id") or "").strip()
        desc = (request.POST.get("description") or "").strip() or "Bank transfer"
        from_id = request.POST.get("from_account_id")
        to_id = request.POST.get("to_account_id")
        try:
            amount = Decimal(str(request.POST.get("amount") or "0"))
        except (InvalidOperation, ValueError):
            amount = Decimal("0")

        if not entry_date:
            error = "Transfer date is required."
        elif amount <= 0:
            error = "Amount must be greater than zero."
        elif not from_id or not to_id or from_id == to_id:
            error = "Select two different bank accounts."
        else:
            from_acc = ChartAccount.objects.using(tenant_db).filter(pk=from_id).first()
            to_acc = ChartAccount.objects.using(tenant_db).filter(pk=to_id).first()
            valid_ids = {ba.account_id for ba in bank_rows}
            if not from_acc or not to_acc or from_acc.id not in valid_ids or to_acc.id not in valid_ids:
                error = "Invalid bank account selection."
            else:
                grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first() if grant_id else None
                try:
                    entry = post_bank_transfer(
                        using=tenant_db,
                        user=user,
                        entry_date=entry_date,
                        amount=amount,
                        description=desc,
                        grant=grant,
                        from_account=from_acc,
                        to_account=to_acc,
                    )
                    messages.success(
                        request,
                        f"Bank transfer posted. Journal {entry.reference or entry.id}.",
                    )
                except Exception as exc:
                    error = str(exc) or "Posting failed."

    if error:
        messages.error(request, error)

    grants = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")

    return render(
        request,
        "tenant_portal/cash/bank_transfer.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "bank_rows": bank_rows,
            "grants": grants,
            "can_post": can_post,
            "active_submenu": "cash",
            "active_item": "cash_bank_transfers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="cashbank:accounts.view")
def cash_cash_transfers_view(request: HttpRequest) -> HttpResponse:
    """
    Cash / on-hand GL transfers (non-bank asset accounts).
    """
    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date
    from django.contrib import messages

    from tenant_finance.models import BankAccount, ChartAccount
    from tenant_grants.models import Grant
    from tenant_finance.services.journal_posting import post_cash_transfer

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_post = user_has_permission(user, "module:finance.manage", using=tenant_db)

    bank_ids = list(
        BankAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .values_list("account_id", flat=True)
    )
    cash_accounts = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True, type=ChartAccount.Type.ASSET)
        .exclude(pk__in=bank_ids or [0])
        .order_by("code")
    )
    if not cash_accounts.exists():
        cash_accounts = ChartAccount.objects.using(tenant_db).filter(
            is_active=True, type=ChartAccount.Type.ASSET
        ).order_by("code")

    error = None
    if request.method == "POST" and can_post:
        entry_date = parse_date(request.POST.get("transfer_date") or "")
        grant_id = (request.POST.get("grant_id") or "").strip()
        desc = (request.POST.get("description") or "").strip() or "Cash transfer"
        from_id = request.POST.get("from_account_id")
        to_id = request.POST.get("to_account_id")
        try:
            amount = Decimal(str(request.POST.get("amount") or "0"))
        except (InvalidOperation, ValueError):
            amount = Decimal("0")

        if not entry_date:
            error = "Transfer date is required."
        elif amount <= 0:
            error = "Amount must be greater than zero."
        elif not from_id or not to_id or from_id == to_id:
            error = "Select two different accounts."
        else:
            from_acc = ChartAccount.objects.using(tenant_db).filter(pk=from_id).first()
            to_acc = ChartAccount.objects.using(tenant_db).filter(pk=to_id).first()
            allowed = {a.id for a in cash_accounts}
            if not from_acc or not to_acc or from_acc.id not in allowed or to_acc.id not in allowed:
                error = "Invalid cash account selection."
            else:
                grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first() if grant_id else None
                try:
                    entry = post_cash_transfer(
                        using=tenant_db,
                        user=user,
                        entry_date=entry_date,
                        amount=amount,
                        description=desc,
                        grant=grant,
                        from_account=from_acc,
                        to_account=to_acc,
                    )
                    messages.success(
                        request,
                        f"Cash transfer posted. Journal {entry.reference or entry.id}.",
                    )
                except Exception as exc:
                    error = str(exc) or "Posting failed."

    if error:
        messages.error(request, error)

    grants = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")

    return render(
        request,
        "tenant_portal/cash/cash_transfer.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "cash_accounts": cash_accounts,
            "grants": grants,
            "can_post": can_post,
            "active_submenu": "cash",
            "active_item": "cash_cash_transfers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_interfund_transfers_view(request: HttpRequest) -> HttpResponse:
    """List inter-fund transfers (operational register) with filters."""
    from decimal import Decimal, InvalidOperation

    from django.db.models import Q
    from django.utils.dateparse import parse_date

    from tenant_finance.db_compat import interfund_tables_present
    from tenant_finance.models import InterFundTransfer
    from tenant_grants.models import Grant, Project

    tenant_db = request.tenant_db
    can_manage = user_has_permission(request.tenant_user, "module:finance.manage", using=tenant_db)
    interfund_ok = interfund_tables_present(tenant_db)

    date_from_s = (request.GET.get("date_from") or "").strip()
    date_to_s = (request.GET.get("date_to") or "").strip()
    fund_q = (request.GET.get("fund") or "").strip()
    status_f = (request.GET.get("status") or "").strip()
    grant_side = (request.GET.get("grant_side") or "").strip()
    grant_id_s = (request.GET.get("grant_id") or "").strip()
    project_side = (request.GET.get("project_side") or "").strip()
    project_id_s = (request.GET.get("project_id") or "").strip()
    amount_min_s = (request.GET.get("amount_min") or "").strip()
    amount_max_s = (request.GET.get("amount_max") or "").strip()

    date_from = parse_date(date_from_s) if date_from_s else None
    date_to = parse_date(date_to_s) if date_to_s else None
    amount_min = None
    amount_max = None
    try:
        if amount_min_s:
            amount_min = Decimal(amount_min_s)
    except (InvalidOperation, ValueError):
        amount_min = None
    try:
        if amount_max_s:
            amount_max = Decimal(amount_max_s)
    except (InvalidOperation, ValueError):
        amount_max = None

    grant_id_filter = None
    if str(grant_id_s).isdigit():
        grant_id_filter = int(grant_id_s)

    project_id_filter = None
    if str(project_id_s).isdigit():
        project_id_filter = int(project_id_s)

    grants_for_filter = []
    projects_for_filter = []
    transfers = []
    if interfund_ok:
        grants_for_filter = list(
            Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")[:500]
        )
        projects_for_filter = list(
            Project.objects.using(tenant_db)
            .filter(status=Project.Status.ACTIVE)
            .order_by("code")[:500]
        )
        qs = InterFundTransfer.objects.using(tenant_db).select_related(
            "rule",
            "created_by",
            "approved_by",
            "posted_journal",
            "posted_journal__posted_by",
            "reversal_journal",
            "currency",
            "from_grant",
            "from_grant__project",
            "to_grant",
            "to_grant__project",
            "from_project",
            "to_project",
            "from_bank_account",
            "from_bank_account__currency",
            "to_bank_account",
            "to_bank_account__currency",
            "donor",
            "reversed_by",
        )
        if date_from:
            qs = qs.filter(transfer_date__gte=date_from)
        if date_to:
            qs = qs.filter(transfer_date__lte=date_to)
        if fund_q:
            fq = fund_q
            qs = qs.filter(
                Q(from_fund_code__icontains=fq)
                | Q(to_fund_code__icontains=fq)
                | Q(from_grant__code__icontains=fq)
                | Q(to_grant__code__icontains=fq)
            )
        if status_f and status_f in {c[0] for c in InterFundTransfer.Status.choices}:
            qs = qs.filter(status=status_f)
        if grant_id_filter:
            if grant_side == "from":
                qs = qs.filter(from_grant_id=grant_id_filter)
            elif grant_side == "to":
                qs = qs.filter(to_grant_id=grant_id_filter)
            else:
                qs = qs.filter(Q(from_grant_id=grant_id_filter) | Q(to_grant_id=grant_id_filter))
        if project_id_filter:
            if project_side == "from":
                qs = qs.filter(from_grant__project_id=project_id_filter)
            elif project_side == "to":
                qs = qs.filter(to_grant__project_id=project_id_filter)
            else:
                qs = qs.filter(
                    Q(from_grant__project_id=project_id_filter)
                    | Q(to_grant__project_id=project_id_filter)
                )
        if amount_min is not None:
            qs = qs.filter(amount__gte=amount_min)
        if amount_max is not None:
            qs = qs.filter(amount__lte=amount_max)
        transfers = list(qs.order_by("-transfer_date", "-id")[:200])

    return render(
        request,
        "tenant_portal/finance/interfund_transfers.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "transfers": transfers,
            "can_manage": can_manage,
            "filter_date_from": date_from_s,
            "filter_date_to": date_to_s,
            "filter_fund": fund_q,
            "filter_status": status_f,
            "filter_grant_id": grant_id_s,
            "filter_grant_side": grant_side,
            "filter_project_id": project_id_s,
            "filter_project_side": project_side,
            "filter_amount_min": amount_min_s,
            "filter_amount_max": amount_max_s,
            "grants_for_filter": grants_for_filter,
            "projects_for_filter": projects_for_filter,
            "status_choices": InterFundTransfer.Status.choices,
            "active_submenu": "core",
            "active_item": "core_interfund",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_interfund_transfer_create_view(request: HttpRequest) -> HttpResponse:
    """Create an inter-fund transfer (uses configured rules); always starts as Draft."""
    from decimal import Decimal, InvalidOperation

    from django.contrib import messages
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.utils.dateparse import parse_date
    from django.utils.translation import gettext as _

    from tenant_finance.db_compat import interfund_tables_present
    from tenant_finance.models import InterFundTransferRule
    from tenant_finance.services.interfund_transfer import InterFundTransferEngine
    from tenant_finance.services.interfund_validation import validate_interfund_transfer_core

    tenant_db = request.tenant_db
    user = request.tenant_user
    interfund_ok = interfund_tables_present(tenant_db)

    if request.method == "GET" and request.GET.get("src") != "post":
        messages.info(
            request,
            _("Inter-fund transfers are created from Post transaction → Inter-Fund Transfer."),
        )
        return redirect(reverse("tenant_portal:finance_post_transaction"))

    from tenant_finance.services.interfund_project_bank import (
        build_payload_from_projects_and_banks,
        projects_payload,
    )

    if interfund_ok:
        engine = InterFundTransferEngine(tenant_db)
        rules = list(
            InterFundTransferRule.objects.using(tenant_db)
            .filter(status=InterFundTransferRule.Status.ACTIVE)
            .order_by("name")
        )
        project_rows, banks_by_project = projects_payload(tenant_db=tenant_db)
        interfund_data = {
            "projects": project_rows,
            "banksByProject": banks_by_project,
            "initial": None,
        }
    else:
        engine = None
        rules = []
        interfund_data = {"projects": [], "banksByProject": {}, "initial": None}

    if request.method == "POST":
        if not interfund_ok:
            messages.error(
                request,
                "Inter-fund tables are missing. Run tenant migrations (tenant_finance 0032 or later), "
                "e.g. python manage.py migrate_tenant --tenant YOUR_SLUG",
            )
        else:
            transfer_date = parse_date(request.POST.get("transfer_date") or "")
            planned_posting_date = parse_date(request.POST.get("planned_posting_date") or "") or None
            reason = (request.POST.get("reason") or "").strip()
            description = (request.POST.get("description") or "").strip()
            try:
                amount = Decimal(str(request.POST.get("amount") or "0"))
            except (InvalidOperation, ValueError):
                amount = Decimal("0")

            from_pid = (request.POST.get("from_project_id") or "").strip()
            to_pid = (request.POST.get("to_project_id") or "").strip()
            from_bid = (request.POST.get("from_bank_account_id") or "").strip()
            to_bid = (request.POST.get("to_bank_account_id") or "").strip()
            reference_no = (request.POST.get("reference_no") or "").strip()

            if not transfer_date:
                messages.error(request, _("Transfer date is required."))
            elif not planned_posting_date:
                messages.error(request, _("Posting date is required."))
            elif not description:
                messages.error(request, _("Description is required."))
            elif amount <= 0:
                messages.error(request, _("Amount must be greater than zero."))
            elif not all(x.isdigit() for x in (from_pid, to_pid, from_bid, to_bid)):
                messages.error(
                    request,
                    _("Source and destination project and bank account are required."),
                )
            else:
                try:
                    payload = build_payload_from_projects_and_banks(
                        tenant_db=tenant_db,
                        from_project_id=int(from_pid),
                        to_project_id=int(to_pid),
                        from_bank_account_id=int(from_bid),
                        to_bank_account_id=int(to_bid),
                    )
                except ValueError as exc:
                    messages.error(request, str(exc))
                    payload = None

                if payload:
                    from_ft = payload["from_fund_type"]
                    to_ft = payload["to_fund_type"]
                    from_code = payload["from_fund_code"]
                    to_code = payload["to_fund_code"]
                    from_grant = payload["from_grant"]
                    to_grant = payload["to_grant"]
                    currency_id = payload["currency_id"]
                    donor_id = payload["donor_id"]

                    rule = engine.select_rule(
                        from_fund_type=from_ft,
                        to_fund_type=to_ft,
                        from_fund_code=from_code,
                        to_fund_code=to_code,
                        transfer_date=transfer_date,
                    )
                    if not rule:
                        messages.error(
                            request,
                            _(
                                "No active inter-fund rule matches PROJECT→PROJECT for this date. "
                                "Add a rule (project → project) under Financial Setup."
                            ),
                        )
                    else:
                        chk = engine.check_transfer(
                            from_fund_type=from_ft,
                            to_fund_type=to_ft,
                            from_fund_code=from_code,
                            to_fund_code=to_code,
                            amount=amount,
                            transfer_date=transfer_date,
                        )
                        if chk.status != "ok":
                            messages.error(request, chk.message or _("Transfer is not allowed."))
                        elif rule.require_reason and not (reason or description).strip():
                            messages.error(
                                request,
                                _("Reason or description is required for this rule."),
                            )
                        else:
                            try:
                                validate_interfund_transfer_core(
                                    from_fund_code=from_code,
                                    to_fund_code=to_code,
                                    amount=amount,
                                    transfer_date=transfer_date,
                                    tenant_db=tenant_db,
                                    user_id=getattr(user, "pk", None),
                                    from_grant=from_grant,
                                    to_grant=to_grant,
                                    require_fiscal_open=True,
                                    fiscal_period_date=planned_posting_date,
                                )
                            except ValueError as exc:
                                messages.error(request, str(exc))
                            else:
                                transfer = engine.create_transfer(
                                    rule=rule,
                                    from_fund_type=from_ft,
                                    to_fund_type=to_ft,
                                    from_fund_code=from_code,
                                    to_fund_code=to_code,
                                    amount=amount,
                                    transfer_date=transfer_date,
                                    reason=reason,
                                    user=user,
                                    description=description,
                                    currency_id=currency_id,
                                    from_grant=from_grant,
                                    to_grant=to_grant,
                                    reference_no=reference_no,
                                    donor_id=donor_id,
                                    planned_posting_date=planned_posting_date,
                                    from_project=payload["from_project"],
                                    to_project=payload["to_project"],
                                    from_bank_account=payload["from_bank_account"],
                                    to_bank_account=payload["to_bank_account"],
                                )
                                if request.FILES.get("attachment"):
                                    transfer.attachment = request.FILES["attachment"]
                                    transfer.save(using=tenant_db, update_fields=["attachment"])
                                transfer.refresh_from_db()
                                ref = (transfer.transfer_no or "").strip() or f"#{transfer.id}"
                                messages.success(
                                    request,
                                    f"Inter-fund transfer {ref} created as {transfer.get_status_display()}.",
                                )
                                return redirect(
                                    "tenant_portal:finance_interfund_transfer_detail",
                                    pk=transfer.pk,
                                )

    return render(
        request,
        "tenant_portal/finance/interfund_transfer_form.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "rules": rules,
            "interfund_data": interfund_data,
            "active_submenu": "core",
            "active_item": "core_interfund",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def finance_interfund_transfer_detail_view(request: HttpRequest, pk: int) -> HttpResponse:
    """View / submit / approve / reject / post / reverse inter-fund transfer (workflow)."""
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    from django.urls import reverse

    from tenant_finance.db_compat import interfund_tables_present
    from tenant_finance.models import InterFundTransfer
    from tenant_finance.services.interfund_workflow import apply_interfund_action

    tenant_db = request.tenant_db
    user = request.tenant_user
    can_manage = user_has_permission(user, "module:finance.manage", using=tenant_db)

    if not interfund_tables_present(tenant_db):
        messages.error(
            request,
            "Inter-fund tables are missing. Run tenant migrations first, then open transfers from the list.",
        )
        return redirect(reverse("tenant_portal:finance_interfund_transfers"))

    transfer = get_object_or_404(
        InterFundTransfer.objects.using(tenant_db).select_related(
            "rule",
            "created_by",
            "approved_by",
            "posted_journal",
            "posted_journal__posted_by",
            "reversal_journal",
            "reversed_by",
            "currency",
            "from_grant",
            "to_grant",
            "from_project",
            "to_project",
            "from_bank_account",
            "to_bank_account",
            "donor",
        ),
        pk=pk,
    )

    if request.method == "POST" and can_manage:
        action = (request.POST.get("action") or "").strip().lower()
        reversal_reason = (request.POST.get("reversal_reason") or "").strip()
        try:
            msg = apply_interfund_action(
                transfer=transfer,
                action=action,
                tenant_db=tenant_db,
                user=user,
                reversal_reason=reversal_reason,
            )
            messages.success(request, msg)
        except ValueError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            messages.error(request, str(exc) or "Action failed.")
        return redirect(reverse("tenant_portal:finance_interfund_transfer_detail", args=[pk]))

    return render(
        request,
        "tenant_portal/finance/interfund_transfer_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "transfer": transfer,
            "can_manage": can_manage,
            "active_submenu": "core",
            "active_item": "core_interfund",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def finance_interfund_transfer_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    """Edit a draft inter-fund transfer only."""
    from decimal import Decimal, InvalidOperation

    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    from django.urls import reverse
    from django.utils.dateparse import parse_date
    from django.utils.translation import gettext as _

    from tenant_finance.db_compat import interfund_tables_present
    from tenant_finance.models import InterFundTransfer, InterFundTransferRule
    from tenant_finance.services.interfund_project_bank import (
        build_payload_from_projects_and_banks,
        projects_payload,
    )
    from tenant_finance.services.interfund_transfer import InterFundTransferEngine
    from tenant_finance.services.interfund_validation import validate_interfund_transfer_core

    tenant_db = request.tenant_db
    user = request.tenant_user

    if not interfund_tables_present(tenant_db):
        messages.error(request, "Inter-fund tables are missing. Run tenant migrations first.")
        return redirect(reverse("tenant_portal:finance_interfund_transfers"))

    transfer = get_object_or_404(
        InterFundTransfer.objects.using(tenant_db).select_related(
            "rule",
            "from_grant",
            "to_grant",
            "from_grant__project",
            "to_grant__project",
            "currency",
            "donor",
            "from_project",
            "to_project",
            "from_bank_account",
            "to_bank_account",
        ),
        pk=pk,
    )
    if transfer.status != InterFundTransfer.Status.DRAFT:
        messages.error(request, "Only draft transfers can be edited.")
        return redirect(reverse("tenant_portal:finance_interfund_transfer_detail", args=[pk]))

    engine = InterFundTransferEngine(tenant_db)
    rules = list(
        InterFundTransferRule.objects.using(tenant_db)
        .filter(status=InterFundTransferRule.Status.ACTIVE)
        .order_by("name")
    )
    project_rows, banks_by_project = projects_payload(tenant_db=tenant_db)

    def _initial_payload() -> dict:
        fp = transfer.from_project_id
        tp = transfer.to_project_id
        fb = transfer.from_bank_account_id
        tb = transfer.to_bank_account_id
        if not fp and transfer.from_grant_id and transfer.from_grant.project_id:
            fp = transfer.from_grant.project_id
        if not tp and transfer.to_grant_id and transfer.to_grant.project_id:
            tp = transfer.to_grant.project_id
        if not fb and transfer.from_grant_id and transfer.from_grant.bank_account_id:
            fb = transfer.from_grant.bank_account_id
        if not tb and transfer.to_grant_id and transfer.to_grant.bank_account_id:
            tb = transfer.to_grant.bank_account_id
        return {
            "fromProjectId": str(fp) if fp else "",
            "toProjectId": str(tp) if tp else "",
            "fromBankId": str(fb) if fb else "",
            "toBankId": str(tb) if tb else "",
        }

    interfund_data = {
        "projects": project_rows,
        "banksByProject": banks_by_project,
        "initial": _initial_payload(),
    }

    if request.method == "POST":
        transfer_date = parse_date(request.POST.get("transfer_date") or "")
        planned_posting_date = parse_date(request.POST.get("planned_posting_date") or "") or None
        reason = (request.POST.get("reason") or "").strip()
        description = (request.POST.get("description") or "").strip()
        try:
            amount = Decimal(str(request.POST.get("amount") or "0"))
        except (InvalidOperation, ValueError):
            amount = Decimal("0")

        from_pid = (request.POST.get("from_project_id") or "").strip()
        to_pid = (request.POST.get("to_project_id") or "").strip()
        from_bid = (request.POST.get("from_bank_account_id") or "").strip()
        to_bid = (request.POST.get("to_bank_account_id") or "").strip()
        reference_no = (request.POST.get("reference_no") or "").strip()

        if not transfer_date:
            messages.error(request, _("Transfer date is required."))
        elif not planned_posting_date:
            messages.error(request, _("Posting date is required."))
        elif not description:
            messages.error(request, _("Description is required."))
        elif amount <= 0:
            messages.error(request, _("Amount must be greater than zero."))
        elif not all(x.isdigit() for x in (from_pid, to_pid, from_bid, to_bid)):
            messages.error(
                request,
                _("Source and destination project and bank account are required."),
            )
        else:
            try:
                payload = build_payload_from_projects_and_banks(
                    tenant_db=tenant_db,
                    from_project_id=int(from_pid),
                    to_project_id=int(to_pid),
                    from_bank_account_id=int(from_bid),
                    to_bank_account_id=int(to_bid),
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                payload = None

            if payload:
                from_ft = payload["from_fund_type"]
                to_ft = payload["to_fund_type"]
                from_code = payload["from_fund_code"]
                to_code = payload["to_fund_code"]
                from_grant = payload["from_grant"]
                to_grant = payload["to_grant"]
                currency_id = payload["currency_id"]
                donor_id = payload["donor_id"]

                rule = engine.select_rule(
                    from_fund_type=from_ft,
                    to_fund_type=to_ft,
                    from_fund_code=from_code,
                    to_fund_code=to_code,
                    transfer_date=transfer_date,
                )
                if not rule:
                    messages.error(
                        request,
                        _(
                            "No active inter-fund rule matches PROJECT→PROJECT for this date. "
                            "Add a rule under Financial Setup."
                        ),
                    )
                else:
                    chk = engine.check_transfer(
                        from_fund_type=from_ft,
                        to_fund_type=to_ft,
                        from_fund_code=from_code,
                        to_fund_code=to_code,
                        amount=amount,
                        transfer_date=transfer_date,
                    )
                    if chk.status != "ok":
                        messages.error(request, chk.message or _("Transfer is not allowed."))
                    elif rule.require_reason and not (reason or description).strip():
                        messages.error(
                            request,
                            _("Reason or description is required for this rule."),
                        )
                    else:
                        try:
                            validate_interfund_transfer_core(
                                from_fund_code=from_code,
                                to_fund_code=to_code,
                                amount=amount,
                                transfer_date=transfer_date,
                                tenant_db=tenant_db,
                                user_id=getattr(user, "pk", None),
                                from_grant=from_grant,
                                to_grant=to_grant,
                                require_fiscal_open=True,
                                fiscal_period_date=planned_posting_date,
                            )
                        except ValueError as exc:
                            messages.error(request, str(exc))
                        else:
                            transfer.rule = rule
                            transfer.transfer_date = transfer_date
                            transfer.from_fund_type = from_ft
                            transfer.to_fund_type = to_ft
                            transfer.from_fund_code = from_code
                            transfer.to_fund_code = to_code
                            transfer.amount = amount
                            transfer.description = description
                            transfer.reason = reason
                            transfer.currency_id = currency_id
                            transfer.from_grant = from_grant
                            transfer.to_grant = to_grant
                            transfer.reference_no = reference_no
                            transfer.donor_id = donor_id
                            transfer.planned_posting_date = planned_posting_date
                            transfer.from_project = payload["from_project"]
                            transfer.to_project = payload["to_project"]
                            transfer.from_bank_account = payload["from_bank_account"]
                            transfer.to_bank_account = payload["to_bank_account"]
                            if request.FILES.get("attachment"):
                                transfer.attachment = request.FILES["attachment"]
                            transfer.save(
                                using=tenant_db,
                                update_fields=[
                                    "rule",
                                    "transfer_date",
                                    "from_fund_type",
                                    "to_fund_type",
                                    "from_fund_code",
                                    "to_fund_code",
                                    "amount",
                                    "description",
                                    "reason",
                                    "currency",
                                    "from_grant",
                                    "to_grant",
                                    "reference_no",
                                    "donor",
                                    "planned_posting_date",
                                    "from_project",
                                    "to_project",
                                    "from_bank_account",
                                    "to_bank_account",
                                    "attachment",
                                    "updated_at",
                                ],
                            )
                            ref = (transfer.transfer_no or "").strip() or f"#{transfer.id}"
                            messages.success(request, f"Inter-fund transfer {ref} updated.")
                            return redirect(
                                "tenant_portal:finance_interfund_transfer_detail",
                                pk=transfer.pk,
                            )

    return render(
        request,
        "tenant_portal/finance/interfund_transfer_form.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "rules": rules,
            "interfund_data": interfund_data,
            "edit_transfer": transfer,
            "active_submenu": "core",
            "active_item": "core_interfund",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="cashbank:petty_cash.create")
def cash_petty_topup_view(request: HttpRequest) -> HttpResponse:
    """
    Petty Cash Top-Up: reuse general journal entry page (petty top-ups as journals).
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_journals") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:petty_cash.view")
def cash_petty_expenses_view(request: HttpRequest) -> HttpResponse:
    """
    Petty Cash Expense Register: reuse general journal entry page filtered by petty accounts.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_journals") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:reconciliation.view")
def cash_bank_import_view(request: HttpRequest) -> HttpResponse:
    """
    Bank Statements Import: temporary redirect to recent transactions page,
    which already shows bank movements with filters and export.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_recent_transactions") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:reconciliation.view")
def cash_bank_recon_view(request: HttpRequest) -> HttpResponse:
    """
    Bank Reconciliation: reuse cash position view as reconciliation summary.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_cash_position") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:reconciliation.view")
def cash_cash_count_view(request: HttpRequest) -> HttpResponse:
    """
    Cash Count & Verification: reuse cash position view for current balances.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_cash_position") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="cashbank:reconciliation.view")
def cash_cash_recon_view(request: HttpRequest) -> HttpResponse:
    """
    Cash Reconciliation & Reports: reuse fund balances page.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_fund_balances") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def outgoing_fund_center_view(request: HttpRequest) -> HttpResponse:
    """Outgoing fund module home: payment KPIs and navigation to payables workflows."""
    import calendar
    from collections import defaultdict
    from datetime import date
    from decimal import Decimal

    from django.db.models import Q, Sum
    from django.utils import timezone
    from django.utils.translation import gettext as _

    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    ps, pe = f["period_start"], f["period_end"]
    today = timezone.now().date()
    month_start = today.replace(day=1)

    def _fmt_money(d: Decimal) -> str:
        d = d.quantize(Decimal("0.01"))
        neg = d < 0
        d = abs(d)
        whole_s, _, frac = f"{d:.2f}".partition(".")
        whole_s = "{:,}".format(int(whole_s))
        return ("-" if neg else "") + whole_s + "." + frac

    payment_entry_q = (
        Q(reference__startswith="PV-")
        | Q(journal_type__iexact="payment_voucher")
        | Q(source_type=JournalEntry.SourceType.PAYMENT_VOUCHER)
    )

    def _expense_debit_sum(entry_ids: list) -> Decimal:
        if not entry_ids:
            return Decimal("0")
        t = (
            JournalLine.objects.using(tenant_db)
            .filter(
                entry_id__in=entry_ids,
                account__type=ChartAccount.Type.EXPENSE,
                debit__gt=0,
            )
            .aggregate(t=Sum("debit"))
            .get("t")
        )
        return t or Decimal("0")

    def _prev_month_range(today_d: date) -> tuple[date, date]:
        y, m = today_d.year, today_d.month
        if m == 1:
            y2, m2 = y - 1, 12
        else:
            y2, m2 = y, m - 1
        start = date(y2, m2, 1)
        last_d = calendar.monthrange(y2, m2)[1]
        return start, date(y2, m2, last_d)

    total_payments_period = Decimal("0")
    expenses_this_month = Decimal("0")
    outstanding_pay = Decimal("0")
    pending_payments_n = 0
    posted_pv_period_n = 0
    over_budget_n = 0
    top_vendors_display = "—"
    expense_mom_display = "—"
    recent_payments: list[dict] = []

    try:
        posted_period = JournalEntry.objects.using(tenant_db).filter(
            payment_entry_q,
            status=JournalEntry.Status.POSTED,
            entry_date__gte=ps,
            entry_date__lte=pe,
        )
        posted_pv_period_n = posted_period.count()
        ids_period = list(posted_period.values_list("id", flat=True))
        total_payments_period = _expense_debit_sum(ids_period)

        expenses_this_month = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account__type=ChartAccount.Type.EXPENSE,
                debit__gt=0,
                entry__status=JournalEntry.Status.POSTED,
                entry__entry_date__gte=month_start,
                entry__entry_date__lte=today,
            )
            .aggregate(t=Sum("debit"))
            .get("t")
        ) or Decimal("0")

        outstanding_pay = _get_global_financial_indicators(request).get("pending_payables") or Decimal("0")

        pending_payments_n = (
            JournalEntry.objects.using(tenant_db)
            .filter(
                payment_entry_q,
                status__in=(
                    JournalEntry.Status.DRAFT,
                    JournalEntry.Status.PENDING_APPROVAL,
                ),
            )
            .count()
        )

        from tenant_grants.models import BudgetLine, Grant

        active_grants_qs = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE)
        budgets_by_grant = {
            r["grant_id"]: r["total"] or Decimal("0")
            for r in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
        }
        spend_by_grant = {
            r["entry__grant_id"]: r["spent"] or Decimal("0")
            for r in JournalLine.objects.using(tenant_db)
            .filter(account__type=ChartAccount.Type.EXPENSE, entry__grant_id__isnull=False)
            .values("entry__grant_id")
            .annotate(spent=Sum("debit"))
        }
        for g in active_grants_qs.only("id", "award_amount"):
            b = budgets_by_grant.get(g.id, Decimal("0"))
            ceiling = b if b > 0 else Decimal(str(g.award_amount or 0))
            if ceiling <= 0:
                continue
            spent = spend_by_grant.get(g.id, Decimal("0"))
            if spent > ceiling:
                over_budget_n += 1

        vendor_totals: defaultdict[str, Decimal] = defaultdict(Decimal)
        for row in posted_period.values("id", "payee_name"):
            eid = row["id"]
            payee = (row.get("payee_name") or "").strip() or str(_("Unknown"))
            vendor_totals[payee] += _expense_debit_sum([eid])
        top_v = sorted(vendor_totals.items(), key=lambda x: x[1], reverse=True)[:3]
        if top_v:
            top_vendors_display = "; ".join(f"{n} ({_fmt_money(a)})" for n, a in top_v)

        p_start, p_end = _prev_month_range(today)
        expense_prev_month = (
            JournalLine.objects.using(tenant_db)
            .filter(
                account__type=ChartAccount.Type.EXPENSE,
                debit__gt=0,
                entry__status=JournalEntry.Status.POSTED,
                entry__entry_date__gte=p_start,
                entry__entry_date__lte=p_end,
            )
            .aggregate(t=Sum("debit"))
            .get("t")
        ) or Decimal("0")
        if expense_prev_month > 0:
            mom = ((expenses_this_month - expense_prev_month) / expense_prev_month) * Decimal("100")
            expense_mom_display = f"{mom.quantize(Decimal('0.1')):+.1f}%"
        elif expenses_this_month > 0:
            expense_mom_display = _("New activity")

        recent_qs = (
            JournalEntry.objects.using(tenant_db)
            .filter(payment_entry_q, status=JournalEntry.Status.POSTED)
            .select_related("grant")
            .order_by("-entry_date", "-id")[:10]
        )
        for e in recent_qs:
            eid = e.id
            amt = _expense_debit_sum([eid])
            payee = (e.payee_name or "").strip() or "—"
            recent_payments.append({
                "reference": (e.reference or f"PV-{e.id:05d}").strip(),
                "date": e.entry_date,
                "amount_fmt": _fmt_money(amt),
                "payee": payee,
            })
    except Exception:
        pass

    outgoing_kpis = [
        {"label": _("Total payments this period"), "value": _fmt_money(total_payments_period)},
        {"label": _("Total expenses this month"), "value": _fmt_money(expenses_this_month)},
        {"label": _("Outstanding payables"), "value": _fmt_money(outstanding_pay)},
        {"label": _("Payments pending approval"), "value": f"{pending_payments_n:,}"},
        {"label": _("Budget exceeded alerts"), "value": f"{over_budget_n:,}"},
        {"label": _("Top vendors"), "value": top_vendors_display},
        {"label": _("Expense vs prior month"), "value": expense_mom_display},
        {"label": _("Payment vouchers this period"), "value": f"{posted_pv_period_n:,}"},
    ]

    return render(
        request,
        "tenant_portal/pay/outgoing_fund_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "payables",
            "active_item": "pay_center_home",
            "outgoing_kpis": outgoing_kpis,
            "recent_payments": recent_payments,
            "outgoing_period_hint": f"{ps.isoformat()} — {pe.isoformat()}",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_payment_vouchers_view(request: HttpRequest) -> HttpResponse:
    """
    Dedicated Payment Voucher screen.

    - Users enter a payment voucher (date, bank/cash account, expense account,
      project/grant, amount, payee, description, attachment).
    - On submit, we POST a balanced JournalEntry:
        * Debit: expense account
        * Credit: payment (bank/cash) account
      and store the voucher as a JournalEntry with reference prefix 'PV-'.
    - The page also lists recent payment vouchers with filters and export.
    """
    from decimal import Decimal, InvalidOperation
    from django.db.models import Sum
    from django.utils import timezone
    from django.utils.dateparse import parse_date

    from tenant_finance.models import (
        ChartAccount,
        JournalEntry,
        JournalLine,
        JournalEntryAttachment,
        AuditLog,
    )
    from tenant_grants.models import Grant, GrantAssignment

    tenant_db = request.tenant_db
    user = request.tenant_user

    # Determine if user is finance manager (can see all projects)
    is_manager = user_has_permission(user, "module:finance.manage", using=tenant_db)

    # Handle new voucher POST (maker creates / saves draft / submits for approval)
    if request.method == "POST":
        errors = []

        # Mandatory fields
        raw_entry_date = (request.POST.get("entry_date") or "").strip()
        payee_type = (request.POST.get("payee_type") or "").strip()
        payee = (request.POST.get("payee") or "").strip()
        payment_method = (request.POST.get("payment_method") or "").strip()
        payment_account_id = request.POST.get("payment_account_id") or ""
        expense_account_id = request.POST.get("expense_account_id") or ""
        grant_id = request.POST.get("grant_id") or None
        description = (request.POST.get("description") or "").strip()

        # Voucher date: required, not in the future
        entry_date = None
        if not raw_entry_date:
            errors.append("Voucher date is required.")
        else:
            entry_date = parse_date(raw_entry_date)
            if not entry_date:
                errors.append("Voucher date is not a valid calendar date.")

        if entry_date and entry_date > timezone.localdate():
            errors.append("Voucher date cannot be in the future.")

        if not payee_type:
            errors.append("Payee type is required.")
        if not payee:
            errors.append("Payee name is required.")
        if not payment_method:
            errors.append("Payment method is required.")
        if not payment_account_id:
            errors.append("Bank / cash account is required.")
        if not expense_account_id:
            errors.append("Expense account is required.")
        if not description:
            errors.append("Purpose of payment is required.")

        # collect detail line amounts (multiple budget lines)
        from decimal import Decimal as _D

        detail_amounts = []
        for key, value in request.POST.items():
            if key.startswith("detail_amount_"):
                try:
                    val = Decimal(str(value or "0"))
                except (InvalidOperation, ValueError):
                    continue
                if val > 0:
                    detail_amounts.append(val)

        # fallback if older single-line field used
        if not detail_amounts:
            try:
                single = Decimal(str(request.POST.get("amount") or "0"))
            except (InvalidOperation, ValueError):
                single = Decimal("0")
            if single > 0:
                detail_amounts.append(single)

        attachment = request.FILES.get("attachment")

        total_amount = sum(detail_amounts) if detail_amounts else Decimal("0")
        if total_amount <= 0:
            errors.append("Total payment amount must be greater than zero and at least one payment detail line is required.")

        # Funding allocation validation (if co-funding is used)
        use_co_funding = request.POST.get("use_co_funding") == "1"
        if use_co_funding and total_amount > 0:
            funding_amounts = []
            for key, value in request.POST.items():
                if key.startswith("funding_amount_"):
                    try:
                        val = Decimal(str(value or "0"))
                    except (InvalidOperation, ValueError):
                        continue
                    if val > 0:
                        funding_amounts.append(val)

            funding_total = sum(funding_amounts) if funding_amounts else Decimal("0")
            if funding_total > total_amount:
                errors.append("Total funding allocation amount cannot exceed the total payment voucher amount.")
            elif funding_total != total_amount:
                errors.append("Total funding allocation amount must equal the total payment voucher amount.")

            # Percentage check (sum of row percentages should be 100%)
            if funding_amounts:
                pct_sum = sum((amt / total_amount) * Decimal("100") for amt in funding_amounts)
                # Allow small rounding tolerance of 0.01%
                if pct_sum.quantize(Decimal("0.01")) != Decimal("100.00"):
                    errors.append("Total funding allocation percentage must be 100%.")

        # If there are validation errors, show them and do not create voucher
        if errors:
            for msg in errors:
                messages.error(request, msg)
        else:
            # Additional runtime validations before creating voucher
            if entry_date:
                try:
                    _finance_assert_open_period(entry_date, tenant_db, getattr(request, "tenant_user_id", None))
                except ValueError as e:
                    messages.error(request, str(e))
                    return redirect(reverse("tenant_portal:pay_payment_vouchers"))

            # Only active projects can accept transactions
            from tenant_grants.models import Grant as _Grant

            if grant_id:
                grant_obj = _Grant.objects.using(tenant_db).filter(pk=grant_id).first()
                from django.utils import timezone as _tz
                today = _tz.localdate()
                if not grant_obj or grant_obj.status != _Grant.Status.ACTIVE or (
                    grant_obj.end_date and grant_obj.end_date < today
                ):
                    messages.error(
                        request,
                        "Payments cannot be recorded for an ended or inactive project.",
                    )
                    return redirect(reverse("tenant_portal:pay_payment_vouchers"))

            payment_account = ChartAccount.objects.using(tenant_db).filter(pk=payment_account_id).first()
            expense_account = ChartAccount.objects.using(tenant_db).filter(pk=expense_account_id).first()
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first() if grant_id else None

            # Finance officer restriction: must have an active assignment for the selected grant
            if not is_manager:
                if not grant:
                    messages.error(request, "You must select an assigned project to post a payment.")
                    grant = None
                else:
                    has_assignment = GrantAssignment.objects.using(tenant_db).filter(
                        grant=grant, officer=user, is_active=True
                    ).exists()
                    if not has_assignment:
                        messages.error(
                            request,
                            "You are not assigned to this project. Please contact the finance manager.",
                        )
                        grant = None

            if total_amount > 0 and payment_account and expense_account and (is_manager or grant):
                # Validate sufficient bank balance for payments before creating voucher
                from tenant_finance.models import BankAccount as _BankAccount

                bank_account = (
                    _BankAccount.objects.using(tenant_db)
                    .filter(account_id=payment_account.id, is_active=True)
                    .first()
                )
                if bank_account:
                    try:
                        _balance = _compute_bank_current_balance(bank_account, tenant_db)
                    except Exception:
                        _balance = None
                    if _balance is not None and _balance < total_amount:
                        messages.error(
                            request,
                            f"Insufficient bank balance. Current balance {_balance:.2f} is "
                            f"less than payment amount {total_amount:.2f}.",
                        )
                        return redirect(reverse("tenant_portal:pay_payment_vouchers"))
                # Determine workflow status: draft or pending approval
                action = (request.POST.get("action") or "").strip()
                if action == "save_draft":
                    status = JournalEntry.Status.DRAFT
                else:
                    status = JournalEntry.Status.PENDING_APPROVAL

                # Create journal entry (voucher header) — system source document for GL register
                entry = JournalEntry.objects.using(tenant_db).create(
                    entry_date=entry_date,
                    memo=description or f"Payment voucher for {payee or 'N/A'}",
                    grant=grant,
                    status=status,
                    created_by=request.tenant_user,
                    payee_name=payee or "",
                    payment_method=payment_method or "",
                    source=JournalEntry.SourceType.PAYMENT_VOUCHER,
                    source_type=JournalEntry.SourceType.PAYMENT_VOUCHER,
                    journal_type="payment_voucher",
                    is_system_generated=True,
                )
                # Set reference with PV prefix
                entry.reference = f"PV-{entry.id:05d}"
                entry.source_document_no = entry.reference
                entry.source_id = entry.pk
                entry.save(
                    using=tenant_db,
                    update_fields=["reference", "source_document_no", "source_id"],
                )

                # Audit trail: creation of payment voucher (draft or pending approval)
                try:
                    AuditLog.objects.using(tenant_db).create(
                        model_name="journalentry",
                        object_id=entry.id,
                        action=AuditLog.Action.CREATE,
                        user_id=getattr(request.tenant_user, "id", None),
                        username=getattr(request.tenant_user, "full_name", "") or getattr(
                            request.tenant_user, "email", ""
                        ),
                        old_data=None,
                        new_data={"status": entry.status, "reference": entry.reference},
                        summary=f"Created payment voucher {entry.reference} ({entry.get_status_display()})",
                    )
                except Exception:
                    # Audit logging should never block the main transaction
                    pass

                # Debit expense per detail line, credit bank/cash with total
                for amt in detail_amounts:
                    JournalLine.objects.using(tenant_db).create(
                        entry=entry,
                        account=expense_account,
                        description=description,
                        debit=amt,
                        credit=Decimal("0"),
                    )
                JournalLine.objects.using(tenant_db).create(
                    entry=entry,
                    account=payment_account,
                    description=description,
                    debit=Decimal("0"),
                    credit=total_amount,
                )

                # Attachment
                if attachment:
                    JournalEntryAttachment.objects.using(tenant_db).create(
                        entry=entry,
                        file=attachment,
                        original_filename=getattr(attachment, "name", "") or "",
                    )

                if status == JournalEntry.Status.DRAFT:
                    messages.success(request, f"Payment voucher {entry.reference} saved as draft.")
                else:
                    messages.success(
                        request,
                        f"Payment voucher {entry.reference} submitted for approval."
                    )
                return redirect(reverse("tenant_portal:pay_payment_vouchers"))

    # Filters for list
    f = _parse_finance_filters(request)
    grant_id = f.get("grant_id")

    vouchers_qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="PV-", entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])
        .select_related("grant")
        .order_by("-entry_date", "-id")
    )
    if grant_id:
        vouchers_qs = vouchers_qs.filter(grant_id=grant_id)

    vouchers = []
    for je in vouchers_qs[:100]:
        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )
        vouchers.append(
            {
                "id": je.id,
                "reference": je.reference or f"PV-{je.id:05d}",
                "date": je.entry_date,
                "project": je.grant.title if je.grant else "",
                "amount": total,
                "memo": je.memo,
            }
        )

    # Export
    export_format = request.GET.get("format") or ""
    if export_format:
        rows = [
            [
                v["date"],
                v["reference"],
                v["project"],
                v["amount"],
                v["memo"],
            ]
            for v in vouchers
        ]
        resp = _export_table_response(
            export_format=export_format,
            filename_base="payment_vouchers",
            title="Payment Vouchers",
            headers=["Date", "Reference", "Project/Grant", "Amount", "Description"],
            rows=rows,
        )
        if resp:
            return resp

    from tenant_finance.models import ChartAccount as CA

    payment_accounts = (
        CA.objects.using(tenant_db)
        .filter(type=CA.Type.ASSET)
        .order_by("code")
    )
    expense_accounts = (
        CA.objects.using(tenant_db)
        .filter(type=CA.Type.EXPENSE)
        .order_by("code")
    )

    if is_manager:
        grants = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")
    else:
        # Officer: only see actively assigned projects
        assigned_ids = GrantAssignment.objects.using(tenant_db).filter(
            officer=user, is_active=True
        ).values_list("grant_id", flat=True)
        grants = Grant.objects.using(tenant_db).filter(
            status=Grant.Status.ACTIVE, id__in=assigned_ids
        ).order_by("code")

    return render(
        request,
        "tenant_portal/pay/payment_vouchers.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "filters": f,
            "vouchers": vouchers,
            "payment_accounts": payment_accounts,
            "expense_accounts": expense_accounts,
            "grants": grants,
            "active_submenu": "payables",
            "active_item": "pay_payment_vouchers",
            "export_csv_url": _grants_export_urls(request)["csv"],
            "export_xlsx_url": _grants_export_urls(request)["xlsx"],
            "export_pdf_url": _grants_export_urls(request)["pdf"],
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_payment_voucher_detail_view(request: HttpRequest, entry_id: int) -> HttpResponse:
    """
    Printable view for a single payment voucher with tenant logo, title and voucher number.
    """
    from decimal import Decimal
    from django.shortcuts import get_object_or_404
    from tenant_finance.models import JournalEntry, JournalLine, ChartAccount

    tenant_db = request.tenant_db
    entry = get_object_or_404(JournalEntry.objects.using(tenant_db), pk=entry_id)

    # Basic safety: ensure this is a payment voucher
    reference = entry.reference or f"PV-{entry.id:05d}"
    if not reference.startswith("PV-"):
        messages.warning(request, "This journal entry is not tagged as a payment voucher.")

    lines = list(JournalLine.objects.using(tenant_db).select_related("account").filter(entry=entry))
    payment_line = next((l for l in lines if l.credit > 0 and l.account.type == ChartAccount.Type.ASSET), None)
    expense_line = next((l for l in lines if l.debit > 0 and l.account.type == ChartAccount.Type.EXPENSE), None)
    total = sum((l.debit - l.credit) for l in lines) if lines else Decimal("0")

    return render(
        request,
        "tenant_portal/pay/payment_voucher_detail.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "entry": entry,
            "reference": reference,
            "lines": lines,
            "payment_line": payment_line,
            "expense_line": expense_line,
            "amount": total,
            "active_submenu": "payables",
            "active_item": "pay_payment_vouchers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def pay_payment_voucher_approve_view(request: HttpRequest, entry_id: int) -> HttpResponse:
    """
    Single-voucher approval view for payment vouchers (PV-...).
    Allows an authorised approver to approve/post or return for correction
    with a mandatory comment for non-approval actions.
    """
    from decimal import Decimal
    from django.shortcuts import get_object_or_404
    from django.utils import timezone as _tz

    from tenant_finance.models import JournalEntry, JournalLine, ChartAccount, AuditLog

    tenant_db = request.tenant_db
    entry = get_object_or_404(JournalEntry.objects.using(tenant_db), pk=entry_id)

    reference = entry.reference or f"PV-{entry.id:05d}"
    if not reference.startswith("PV-"):
        messages.error(request, "This entry is not a payment voucher.")
        return redirect(reverse("tenant_portal:finance_journal_approval"))

    if request.method == "POST":
        from rbac.models import user_has_permission as _uhp
        cached = getattr(request, "rbac_permission_codes", None)

        def _has(code: str) -> bool:
            if isinstance(cached, set):
                return ("*" in cached) or (code in cached)
            return _uhp(request.tenant_user, code, using=tenant_db)

        action = (request.POST.get("action") or "").strip()
        comment = (request.POST.get("comment") or "").strip()

        if action in {"return", "reject"} and not comment:
            messages.error(request, "Rejection / correction comment is required.")
        else:
            old_status = entry.status

            if action == "approve":
                if not _has("finance:vouchers.approve"):
                    messages.error(request, "You do not have permission to approve vouchers.")
                    return redirect(reverse("tenant_portal:pay_payment_vouchers"))
                # Data-level access: grant-linked vouchers require assignment unless scope allows all
                if entry.grant_id and not _has("finance:scope.all_grants"):
                    try:
                        if not request.tenant_user.assigned_grants.using(tenant_db).filter(id=entry.grant_id).exists():
                            messages.error(request, "You do not have access to this grant/project.")
                            return redirect(reverse("tenant_portal:pay_payment_vouchers"))
                    except Exception:
                        messages.error(request, "You do not have access to this grant/project.")
                        return redirect(reverse("tenant_portal:pay_payment_vouchers"))
                # Maker-checker: maker cannot approve/post their own voucher unless override
                if (
                    entry.created_by_id
                    and getattr(request.tenant_user, "id", None)
                    and entry.created_by_id == request.tenant_user.id
                    and not _has("finance:journals.override_maker_checker")
                    and not _has("finance:vouchers.post")
                ):
                    messages.error(request, "Maker-checker is enforced: you cannot approve a voucher you created.")
                    return redirect(reverse("tenant_portal:pay_payment_voucher_approve", args=[entry.id]))

                # Post the voucher
                from tenant_finance.services.journal_posting import post_payment_voucher

                entry.status = JournalEntry.Status.POSTED
                entry.posted_at = _tz.now()
                entry.approved_by_id = getattr(request.tenant_user, "id", None)
                entry.posted_by_id = getattr(request.tenant_user, "id", None)
                entry.save(
                    update_fields=[
                        "status",
                        "posted_at",
                        "approved_by_id",
                        "posted_by_id",
                    ]
                )
                post_payment_voucher(using=tenant_db, entry=entry, user=request.tenant_user)

                try:
                    AuditLog.objects.using(tenant_db).create(
                        model_name="journalentry",
                        object_id=entry.id,
                        action=AuditLog.Action.UPDATE,
                        user_id=getattr(request.tenant_user, "id", None),
                        username=getattr(request.tenant_user, "full_name", "")
                        or getattr(request.tenant_user, "email", ""),
                        old_data={"status": old_status},
                        new_data={"status": entry.status},
                        summary="Payment voucher approved and posted.",
                    )
                except Exception:
                    pass

                messages.success(request, f"Payment voucher {reference} approved and posted.")
                return redirect(reverse("tenant_portal:pay_payment_vouchers"))

            elif action in {"return", "reject"}:
                if not _has("finance:vouchers.approve"):
                    messages.error(request, "You do not have permission to return/reject vouchers.")
                    return redirect(reverse("tenant_portal:pay_payment_vouchers"))
                # Return to draft for correction by maker
                entry.status = JournalEntry.Status.DRAFT
                entry.save(update_fields=["status"])

                try:
                    AuditLog.objects.using(tenant_db).create(
                        model_name="journalentry",
                        object_id=entry.id,
                        action=AuditLog.Action.UPDATE,
                        user_id=getattr(request.tenant_user, "id", None),
                        username=getattr(request.tenant_user, "full_name", "")
                        or getattr(request.tenant_user, "email", ""),
                        old_data={"status": old_status},
                        new_data={"status": entry.status},
                        summary=f"Payment voucher returned for correction. Comment: {comment}"[:255],
                    )
                except Exception:
                    pass

                messages.success(request, f"Payment voucher {reference} returned for correction.")
                return redirect(reverse("tenant_portal:pay_payment_vouchers"))

    # Build detail context (re-use detail layout semantics)
    lines = list(
        JournalLine.objects.using(tenant_db)
        .select_related("account")
        .filter(entry=entry)
    )
    payment_line = next(
        (l for l in lines if l.credit > 0 and l.account.type == ChartAccount.Type.ASSET),
        None,
    )
    expense_line = next(
        (l for l in lines if l.debit > 0 and l.account.type == ChartAccount.Type.EXPENSE),
        None,
    )
    total = sum((l.debit - l.credit) for l in lines) if lines else Decimal("0")

    return render(
        request,
        "tenant_portal/pay/payment_voucher_approve.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "entry": entry,
            "reference": reference,
            "lines": lines,
            "payment_line": payment_line,
            "expense_line": expense_line,
            "amount": total,
            "active_submenu": "payables",
            "active_item": "pay_payment_vouchers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def pay_payment_voucher_bulk_approval_view(request: HttpRequest) -> HttpResponse:
    """
    Bulk approval view for payment vouchers.

    Shows pending payment vouchers (status=PENDING_APPROVAL, PV-...) and allows
    authorised approvers to approve, return for correction, or reject in bulk.
    """
    from django.utils import timezone as _tz

    from tenant_finance.models import JournalEntry, AuditLog

    tenant_db = request.tenant_db
    from rbac.models import user_has_permission as _uhp
    cached = getattr(request, "rbac_permission_codes", None)

    def _has(code: str) -> bool:
        if isinstance(cached, set):
            return ("*" in cached) or (code in cached)
        return _uhp(request.tenant_user, code, using=tenant_db)

    if not _has("finance:vouchers.approve"):
        return render(
            request,
            "tenant_portal/forbidden.html",
            {"tenant": request.tenant, "tenant_user": request.tenant_user, "reason": "You do not have permission to approve vouchers."},
            status=403,
        )

    # Pending vouchers (maker has submitted, not yet posted)
    pending_qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(status=JournalEntry.Status.PENDING_APPROVAL)
        .filter(reference__startswith="PV-")
        .select_related("grant", "created_by")
        .order_by("entry_date", "id")
    )

    if request.method == "POST":
        action = (request.POST.get("bulk_action") or "").strip()
        selected_ids = request.POST.getlist("entry_ids")
        comment = (request.POST.get("bulk_comment") or "").strip()

        if not selected_ids:
            messages.error(request, "Please select at least one payment voucher.")
        elif action not in {"approve", "return", "reject"}:
            messages.error(request, "Please choose a bulk action.")
        elif action in {"return", "reject"} and not comment:
            messages.error(request, "Rejection / correction comment is required for bulk actions.")
        else:
            entries = list(
                pending_qs.filter(pk__in=selected_ids)
            )  # already filtered to pending PV

            if not entries:
                messages.error(request, "No matching pending vouchers were found for your selection.")
            else:
                for entry in entries:
                    old_status = entry.status
                    if action == "approve":
                        # Data-level access: grant-linked vouchers require assignment unless scope allows all
                        if entry.grant_id and not _has("finance:scope.all_grants"):
                            try:
                                if not request.tenant_user.assigned_grants.using(tenant_db).filter(id=entry.grant_id).exists():
                                    continue
                            except Exception:
                                continue
                        # Maker-checker: maker cannot approve/post their own voucher unless override
                        if (
                            entry.created_by_id
                            and getattr(request.tenant_user, "id", None)
                            and entry.created_by_id == request.tenant_user.id
                            and not _has("finance:journals.override_maker_checker")
                            and not _has("finance:vouchers.post")
                        ):
                            continue
                        from tenant_finance.services.journal_posting import post_payment_voucher

                        entry.status = JournalEntry.Status.POSTED
                        entry.posted_at = _tz.now()
                        entry.approved_by_id = getattr(request.tenant_user, "id", None)
                        entry.posted_by_id = getattr(request.tenant_user, "id", None)
                        entry.save(
                            update_fields=[
                                "status",
                                "posted_at",
                                "approved_by_id",
                                "posted_by_id",
                            ]
                        )
                        post_payment_voucher(using=tenant_db, entry=entry, user=request.tenant_user)
                        summary = "Payment voucher approved and posted (bulk)."
                    else:
                        # Return to draft for correction by maker
                        entry.status = JournalEntry.Status.DRAFT
                        entry.save(update_fields=["status"])
                        summary = f"Payment voucher returned for correction (bulk). Comment: {comment}"

                    try:
                        AuditLog.objects.using(tenant_db).create(
                            model_name="journalentry",
                            object_id=entry.id,
                            action=AuditLog.Action.UPDATE,
                            user_id=getattr(request.tenant_user, "id", None),
                            username=getattr(request.tenant_user, "full_name", "")
                            or getattr(request.tenant_user, "email", ""),
                            old_data={"status": old_status},
                            new_data={"status": entry.status},
                            summary=summary[:255],
                        )
                    except Exception:
                        pass

                if action == "approve":
                    messages.success(
                        request,
                        f"{len(entries)} payment voucher(s) approved and posted.",
                    )
                else:
                    messages.success(
                        request,
                        f"{len(entries)} payment voucher(s) returned for correction.",
                    )

                return redirect(reverse("tenant_portal:pay_payment_voucher_bulk_approval"))

    return render(
        request,
        "tenant_portal/pay/payment_voucher_bulk_approval.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "pending_vouchers": pending_qs,
            "active_submenu": "payables",
            "active_item": "pay_payment_vouchers",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_disbursement_list_view(request: HttpRequest) -> HttpResponse:
    """
    Disbursement form list: approved (posted) payment vouchers that are not yet paid.
    Includes KPI cards, filters (date, project, payment method), search, and table.
    """
    from decimal import Decimal
    from django.db.models import Sum, Count
    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine, BankAccount
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db
    f = _parse_finance_filters(request)
    search_q = (request.GET.get("search") or "").strip()
    payment_method_filter = (request.GET.get("payment_method") or "").strip()

    qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(
            reference__startswith="PV-",
            status=JournalEntry.Status.POSTED,
            payment_status=JournalEntry.PaymentStatus.UNPAID,
        )
        .filter(entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])
        .select_related("grant", "grant__donor", "grant__bank_account")
        .order_by("-entry_date", "-id")
    )
    if f.get("grant_id"):
        qs = qs.filter(grant_id=f["grant_id"])
    if payment_method_filter:
        qs = qs.filter(payment_method=payment_method_filter)
    if search_q:
        from django.db.models import Q
        qs = qs.filter(
            Q(reference__icontains=search_q)
            | Q(payee_name__icontains=search_q)
            | Q(memo__icontains=search_q)
        )

    # Build voucher list and compute totals for KPIs (from same base qs before search for KPIs)
    qs_for_kpis = (
        JournalEntry.objects.using(tenant_db)
        .filter(
            reference__startswith="PV-",
            status=JournalEntry.Status.POSTED,
            payment_status=JournalEntry.PaymentStatus.UNPAID,
        )
        .filter(entry_date__gte=f["period_start"], entry_date__lte=f["period_end"])
    )
    if f.get("grant_id"):
        qs_for_kpis = qs_for_kpis.filter(grant_id=f["grant_id"])
    if payment_method_filter:
        qs_for_kpis = qs_for_kpis.filter(payment_method=payment_method_filter)

    vouchers = []
    total_pending = Decimal("0")
    largest = Decimal("0")
    grant_ids = set()
    for je in qs[:200]:
        lines = list(
            JournalLine.objects.using(tenant_db)
            .select_related("account")
            .filter(entry=je)
        )
        total = sum((l.debit - l.credit) for l in lines) or Decimal("0")
        total_pending += total
        if total > largest:
            largest = total
        if je.grant_id:
            grant_ids.add(je.grant_id)
        payment_line = next(
            (l for l in lines if l.credit > 0 and l.account.type == ChartAccount.Type.ASSET),
            None,
        )
        bank_account = None
        if payment_line:
            bank_account = (
                BankAccount.objects.using(tenant_db)
                .filter(account=payment_line.account, is_active=True)
                .first()
            )
        bank_display = (
            f"{bank_account.bank_name} — {bank_account.account_number}"
            if bank_account
            else (payment_line.account.code + " — " + payment_line.account.name if payment_line else "—")
        )
        payee = getattr(je, "payee_name", None) or (je.memo or "")[:80]
        payment_method = getattr(je, "payment_method", None) or ""
        method_display = payment_method.replace("_", " ").title() if payment_method else "—"
        narration = (je.memo or "")[:120] if je.memo else ""
        donor_display = ""
        if je.grant and je.grant.donor:
            donor_display = je.grant.donor.name
        vouchers.append(
            {
                "id": je.id,
                "reference": je.reference or f"PV-{je.id:05d}",
                "entry_date": je.entry_date,
                "grant": je.grant,
                "payee": payee,
                "narration": narration,
                "amount": total,
                "bank_account": bank_display,
                "payment_method": method_display,
                "donor_display": donor_display,
            }
        )

    approved_count = qs_for_kpis.count()
    kpi_total_pending = Decimal("0")
    kpi_largest = Decimal("0")
    kpi_grant_ids = set()
    for je in qs_for_kpis[:500]:
        lines = list(
            JournalLine.objects.using(tenant_db).filter(entry=je).values_list("debit", "credit")
        )
        total = sum((d - c) for d, c in lines) or Decimal("0")
        kpi_total_pending += total
        if total > kpi_largest:
            kpi_largest = total
        if je.grant_id:
            kpi_grant_ids.add(je.grant_id)

    grants = Grant.objects.using(tenant_db).filter(status=Grant.Status.ACTIVE).order_by("code")
    try:
        selected_grant_id = int(f["grant_id"]) if f.get("grant_id") else None
    except (ValueError, TypeError):
        selected_grant_id = None

    return render(
        request,
        "tenant_portal/pay/disbursement_list.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "vouchers": vouchers,
            "grants": grants,
            "filters": f,
            "selected_grant_id": selected_grant_id,
            "search_q": search_q,
            "payment_method_filter": payment_method_filter,
            "kpi_approved_count": approved_count,
            "kpi_total_pending": kpi_total_pending,
            "kpi_largest": kpi_largest,
            "kpi_projects_count": len(kpi_grant_ids),
            "active_submenu": "payables",
            "active_item": "pay_disbursement_form",
        },
    )


DISBURSEMENT_TRANSACTION_TYPES = [
    ("payment_voucher", "Payment Voucher"),
    ("bank_transfer", "Bank Transfer"),
    ("cheque", "Cheque"),
    ("cash", "Cash"),
]


@tenant_view(require_module="finance_grants", require_perm="module:finance.manage")
def pay_disbursement_form_view(request: HttpRequest, entry_id: int) -> HttpResponse:
    """
    Disbursement form: open approved unpaid PV, prepare cheque/transfer details,
    set Post Transaction type, then Mark as Paid. When paid: post to GL (already posted),
    bank/vendor ledgers (via existing lines), payment register; record is locked.
    """
    from decimal import Decimal
    from django.shortcuts import get_object_or_404
    from django.utils import timezone

    from tenant_finance.models import (
        JournalEntry,
        JournalLine,
        ChartAccount,
        AuditLog,
        PaymentRegister,
    )
    from tenant_grants.models import Grant, BudgetLine

    tenant_db = request.tenant_db
    entry = get_object_or_404(JournalEntry.objects.using(tenant_db), pk=entry_id)

    reference = entry.reference or f"PV-{entry.id:05d}"
    if not reference.startswith("PV-"):
        messages.error(request, "This entry is not a payment voucher.")
        return redirect(reverse("tenant_portal:pay_disbursement_list"))

    if entry.status != JournalEntry.Status.POSTED:
        messages.warning(request, "Only approved (posted) payment vouchers can be disbursed.")
        return redirect(reverse("tenant_portal:pay_disbursement_list"))

    if entry.payment_status == JournalEntry.PaymentStatus.PAID:
        messages.info(request, f"Voucher {reference} is already paid (record locked).")

    lines = list(
        JournalLine.objects.using(tenant_db).select_related("account").filter(entry=entry)
    )
    payment_line = next(
        (l for l in lines if l.credit > 0 and l.account.type == ChartAccount.Type.ASSET),
        None,
    )
    expense_line = next(
        (l for l in lines if l.debit > 0 and l.account.type == ChartAccount.Type.EXPENSE),
        None,
    )
    total = sum((l.debit - l.credit) for l in lines) if lines else Decimal("0")

    grant = entry.grant
    budget_lines = []
    if grant:
        budget_lines = list(
            BudgetLine.objects.using(tenant_db)
            .filter(grant=grant)
            .select_related("account")
            .order_by("id")[:50]
        )
    payee_display = entry.payee_name or (entry.memo or "").strip() or "—"
    project_bank = grant.bank_account if grant else None

    has_project = grant is not None
    has_payee = bool(entry.payee_name or (entry.memo or "").strip())
    attachments = list(entry.attachments.all())
    has_attachments = len(attachments) > 0
    validation_ok = has_project and has_payee

    # Payment register record (when already paid)
    payment_register = (
        PaymentRegister.objects.using(tenant_db).filter(entry=entry).first()
    )

    # POST: Mark as Paid
    if request.method == "POST" and request.POST.get("action") == "mark_paid":
        if entry.payment_status == JournalEntry.PaymentStatus.PAID:
            messages.warning(request, "This voucher is already paid and locked.")
        elif not validation_ok:
            messages.error(
                request,
                "Cannot mark as paid: project and payee are required.",
            )
        else:
            paid_at = timezone.now()
            transaction_type = (request.POST.get("transaction_type") or "payment_voucher").strip()
            if transaction_type not in [t[0] for t in DISBURSEMENT_TRANSACTION_TYPES]:
                transaction_type = "payment_voucher"
            cheque_number = (request.POST.get("cheque_number") or "").strip()
            transfer_reference = (request.POST.get("transfer_reference") or "").strip()
            payment_method = (request.POST.get("payment_method") or "").strip()

            entry.payment_status = JournalEntry.PaymentStatus.PAID
            entry.paid_at = paid_at
            entry.save(update_fields=["payment_status", "paid_at"])

            PaymentRegister.objects.using(tenant_db).create(
                entry=entry,
                paid_at=paid_at,
                transaction_type=transaction_type,
                cheque_number=cheque_number,
                transfer_reference=transfer_reference,
                payment_method=payment_method,
                created_by=request.tenant_user,
            )

            try:
                AuditLog.objects.using(tenant_db).create(
                    model_name="journalentry",
                    object_id=entry.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=getattr(request.tenant_user, "id", None),
                    username=getattr(request.tenant_user, "full_name", "")
                    or getattr(request.tenant_user, "email", ""),
                    old_data={"payment_status": JournalEntry.PaymentStatus.UNPAID},
                    new_data={"payment_status": entry.payment_status, "paid_at": str(entry.paid_at)},
                    summary=f"Disbursement: {reference} marked as paid (transaction type: {transaction_type}).",
                )
            except Exception:
                pass
            messages.success(
                request,
                f"Voucher {reference} marked as paid. Payment recorded in payment register and the record is now locked.",
            )
            return redirect(reverse("tenant_portal:pay_disbursement_list"))

    return render(
        request,
        "tenant_portal/pay/disbursement_form.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "entry": entry,
            "reference": reference,
            "lines": lines,
            "payment_line": payment_line,
            "expense_line": expense_line,
            "amount": total,
            "grant": grant,
            "budget_lines": budget_lines,
            "payee_display": payee_display,
            "project_bank": project_bank,
            "attachments": attachments,
            "has_project": has_project,
            "has_payee": has_payee,
            "has_attachments": has_attachments,
            "validation_ok": validation_ok,
            "is_paid": entry.payment_status == JournalEntry.PaymentStatus.PAID,
            "payment_register": payment_register,
            "transaction_types": DISBURSEMENT_TRANSACTION_TYPES,
            "active_submenu": "payables",
            "active_item": "pay_disbursement_form",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_vendor_payments_view(request: HttpRequest) -> HttpResponse:
    """
    Vendor payments: list payment vouchers where payee_type = Vendor.
    """
    from decimal import Decimal, InvalidOperation
    from django.db.models import Sum
    from tenant_finance.models import ChartAccount, JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    # Base queryset: all PV- journal entries
    vouchers_qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="PV-")
        .select_related("grant")
        .order_by("-entry_date", "-id")
    )

    # Primary filters
    ref = (request.GET.get("ref") or "").strip()
    project = (request.GET.get("project") or "").strip()
    fund_name = (request.GET.get("fund") or "").strip()
    status_code = (request.GET.get("status") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None

    if ref:
        vouchers_qs = vouchers_qs.filter(reference__icontains=ref)
    if project:
        vouchers_qs = vouchers_qs.filter(grant__title__icontains=project)
    if fund_name:
        vouchers_qs = vouchers_qs.filter(grant__donor__name__icontains=fund_name)
    if status_code:
        vouchers_qs = vouchers_qs.filter(status=status_code)
    if from_date:
        vouchers_qs = vouchers_qs.filter(entry_date__gte=from_date)
    if to_date:
        vouchers_qs = vouchers_qs.filter(entry_date__lte=to_date)

    # Amount range (applied after total computation)
    min_amount = None
    max_amount = None
    raw_min = (request.GET.get("amount_from") or "").strip()
    raw_max = (request.GET.get("amount_to") or "").strip()
    try:
        if raw_min:
            min_amount = Decimal(raw_min.replace(",", ""))
    except InvalidOperation:
        min_amount = None
    try:
        if raw_max:
            max_amount = Decimal(raw_max.replace(",", ""))
    except InvalidOperation:
        max_amount = None

    rows = []
    for je in vouchers_qs[:200]:
        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        if min_amount is not None and total < min_amount:
            continue
        if max_amount is not None and total > max_amount:
            continue

        rows.append(
            {
                "id": je.id,
                "reference": je.reference or f"PV-{je.id:05d}",
                "date": je.entry_date,
                "project": je.grant.title if je.grant else "",
                "fund": getattr(je.grant.donor, "name", "") if je.grant and getattr(je.grant, "donor", None) else "",
                "amount": total,
                "status": je.get_status_display(),
                "memo": (je.memo or ""),
            }
        )

    return render(
        request,
        "tenant_portal/pay/vendor_payments.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": JournalEntry.Status.choices,
            "active_submenu": "payables",
            "active_item": "pay_vendor_payments",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_non_vendor_payments_view(request: HttpRequest) -> HttpResponse:
    """
    Non-vendor payments: list payment vouchers not related to vendor accounts.
    """
    from decimal import Decimal, InvalidOperation
    from django.db.models import Sum
    from tenant_finance.models import JournalEntry, JournalLine
    from tenant_grants.models import Grant

    tenant_db = request.tenant_db

    vouchers_qs = (
        JournalEntry.objects.using(tenant_db)
        .filter(reference__startswith="PV-")
        .select_related("grant")
        .order_by("-entry_date", "-id")
    )

    # Primary filters
    ref = (request.GET.get("ref") or "").strip()
    payee_name = (request.GET.get("payee") or "").strip()
    project = (request.GET.get("project") or "").strip()
    fund_name = (request.GET.get("fund") or "").strip()
    status_code = (request.GET.get("status") or "").strip()
    raw_from = (request.GET.get("from") or "").strip()
    raw_to = (request.GET.get("to") or "").strip()

    from_date = parse_date(raw_from) if raw_from else None
    to_date = parse_date(raw_to) if raw_to else None

    if ref:
        vouchers_qs = vouchers_qs.filter(reference__icontains=ref)
    if project:
        vouchers_qs = vouchers_qs.filter(grant__title__icontains=project)
    if fund_name:
        vouchers_qs = vouchers_qs.filter(grant__donor__name__icontains=fund_name)
    if status_code:
        vouchers_qs = vouchers_qs.filter(status=status_code)
    if from_date:
        vouchers_qs = vouchers_qs.filter(entry_date__gte=from_date)
    if to_date:
        vouchers_qs = vouchers_qs.filter(entry_date__lte=to_date)

    # NOTE: payee_name not yet stored on JournalEntry; reserved for future.

    # Amount range (applied after total computation)
    min_amount = None
    max_amount = None
    raw_min = (request.GET.get("amount_from") or "").strip()
    raw_max = (request.GET.get("amount_to") or "").strip()
    try:
        if raw_min:
            min_amount = Decimal(raw_min.replace(",", ""))
    except InvalidOperation:
        min_amount = None
    try:
        if raw_max:
            max_amount = Decimal(raw_max.replace(",", ""))
    except InvalidOperation:
        max_amount = None

    rows = []
    for je in vouchers_qs[:200]:
        total = (
            JournalLine.objects.using(tenant_db)
            .filter(entry=je)
            .aggregate(t=Sum("debit") - Sum("credit"))
            .get("t")
            or Decimal("0")
        )

        if min_amount is not None and total < min_amount:
            continue
        if max_amount is not None and total > max_amount:
            continue

        rows.append(
            {
                "id": je.id,
                "reference": je.reference or f"PV-{je.id:05d}",
                "date": je.entry_date,
                "project": je.grant.title if je.grant else "",
                "fund": getattr(je.grant.donor, "name", "") if je.grant and getattr(je.grant, "donor", None) else "",
                "amount": total,
                "status": je.get_status_display(),
                "memo": (je.memo or ""),
            }
        )

    return render(
        request,
        "tenant_portal/pay/non_vendor_payments.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "rows": rows,
            "status_choices": JournalEntry.Status.choices,
            "active_submenu": "payables",
            "active_item": "pay_non_vendor_payments",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_expense_register_view(request: HttpRequest) -> HttpResponse:
    """
    Expense register: reuse recent transactions (expense-focused) with filters and exports.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_recent_transactions") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_expense_allocation_view(request: HttpRequest) -> HttpResponse:
    """
    Expense allocation: reuse grant utilization / fund balances views which already
    track allocation across grants/projects.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_fund_balances") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_budget_checks_view(request: HttpRequest) -> HttpResponse:
    """
    Budget control checks: reuse budget vs actual analysis which enforces and reports
    on overspend and variances.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_budget_vs_actual") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_ledger_view(request: HttpRequest) -> HttpResponse:
    """
    Payables ledger: reuse account ledger focused on payable accounts.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_account_ledger") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def pay_outstanding_view(request: HttpRequest) -> HttpResponse:
    """
    Outstanding payables: reuse financial alerts which already show overdue / risk items.
    """
    from django.urls import reverse
    from django.http import HttpResponseRedirect

    return HttpResponseRedirect(reverse("tenant_portal:finance_financial_alerts") + ("?" + request.GET.urlencode() if request.GET else ""))


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_approvals_view(request: HttpRequest) -> HttpResponse:
    from django.utils import timezone
    from tenant_grants.models import Grant, GrantApproval

    tenant_db = request.tenant_db
    user = request.tenant_user

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "request":
            grant_id = request.POST.get("grant_id")
            note = (request.POST.get("note") or "").strip()
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            if not grant:
                messages.error(request, "Invalid grant.")
            else:
                GrantApproval.objects.using(tenant_db).create(grant=grant, requested_by=user, note=note)
                messages.success(request, "Approval requested.")
                return redirect(reverse("tenant_portal:grants_approvals"))

        if action in {"approve", "reject"}:
            if not user_has_permission(user, "module:grants.approve", using=tenant_db):
                return render(
                    request,
                    "tenant_portal/forbidden.html",
                    {"tenant": request.tenant, "tenant_user": user, "reason": "You do not have approval permission."},
                    status=403,
                )
            approval_id = request.POST.get("approval_id")
            approval = GrantApproval.objects.using(tenant_db).select_related("grant").filter(pk=approval_id).first()
            if not approval:
                messages.error(request, "Invalid approval request.")
            else:
                approval.status = GrantApproval.Status.APPROVED if action == "approve" else GrantApproval.Status.REJECTED
                approval.decided_by = user
                approval.decided_at = timezone.now()
                approval.save(using=tenant_db)
                messages.success(request, f"Request {approval.status}.")
                return redirect(reverse("tenant_portal:grants_approvals"))

    approvals = GrantApproval.objects.using(tenant_db).select_related("grant", "requested_by", "decided_by").order_by("-created_at")[:100]
    grants = Grant.objects.using(tenant_db).order_by("-created_at")[:100]
    can_decide = user_has_permission(user, "module:grants.approve", using=tenant_db)
    return render(
        request,
        "tenant_portal/grants/approvals.html",
        {
            "tenant": request.tenant,
            "tenant_user": user,
            "approvals": approvals,
            "grants": grants,
            "can_decide": can_decide,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:grants.view")
def grants_reports_view(request: HttpRequest) -> HttpResponse:
    from django.db.models import Sum
    from tenant_grants.models import BudgetLine, Donor, Grant
    from tenant_finance.models import JournalLine

    tenant_db = request.tenant_db

    start = request.GET.get("start") or ""
    end = request.GET.get("end") or ""
    grant_filter = (request.GET.get("grant_id") or "").strip()
    donor_filter = (request.GET.get("donor_id") or "").strip()

    grants_master = list(Grant.objects.using(tenant_db).select_related("donor").order_by("-created_at")[:200])
    grants = grants_master
    if grant_filter.isdigit():
        grants = [g for g in grants_master if str(g.pk) == grant_filter]
    elif donor_filter.isdigit():
        grants = [g for g in grants_master if g.donor_id and str(g.donor_id) == donor_filter]

    budget_by_grant = {
        row["grant_id"]: row["total"] or 0
        for row in BudgetLine.objects.using(tenant_db).values("grant_id").annotate(total=Sum("amount"))
    }

    entry_filter = {"entry__grant_id__isnull": False, "account__type": "expense"}
    if start:
        entry_filter["entry__entry_date__gte"] = start
    if end:
        entry_filter["entry__entry_date__lte"] = end

    line_filter = dict(entry_filter)
    if grant_filter.isdigit():
        line_filter["entry__grant_id"] = int(grant_filter)
    elif donor_filter.isdigit():
        line_filter["entry__grant__donor_id"] = int(donor_filter)

    spend_by_grant = {
        row["entry__grant_id"]: row["spent"] or 0
        for row in JournalLine.objects.using(tenant_db).filter(**entry_filter).values("entry__grant_id").annotate(spent=Sum("debit"))
    }

    # Breakdown: spend by expense account (respects date + optional grant/donor scope)
    spend_by_account = list(
        JournalLine.objects.using(tenant_db)
        .filter(**line_filter)
        .values("account__code", "account__name")
        .annotate(spent=Sum("debit"))
        .order_by("-spent")[:20]
    )

    bl_scope = BudgetLine.objects.using(tenant_db)
    if grant_filter.isdigit():
        bl_scope = bl_scope.filter(grant_id=int(grant_filter))
    elif donor_filter.isdigit():
        bl_scope = bl_scope.filter(grant__donor_id=int(donor_filter))
    budget_by_category = list(
        bl_scope.values("category").annotate(total=Sum("amount")).order_by("-total")[:20]
    )

    rows = []
    for g in grants:
        budget = budget_by_grant.get(g.id, 0)
        spent = spend_by_grant.get(g.id, 0)
        remaining = (g.award_amount or 0) - spent
        rows.append(
            {
                "grant": g,
                "budget": budget,
                "spent": spent,
                "remaining": remaining,
            }
        )

    if request.GET.get("format") == "csv":
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="grant_financial_overview.csv"'
        w = csv.writer(response)
        if start or end:
            _period = f"{start or '—'} to {end or '—'}"
        else:
            _period = "All dates (spend breakdowns may still reflect date filters when applied)"
        _official_csv_preamble(
            w,
            request,
            "Grant financial overview",
            [
                ("Activity period", _period),
                ("Grant filter", grant_filter or "All grants"),
                ("Donor filter", donor_filter or "All donors"),
            ],
        )
        w.writerow(
            ["Grant code", "Grant title", "Donor", "Award", "Budget", "Spent", "Remaining", "Status"]
        )
        for r in rows:
            g = r["grant"]
            w.writerow(
                [
                    g.code,
                    g.title,
                    g.donor.name if g.donor_id else "",
                    str(g.award_amount or 0),
                    str(r["budget"]),
                    str(r["spent"]),
                    str(r["remaining"]),
                    g.status,
                ]
            )
        return response

    q = request.GET.copy()
    q["format"] = "csv"
    export_csv_url = request.path + "?" + q.urlencode()

    period_line = ""
    if start or end:
        period_line = f"Activity period: {start or '—'} to {end or '—'}"

    grants_select = (
        Grant.objects.using(tenant_db)
        .select_related("donor")
        .filter(status__in=[Grant.Status.ACTIVE, Grant.Status.CLOSED])
        .order_by("code")[:400]
    )
    donors_select = Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")

    return render(
        request,
        "tenant_portal/grants/reports.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "reports",
            "active_item": "report_grant",
            "rows": rows,
            "start": start,
            "end": end,
            "grant_id": grant_filter,
            "donor_id": donor_filter,
            "grants_select": grants_select,
            "donors_select": donors_select,
            "spend_by_account": spend_by_account,
            "budget_by_category": budget_by_category,
            "export_csv_url": export_csv_url,
            "official_report_period_line": period_line,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def core_accounting_center_view(request: HttpRequest) -> HttpResponse:
    """Core Accounting module home: dashboard for chart, journals, ledgers, statements, and controls."""
    from django.db.models import Max
    from django.utils import timezone

    from tenant_finance.models import ChartAccount, FiscalPeriod, JournalEntry

    tenant_db = request.tenant_db
    kpi_accounts = 0
    kpi_open_periods = 0
    kpi_drafts = 0
    kpi_pending = 0
    kpi_posted_month = 0
    last_posting_at = None
    try:
        kpi_accounts = ChartAccount.objects.using(tenant_db).filter(is_active=True).count()
        kpi_open_periods = FiscalPeriod.objects.using(tenant_db).filter(
            is_closed=False, status=FiscalPeriod.Status.OPEN
        ).count()
        kpi_drafts = JournalEntry.objects.using(tenant_db).filter(status=JournalEntry.Status.DRAFT).count()
        kpi_pending = JournalEntry.objects.using(tenant_db).filter(
            status=JournalEntry.Status.PENDING_APPROVAL
        ).count()
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        kpi_posted_month = JournalEntry.objects.using(tenant_db).filter(
            status=JournalEntry.Status.POSTED,
            posted_at__gte=month_start,
        ).count()
        last_posting_at = (
            JournalEntry.objects.using(tenant_db)
            .filter(status=JournalEntry.Status.POSTED, posted_at__isnull=False)
            .aggregate(m=Max("posted_at"))
            .get("m")
        )
    except Exception:
        pass

    core_kpis = [
        {"label": "Active accounts", "value": kpi_accounts},
        {"label": "Open periods", "value": kpi_open_periods},
        {"label": "Draft journals", "value": kpi_drafts},
        {"label": "Pending approval", "value": kpi_pending},
        {"label": "Posted this month", "value": kpi_posted_month},
        {
            "label": "Last posting",
            "value": last_posting_at.strftime("%Y-%m-%d %H:%M") if last_posting_at else "—",
        },
    ]

    return render(
        request,
        "tenant_portal/finance/core_accounting_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "core",
            "active_item": "core_center_home",
            "core_kpis": core_kpis,
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def reporting_center_view(request: HttpRequest) -> HttpResponse:
    """Reporting module home: dashboard of links to financial, budget, grant, and donor reports."""
    return render(
        request,
        "tenant_portal/reporting/reporting_center.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "reports",
            "active_item": "report_center_home",
        },
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def reporting_export_tools_view(request: HttpRequest) -> HttpResponse:
    return render(
        request,
        "tenant_portal/grants/reporting_export_tools.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "active_submenu": "reports",
            "active_item": "report_export_tools",
        },
    )


@tenant_view(require_module="integrations", require_perm="module:integrations.manage")
def integrations_home_view(request: HttpRequest) -> HttpResponse:
    from tenant_integrations.models import OutboundWebhook, ErpConnection

    tenant_db = request.tenant_db
    return render(
        request,
        "tenant_portal/integrations/home.html",
        {
            "tenant": request.tenant,
            "tenant_user": request.tenant_user,
            "webhooks_count": OutboundWebhook.objects.using(tenant_db).count(),
            "erp_count": ErpConnection.objects.using(tenant_db).count(),
        },
    )


@tenant_view(require_module="integrations", require_perm="module:integrations.manage")
def integrations_webhooks_view(request: HttpRequest) -> HttpResponse:
    from tenant_integrations.models import OutboundWebhook

    tenant_db = request.tenant_db
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        url = (request.POST.get("url") or "").strip()
        secret = (request.POST.get("secret") or "").strip()
        if not name or not url:
            messages.error(request, "Name and URL are required.")
        else:
            wh = OutboundWebhook.objects.using(tenant_db).create(name=name, url=url, is_active=True)
            if secret:
                wh.set_secret(secret)
                wh.save(using=tenant_db)
            messages.success(request, "Webhook created.")
            return redirect(reverse("tenant_portal:integrations_webhooks"))

    webhooks = OutboundWebhook.objects.using(tenant_db).order_by("-created_at")[:100]
    return render(
        request,
        "tenant_portal/integrations/webhooks.html",
        {"tenant": request.tenant, "tenant_user": request.tenant_user, "webhooks": webhooks},
    )


@tenant_view(require_module="integrations", require_perm="module:integrations.manage")
def integrations_erp_view(request: HttpRequest) -> HttpResponse:
    from tenant_integrations.models import ErpConnection

    tenant_db = request.tenant_db
    if request.method == "POST":
        provider = (request.POST.get("provider") or "").strip() or ErpConnection.Provider.GENERIC
        name = (request.POST.get("name") or "").strip()
        base_url = (request.POST.get("base_url") or "").strip()
        api_key = (request.POST.get("api_key") or "").strip()
        if not name:
            messages.error(request, "Connection name is required.")
        else:
            conn = ErpConnection.objects.using(tenant_db).create(provider=provider, name=name, base_url=base_url, is_active=True)
            if api_key:
                conn.set_api_key(api_key)
                conn.save(using=tenant_db)
            messages.success(request, "ERP connection created.")
            return redirect(reverse("tenant_portal:integrations_erp"))

    conns = ErpConnection.objects.using(tenant_db).order_by("-created_at")[:100]
    return render(
        request,
        "tenant_portal/integrations/erp.html",
        {"tenant": request.tenant, "tenant_user": request.tenant_user, "conns": conns, "providers": ErpConnection.Provider},
    )


# Financial Setup views (configuration center)
from tenant_portal.views_setup import *  # noqa: F401, F403
