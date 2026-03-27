"""
Financial Setup views: configuration center for dimensions, currencies, numbering, etc.
Access: module:finance.view for read; module:finance.manage for add/edit/delete.
"""
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import connections
from django.db.models import Q
from django.http import HttpRequest, HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from rbac.models import user_has_permission

from tenant_portal.decorators import tenant_view


# region agent log helper
def _agent_debug_log(hypothesis_id: str, location: str, message: str, data: dict) -> None:
    """
    Lightweight NDJSON logger for debug session 9554db.
    Writes to debug-9554db.log without impacting main flow.
    """
    try:
        import json
        import time

        payload = {
            "sessionId": "9554db",
            "id": f"log_{int(time.time() * 1000)}",
            "timestamp": int(time.time() * 1000),
            "runId": "pre-fix",
            "hypothesisId": hypothesis_id,
            "location": location,
            "message": message,
            "data": data,
        }
        with open("debug-9554db.log", "a", encoding="utf-8") as f:
            f.write(json.dumps(payload, default=str) + "\n")
    except Exception:
        # Never let logging interfere with request handling.
        pass


# endregion

PAGE_SIZE = 15


def _ensure_documentseries_schema(tenant_db: str) -> None:
    """
    Ensure tenant_finance_documentseries in the given tenant DB has
    the columns expected by the current DocumentSeries model.
    """
    try:
        from tenant_finance.models import DocumentSeries

        conn = connections[tenant_db]
        table = DocumentSeries._meta.db_table

        with conn.cursor() as cursor:
            # Align extra columns introduced in later migrations.
            cursor.execute(
                f"""
                ALTER TABLE {table}
                ADD COLUMN IF NOT EXISTS number_format varchar(80) DEFAULT %s,
                ADD COLUMN IF NOT EXISTS reset_frequency varchar(20) DEFAULT %s,
                ADD COLUMN IF NOT EXISTS status varchar(20) DEFAULT %s,
                ADD COLUMN IF NOT EXISTS notes text;
                """,
                ["{prefix}{year}-{seq:05d}", "yearly", "active"],
            )

            # Backward-compatibility: some older tenant DBs have a NOT NULL year_format
            # column without a default, which breaks inserts from the current model
            # (year_format field was removed). Ensure a safe default and fix nulls.
            cursor.execute(
                f"""
                DO $$
                BEGIN
                    IF EXISTS (
                        SELECT 1
                        FROM information_schema.columns
                        WHERE table_name = %s
                          AND column_name = 'year_format'
                    ) THEN
                        BEGIN
                            EXECUTE format(
                                'ALTER TABLE %I ALTER COLUMN year_format SET DEFAULT %L',
                                %s,
                                'YYYY'
                            );
                        EXCEPTION
                            WHEN undefined_column THEN
                                -- Column disappeared between checks; ignore.
                                NULL;
                        END;

                        BEGIN
                            EXECUTE format(
                                'UPDATE %I SET year_format = %L WHERE year_format IS NULL',
                                %s,
                                'YYYY'
                            );
                        EXCEPTION
                            WHEN undefined_column THEN
                                NULL;
                        END;
                    END IF;
                END
                $$;
                """,
                [table, table, table],
            )

        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_documentseries_schema",
            message="Ensured DocumentSeries schema for tenant DB",
            data={"tenant_db": tenant_db, "table": table},
        )
    except Exception as exc:  # pragma: no cover - defensive
        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_documentseries_schema",
            message="Failed to ensure DocumentSeries schema",
            data={"tenant_db": tenant_db, "error": str(exc)},
        )


def _ensure_grantcompliancerule_schema(tenant_db: str) -> None:
    """
    Ensure tenant_finance_grantcompliancerule and its M2M tables exist for the
    current GrantComplianceRule model. Some environments applied older
    migrations that did not create this table, so we create it defensively.
    """
    try:
        from tenant_finance.models import GrantComplianceRule, AccountCategory

        conn = connections[tenant_db]
        rule_table = GrantComplianceRule._meta.db_table
        allowed_m2m = (
            GrantComplianceRule._meta.get_field("allowed_account_categories")
            .remote_field.through._meta.db_table
        )
        disallowed_m2m = (
            GrantComplianceRule._meta.get_field("disallowed_account_categories")
            .remote_field.through._meta.db_table
        )
        account_table = AccountCategory._meta.db_table

        with conn.cursor() as cursor:
            # Main GrantComplianceRule table
            cursor.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {rule_table} (
                    id BIGSERIAL PRIMARY KEY,
                    name varchar(160) NOT NULL,
                    donor_id bigint NULL,
                    grant_id bigint NULL,
                    project_id bigint NULL,
                    effective_from date NOT NULL,
                    effective_to date NOT NULL,
                    maximum_admin_cost_percent numeric(5, 2) NULL,
                    require_attachments boolean NOT NULL DEFAULT FALSE,
                    require_procurement_compliance boolean NOT NULL DEFAULT FALSE,
                    require_budget_check boolean NOT NULL DEFAULT TRUE,
                    allow_posting_outside_grant_period boolean NOT NULL DEFAULT FALSE,
                    require_additional_approval boolean NOT NULL DEFAULT FALSE,
                    additional_approval_role varchar(120) NOT NULL DEFAULT '',
                    mode varchar(10) NOT NULL DEFAULT 'block',
                    status varchar(20) NOT NULL DEFAULT 'active',
                    notes text NULL,
                    created_at timestamp with time zone DEFAULT now(),
                    updated_at timestamp with time zone DEFAULT now()
                );
                """
            )

            # Allowed account categories M2M table
            cursor.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {allowed_m2m} (
                    id BIGSERIAL PRIMARY KEY,
                    grantcompliancerule_id bigint NOT NULL
                        REFERENCES {rule_table}(id) DEFERRABLE INITIALLY DEFERRED,
                    accountcategory_id bigint NOT NULL
                        REFERENCES {account_table}(id) DEFERRABLE INITIALLY DEFERRED
                );
                """
            )

            # Disallowed account categories M2M table
            cursor.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {disallowed_m2m} (
                    id BIGSERIAL PRIMARY KEY,
                    grantcompliancerule_id bigint NOT NULL
                        REFERENCES {rule_table}(id) DEFERRABLE INITIALLY DEFERRED,
                    accountcategory_id bigint NOT NULL
                        REFERENCES {account_table}(id) DEFERRABLE INITIALLY DEFERRED
                );
                """
            )

        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_grantcompliancerule_schema",
            message="Ensured GrantComplianceRule schema for tenant DB",
            data={
                "tenant_db": tenant_db,
                "rule_table": rule_table,
                "allowed_m2m": allowed_m2m,
                "disallowed_m2m": disallowed_m2m,
            },
        )
    except Exception as exc:  # pragma: no cover - defensive
        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_grantcompliancerule_schema",
            message="Failed to ensure GrantComplianceRule schema",
            data={"tenant_db": tenant_db, "error": str(exc)},
        )


def _ensure_interfundtransferrule_schema(tenant_db: str) -> None:
    """
    Ensure tenant_finance_interfundtransferrule exists for tenants that were
    created before the InterFundTransferRule model was introduced.
    """
    try:
        from tenant_finance.models import InterFundTransferRule, ChartAccount

        conn = connections[tenant_db]
        rule_table = InterFundTransferRule._meta.db_table
        account_table = ChartAccount._meta.db_table

        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {rule_table} (
                    id BIGSERIAL PRIMARY KEY,
                    name varchar(160) NOT NULL,
                    from_fund_type varchar(40) NOT NULL,
                    to_fund_type varchar(40) NOT NULL,
                    specific_from_fund_code varchar(80) NULL,
                    specific_to_fund_code varchar(80) NULL,
                    allow_transfer boolean NOT NULL DEFAULT TRUE,
                    require_approval boolean NOT NULL DEFAULT TRUE,
                    approval_role varchar(120) NULL,
                    require_reason boolean NOT NULL DEFAULT TRUE,
                    maximum_transfer_amount numeric(18,2) NULL,
                    transfer_account_id bigint NOT NULL REFERENCES {account_table}(id)
                        DEFERRABLE INITIALLY DEFERRED,
                    effective_from date NOT NULL,
                    effective_to date NOT NULL,
                    status varchar(20) NOT NULL DEFAULT 'active',
                    notes text NULL,
                    created_at timestamp with time zone DEFAULT now(),
                    updated_at timestamp with time zone DEFAULT now()
                );
                """
            )

        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_interfundtransferrule_schema",
            message="Ensured InterFundTransferRule schema for tenant DB",
            data={
                "tenant_db": tenant_db,
                "rule_table": rule_table,
            },
        )
    except Exception as exc:  # pragma: no cover - defensive
        _agent_debug_log(
            hypothesis_id="H_schema_autofix",
            location="tenant_portal/views_setup.py:_ensure_interfundtransferrule_schema",
            message="Failed to ensure InterFundTransferRule schema",
            data={"tenant_db": tenant_db, "error": str(exc)},
        )


def _setup_context(request: HttpRequest) -> dict:
    tenant_db = getattr(request, "tenant_db", None)
    can_manage = bool(
        tenant_db
        and getattr(request, "tenant_user", None)
        and user_has_permission(request.tenant_user, "module:finance.manage", using=tenant_db)
    )
    return {
        "tenant": request.tenant,
        "tenant_user": getattr(request, "tenant_user", None),
        "can_manage": can_manage,
        "active_submenu": "setup",
    }


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_home_view(request: HttpRequest) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_home"
    return render(request, "tenant_portal/setup/home.html", ctx)


# ----- Dimensions & Cost Centers -----
def _dimensions_queryset(tenant_db: str, request: HttpRequest):
    from tenant_finance.models import FinancialDimension

    qs = FinancialDimension.objects.using(tenant_db).all().order_by("dimension_code")
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(dimension_code__icontains=q)
            | Q(dimension_name__icontains=q)
            | Q(description__icontains=q)
        )
    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)
    else:
        # By default, only active dimensions should appear in lists/dropdowns
        qs = qs.filter(status=FinancialDimension.Status.ACTIVE)
    dim_type = (request.GET.get("type") or "").strip()
    if dim_type:
        qs = qs.filter(dimension_type=dim_type)
    return qs


def _cost_centers_queryset(tenant_db: str, request: HttpRequest):
    from tenant_finance.models import CostCenter

    qs = CostCenter.objects.using(tenant_db).select_related("parent", "manager").order_by("code")
    q_cc = (request.GET.get("q_cc") or "").strip()
    if q_cc:
        qs = qs.filter(
            Q(code__icontains=q_cc) | Q(name__icontains=q_cc) | Q(description__icontains=q_cc)
        )
    status_cc = (request.GET.get("status_cc") or "").strip()
    if status_cc:
        qs = qs.filter(status=status_cc)
    else:
        # By default, only active cost centers should appear
        qs = qs.filter(status=CostCenter.Status.ACTIVE)
    return qs


def _program_dimension_values_queryset(tenant_db: str, request: HttpRequest):
    from tenant_finance.models import FinancialDimensionValue

    qs = (
        FinancialDimensionValue.objects.using(tenant_db)
        .select_related("dimension")
        .filter(dimension__dimension_code="PROG")
        .order_by("code")
    )
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(description__icontains=q))
    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)
    return qs


def _sector_dimension_values_queryset(tenant_db: str, request: HttpRequest):
    from tenant_finance.models import FinancialDimensionValue

    qs = (
        FinancialDimensionValue.objects.using(tenant_db)
        .select_related("dimension")
        .filter(dimension__dimension_code="SECTOR")
        .order_by("code")
    )
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q) | Q(description__icontains=q))
    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)
    return qs


def _ensure_program_dimension_defaults(tenant_db: str, tenant_user):
    from tenant_finance.models import FinancialDimension, FinancialDimensionValue

    prog_dim, _ = FinancialDimension.objects.using(tenant_db).get_or_create(
        dimension_code="PROG",
        defaults={
            "dimension_name": "Program",
            "dimension_type": FinancialDimension.DimensionType.PROGRAM,
            "description": "Program dimension for funding/program categories.",
            "status": FinancialDimension.Status.ACTIVE,
            "created_by": tenant_user,
        },
    )
    if prog_dim.status != FinancialDimension.Status.ACTIVE:
        prog_dim.status = FinancialDimension.Status.ACTIVE
        prog_dim.save(using=tenant_db, update_fields=["status"])

    defaults = [
        ("PRG-01", "Project grant"),
        ("PRG-02", "Core / institutional"),
        ("PRG-03", "Emergency"),
        ("PRG-04", "Institutional"),
        ("PRG-05", "Other"),
    ]
    for code, name in defaults:
        obj, _ = FinancialDimensionValue.objects.using(tenant_db).get_or_create(
            dimension=prog_dim,
            code=code,
            defaults={
                "name": name,
                "description": "",
                "status": FinancialDimensionValue.Status.ACTIVE,
                "created_by": tenant_user,
            },
        )
        changed = False
        if obj.name != name:
            obj.name = name
            changed = True
        if obj.status != FinancialDimensionValue.Status.ACTIVE:
            obj.status = FinancialDimensionValue.Status.ACTIVE
            changed = True
        if changed:
            obj.save(using=tenant_db, update_fields=["name", "status"])

    legacy_code_map = {
        "PROJECT_GRANT": ("PRG-01", "Project grant"),
        "CORE_INSTITUTIONAL": ("PRG-02", "Core / institutional"),
        "EMERGENCY": ("PRG-03", "Emergency"),
        "INSTITUTIONAL": ("PRG-04", "Institutional"),
        "OTHER": ("PRG-05", "Other"),
    }
    for legacy_code, (new_code, new_name) in legacy_code_map.items():
        legacy = (
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension=prog_dim, code=legacy_code)
            .first()
        )
        if not legacy:
            continue
        target = (
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension=prog_dim, code=new_code)
            .first()
        )
        if target and target.pk != legacy.pk:
            continue
        legacy.code = new_code
        legacy.name = new_name
        legacy.status = FinancialDimensionValue.Status.ACTIVE
        legacy.save(using=tenant_db, update_fields=["code", "name", "status"])
    return prog_dim


def _ensure_sector_dimension_defaults(tenant_db: str, tenant_user):
    from tenant_finance.models import FinancialDimension, FinancialDimensionValue

    sector_dim, _ = FinancialDimension.objects.using(tenant_db).get_or_create(
        dimension_code="SECTOR",
        defaults={
            "dimension_name": "Program sector",
            "dimension_type": FinancialDimension.DimensionType.CLASSIFICATION,
            "description": "Program sector classification values.",
            "status": FinancialDimension.Status.ACTIVE,
            "created_by": tenant_user,
        },
    )
    if sector_dim.status != FinancialDimension.Status.ACTIVE:
        sector_dim.status = FinancialDimension.Status.ACTIVE
        sector_dim.save(using=tenant_db, update_fields=["status"])

    defaults = [
        ("SEC-01", "Health"),
        ("SEC-02", "WASH"),
        ("SEC-03", "Education"),
        ("SEC-04", "Protection"),
        ("SEC-05", "Nutrition"),
        ("SEC-06", "Livelihood"),
        ("SEC-07", "Food Security"),
        ("SEC-08", "Shelter"),
        ("SEC-09", "GBV"),
        ("SEC-10", "Child Protection"),
        ("SEC-11", "Governance"),
        ("SEC-12", "Capacity building"),
        ("SEC-13", "Multi-sector"),
        ("SEC-14", "Other"),
    ]
    for code, name in defaults:
        obj, _ = FinancialDimensionValue.objects.using(tenant_db).get_or_create(
            dimension=sector_dim,
            code=code,
            defaults={
                "name": name,
                "description": "",
                "status": FinancialDimensionValue.Status.ACTIVE,
                "created_by": tenant_user,
            },
        )
        changed = False
        if obj.name != name:
            obj.name = name
            changed = True
        if obj.status != FinancialDimensionValue.Status.ACTIVE:
            obj.status = FinancialDimensionValue.Status.ACTIVE
            changed = True
        if changed:
            obj.save(using=tenant_db, update_fields=["name", "status"])
    return sector_dim


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_dimensions_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"

    dimensions_qs = _dimensions_queryset(tenant_db, request)
    paginator_dim = Paginator(dimensions_qs, PAGE_SIZE)
    page_dim = request.GET.get("page_dim", "1")
    ctx["dimensions_page"] = paginator_dim.get_page(page_dim)
    ctx["dimensions_filter_q"] = request.GET.get("q", "")
    ctx["dimensions_filter_status"] = request.GET.get("status", "")
    ctx["dimensions_filter_type"] = request.GET.get("type", "")

    # Query strings for pagination (preserve other params)
    get = request.GET.copy()
    if "page_dim" in get:
        get.pop("page_dim")
    ctx["dimensions_base_query"] = get.urlencode()

    from tenant_finance.models import FinancialDimension

    ctx["dimension_types"] = FinancialDimension.DimensionType.choices
    return render(request, "tenant_portal/setup/dimensions_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_program_dimension_values_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"

    from tenant_finance.models import FinancialDimensionValue

    prog_dim = _ensure_program_dimension_defaults(tenant_db, request.tenant_user)

    values_qs = _program_dimension_values_queryset(tenant_db, request)
    paginator = Paginator(values_qs, PAGE_SIZE)
    page = request.GET.get("page", "1")
    ctx["values_page"] = paginator.get_page(page)
    ctx["filter_q"] = request.GET.get("q", "")
    ctx["filter_status"] = request.GET.get("status", "")
    ctx["prog_dimension"] = prog_dim
    ctx["status_choices"] = FinancialDimensionValue.Status.choices

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()
    return render(request, "tenant_portal/setup/program_dimension_values_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_program_dimension_values_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to create program values.")
        return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    prog_dim = _ensure_program_dimension_defaults(tenant_db, request.tenant_user)

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimensionValue.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Code is required.")
        if not name:
            errors.append("Name is required.")
        if code and FinancialDimensionValue.objects.using(tenant_db).filter(dimension=prog_dim, code__iexact=code).exists():
            errors.append("A program value with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            FinancialDimensionValue.objects.using(tenant_db).create(
                dimension=prog_dim,
                code=code,
                name=name,
                description=description,
                status=status,
                created_by=request.tenant_user,
            )
            messages.success(request, "Program value created.")
            return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    ctx["prog_dimension"] = prog_dim
    ctx["status_choices"] = FinancialDimensionValue.Status.choices
    ctx["form_title"] = "Add Program Value"
    return render(request, "tenant_portal/setup/program_dimension_values_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_program_dimension_values_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit program values.")
        return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    obj = get_object_or_404(
        FinancialDimensionValue.objects.using(tenant_db).select_related("dimension"),
        pk=pk,
        dimension__dimension_code="PROG",
    )

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimensionValue.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Code is required.")
        if not name:
            errors.append("Name is required.")
        if (
            code
            and FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension=obj.dimension, code__iexact=code)
            .exclude(pk=pk)
            .exists()
        ):
            errors.append("A program value with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj.code = code
            obj.name = name
            obj.description = description
            obj.status = status
            obj.save(using=tenant_db)
            messages.success(request, "Program value updated.")
            return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    ctx["value"] = obj
    ctx["prog_dimension"] = obj.dimension
    ctx["status_choices"] = FinancialDimensionValue.Status.choices
    ctx["form_title"] = "Edit Program Value"
    return render(request, "tenant_portal/setup/program_dimension_values_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_program_dimension_values_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete program values.")
        return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    obj = get_object_or_404(
        FinancialDimensionValue.objects.using(tenant_db).select_related("dimension"),
        pk=pk,
        dimension__dimension_code="PROG",
    )
    if request.method == "POST":
        obj.delete(using=tenant_db)
        messages.success(request, "Program value deleted.")
        return redirect(reverse("tenant_portal:setup_program_dimension_values_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Program value {obj.code} — {obj.name}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_program_dimension_values_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_program_dimension_values_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_sector_dimension_values_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"

    from tenant_finance.models import FinancialDimensionValue

    sector_dim = _ensure_sector_dimension_defaults(tenant_db, request.tenant_user)

    values_qs = _sector_dimension_values_queryset(tenant_db, request)
    paginator = Paginator(values_qs, PAGE_SIZE)
    page = request.GET.get("page", "1")
    ctx["values_page"] = paginator.get_page(page)
    ctx["filter_q"] = request.GET.get("q", "")
    ctx["filter_status"] = request.GET.get("status", "")
    ctx["dimension_label"] = "Program sector"
    ctx["dimension_code"] = sector_dim.dimension_code
    ctx["status_choices"] = FinancialDimensionValue.Status.choices
    ctx["list_url_name"] = "tenant_portal:setup_sector_dimension_values_list"
    ctx["add_url_name"] = "tenant_portal:setup_sector_dimension_values_add"
    ctx["edit_url_name"] = "tenant_portal:setup_sector_dimension_values_edit"
    ctx["delete_url_name"] = "tenant_portal:setup_sector_dimension_values_delete"

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()
    return render(request, "tenant_portal/setup/dimension_values_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_sector_dimension_values_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to create sector values.")
        return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    sector_dim = _ensure_sector_dimension_defaults(tenant_db, request.tenant_user)

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimensionValue.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Code is required.")
        if not name:
            errors.append("Name is required.")
        if code and FinancialDimensionValue.objects.using(tenant_db).filter(dimension=sector_dim, code__iexact=code).exists():
            errors.append("A sector value with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            FinancialDimensionValue.objects.using(tenant_db).create(
                dimension=sector_dim,
                code=code,
                name=name,
                description=description,
                status=status,
                created_by=request.tenant_user,
            )
            messages.success(request, "Program sector value created.")
            return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    ctx["dimension_label"] = "Program sector"
    ctx["dimension_code"] = sector_dim.dimension_code
    ctx["status_choices"] = FinancialDimensionValue.Status.choices
    ctx["form_title"] = "Add Program Sector Value"
    ctx["list_url_name"] = "tenant_portal:setup_sector_dimension_values_list"
    return render(request, "tenant_portal/setup/dimension_values_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_sector_dimension_values_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit sector values.")
        return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    obj = get_object_or_404(
        FinancialDimensionValue.objects.using(tenant_db).select_related("dimension"),
        pk=pk,
        dimension__dimension_code="SECTOR",
    )

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimensionValue.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Code is required.")
        if not name:
            errors.append("Name is required.")
        if (
            code
            and FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension=obj.dimension, code__iexact=code)
            .exclude(pk=pk)
            .exists()
        ):
            errors.append("A sector value with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj.code = code
            obj.name = name
            obj.description = description
            obj.status = status
            obj.save(using=tenant_db)
            messages.success(request, "Program sector value updated.")
            return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    ctx["value"] = obj
    ctx["dimension_label"] = "Program sector"
    ctx["dimension_code"] = obj.dimension.dimension_code
    ctx["status_choices"] = FinancialDimensionValue.Status.choices
    ctx["form_title"] = "Edit Program Sector Value"
    ctx["list_url_name"] = "tenant_portal:setup_sector_dimension_values_list"
    return render(request, "tenant_portal/setup/dimension_values_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_sector_dimension_values_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete sector values.")
        return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    from tenant_finance.models import FinancialDimensionValue

    obj = get_object_or_404(
        FinancialDimensionValue.objects.using(tenant_db).select_related("dimension"),
        pk=pk,
        dimension__dimension_code="SECTOR",
    )
    if request.method == "POST":
        obj.delete(using=tenant_db)
        messages.success(request, "Program sector value deleted.")
        return redirect(reverse("tenant_portal:setup_sector_dimension_values_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Program sector value {obj.code} — {obj.name}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_sector_dimension_values_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_sector_dimension_values_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_cost_centers_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_cost_centers"

    cost_centers_qs = _cost_centers_queryset(tenant_db, request)
    paginator_cc = Paginator(cost_centers_qs, PAGE_SIZE)
    page_cc = request.GET.get("page_cc", "1")
    ctx["cost_centers_page"] = paginator_cc.get_page(page_cc)
    ctx["cost_centers_filter_q"] = request.GET.get("q_cc", "")
    ctx["cost_centers_filter_status"] = request.GET.get("status_cc", "")

    get = request.GET.copy()
    if "page_cc" in get:
        get.pop("page_cc")
    ctx["cost_centers_base_query"] = get.urlencode()

    from tenant_finance.models import CostCenter

    ctx["status_choices"] = CostCenter.Status.choices
    return render(request, "tenant_portal/setup/cost_centers_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_dimensions_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to create dimensions.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import FinancialDimension

    if request.method == "POST":
        code = (request.POST.get("dimension_code") or "").strip()
        name = (request.POST.get("dimension_name") or "").strip()
        dim_type = (request.POST.get("dimension_type") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimension.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Dimension code is required.")
        if not name:
            errors.append("Dimension name is required.")
        if not dim_type:
            errors.append("Dimension type is required.")
        if code and FinancialDimension.objects.using(tenant_db).filter(dimension_code__iexact=code).exists():
            errors.append("A dimension with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            FinancialDimension.objects.using(tenant_db).create(
                dimension_code=code,
                dimension_name=name,
                dimension_type=dim_type,
                description=description,
                status=status,
                created_by=request.tenant_user,
            )
            messages.success(request, "Dimension created.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))

    ctx["dimension_types"] = FinancialDimension.DimensionType.choices
    ctx["status_choices"] = FinancialDimension.Status.choices
    return render(request, "tenant_portal/setup/dimensions_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_dimensions_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit dimensions.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import FinancialDimension

    obj = get_object_or_404(FinancialDimension.objects.using(tenant_db), pk=pk)
    ctx["dimension"] = obj

    if request.method == "POST":
        code = (request.POST.get("dimension_code") or "").strip()
        name = (request.POST.get("dimension_name") or "").strip()
        dim_type = (request.POST.get("dimension_type") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or FinancialDimension.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Dimension code is required.")
        if not name:
            errors.append("Dimension name is required.")
        if not dim_type:
            errors.append("Dimension type is required.")
        if code and FinancialDimension.objects.using(tenant_db).filter(dimension_code__iexact=code).exclude(pk=pk).exists():
            errors.append("A dimension with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj.dimension_code = code
            obj.dimension_name = name
            obj.dimension_type = dim_type
            obj.description = description
            obj.status = status
            obj.save(using=tenant_db)
            messages.success(request, "Dimension updated.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))

    ctx["dimension_types"] = FinancialDimension.DimensionType.choices
    ctx["status_choices"] = FinancialDimension.Status.choices
    return render(request, "tenant_portal/setup/dimensions_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_dimensions_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete dimensions.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import FinancialDimension, JournalEntry

    obj = get_object_or_404(FinancialDimension.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if JournalEntry.objects.using(tenant_db).filter(dimension_id=pk).exists():
            messages.error(request, "Cannot delete: this dimension is used in transactions.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))
        obj.delete(using=tenant_db)
        messages.success(request, "Dimension deleted.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Dimension {obj.dimension_code} — {obj.dimension_name}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_dimensions_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_dimensions_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


# ----- Cost Centers -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_cost_centers_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to create cost centers.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import CostCenter
    from tenant_users.models import TenantUser

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        parent_id = (request.POST.get("parent_id") or "").strip()
        manager_id = (request.POST.get("manager_id") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or CostCenter.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Cost center code is required.")
        if not name:
            errors.append("Cost center name is required.")
        if code and CostCenter.objects.using(tenant_db).filter(code__iexact=code).exists():
            errors.append("A cost center with this code already exists.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            parent = None
            if parent_id:
                parent = CostCenter.objects.using(tenant_db).filter(pk=parent_id).first()
            manager = None
            if manager_id:
                manager = TenantUser.objects.using(tenant_db).filter(pk=manager_id).first()
            CostCenter.objects.using(tenant_db).create(
                code=code,
                name=name,
                parent=parent,
                manager=manager,
                description=description,
                status=status,
                created_by=request.tenant_user,
            )
            messages.success(request, "Cost center created.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))

    cost_centers = list(CostCenter.objects.using(tenant_db).filter(status=CostCenter.Status.ACTIVE).order_by("code"))
    users = list(TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("email"))
    ctx["cost_centers"] = cost_centers
    ctx["users"] = users
    ctx["status_choices"] = CostCenter.Status.choices
    ctx["form_title"] = "Create Cost Center"
    return render(request, "tenant_portal/setup/cost_centers_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_cost_centers_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_dimensions"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit cost centers.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import CostCenter
    from tenant_users.models import TenantUser

    obj = get_object_or_404(CostCenter.objects.using(tenant_db), pk=pk)
    ctx["cost_center"] = obj

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip()
        name = (request.POST.get("name") or "").strip()
        parent_id = (request.POST.get("parent_id") or "").strip()
        manager_id = (request.POST.get("manager_id") or "").strip()
        description = (request.POST.get("description") or "").strip()
        status = (request.POST.get("status") or "").strip() or CostCenter.Status.ACTIVE
        errors = []
        if not code:
            errors.append("Cost center code is required.")
        if not name:
            errors.append("Cost center name is required.")
        if code and CostCenter.objects.using(tenant_db).filter(code__iexact=code).exclude(pk=pk).exists():
            errors.append("A cost center with this code already exists.")
        if parent_id and int(parent_id) == pk:
            errors.append("Cost center cannot be its own parent.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            parent = None
            if parent_id:
                parent = CostCenter.objects.using(tenant_db).filter(pk=parent_id).first()
            manager = None
            if manager_id:
                manager = TenantUser.objects.using(tenant_db).filter(pk=manager_id).first()
            obj.code = code
            obj.name = name
            obj.parent = parent
            obj.manager = manager
            obj.description = description
            obj.status = status
            obj.save(using=tenant_db)
            messages.success(request, "Cost center updated.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))

    cost_centers = list(CostCenter.objects.using(tenant_db).exclude(pk=pk).order_by("code"))
    users = list(TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("email"))
    ctx["cost_centers"] = cost_centers
    ctx["users"] = users
    ctx["status_choices"] = CostCenter.Status.choices
    ctx["form_title"] = "Edit Cost Center"
    return render(request, "tenant_portal/setup/cost_centers_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_cost_centers_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete cost centers.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    from tenant_finance.models import CostCenter, JournalEntry

    obj = get_object_or_404(CostCenter.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if JournalEntry.objects.using(tenant_db).filter(cost_center_id=pk).exists():
            messages.error(request, "Cannot delete: this cost center is used in transactions.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))
        if obj.children.exists():
            messages.error(request, "Cannot delete: this cost center has child cost centers. Remove or reassign them first.")
            return redirect(reverse("tenant_portal:setup_dimensions_list"))
        obj.delete(using=tenant_db)
        messages.success(request, "Cost center deleted.")
        return redirect(reverse("tenant_portal:setup_dimensions_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Cost center {obj.code} — {obj.name}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_dimensions_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_cost_centers_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


# ----- Project / Grant dimensions -----
PROJECTS_LIST_PAGE_SIZE = 25


def _project_dimensions_queryset(tenant_db: str, request: HttpRequest):
    from datetime import datetime

    from django.db.models.functions import Coalesce

    from tenant_grants.models import Project

    qs = Project.objects.using(tenant_db).select_related("donor", "project_manager", "currency")
    qs = qs.annotate(effective_end=Coalesce("revised_end_date", "original_end_date", "end_date"))

    q = (request.GET.get("q_proj") or request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(name__icontains=q))
    title = (request.GET.get("title") or "").strip()
    if title:
        qs = qs.filter(name__icontains=title)
    code = (request.GET.get("code") or "").strip()
    if code:
        qs = qs.filter(code__icontains=code)
    status = (request.GET.get("status_proj") or request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)
    donor_id = (request.GET.get("donor_id") or "").strip()
    if donor_id.isdigit():
        qs = qs.filter(donor_id=int(donor_id))
    pm_id = (request.GET.get("project_manager_id") or "").strip()
    if pm_id.isdigit():
        qs = qs.filter(project_manager_id=int(pm_id))
    currency_id = (request.GET.get("currency_id") or "").strip()
    if currency_id.isdigit():
        qs = qs.filter(currency_id=int(currency_id))
    prog = (request.GET.get("program") or "").strip()
    if prog:
        qs = qs.filter(program_sector__icontains=prog)
    loc = (request.GET.get("location") or "").strip()
    if loc:
        qs = qs.filter(location__icontains=loc)
    phase = (request.GET.get("phase") or "").strip()
    if phase:
        today = timezone.now().date()
        if phase == "upcoming":
            qs = qs.filter(start_date__gt=today)
        elif phase == "ended":
            qs = qs.filter(effective_end__lt=today, effective_end__isnull=False)
        elif phase == "in_period":
            qs = qs.filter(
                start_date__lte=today,
                start_date__isnull=False,
                effective_end__gte=today,
                effective_end__isnull=False,
            )
    start_after = (request.GET.get("start_after") or "").strip()
    if start_after:
        try:
            sa = datetime.strptime(start_after, "%Y-%m-%d").date()
            qs = qs.filter(Q(start_date__isnull=True) | Q(start_date__gte=sa))
        except ValueError:
            pass
    start_before = (request.GET.get("start_before") or "").strip()
    if start_before:
        try:
            sb = datetime.strptime(start_before, "%Y-%m-%d").date()
            qs = qs.filter(start_date__lte=sb, start_date__isnull=False)
        except ValueError:
            pass
    end_before = (request.GET.get("end_before") or "").strip()
    if end_before:
        try:
            eb = datetime.strptime(end_before, "%Y-%m-%d").date()
            qs = qs.filter(effective_end__lte=eb, effective_end__isnull=False)
        except ValueError:
            pass
    end_after = (request.GET.get("end_after") or "").strip()
    if end_after:
        try:
            ea = datetime.strptime(end_after, "%Y-%m-%d").date()
            qs = qs.filter(effective_end__gte=ea, effective_end__isnull=False)
        except ValueError:
            pass

    sort = (request.GET.get("sort_proj") or request.GET.get("sort") or "code").strip()
    order_map = {
        "code": "code",
        "-code": "-code",
        "name": "name",
        "-name": "-name",
        "start": "start_date",
        "-start": "-start_date",
        "end": "effective_end",
        "-end": "-effective_end",
        "status": "status",
        "-status": "-status",
        "donor": "donor__name",
        "-donor": "-donor__name",
        "updated": "-updated_at",
        "-updated": "updated_at",
    }
    qs = qs.order_by(order_map.get(sort, "code"))
    return qs


def _grant_dimensions_queryset(tenant_db: str, request: HttpRequest):
    from tenant_grants.models import Grant

    qs = (
        Grant.objects.using(tenant_db)
        .select_related("donor", "project", "source_tracking", "bank_account")
        .order_by("code")
    )
    q = (request.GET.get("q_grant") or "").strip()
    if q:
        qs = qs.filter(
            Q(code__icontains=q)
            | Q(title__icontains=q)
            | Q(donor__name__icontains=q)
            | Q(donor__code__icontains=q)
        )
    status = (request.GET.get("status_grant") or "").strip()
    if status:
        qs = qs.filter(status=status)
    return qs


def _mapping_queryset(tenant_db: str, request: HttpRequest):
    from tenant_finance.models import ProjectDimensionMapping

    return (
        ProjectDimensionMapping.objects.using(tenant_db)
        .select_related("project", "cost_center", "bank_account", "donor")
        .order_by("project__code")
    )


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimensions_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    ctx["page_title"] = "Projects & Grants"

    projects_qs = _project_dimensions_queryset(tenant_db, request)
    grants_qs = _grant_dimensions_queryset(tenant_db, request)
    mapping_qs = _mapping_queryset(tenant_db, request)

    paginator_proj = Paginator(projects_qs, PAGE_SIZE)
    paginator_grant = Paginator(grants_qs, PAGE_SIZE)
    paginator_map = Paginator(mapping_qs, PAGE_SIZE)

    page_proj = request.GET.get("page_proj", "1")
    page_grant = request.GET.get("page_grant", "1")
    page_map = request.GET.get("page_map", "1")

    projects_page = paginator_proj.get_page(page_proj)
    grants_page = paginator_grant.get_page(page_grant)
    ctx["projects_page"] = projects_page
    ctx["grants_page"] = grants_page
    ctx["mapping_page"] = paginator_map.get_page(page_map)

    from tenant_grants.services.project_financials import attach_project_financials

    attach_project_financials(list(projects_page.object_list), tenant_db)

    ctx["filter_q_proj"] = request.GET.get("q_proj", "")
    ctx["filter_status_proj"] = request.GET.get("status_proj", "")
    ctx["filter_donor_id"] = request.GET.get("donor_id", "")
    ctx["filter_program"] = request.GET.get("program", "")
    ctx["filter_location"] = request.GET.get("location", "")
    ctx["filter_sort_proj"] = request.GET.get("sort_proj", "") or request.GET.get("sort", "code")
    ctx["filter_start_after"] = request.GET.get("start_after", "")
    ctx["filter_end_before"] = request.GET.get("end_before", "")
    from tenant_grants.models import Donor

    ctx["donors_for_project_filter"] = (
        Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    )
    ctx["filter_q_grant"] = request.GET.get("q_grant", "")
    ctx["filter_status_grant"] = request.GET.get("status_grant", "")

    # Grant spend / remaining for UI (Budget, Spent, Remaining) and utilization colour
    from decimal import Decimal
    from django.db.models import Sum
    from tenant_finance.models import get_grant_posted_expense_total
    from tenant_grants.models import BudgetLine

    for g in grants_page.object_list:
        if not g.pk:
            g.spent_amount = Decimal("0")
            budget_total = getattr(g, "award_amount", None) or getattr(g, "amount_awarded", None) or Decimal("0")
            g.remaining_amount = budget_total
            continue
        spent = get_grant_posted_expense_total(g.pk, tenant_db)
        g.spent_amount = spent
        budget_total = getattr(g, "award_amount", None) or getattr(g, "amount_awarded", None) or Decimal("0")
        if not budget_total or budget_total <= 0:
            budget_total = (
                BudgetLine.objects.using(tenant_db)
                .filter(grant_id=g.pk)
                .aggregate(t=Sum("amount"))
                .get("t")
                or Decimal("0")
            )
        g.remaining_amount = (budget_total or Decimal("0")) - (spent or Decimal("0"))
        # Utilization ratio for colour coding
        if budget_total and budget_total > 0:
            ratio = ((spent or Decimal("0")) / budget_total) if spent is not None else Decimal("0")
        else:
            ratio = Decimal("0")
        if ratio >= Decimal("1"):
            g.utilization_class = "dyn-pill-red"
        elif ratio >= Decimal("0.8"):
            g.utilization_class = "dyn-pill-orange"
        else:
            g.utilization_class = "dyn-pill-green"

    get = request.GET.copy()
    for key in ("page_proj", "page_grant", "page_map"):
        if key in get:
            get.pop(key)
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/project_grant_dimensions.html", ctx)


def _projects_list_export_rows(projects: list) -> list[list]:
    """Eight columns for print / CSV / Excel: code, name, donor, location, budget, beneficiaries, start, end."""
    rows_out: list[list] = []
    for p in projects:
        fin = getattr(p, "fin", {}) or {}
        donor = p.donor.name if p.donor_id else ""
        end_eff = ""
        if getattr(p, "effective_end", None):
            end_eff = p.effective_end.isoformat()
        rows_out.append(
            [
                p.code,
                p.name,
                donor,
                p.location or "",
                fin.get("total_budget") if fin.get("total_budget") is not None else "",
                p.total_beneficiaries,
                p.start_date.isoformat() if p.start_date else "",
                end_eff,
            ]
        )
    return rows_out


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def projects_list_view(request: HttpRequest) -> HttpResponse:
    import csv
    import os
    from io import BytesIO, StringIO

    from django.db.models import Count
    from django.http import HttpResponse
    from django.utils import timezone as dj_tz

    from tenant_finance.models import OrganizationSettings
    from tenant_grants.models import Donor, Project
    from tenant_grants.services.project_financials import attach_project_financials

    tenant_db = request.tenant_db

    export_fmt = (request.GET.get("export") or "").strip().lower()
    if export_fmt in {"csv", "xlsx"}:
        projects_qs = list(_project_dimensions_queryset(tenant_db, request))
        attach_project_financials(projects_qs, tenant_db)
        org_settings = OrganizationSettings.objects.using(tenant_db).first()
        legal_name = (
            (org_settings.organization_name or "").strip()
            if org_settings and getattr(org_settings, "organization_name", None)
            else (getattr(request.tenant, "name", None) or "")
        )
        headers = [
            "Project code",
            "Project name",
            "Donor",
            "Location",
            "Total budget",
            "Number of beneficiaries",
            "Start date",
            "End date",
        ]
        data_rows = _projects_list_export_rows(projects_qs)

        if export_fmt == "csv":
            buf = StringIO()
            buf.write("\ufeff")
            w = csv.writer(buf)
            w.writerow([legal_name])
            w.writerow([])
            w.writerow(["Project List"])
            w.writerow([])
            w.writerow(headers)
            for row in data_rows:
                w.writerow(row)
            resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
            resp["Content-Disposition"] = 'attachment; filename="project_list.csv"'
            return resp

        # xlsx
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Project List"
        ncols = len(headers)
        last_col = get_column_letter(ncols)

        def _merge_title(r: int, text: str, *, bold: bool = True, size: int = 14) -> None:
            ws.merge_cells(f"A{r}:{last_col}{r}")
            c = ws.cell(row=r, column=1)
            c.value = text
            c.font = Font(size=size, bold=bold)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_cursor = 1
        if org_settings:
            logo_field = getattr(org_settings, "report_logo", None) or getattr(
                org_settings, "organization_logo", None
            )
            if logo_field:
                try:
                    logo_path = logo_field.path
                    if logo_path and os.path.isfile(logo_path):
                        img = XLImage(logo_path)
                        img.height = 52
                        img.width = 170
                        center_idx = max(1, (ncols // 2))
                        img.anchor = f"{get_column_letter(center_idx)}1"
                        ws.add_image(img)
                        row_cursor = 5
                except Exception:
                    row_cursor = 1
        _merge_title(row_cursor, legal_name, size=13)
        row_cursor += 1
        _merge_title(row_cursor, "Project List", size=14)
        row_cursor += 1
        ws.merge_cells(f"A{row_cursor}:{last_col}{row_cursor}")
        gen = ws.cell(row=row_cursor, column=1)
        gen.value = f"Generated: {dj_tz.localtime(dj_tz.now()).strftime('%Y-%m-%d %H:%M')}"
        gen.font = Font(size=10)
        gen.alignment = Alignment(horizontal="center")
        row_cursor += 2

        hdr_row = row_cursor
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=col_idx)
            c.value = h
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_start = hdr_row + 1
        for i, row in enumerate(data_rows):
            for j, val in enumerate(row, start=1):
                ws.cell(row=data_start + i, column=j, value=val)

        # Column widths (Excel character units) — proportional to print layout; avoids wide empty columns
        _pl_xlsx_col_widths = (11, 22, 18, 16, 12, 12, 10, 10)
        for i, w in enumerate(_pl_xlsx_col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(
            out.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="project_list.xlsx"'
        return resp

    ctx = _setup_context(request)
    ctx["active_submenu"] = "funds"
    ctx["active_item"] = "funds_projects_list"
    ctx["page_title"] = "Project list"

    from tenant_grants.services.project_financials import aggregate_project_financial_rollups, project_financial_rollups

    projects_qs = _project_dimensions_queryset(tenant_db, request)
    ctx["projects_filtered_count"] = projects_qs.count()
    status_counts_filtered = dict(
        projects_qs.values("status").annotate(c=Count("id")).values_list("status", "c")
    )
    all_pids = list(projects_qs.values_list("pk", flat=True))
    roll_all = project_financial_rollups(tenant_db, all_pids)
    ctx["projects_totals"] = aggregate_project_financial_rollups(roll_all)

    projects_export_all = list(projects_qs)
    attach_project_financials(projects_export_all, tenant_db)
    ctx["projects_export_all"] = projects_export_all

    paginator = Paginator(projects_qs, PROJECTS_LIST_PAGE_SIZE)
    page_num = request.GET.get("page") or request.GET.get("page_proj") or "1"
    projects_page = paginator.get_page(page_num)
    attach_project_financials(list(projects_page.object_list), tenant_db)

    from tenant_grants.services.project_end_schedule import project_end_alert_state

    for _p in projects_page.object_list:
        _p.end_date_alert_state = project_end_alert_state(_p)

    ctx["projects_page"] = projects_page
    ctx["kpi_cards"] = [
        {"key": "all", "label": "All projects", "count": ctx["projects_filtered_count"]},
        {"key": "draft", "label": "Draft projects", "count": status_counts_filtered.get(Project.Status.DRAFT, 0)},
        {"key": "planning", "label": "Planning stage", "count": status_counts_filtered.get(Project.Status.PLANNING, 0)},
        {"key": "active", "label": "Active projects", "count": status_counts_filtered.get(Project.Status.ACTIVE, 0)},
        {"key": "on_hold", "label": "On hold", "count": status_counts_filtered.get(Project.Status.ON_HOLD, 0)},
        {"key": "closed", "label": "Closed", "count": status_counts_filtered.get(Project.Status.CLOSED, 0)},
        {"key": "completed", "label": "Completed", "count": status_counts_filtered.get(Project.Status.COMPLETED, 0)},
    ]

    ctx["filter_q"] = request.GET.get("q") or request.GET.get("q_proj") or ""
    ctx["filter_title"] = request.GET.get("title") or ""
    ctx["filter_code"] = request.GET.get("code") or ""
    ctx["filter_status"] = request.GET.get("status") or request.GET.get("status_proj") or ""
    ctx["filter_donor_id"] = request.GET.get("donor_id", "")
    ctx["filter_project_manager_id"] = request.GET.get("project_manager_id", "")
    ctx["filter_currency_id"] = request.GET.get("currency_id", "")
    ctx["filter_phase"] = request.GET.get("phase", "")
    ctx["filter_program"] = request.GET.get("program", "")
    ctx["filter_location"] = request.GET.get("location", "")
    ctx["filter_start_after"] = request.GET.get("start_after", "")
    ctx["filter_start_before"] = request.GET.get("start_before", "")
    ctx["filter_end_before"] = request.GET.get("end_before", "")
    ctx["filter_end_after"] = request.GET.get("end_after", "")
    ctx["filter_sort"] = request.GET.get("sort") or request.GET.get("sort_proj") or "code"
    ctx["donors_for_project_filter"] = (
        Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    )
    ctx["project_status_choices"] = Project.Status.choices
    ctx["phase_filter_choices"] = (
        ("upcoming", "Upcoming"),
        ("in_period", "In period"),
        ("ended", "Ended"),
    )

    from tenant_finance.models import Currency, OrganizationSettings
    from tenant_users.models import TenantUser

    org_settings = OrganizationSettings.objects.using(tenant_db).first()
    ctx["org_settings"] = org_settings
    ctx["tenant_legal_name"] = (
        (org_settings.organization_name or "").strip()
        if org_settings and getattr(org_settings, "organization_name", None)
        else (getattr(request.tenant, "name", None) or "")
    )

    ctx["currencies_for_project_filter"] = (
        Currency.objects.using(tenant_db).filter(status=Currency.Status.ACTIVE).order_by("code")
    )
    mgr_ids = (
        Project.objects.using(tenant_db)
        .exclude(project_manager_id__isnull=True)
        .values_list("project_manager_id", flat=True)
        .distinct()
    )
    ctx["project_managers_for_filter"] = (
        TenantUser.objects.using(tenant_db).filter(pk__in=mgr_ids).order_by("email")
    )

    get_params = request.GET.copy()
    for _k in ("page", "page_proj", "export"):
        get_params.pop(_k, None)
    ctx["has_active_filters"] = bool(get_params)

    get = request.GET.copy()
    for key in ("page", "page_proj", "export"):
        if key in get:
            get.pop(key)
    ctx["base_query"] = get.urlencode()
    export_get = request.GET.copy()
    for key in ("page", "page_proj", "export"):
        export_get.pop(key, None)
    export_get_csv = export_get.copy()
    export_get_csv["export"] = "csv"
    export_get_xlsx = export_get.copy()
    export_get_xlsx["export"] = "xlsx"
    ctx["projects_export_query"] = export_get_xlsx.urlencode()
    ctx["projects_export_csv_query"] = export_get_csv.urlencode()

    return render(request, "tenant_portal/grants/projects_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimensions_add_view(request: HttpRequest) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    ctx["page_title"] = "Add grant dimension"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimensions_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    ctx["page_title"] = "Edit grant dimension"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimensions_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))


def _save_project_from_post(
    *,
    tenant_db: str,
    post,
    existing=None,
    tenant_user=None,
) -> tuple[list[str], object | None]:
    """Parse POST into a Project, full_clean, save. Returns (errors, saved_instance_or_none)."""
    from datetime import datetime

    from tenant_finance.models import Currency, FinancialDimension, FinancialDimensionValue
    from tenant_grants.models import Donor, FundingSource, Project
    from tenant_grants.services.project_end_schedule import (
        project_closure_blockers,
        user_can_manage_project_closure_or_extension,
    )
    from tenant_users.models import TenantUser

    errors: list[str] = []
    code = (post.get("code") or "").strip()
    name = (post.get("name") or "").strip()
    donor_id = (post.get("donor_id") or "").strip()
    pm_id = (post.get("project_manager_id") or "").strip()
    currency_id = (post.get("currency_id") or "").strip()
    funding_modality_id = (post.get("funding_modality_id") or "").strip()
    location = (post.get("location") or "").strip()
    program_sector = (post.get("program_sector") or "").strip()
    funding_type = (post.get("funding_type") or "").strip()
    status = (post.get("status") or "").strip() or Project.Status.PLANNING
    start_date = (post.get("start_date") or "").strip() or None
    end_date = (post.get("end_date") or "").strip() or None
    original_end_date = (post.get("original_end_date") or "").strip() or None
    revised_end_date = (post.get("revised_end_date") or "").strip() or None
    tb_raw = (post.get("total_beneficiaries") or "").strip()

    if not code:
        errors.append("Project code is required.")
    if not name:
        errors.append("Project title is required.")
    valid_status = {c[0] for c in Project.Status.choices}
    if status not in valid_status:
        errors.append("Invalid workflow status.")

    prog_dim = _ensure_program_dimension_defaults(tenant_db, tenant_user)
    valid_program_codes: set[str] = set()
    if prog_dim:
        valid_program_codes = set(
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension_id=prog_dim.pk, status=FinancialDimensionValue.Status.ACTIVE)
            .values_list("code", flat=True)
        )
    if funding_type and funding_type not in valid_program_codes:
        errors.append("Program category must be selected from Program dimension values.")

    sector_dim = _ensure_sector_dimension_defaults(tenant_db, tenant_user)
    valid_sector_codes: set[str] = set()
    if sector_dim:
        valid_sector_codes = set(
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension_id=sector_dim.pk, status=FinancialDimensionValue.Status.ACTIVE)
            .values_list("code", flat=True)
        )
    if program_sector and program_sector not in valid_sector_codes:
        errors.append("Program / sector must be selected from Program sector dimension values.")

    if tb_raw:
        try:
            tb = int(tb_raw)
        except ValueError:
            errors.append("Planned beneficiaries must be a whole number.")
            tb = 0
    else:
        tb = existing.total_beneficiaries if existing is not None else 0

    if tb < 0:
        errors.append("Planned beneficiaries cannot be negative.")
    elif tb < 1:
        errors.append("Planned beneficiaries must be a positive whole number (at least 1).")

    def _parse_date(label: str, raw: str | None):
        if not raw:
            return None, None
        try:
            return datetime.strptime(raw, "%Y-%m-%d").date(), None
        except ValueError:
            return None, f"Invalid {label}. Use YYYY-MM-DD."

    sd, e = _parse_date("start date", start_date)
    if e:
        errors.append(e)
    ed, e = _parse_date("operational end date", end_date)
    if e:
        errors.append(e)
    oed, e = _parse_date("original end date", original_end_date)
    if e:
        errors.append(e)
    red, e = _parse_date("revised end date", revised_end_date)
    if e:
        errors.append(e)

    if not start_date:
        errors.append("Start date is required.")
    if not end_date:
        errors.append("Operational end date is required.")
    if sd is not None and ed is not None and sd > ed:
        errors.append("Operational end date must be on or after the start date.")
    if oed is not None and sd is not None and oed < sd:
        errors.append("Original end date must be on or after the start date.")
    if red is not None and oed is not None and red < oed:
        errors.append("Revised end date must be on or after original end date.")

    if currency_id:
        if not currency_id.isdigit():
            errors.append("Invalid reporting currency.")
        else:
            _cur = (
                Currency.objects.using(tenant_db)
                .filter(pk=int(currency_id), status=Currency.Status.ACTIVE)
                .first()
            )
            if not _cur:
                errors.append(
                    "Reporting currency must be an active currency defined in the system."
                )

    if funding_modality_id:
        if not funding_modality_id.isdigit():
            errors.append("Invalid funding modality.")
        else:
            _fm = (
                FundingSource.objects.using(tenant_db)
                .filter(pk=int(funding_modality_id), is_active=True)
                .first()
            )
            if not _fm:
                errors.append("Funding modality must be an active entry in the catalog.")
            else:
                from tenant_grants.services.payment_modality import has_complete_gl_mapping

                if not has_complete_gl_mapping(
                    using=tenant_db,
                    funding_source=_fm,
                ):
                    errors.append("GL account mapping is missing for selected funding modality.")

    if status == Project.Status.ACTIVE:
        if not donor_id:
            errors.append("Donor is required before project status can be set to Active.")
        if not currency_id:
            errors.append("Reporting currency is required before project status can be set to Active.")
        if not funding_type:
            errors.append("Program category is required before project status can be set to Active.")
        if not funding_modality_id:
            errors.append("Funding modality is required before project status can be set to Active.")

    if existing is None:
        if code and Project.objects.using(tenant_db).filter(code__iexact=code).exists():
            errors.append("A project with this code already exists.")
    elif code and Project.objects.using(tenant_db).filter(code__iexact=code).exclude(pk=existing.pk).exists():
        errors.append("A project with this code already exists.")

    if existing is not None and tenant_user is not None:
        closing_now = status in (Project.Status.CLOSED, Project.Status.COMPLETED) and existing.status not in (
            Project.Status.CLOSED,
            Project.Status.COMPLETED,
        )

        def _dn(d1, d2):
            return (d1 or None) != (d2 or None)

        schedule_changed = _dn(ed, existing.end_date) or _dn(red, existing.revised_end_date) or _dn(
            oed, existing.original_end_date
        )
        if closing_now or schedule_changed:
            if not user_can_manage_project_closure_or_extension(tenant_user, existing, tenant_db):
                errors.append(
                    "Only the project manager or an administrator may close the project or change operational or revised end dates (extensions)."
                )
            elif closing_now:
                for msg in project_closure_blockers(existing, tenant_db):
                    errors.append(msg)

    if errors:
        return errors, None

    donor = Donor.objects.using(tenant_db).filter(pk=int(donor_id)).first() if donor_id.isdigit() else None
    pm = TenantUser.objects.using(tenant_db).filter(pk=int(pm_id)).first() if pm_id.isdigit() else None
    currency = (
        Currency.objects.using(tenant_db)
        .filter(pk=int(currency_id), status=Currency.Status.ACTIVE)
        .first()
        if currency_id.isdigit()
        else None
    )
    funding_modality = (
        FundingSource.objects.using(tenant_db)
        .filter(pk=int(funding_modality_id), is_active=True)
        .first()
        if funding_modality_id.isdigit()
        else None
    )

    if existing is not None:
        obj = existing
        obj.code = code
        obj.name = name
    else:
        obj = Project(code=code, name=name)

    obj.donor = donor
    obj.project_manager = pm
    obj.currency = currency
    obj.location = location
    obj.program_sector = program_sector
    obj.funding_type = funding_type
    obj.funding_modality = funding_modality
    obj.status = status
    obj.is_active = status == Project.Status.ACTIVE
    obj.start_date = sd
    obj.end_date = ed
    if "original_end_date" in post:
        obj.original_end_date = oed
    if "revised_end_date" in post:
        obj.revised_end_date = red

    obj.total_beneficiaries = tb

    try:
        obj.full_clean()
    except ValidationError as exc:
        for _f, msgs in exc.error_dict.items():
            errors.extend(str(m) for m in msgs)
        if exc.error_list:
            errors.extend(str(m) for m in exc.error_list)
        return errors, None

    obj.save(using=tenant_db)
    return [], obj


def _create_funding_modality_template_from_post(*, tenant_db: str, post, request) -> tuple[list[str], object | None]:
    from decimal import Decimal, InvalidOperation

    from django.core.exceptions import ValidationError
    from django.db import transaction

    from tenant_grants.models import FundingSource
    from tenant_grants.services.funding_source_structure import (
        replace_funding_source_payment_structure_from_post,
    )

    errors: list[str] = []
    name = (post.get("modality_name") or "").strip()
    modality_type = (post.get("modality_type") or "").strip()
    retention_raw = (post.get("modality_retention_percentage") or "").strip()
    allow_instalments = post.get("modality_allow_instalments") == "on"
    requires_reporting_before_next_payment = post.get("modality_requires_reporting_before_next_payment") == "on"
    description = (post.get("modality_description") or "").strip()
    is_active = post.get("modality_is_active") == "on"

    valid_mt = {c[0] for c in FundingSource.ModalityType.choices}
    retention_pct = None
    if retention_raw:
        try:
            retention_pct = Decimal(retention_raw.replace(",", ""))
            if retention_pct < 0 or retention_pct > 100:
                errors.append("Retention percentage must be between 0 and 100.")
        except (InvalidOperation, ValueError):
            errors.append("Retention percentage must be a valid number.")

    if not name:
        errors.append("Funding modality template name is required.")
    if modality_type not in valid_mt:
        errors.append("Funding modality type is required.")
    if errors:
        return errors, None

    obj = FundingSource(
        name=name,
        modality_type=modality_type,
        retention_percentage=retention_pct,
        allow_instalments=allow_instalments,
        requires_reporting_before_next_payment=requires_reporting_before_next_payment,
        description=description,
        is_active=is_active,
        donor=None,  # Template is global; donor is not controlling link.
    )
    try:
        obj.full_clean()
    except ValidationError as exc:
        for _f, msgs in getattr(exc, "error_dict", {}).items():
            errors.extend(str(m) for m in msgs)
        for m in getattr(exc, "error_list", []):
            errors.append(str(m))
        if not errors:
            errors.extend(getattr(exc, "messages", []))
        return errors, None

    try:
        with transaction.atomic(using=tenant_db):
            obj.save(using=tenant_db)
            if obj.modality_type == FundingSource.ModalityType.MIXED_MODALITY:
                replace_funding_source_payment_structure_from_post(request, obj, tenant_db)
                obj.refresh_from_db()
                obj.full_clean()
    except ValidationError as exc:
        for _f, msgs in getattr(exc, "error_dict", {}).items():
            errors.extend(str(m) for m in msgs)
        for m in getattr(exc, "error_list", []):
            errors.append(str(m))
        if not errors:
            errors.extend(getattr(exc, "messages", []))
        return errors, None
    return [], obj


def _project_dimension_form_context(tenant_db: str, ctx: dict, project=None) -> dict:
    from tenant_finance.models import Currency, FinancialDimension, FinancialDimensionValue
    from tenant_grants.models import Donor, FundingSource, Project
    from tenant_grants.services.payment_modality import has_complete_gl_mapping
    from tenant_grants.services.project_end_schedule import user_can_manage_project_closure_or_extension
    from tenant_users.models import TenantUser
    import json

    ctx["donors"] = list(
        Donor.objects.using(tenant_db).filter(status=Donor.Status.ACTIVE).order_by("name")
    )
    ctx["users"] = list(TenantUser.objects.using(tenant_db).filter(is_active=True).order_by("email"))
    ctx["currencies"] = list(
        Currency.objects.using(tenant_db).filter(status=Currency.Status.ACTIVE).order_by("code")
    )
    ctx["project_statuses"] = Project.Status.choices
    prog_dim = _ensure_program_dimension_defaults(tenant_db, ctx.get("tenant_user"))
    ctx["program_dimension_values"] = (
        list(
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension_id=prog_dim.pk, status=FinancialDimensionValue.Status.ACTIVE)
            .order_by("name")
        )
        if prog_dim
        else []
    )
    ctx["funding_modalities"] = list(
        FundingSource.objects.using(tenant_db).filter(is_active=True).prefetch_related("payment_structure").order_by("name")
    )
    sector_dim = _ensure_sector_dimension_defaults(tenant_db, ctx.get("tenant_user"))
    ctx["sector_dimension_values"] = (
        list(
            FinancialDimensionValue.objects.using(tenant_db)
            .filter(dimension_id=sector_dim.pk, status=FinancialDimensionValue.Status.ACTIVE)
            .order_by("name")
        )
        if sector_dim
        else []
    )
    ctx["modality_types"] = FundingSource.ModalityType.choices
    ctx["mixed_modality_value"] = FundingSource.ModalityType.MIXED_MODALITY
    modality_details = {}
    for m in ctx["funding_modalities"]:
        lines = []
        if m.modality_type == FundingSource.ModalityType.MIXED_MODALITY:
            lines = [
                f"{ln.get_component_type_display()} {ln.percentage}% ({ln.get_payment_trigger_display()})"
                for ln in m.payment_structure.all().order_by("sort_order", "id")
            ]
        modality_details[str(m.pk)] = {
            "modality_type_display": m.get_modality_type_display(),
            "allow_instalments": bool(m.allow_instalments),
            "requires_reporting_before_next_payment": bool(m.requires_reporting_before_next_payment),
            "retention_percentage": str(m.retention_percentage) if m.retention_percentage is not None else "",
            "structure": lines,
            "gl_mapping_complete": has_complete_gl_mapping(using=tenant_db, funding_source=m),
        }
    ctx["funding_modality_details_json"] = json.dumps(modality_details)
    ctx["show_create_modality_modal"] = bool(ctx.get("show_create_modality_modal"))
    ctx["project"] = project
    ctx["projects_list_url"] = reverse("tenant_portal:grants_projects_list")
    tu = ctx.get("tenant_user")

    if project is None:
        ctx["can_manage_project_schedule"] = bool(
            tu and user_has_permission(tu, "module:finance.manage", using=tenant_db)
        )
    else:
        ctx["can_manage_project_schedule"] = bool(
            tu and user_can_manage_project_closure_or_extension(tu, project, tenant_db)
        )
    return ctx


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_dimension_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to add projects.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        if action == "create_modality":
            errs, created = _create_funding_modality_template_from_post(
                tenant_db=tenant_db, post=request.POST, request=request
            )
            _project_dimension_form_context(tenant_db, ctx, project=None)
            post_data = request.POST.copy()
            post_data["funding_modality_id"] = str(created.pk) if created else (post_data.get("funding_modality_id") or "")
            ctx["form_post"] = post_data
            ctx["show_create_modality_modal"] = True
            ctx["form_title"] = "Add project"
            if errs:
                for e in errs:
                    messages.error(request, e)
            else:
                messages.success(request, "Funding modality template created and selected.")
                ctx["show_create_modality_modal"] = False
            return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)
        errs, _saved = _save_project_from_post(
            tenant_db=tenant_db, post=request.POST, existing=None, tenant_user=request.tenant_user
        )
        if errs:
            for e in errs:
                messages.error(request, e)
            _project_dimension_form_context(tenant_db, ctx, project=None)
            ctx["form_post"] = request.POST
            ctx["form_title"] = "Add project"
            return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)
        messages.success(request, "Project created.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    _project_dimension_form_context(tenant_db, ctx, project=None)
    ctx["form_title"] = "Add project"
    return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_dimension_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit projects.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    from tenant_grants.models import Project

    obj = get_object_or_404(Project.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        if action == "create_modality":
            errs, created = _create_funding_modality_template_from_post(
                tenant_db=tenant_db, post=request.POST, request=request
            )
            _project_dimension_form_context(tenant_db, ctx, project=obj)
            post_data = request.POST.copy()
            post_data["funding_modality_id"] = str(created.pk) if created else (post_data.get("funding_modality_id") or "")
            ctx["form_post"] = post_data
            ctx["show_create_modality_modal"] = True
            ctx["form_title"] = "Edit project"
            if errs:
                for e in errs:
                    messages.error(request, e)
            else:
                messages.success(request, "Funding modality template created and selected.")
                ctx["show_create_modality_modal"] = False
            return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)
        errs, _saved = _save_project_from_post(
            tenant_db=tenant_db, post=request.POST, existing=obj, tenant_user=request.tenant_user
        )
        if errs:
            for e in errs:
                messages.error(request, e)
            _project_dimension_form_context(tenant_db, ctx, project=obj)
            ctx["form_post"] = request.POST
            ctx["form_title"] = "Edit project"
            return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)
        messages.success(request, "Project updated.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    _project_dimension_form_context(tenant_db, ctx, project=obj)
    ctx["form_title"] = "Edit project"
    return render(request, "tenant_portal/setup/project_dimension_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_dimension_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete projects.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    from tenant_finance.models import ProjectDimensionMapping
    from tenant_grants.models import Grant, Project

    obj = get_object_or_404(Project.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if Grant.objects.using(tenant_db).filter(project_id=pk).exists():
            messages.error(request, "Cannot delete: grants are linked to this project. Unlink or delete them first.")
            return redirect(reverse("tenant_portal:grants_projects_list"))
        ProjectDimensionMapping.objects.using(tenant_db).filter(project_id=pk).delete()
        obj.delete(using=tenant_db)
        messages.success(request, "Project deleted.")
        return redirect(reverse("tenant_portal:grants_projects_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Project {obj.code} — {obj.name}"
    ctx["cancel_url"] = reverse("tenant_portal:grants_projects_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_project_dimension_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimension_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to add grants.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import Currency
    from tenant_grants.models import (
        Donor,
        DonorAgreement,
        DonorAgreementGrant,
        DonorAgreementProject,
        FundingSource,
        Grant,
        GrantTranche,
        Project,
    )

    if request.method == "POST":
        from decimal import Decimal, InvalidOperation

        code = (request.POST.get("code") or "").strip()
        title = (request.POST.get("title") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        donor_contract_id = (request.POST.get("donor_contract_id") or "").strip()
        project_id = (request.POST.get("project_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        raw_ceiling = (request.POST.get("grant_ceiling") or request.POST.get("award_amount") or "").strip() or "0"
        raw_eligible = (request.POST.get("eligible_receivable_amount") or "").strip()
        receivable_basis_note = (request.POST.get("receivable_basis_note") or "").strip()
        status = (request.POST.get("status") or "").strip() or Grant.Status.DRAFT
        errors = []
        if not code:
            errors.append("Grant code is required.")
        if not title:
            errors.append("Grant name is required.")
        if not donor_id:
            errors.append("Donor is required.")
        if not donor_contract_id:
            errors.append("Donor contract is required.")
        if not project_id:
            errors.append("Grant must belong to a project.")
        if code and Grant.objects.using(tenant_db).filter(code__iexact=code).exists():
            errors.append("A grant with this code already exists.")
        grant_ceiling = Decimal("0")
        try:
            grant_ceiling = Decimal(raw_ceiling.replace(",", ""))
            if grant_ceiling <= 0:
                errors.append("Budget (grant ceiling) must be greater than zero.")
        except (InvalidOperation, ValueError):
            errors.append("Grant ceiling must be a valid number.")
        eligible_receivable_amount = grant_ceiling
        if raw_eligible:
            try:
                eligible_receivable_amount = Decimal(raw_eligible.replace(",", ""))
            except (InvalidOperation, ValueError):
                errors.append("Eligible receivable must be a valid number.")
        if not errors and eligible_receivable_amount > grant_ceiling:
            errors.append("Eligible receivable cannot exceed grant ceiling.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            from django.core.exceptions import ValidationError
            from django.db import transaction

            from tenant_grants.services.grant_tranches import replace_grant_tranches_from_post

            donor = get_object_or_404(Donor.objects.using(tenant_db), pk=donor_id)
            donor_contract = get_object_or_404(DonorAgreement.objects.using(tenant_db), pk=donor_contract_id)
            if donor_contract.donor_id != donor.pk:
                messages.error(request, "Selected donor contract must belong to the selected donor.")
                return redirect(reverse("tenant_portal:setup_grant_dimension_add"))
            project = get_object_or_404(Project.objects.using(tenant_db), pk=project_id)
            currency = Currency.objects.using(tenant_db).filter(pk=currency_id).first() if currency_id else None
            raw_fm = (request.POST.get("funding_method") or "").strip()
            valid_fm = {c[0] for c in Grant.FundingMethod.choices}
            funding_method = raw_fm if raw_fm in valid_fm else ""
            funding_modality_id = (request.POST.get("funding_modality_id") or "").strip()
            funding_modality = (
                FundingSource.objects.using(tenant_db)
                .filter(pk=int(funding_modality_id), is_active=True)
                .first()
                if funding_modality_id.isdigit()
                else None
            )
            if not funding_modality:
                messages.error(request, "Funding modality is required.")
                return redirect(reverse("tenant_portal:setup_grant_dimension_add"))
            if funding_modality is not None:
                from tenant_grants.services.payment_modality import has_complete_gl_mapping

                if not has_complete_gl_mapping(
                    using=tenant_db,
                    funding_source=funding_modality,
                ):
                    messages.error(request, "GL account mapping is missing for selected funding modality.")
                    return redirect(reverse("tenant_portal:setup_grant_dimension_add"))
            expense_report_approved = request.POST.get("expense_report_approved") == "on"
            audit_approved = request.POST.get("audit_approved") == "on"
            final_report_approved = request.POST.get("final_report_approved") == "on"
            try:
                with transaction.atomic(using=tenant_db):
                    grant = Grant.objects.using(tenant_db).create(
                        code=code,
                        title=title,
                        donor=donor,
                        project=project,
                        currency=currency,
                        award_amount=grant_ceiling,
                        grant_ceiling=grant_ceiling,
                        eligible_receivable_amount=eligible_receivable_amount,
                        receivable_basis_note=receivable_basis_note,
                        funding_modality=funding_modality,
                        funding_method=funding_method,
                        expense_report_approved=expense_report_approved,
                        audit_approved=audit_approved,
                        final_report_approved=final_report_approved,
                        status=status,
                    )
                    DonorAgreementGrant.objects.using(tenant_db).update_or_create(
                        grant=grant,
                        defaults={"agreement": donor_contract},
                    )
                    DonorAgreementProject.objects.using(tenant_db).update_or_create(
                        agreement=donor_contract,
                        project=project,
                    )
                    replace_grant_tranches_from_post(request, grant, tenant_db)
            except ValidationError as ve:
                for _f, msgs in getattr(ve, "error_dict", {}).items():
                    for m in msgs:
                        messages.error(request, m)
                if not getattr(ve, "error_dict", None):
                    for m in getattr(ve, "messages", []):
                        messages.error(request, m)
            else:
                messages.success(request, "Project grant created.")
                return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    donors = list(Donor.objects.using(tenant_db).filter(status="active").order_by("name"))
    donor_contracts = list(
        DonorAgreement.objects.using(tenant_db)
        .select_related("donor", "funding_source")
        .exclude(status=DonorAgreement.Status.CLOSED)
        .order_by("donor__name", "agreement_code")
    )
    projects = list(Project.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    currencies = list(Currency.objects.using(tenant_db).filter(status="active").order_by("code"))
    ctx["donors"] = donors
    ctx["donor_contracts"] = donor_contracts
    ctx["projects"] = projects
    ctx["currencies"] = currencies
    ctx["status_choices"] = Grant.Status.choices
    ctx["funding_method_choices"] = Grant.FundingMethod.choices
    import json

    fm_list = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    ctx["funding_modalities"] = fm_list
    ctx["funding_modality_type_map_json"] = json.dumps({str(m.pk): m.modality_type for m in fm_list})
    ctx["mixed_modality_value"] = FundingSource.ModalityType.MIXED_MODALITY
    ctx["grant_uses_mixed_modality"] = False
    ctx["tranche_payment_choices"] = GrantTranche.PaymentType.choices
    ctx["tranche_trigger_choices"] = GrantTranche.TriggerCondition.choices
    ctx["tranches"] = []
    ctx["form_title"] = "Add Project Grant"
    return render(request, "tenant_portal/setup/grant_dimension_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimension_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit grants.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import Currency
    from tenant_grants.models import (
        Donor,
        DonorAgreement,
        DonorAgreementGrant,
        DonorAgreementProject,
        FundingSource,
        Grant,
        GrantTranche,
        Project,
    )

    obj = get_object_or_404(
        Grant.objects.using(tenant_db).select_related(
            "donor", "project", "source_tracking", "bank_account", "funding_modality"
        ),
        pk=pk,
    )
    ctx["grant"] = obj

    if request.method == "POST":
        from decimal import Decimal, InvalidOperation

        from django.core.exceptions import ValidationError
        from django.db import transaction

        from tenant_grants.services.grant_tranches import replace_grant_tranches_from_post

        code = (request.POST.get("code") or "").strip()
        title = (request.POST.get("title") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        donor_contract_id = (request.POST.get("donor_contract_id") or "").strip()
        project_id = (request.POST.get("project_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        raw_ceiling = (request.POST.get("grant_ceiling") or request.POST.get("award_amount") or "").strip() or "0"
        raw_eligible = (request.POST.get("eligible_receivable_amount") or "").strip()
        receivable_basis_note = (request.POST.get("receivable_basis_note") or "").strip()
        status = (request.POST.get("status") or "").strip() or Grant.Status.DRAFT
        errors = []
        if not code:
            errors.append("Grant code is required.")
        if not title:
            errors.append("Grant name is required.")
        if not donor_id:
            errors.append("Donor is required.")
        if not donor_contract_id:
            errors.append("Donor contract is required.")
        if not project_id:
            errors.append("Grant must belong to a project.")
        if code and Grant.objects.using(tenant_db).filter(code__iexact=code).exclude(pk=pk).exists():
            errors.append("A grant with this code already exists.")
        grant_ceiling = Decimal("0")
        try:
            grant_ceiling = Decimal(raw_ceiling.replace(",", ""))
            if grant_ceiling <= 0:
                errors.append("Budget (grant ceiling) must be greater than zero.")
        except (InvalidOperation, ValueError):
            errors.append("Grant ceiling must be a valid number.")
        eligible_receivable_amount = grant_ceiling
        if raw_eligible:
            try:
                eligible_receivable_amount = Decimal(raw_eligible.replace(",", ""))
            except (InvalidOperation, ValueError):
                errors.append("Eligible receivable must be a valid number.")
        if not errors and eligible_receivable_amount > grant_ceiling:
            errors.append("Eligible receivable cannot exceed grant ceiling.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            donor = get_object_or_404(Donor.objects.using(tenant_db), pk=donor_id)
            donor_contract = get_object_or_404(DonorAgreement.objects.using(tenant_db), pk=donor_contract_id)
            if donor_contract.donor_id != donor.pk:
                messages.error(request, "Selected donor contract must belong to the selected donor.")
                return redirect(reverse("tenant_portal:setup_grant_dimension_edit", args=[pk]))
            project = get_object_or_404(Project.objects.using(tenant_db), pk=project_id)
            currency = Currency.objects.using(tenant_db).filter(pk=currency_id).first() if currency_id else None
            raw_fm = (request.POST.get("funding_method") or "").strip()
            valid_fm = {c[0] for c in Grant.FundingMethod.choices}
            funding_method = raw_fm if raw_fm in valid_fm else ""
            funding_modality_id = (request.POST.get("funding_modality_id") or "").strip()
            funding_modality = (
                FundingSource.objects.using(tenant_db)
                .filter(pk=int(funding_modality_id), is_active=True)
                .first()
                if funding_modality_id.isdigit()
                else None
            )
            if not funding_modality:
                messages.error(request, "Funding modality is required.")
                return redirect(reverse("tenant_portal:setup_grant_dimension_edit", args=[pk]))
            if funding_modality is not None:
                from tenant_grants.services.payment_modality import has_complete_gl_mapping

                if not has_complete_gl_mapping(
                    using=tenant_db,
                    funding_source=funding_modality,
                ):
                    messages.error(request, "GL account mapping is missing for selected funding modality.")
                    return redirect(reverse("tenant_portal:setup_grant_dimension_edit", args=[pk]))
            expense_report_approved = request.POST.get("expense_report_approved") == "on"
            audit_approved = request.POST.get("audit_approved") == "on"
            final_report_approved = request.POST.get("final_report_approved") == "on"
            try:
                with transaction.atomic(using=tenant_db):
                    obj.code = code
                    obj.title = title
                    obj.donor = donor
                    obj.project = project
                    obj.currency = currency
                    obj.grant_ceiling = grant_ceiling
                    obj.eligible_receivable_amount = eligible_receivable_amount
                    obj.receivable_basis_note = receivable_basis_note
                    obj.award_amount = grant_ceiling
                    obj.status = status
                    obj.funding_modality = funding_modality
                    obj.funding_method = funding_method
                    obj.expense_report_approved = expense_report_approved
                    obj.audit_approved = audit_approved
                    obj.final_report_approved = final_report_approved
                    obj.save(using=tenant_db)
                    DonorAgreementGrant.objects.using(tenant_db).update_or_create(
                        grant=obj,
                        defaults={"agreement": donor_contract},
                    )
                    DonorAgreementProject.objects.using(tenant_db).update_or_create(
                        agreement=donor_contract,
                        project=project,
                    )
                    replace_grant_tranches_from_post(request, obj, tenant_db)
            except ValidationError as ve:
                for _f, msgs in getattr(ve, "error_dict", {}).items():
                    for m in msgs:
                        messages.error(request, m)
                if not getattr(ve, "error_dict", None):
                    for m in getattr(ve, "messages", []):
                        messages.error(request, m)
            else:
                messages.success(request, "Project grant updated.")
                return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    donors = list(Donor.objects.using(tenant_db).filter(status="active").order_by("name"))
    donor_contracts = list(
        DonorAgreement.objects.using(tenant_db)
        .select_related("donor", "funding_source")
        .exclude(status=DonorAgreement.Status.CLOSED)
        .order_by("donor__name", "agreement_code")
    )
    projects = list(Project.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    currencies = list(Currency.objects.using(tenant_db).filter(status="active").order_by("code"))
    ctx["donors"] = donors
    ctx["donor_contracts"] = donor_contracts
    current_contract_link = (
        DonorAgreementGrant.objects.using(tenant_db)
        .filter(grant_id=obj.pk)
        .select_related("agreement")
        .first()
    )
    ctx["current_donor_contract_id"] = current_contract_link.agreement_id if current_contract_link else ""
    ctx["projects"] = projects
    ctx["currencies"] = currencies
    ctx["status_choices"] = Grant.Status.choices
    ctx["funding_method_choices"] = Grant.FundingMethod.choices
    import json

    fm_list = list(FundingSource.objects.using(tenant_db).filter(is_active=True).order_by("name"))
    ctx["funding_modalities"] = fm_list
    ctx["funding_modality_type_map_json"] = json.dumps({str(m.pk): m.modality_type for m in fm_list})
    ctx["mixed_modality_value"] = FundingSource.ModalityType.MIXED_MODALITY
    ctx["grant_uses_mixed_modality"] = bool(
        obj.funding_modality_id
        and obj.funding_modality.modality_type == FundingSource.ModalityType.MIXED_MODALITY
    )
    ctx["tranche_payment_choices"] = GrantTranche.PaymentType.choices
    ctx["tranche_trigger_choices"] = GrantTranche.TriggerCondition.choices
    ctx["tranches"] = list(obj.tranches.order_by("sort_order", "tranche_no"))
    ctx["form_title"] = "Edit Project Grant"
    return render(request, "tenant_portal/setup/grant_dimension_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_dimension_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete grants.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import JournalEntry
    from tenant_grants.models import Grant

    obj = get_object_or_404(Grant.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if JournalEntry.objects.using(tenant_db).filter(grant_id=pk).exists():
            messages.error(request, "Cannot delete: this grant is used in transactions.")
            return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))
        obj.delete(using=tenant_db)
        messages.success(request, "Grant deleted.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Grant {obj.code} — {obj.title}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_grant_dimensions_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_grant_dimension_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_mapping_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to add mappings.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import BankAccount, ChartAccount, CostCenter, Currency, ProjectDimensionMapping
    from tenant_grants.models import BudgetLine, Donor, Project

    if request.method == "POST":
        project_id = (request.POST.get("project_id") or "").strip()
        cost_center_id = (request.POST.get("cost_center_id") or "").strip()
        bank_account_id = (request.POST.get("bank_account_id") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        budget_line_id = (request.POST.get("budget_line_id") or "").strip()
        default_debit_account_id = (request.POST.get("default_debit_account_id") or "").strip()
        default_credit_account_id = (request.POST.get("default_credit_account_id") or "").strip()
        active_from = (request.POST.get("active_from") or "").strip() or None
        active_to = (request.POST.get("active_to") or "").strip() or None
        status = (request.POST.get("status") or "").strip() or ProjectDimensionMapping.Status.ACTIVE
        errors = []
        if not project_id:
            errors.append("Project is required.")
        if project_id and ProjectDimensionMapping.objects.using(tenant_db).filter(project_id=project_id).exists():
            errors.append("This project already has a mapping. Edit the existing one or delete it first.")
        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            project = get_object_or_404(Project.objects.using(tenant_db), pk=project_id)
            cost_center = CostCenter.objects.using(tenant_db).filter(pk=cost_center_id).first() if cost_center_id else None
            bank_account = BankAccount.objects.using(tenant_db).filter(pk=bank_account_id).first() if bank_account_id else None
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
            currency = Currency.objects.using(tenant_db).filter(pk=currency_id).first() if currency_id else None
            budget_line = BudgetLine.objects.using(tenant_db).filter(pk=budget_line_id).first() if budget_line_id else None
            default_debit_account = ChartAccount.objects.using(tenant_db).filter(pk=default_debit_account_id).first() if default_debit_account_id else None
            default_credit_account = ChartAccount.objects.using(tenant_db).filter(pk=default_credit_account_id).first() if default_credit_account_id else None
            obj = ProjectDimensionMapping(
                project=project,
                cost_center=cost_center,
                bank_account=bank_account,
                donor=donor,
                currency=currency,
                budget_line=budget_line,
                default_debit_account=default_debit_account,
                default_credit_account=default_credit_account,
                active_from=active_from,
                active_to=active_to,
                status=status,
            )
            obj.full_clean()
            obj.save(using=tenant_db)
            messages.success(request, "Project dimension mapping created.")
            return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    mapped_project_ids = set(
        ProjectDimensionMapping.objects.using(tenant_db).values_list("project_id", flat=True)
    )
    projects = [
        p
        for p in Project.objects.using(tenant_db).filter(is_active=True).order_by("code")
        if p.pk not in mapped_project_ids
    ]
    cost_centers = list(CostCenter.objects.using(tenant_db).filter(status="active").order_by("code"))
    bank_accounts = list(BankAccount.objects.using(tenant_db).filter(is_active=True).order_by("bank_name"))
    donors = list(Donor.objects.using(tenant_db).filter(status="active").order_by("name"))
    currencies = list(Currency.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    budget_lines = list(BudgetLine.objects.using(tenant_db).order_by("id")[:200])
    accounts = list(ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    ctx["projects"] = projects
    ctx["cost_centers"] = cost_centers
    ctx["bank_accounts"] = bank_accounts
    ctx["donors"] = donors
    ctx["currencies"] = currencies
    ctx["budget_lines"] = budget_lines
    ctx["accounts"] = accounts
    ctx["form_title"] = "Add Default Mapping"
    return render(request, "tenant_portal/setup/project_mapping_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_mapping_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_project_grant_dims"
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to edit mappings.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import BankAccount, ChartAccount, CostCenter, Currency, ProjectDimensionMapping
    from tenant_grants.models import BudgetLine, Donor, Project

    obj = get_object_or_404(
        ProjectDimensionMapping.objects.using(tenant_db).select_related("project", "cost_center", "bank_account", "donor"),
        pk=pk,
    )
    ctx["mapping"] = obj

    if request.method == "POST":
        cost_center_id = (request.POST.get("cost_center_id") or "").strip()
        bank_account_id = (request.POST.get("bank_account_id") or "").strip()
        donor_id = (request.POST.get("donor_id") or "").strip()
        currency_id = (request.POST.get("currency_id") or "").strip()
        budget_line_id = (request.POST.get("budget_line_id") or "").strip()
        default_debit_account_id = (request.POST.get("default_debit_account_id") or "").strip()
        default_credit_account_id = (request.POST.get("default_credit_account_id") or "").strip()
        active_from = (request.POST.get("active_from") or "").strip() or None
        active_to = (request.POST.get("active_to") or "").strip() or None
        status = (request.POST.get("status") or "").strip() or ProjectDimensionMapping.Status.ACTIVE
        cost_center = CostCenter.objects.using(tenant_db).filter(pk=cost_center_id).first() if cost_center_id else None
        bank_account = BankAccount.objects.using(tenant_db).filter(pk=bank_account_id).first() if bank_account_id else None
        donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first() if donor_id else None
        currency = Currency.objects.using(tenant_db).filter(pk=currency_id).first() if currency_id else None
        budget_line = BudgetLine.objects.using(tenant_db).filter(pk=budget_line_id).first() if budget_line_id else None
        default_debit_account = ChartAccount.objects.using(tenant_db).filter(pk=default_debit_account_id).first() if default_debit_account_id else None
        default_credit_account = ChartAccount.objects.using(tenant_db).filter(pk=default_credit_account_id).first() if default_credit_account_id else None
        obj.cost_center = cost_center
        obj.bank_account = bank_account
        obj.donor = donor
        obj.currency = currency
        obj.budget_line = budget_line
        obj.default_debit_account = default_debit_account
        obj.default_credit_account = default_credit_account
        obj.active_from = active_from
        obj.active_to = active_to
        obj.status = status
        obj.full_clean()
        obj.save(using=tenant_db)
        messages.success(request, "Project dimension mapping updated.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    cost_centers = list(CostCenter.objects.using(tenant_db).filter(status="active").order_by("code"))
    bank_accounts = list(BankAccount.objects.using(tenant_db).filter(is_active=True).order_by("bank_name"))
    donors = list(Donor.objects.using(tenant_db).filter(status="active").order_by("name"))
    currencies = list(Currency.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    budget_lines = list(BudgetLine.objects.using(tenant_db).order_by("id")[:200])
    accounts = list(ChartAccount.objects.using(tenant_db).filter(is_active=True).order_by("code"))
    ctx["cost_centers"] = cost_centers
    ctx["bank_accounts"] = bank_accounts
    ctx["donors"] = donors
    ctx["currencies"] = currencies
    ctx["budget_lines"] = budget_lines
    ctx["accounts"] = accounts
    ctx["form_title"] = "Edit Default Mapping"
    return render(request, "tenant_portal/setup/project_mapping_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_project_mapping_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete mappings.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    from tenant_finance.models import ProjectDimensionMapping

    obj = get_object_or_404(ProjectDimensionMapping.objects.using(tenant_db).select_related("project"), pk=pk)
    if request.method == "POST":
        obj.delete(using=tenant_db)
        messages.success(request, "Project dimension mapping deleted.")
        return redirect(reverse("tenant_portal:setup_grant_dimensions_list"))

    ctx["object"] = obj
    ctx["object_label"] = f"Mapping: {obj.project.code} → Cost Center / Bank Account / Donor"
    ctx["cancel_url"] = reverse("tenant_portal:setup_grant_dimensions_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_project_mapping_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


# ----- Currencies -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_currencies_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_currencies"
    ctx["page_title"] = "Currencies"

    from tenant_finance.models import Currency, ensure_default_currencies

    # Seed base currencies for this tenant if missing
    ensure_default_currencies(using=tenant_db)

    qs = Currency.objects.using(tenant_db).all().order_by("code")

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(models.Q(code__icontains=q) | models.Q(name__icontains=q))

    status = (request.GET.get("status") or "").strip()
    if status:
        qs = qs.filter(status=status)

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    page_obj = paginator.get_page(page_number)

    ctx["currencies_page"] = page_obj
    ctx["filter_q"] = q
    ctx["filter_status"] = status
    ctx["status_choices"] = Currency.Status.choices

    # Preserve other query params for pagination links
    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/currencies_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_currencies_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_currencies"
    ctx["page_title"] = "Add currency"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage currencies.")
        return redirect(reverse("tenant_portal:setup_currencies_list"))

    from tenant_finance.models import Currency

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip().upper()
        name = (request.POST.get("name") or "").strip()
        symbol = (request.POST.get("symbol") or "").strip()
        decimal_places_raw = (request.POST.get("decimal_places") or "").strip()
        status = (request.POST.get("status") or "").strip() or Currency.Status.ACTIVE

        errors: list[str] = []
        if not code:
            errors.append("Currency code is required.")
        if code and len(code) != 3:
            errors.append("Currency code should be a 3-letter ISO code (e.g. USD, EUR).")
        if decimal_places_raw:
            try:
                decimal_places = int(decimal_places_raw)
                if decimal_places < 0 or decimal_places > 6:
                    errors.append("Decimal places must be between 0 and 6.")
            except ValueError:
                errors.append("Decimal places must be a number.")
        else:
            decimal_places = 2

        if code and Currency.objects.using(tenant_db).filter(code__iexact=code).exists():
            errors.append("A currency with this code already exists.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            Currency.objects.using(tenant_db).create(
                code=code,
                name=name,
                symbol=symbol,
                decimal_places=decimal_places,
                status=status,
                created_by=request.tenant_user,
            )
            messages.success(request, "Currency created.")
            return redirect(reverse("tenant_portal:setup_currencies_list"))

        ctx["form_initial"] = {
            "code": code,
            "name": name,
            "symbol": symbol,
            "decimal_places": decimal_places_raw,
            "status": status,
        }

    from tenant_finance.models import Currency as CurrencyModel

    ctx["status_choices"] = CurrencyModel.Status.choices
    ctx["form_mode"] = "create"
    return render(request, "tenant_portal/setup/currency_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_currencies_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_currencies"
    ctx["page_title"] = "Edit currency"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage currencies.")
        return redirect(reverse("tenant_portal:setup_currencies_list"))

    from tenant_finance.models import Currency

    currency = get_object_or_404(Currency.objects.using(tenant_db), pk=pk)
    ctx["currency"] = currency

    if request.method == "POST":
        code = (request.POST.get("code") or "").strip().upper()
        name = (request.POST.get("name") or "").strip()
        symbol = (request.POST.get("symbol") or "").strip()
        decimal_places_raw = (request.POST.get("decimal_places") or "").strip()
        status = (request.POST.get("status") or "").strip() or Currency.Status.ACTIVE

        errors: list[str] = []
        if not code:
            errors.append("Currency code is required.")
        if code and len(code) != 3:
            errors.append("Currency code should be a 3-letter ISO code (e.g. USD, EUR).")

        if decimal_places_raw:
            try:
                decimal_places = int(decimal_places_raw)
                if decimal_places < 0 or decimal_places > 6:
                    errors.append("Decimal places must be between 0 and 6.")
            except ValueError:
                errors.append("Decimal places must be a number.")
        else:
            decimal_places = 2

        if (
            code
            and Currency.objects.using(tenant_db)
            .filter(code__iexact=code)
            .exclude(pk=pk)
            .exists()
        ):
            errors.append("A currency with this code already exists.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            currency.code = code
            currency.name = name
            currency.symbol = symbol
            currency.decimal_places = decimal_places
            currency.status = status
            currency.save(using=tenant_db)
            messages.success(request, "Currency updated.")
            return redirect(reverse("tenant_portal:setup_currencies_list"))

    ctx["form_initial"] = {
        "code": currency.code,
        "name": currency.name,
        "symbol": currency.symbol,
        "decimal_places": currency.decimal_places,
        "status": currency.status,
    }
    ctx["status_choices"] = Currency.Status.choices
    ctx["form_mode"] = "edit"
    return render(request, "tenant_portal/setup/currency_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_currencies_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage currencies.")
        return redirect(reverse("tenant_portal:setup_currencies_list"))

    from tenant_finance.models import Currency, OrganizationSettings

    currency = get_object_or_404(Currency.objects.using(tenant_db), pk=pk)

    # Prevent deactivating the default currency
    org_settings = OrganizationSettings.objects.using(tenant_db).first()
    if org_settings and org_settings.default_currency_id == currency.id:
        messages.error(request, "You cannot deactivate the default currency. Change the default first.")
        return redirect(reverse("tenant_portal:setup_currencies_list"))

    new_status = Currency.Status.INACTIVE if currency.status == Currency.Status.ACTIVE else Currency.Status.ACTIVE
    currency.status = new_status
    currency.save(using=tenant_db)
    if new_status == Currency.Status.ACTIVE:
        messages.success(request, f"Currency {currency.code} reactivated.")
    else:
        messages.success(request, f"Currency {currency.code} deactivated.")
    return redirect(reverse("tenant_portal:setup_currencies_list"))


# ----- Exchange rates -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_exchange_rates_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_exchange_rates"
    ctx["page_title"] = "Exchange rates"

    from tenant_finance.models import Currency, ExchangeRate

    qs = ExchangeRate.objects.using(tenant_db).select_related("currency", "base_currency")

    from_currency_code = (request.GET.get("from_currency") or "").strip().upper()
    date_str = (request.GET.get("date") or "").strip()
    status = (request.GET.get("status") or "").strip()
    sort = (request.GET.get("sort") or "").strip()

    if from_currency_code:
        qs = qs.filter(currency__code__iexact=from_currency_code)
    if date_str:
        from django.utils.dateparse import parse_date

        d = parse_date(date_str)
        if d:
            qs = qs.filter(effective_date=d)
    if status:
        qs = qs.filter(status=status)

    # Sorting
    if sort == "currency":
        qs = qs.order_by("currency__code", "-effective_date")
    elif sort == "date":
        qs = qs.order_by("-effective_date", "currency__code")
    elif sort == "status":
        qs = qs.order_by("status", "-effective_date", "currency__code")
    else:
        qs = qs.order_by("-effective_date", "currency__code")

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    page_obj = paginator.get_page(page_number)

    ctx["rates_page"] = page_obj
    ctx["filter_from_currency"] = from_currency_code
    ctx["filter_date"] = date_str
    ctx["filter_status"] = status
    ctx["current_sort"] = sort
    ctx["rate_type_choices"] = ExchangeRate.RateType.choices
    ctx["status_choices"] = ExchangeRate.Status.choices

    return render(request, "tenant_portal/setup/exchange_rates_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_exchange_rates_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_exchange_rates"
    ctx["page_title"] = "Add exchange rate"

    from tenant_finance.models import Currency, ExchangeRate, OrganizationSettings

    base_currency = (
        OrganizationSettings.objects.using(tenant_db)
        .select_related("default_currency")
        .first()
    )
    base_currency = base_currency.default_currency if base_currency else None

    currencies = list(
        Currency.objects.using(tenant_db)
        .filter(status=Currency.Status.ACTIVE)
        .order_by("code")
    )

    if request.method == "POST":
        from_code = (request.POST.get("from_currency") or "").strip().upper()
        rate = (request.POST.get("rate") or "").strip()
        eff_date = (request.POST.get("effective_date") or "").strip()
        rate_type = (request.POST.get("rate_type") or "").strip() or ExchangeRate.RateType.SPOT
        source = (request.POST.get("source") or "").strip() or ExchangeRate.Source.MANUAL
        status = (request.POST.get("status") or "").strip() or ExchangeRate.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []
        if not base_currency:
            errors.append("Base currency is not configured in Organization settings.")
        if not from_code:
            errors.append("From currency is required.")
        if not rate:
            errors.append("Exchange rate is required.")
        if not eff_date:
            errors.append("Effective date is required.")

        from decimal import Decimal, InvalidOperation
        from django.utils.dateparse import parse_date

        try:
            rate_val = Decimal(rate)
        except (InvalidOperation, ValueError):
            errors.append("Exchange rate must be a valid number.")
            rate_val = None  # type: ignore

        eff = parse_date(eff_date) if eff_date else None
        if not eff:
            errors.append("Effective date format is invalid.")

        currency = (
            Currency.objects.using(tenant_db).filter(code__iexact=from_code).first()
            if from_code
            else None
        )
        if not currency:
            errors.append("From currency not found.")
        elif base_currency and currency.id == base_currency.id:
            errors.append("From currency cannot be the same as the base currency.")

        if not errors and rate_val is not None and eff and currency and base_currency:
            # Prevent duplicates for same pair/date
            exists = ExchangeRate.objects.using(tenant_db).filter(
                currency=currency, base_currency=base_currency, effective_date=eff
            ).exists()
            if exists:
                errors.append("An exchange rate already exists for this currency pair and date.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rate_obj = ExchangeRate.objects.using(tenant_db).create(
                currency=currency,
                base_currency=base_currency,
                rate=rate_val,
                effective_date=eff,
                rate_type=rate_type,
                source=source,
                status=status,
                notes=notes,
            )
            from tenant_finance.models import AuditLog

            AuditLog.objects.using(tenant_db).create(
                model_name="exchangerate",
                object_id=rate_obj.id,
                action=AuditLog.Action.CREATE,
                user_id=getattr(request.tenant_user, "id", None),
                username=getattr(request.tenant_user, "email", "") or "",
                summary=(
                    f"Created exchange rate {rate_obj.currency.code}/{rate_obj.base_currency.code} "
                    f"on {rate_obj.effective_date} = {rate_obj.rate}"
                ),
            )
            messages.success(request, "Exchange rate created.")
            return redirect(reverse("tenant_portal:setup_exchange_rates_list"))

    ctx["currencies"] = currencies
    ctx["base_currency"] = base_currency
    ctx["rate_type_choices"] = ExchangeRate.RateType.choices
    ctx["source_choices"] = ExchangeRate.Source.choices
    ctx["status_choices"] = ExchangeRate.Status.choices
    ctx["form_title"] = "New exchange rate"
    return render(request, "tenant_portal/setup/exchange_rates_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_exchange_rates_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_exchange_rates"
    ctx["page_title"] = "Edit exchange rate"

    from decimal import Decimal, InvalidOperation
    from django.utils.dateparse import parse_date
    from tenant_finance.models import Currency, ExchangeRate, OrganizationSettings, FiscalPeriod, AuditLog

    rate_obj = get_object_or_404(
        ExchangeRate.objects.using(tenant_db).select_related("currency", "base_currency"), pk=pk
    )

    base_currency = rate_obj.base_currency
    currencies = list(
        Currency.objects.using(tenant_db)
        .filter(status=Currency.Status.ACTIVE)
        .order_by("code")
    )

    # Prevent editing rates in closed accounting periods
    closed_period = None
    if rate_obj.effective_date:
        closed_period = (
            FiscalPeriod.objects.using(tenant_db)
            .filter(
                start_date__lte=rate_obj.effective_date,
                end_date__gte=rate_obj.effective_date,
                is_closed=True,
            )
            .first()
        )

    if request.method == "POST" and not closed_period:
        from_code = (request.POST.get("from_currency") or "").strip().upper()
        rate = (request.POST.get("rate") or "").strip()
        eff_date = (request.POST.get("effective_date") or "").strip()
        rate_type = (request.POST.get("rate_type") or "").strip() or rate_obj.rate_type
        source = (request.POST.get("source") or "").strip() or rate_obj.source
        status = (request.POST.get("status") or "").strip() or rate_obj.status
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []
        if not from_code:
            errors.append("From currency is required.")
        if not rate:
            errors.append("Exchange rate is required.")
        if not eff_date:
            errors.append("Effective date is required.")

        try:
            rate_val = Decimal(rate)
        except (InvalidOperation, ValueError):
            errors.append("Exchange rate must be a valid number.")
            rate_val = None  # type: ignore

        eff = parse_date(eff_date) if eff_date else None
        if not eff:
            errors.append("Effective date format is invalid.")

        currency = (
            Currency.objects.using(tenant_db).filter(code__iexact=from_code).first()
            if from_code
            else None
        )
        if not currency:
            errors.append("From currency not found.")

        if (
            not errors
            and rate_val is not None
            and eff
            and currency
            and base_currency
            and (currency != rate_obj.currency or eff != rate_obj.effective_date)
        ):
            exists = ExchangeRate.objects.using(tenant_db).filter(
                currency=currency, base_currency=base_currency, effective_date=eff
            ).exclude(pk=pk)
            if exists.exists():
                errors.append("An exchange rate already exists for this currency pair and date.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rate_obj.currency = currency
            rate_obj.rate = rate_val
            rate_obj.effective_date = eff
            rate_obj.rate_type = rate_type
            rate_obj.source = source
            rate_obj.status = status
            rate_obj.notes = notes
            rate_obj.save(using=tenant_db)
            AuditLog.objects.using(tenant_db).create(
                model_name="exchangerate",
                object_id=rate_obj.id,
                action=AuditLog.Action.UPDATE,
                user_id=getattr(request.tenant_user, "id", None),
                username=getattr(request.tenant_user, "email", "") or "",
                summary=(
                    f"Updated exchange rate {rate_obj.currency.code}/{rate_obj.base_currency.code} "
                    f"on {rate_obj.effective_date} to {rate_obj.rate}"
                ),
            )
            messages.success(request, "Exchange rate updated.")
            return redirect(reverse("tenant_portal:setup_exchange_rates_list"))

    ctx["rate"] = rate_obj
    ctx["currencies"] = currencies
    ctx["base_currency"] = base_currency
    ctx["rate_type_choices"] = ExchangeRate.RateType.choices
    ctx["source_choices"] = ExchangeRate.Source.choices
    ctx["status_choices"] = ExchangeRate.Status.choices
    ctx["form_title"] = "Edit exchange rate"
    return render(request, "tenant_portal/setup/exchange_rates_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_exchange_rates_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    from tenant_finance.models import ExchangeRate, FiscalPeriod, AuditLog

    obj = get_object_or_404(ExchangeRate.objects.using(tenant_db), pk=pk)

    # Do not allow deleting rates in closed accounting periods
    closed_period = None
    if obj.effective_date:
        closed_period = (
            FiscalPeriod.objects.using(tenant_db)
            .filter(
                start_date__lte=obj.effective_date,
                end_date__gte=obj.effective_date,
                is_closed=True,
            )
            .first()
        )

    if request.method == "POST":
        if closed_period:
            messages.error(
                request,
                "Exchange rates for closed accounting periods cannot be deleted.",
            )
            return redirect(reverse("tenant_portal:setup_exchange_rates_list"))
        entry_label = f"{obj.currency.code}/{obj.base_currency.code} on {obj.effective_date}"
        obj_id = obj.id
        obj.delete(using=tenant_db)
        AuditLog.objects.using(tenant_db).create(
            model_name="exchangerate",
            object_id=obj_id,
            action=AuditLog.Action.DELETE,
            user_id=getattr(request.tenant_user, "id", None),
            username=getattr(request.tenant_user, "email", "") or "",
            summary=f"Deleted exchange rate {entry_label}",
        )
        messages.success(request, "Exchange rate deleted.")
        return redirect(reverse("tenant_portal:setup_exchange_rates_list"))

    ctx = _setup_context(request)
    ctx["object"] = obj
    ctx["object_label"] = (
        f"Exchange rate {obj.currency.code}/{obj.base_currency.code} on {obj.effective_date}"
    )
    ctx["cancel_url"] = reverse("tenant_portal:setup_exchange_rates_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_exchange_rates_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


# ----- Numbering -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_numbering_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_numbering"
    ctx["page_title"] = "Numbering & document series"

    from tenant_finance.models import DocumentSeries, FiscalYear

    qs = DocumentSeries.objects.using(tenant_db).select_related("fiscal_year").all()

    doc_type = (request.GET.get("document_type") or "").strip()
    fy_id = (request.GET.get("fiscal_year") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if doc_type:
        qs = qs.filter(document_type=doc_type)
    if fy_id:
        qs = qs.filter(fiscal_year_id=fy_id)
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("document_type", "prefix")

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["series_page"] = paginator.get_page(page_number)

    ctx["filter_document_type"] = doc_type
    ctx["filter_fiscal_year"] = fy_id
    ctx["filter_status"] = status

    ctx["document_type_choices"] = DocumentSeries.DocumentType.choices
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")

    # Preserve filters in pagination links
    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/numbering_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_numbering_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_documentseries_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_numbering"
    ctx["page_title"] = "Add document series"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage numbering series.")
        return redirect(reverse("tenant_portal:setup_numbering_list"))

    from datetime import date

    from tenant_finance.models import DocumentSeries, FiscalYear
    from tenant_finance.services.numbering import preview_number

    if request.method == "POST":
        document_type = (request.POST.get("document_type") or "").strip()
        prefix = (request.POST.get("prefix") or "").strip()
        start_number_raw = (request.POST.get("start_number") or "").strip()
        current_number_raw = (request.POST.get("current_number") or "").strip()
        number_format = (request.POST.get("number_format") or "").strip()
        fiscal_year_id = (request.POST.get("fiscal_year") or "").strip()
        reset_frequency = (request.POST.get("reset_frequency") or "").strip() or DocumentSeries.ResetFrequency.YEARLY
        status = (request.POST.get("status") or "").strip() or DocumentSeries.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []

        if not document_type:
            errors.append("Document type is required.")
        if not prefix:
            errors.append("Prefix is required.")

        try:
            start_number = int(start_number_raw or "1")
        except ValueError:
            errors.append("Starting number must be a valid integer.")
            start_number = 1

        try:
            current_number = int(current_number_raw or "0")
        except ValueError:
            errors.append("Current number must be a valid integer.")
            current_number = 0

        fiscal_year = None
        if fiscal_year_id:
            fiscal_year = (
                FiscalYear.objects.using(tenant_db).filter(pk=fiscal_year_id).first()
            )
            if not fiscal_year:
                errors.append("Selected fiscal year does not exist.")

        # Prevent duplicate series with same type, fiscal year and prefix.
        if (
            document_type
            and prefix
            and DocumentSeries.objects.using(tenant_db)
            .filter(
                document_type=document_type,
                fiscal_year=fiscal_year,
                prefix__iexact=prefix,
            )
            .exists()
        ):
            errors.append(
                "A series with this document type, fiscal year and prefix already exists."
            )

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj = DocumentSeries(
                document_type=document_type,
                prefix=prefix,
                start_number=start_number,
                current_number=current_number,
                number_format=number_format or "{prefix}{year}-{seq:05d}",
                fiscal_year=fiscal_year,
                reset_frequency=reset_frequency,
                status=status,
                notes=notes,
            )
            try:
                obj.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                obj.save(using=tenant_db)
                messages.success(request, "Document series created.")
                return redirect(reverse("tenant_portal:setup_numbering_list"))

    # Preview
    try:
        today = date.today()
        fmt = (request.POST.get("number_format") if request.method == "POST" else None) or "{prefix}{year}-{seq:05d}"
        pfx = (request.POST.get("prefix") if request.method == "POST" else None) or "PV-"
        start_raw = (request.POST.get("start_number") if request.method == "POST" else None) or "1"
        current_raw = (request.POST.get("current_number") if request.method == "POST" else None) or "0"
        start = int(start_raw)
        current = int(current_raw)
        next_seq = max(current, start - 1) + 1
        ctx["preview_value"] = preview_number(fmt=fmt, prefix=pfx, entry_date=today, seq=next_seq, fiscal_year="")
    except Exception:
        ctx["preview_value"] = ""

    ctx["document_type_choices"] = DocumentSeries.DocumentType.choices
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["form_initial"] = {
        "start_number": 1,
        "current_number": 0,
        "reset_frequency": DocumentSeries.ResetFrequency.YEARLY,
        "status": DocumentSeries.Status.ACTIVE,
        "number_format": "{prefix}{year}-{seq:05d}",
    }
    return render(request, "tenant_portal/setup/numbering_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_numbering_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_numbering"
    ctx["page_title"] = "Edit document series"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage numbering series.")
        return redirect(reverse("tenant_portal:setup_numbering_list"))

    from datetime import date

    from tenant_finance.models import DocumentSeries, FiscalYear
    from tenant_finance.services.numbering import preview_number

    obj = get_object_or_404(DocumentSeries.objects.using(tenant_db), pk=pk)
    ctx["series"] = obj

    if request.method == "POST":
        document_type = (request.POST.get("document_type") or "").strip()
        prefix = (request.POST.get("prefix") or "").strip()
        start_number_raw = (request.POST.get("start_number") or "").strip()
        current_number_raw = (request.POST.get("current_number") or "").strip()
        number_format = (request.POST.get("number_format") or "").strip()
        fiscal_year_id = (request.POST.get("fiscal_year") or "").strip()
        reset_frequency = (request.POST.get("reset_frequency") or "").strip() or DocumentSeries.ResetFrequency.YEARLY
        status = (request.POST.get("status") or "").strip() or DocumentSeries.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []

        if not document_type:
            errors.append("Document type is required.")
        if not prefix:
            errors.append("Prefix is required.")

        try:
            start_number = int(start_number_raw or "1")
        except ValueError:
            errors.append("Starting number must be a valid integer.")
            start_number = obj.start_number or 1

        try:
            current_number = int(current_number_raw or "0")
        except ValueError:
            errors.append("Current number must be a valid integer.")
            current_number = obj.current_number or 0

        fiscal_year = None
        if fiscal_year_id:
            fiscal_year = (
                FiscalYear.objects.using(tenant_db).filter(pk=fiscal_year_id).first()
            )
            if not fiscal_year:
                errors.append("Selected fiscal year does not exist.")

        # Prevent duplicate series with same type, fiscal year and prefix.
        if (
            document_type
            and prefix
            and DocumentSeries.objects.using(tenant_db)
            .filter(
                document_type=document_type,
                fiscal_year=fiscal_year,
                prefix__iexact=prefix,
            )
            .exclude(pk=obj.pk)
            .exists()
        ):
            errors.append(
                "A series with this document type, fiscal year and prefix already exists."
            )

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj.document_type = document_type
            obj.prefix = prefix
            obj.start_number = start_number
            obj.current_number = current_number
            obj.number_format = number_format or "{prefix}{year}-{seq:05d}"
            obj.fiscal_year = fiscal_year
            obj.reset_frequency = reset_frequency
            obj.status = status
            obj.notes = notes

            try:
                obj.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                obj.save(using=tenant_db)
                messages.success(request, "Document series updated.")
                return redirect(reverse("tenant_portal:setup_numbering_list"))

    # Preview
    try:
        today = date.today()
        fmt = (request.POST.get("number_format") if request.method == "POST" else None) or obj.number_format
        pfx = (request.POST.get("prefix") if request.method == "POST" else None) or obj.prefix
        start_raw = (request.POST.get("start_number") if request.method == "POST" else None) or str(obj.start_number or 1)
        current_raw = (request.POST.get("current_number") if request.method == "POST" else None) or str(obj.current_number or 0)
        start = int(start_raw)
        current = int(current_raw)
        next_seq = max(current, start - 1) + 1
        ctx["preview_value"] = preview_number(fmt=fmt, prefix=pfx, entry_date=today, seq=next_seq, fiscal_year=(obj.fiscal_year.name if obj.fiscal_year_id else ""))
    except Exception:
        ctx["preview_value"] = ""

    ctx["document_type_choices"] = DocumentSeries.DocumentType.choices
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["form_initial"] = {
        "document_type": obj.document_type,
        "prefix": obj.prefix,
        "start_number": obj.start_number,
        "current_number": obj.current_number,
        "number_format": obj.number_format,
        "fiscal_year_id": obj.fiscal_year_id,
        "reset_frequency": obj.reset_frequency,
        "status": obj.status,
        "notes": obj.notes,
    }
    return render(request, "tenant_portal/setup/numbering_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_numbering_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage numbering series.")
        return redirect(reverse("tenant_portal:setup_numbering_list"))

    from tenant_finance.models import DocumentSeries

    obj = get_object_or_404(DocumentSeries.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        # Toggle active/inactive to support deactivation without data loss.
        if obj.status == DocumentSeries.Status.ACTIVE:
            obj.status = DocumentSeries.Status.INACTIVE
            msg = "Document series deactivated."
        else:
            obj.status = DocumentSeries.Status.ACTIVE
            msg = "Document series activated."
        try:
            obj.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            obj.save(using=tenant_db, update_fields=["status"])
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_numbering_list"))

    # For non-POST requests, just redirect back to the list.
    return redirect(reverse("tenant_portal:setup_numbering_list"))


# ----- Grant compliance rules -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_compliance_rules_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_grantcompliancerule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_grant_compliance_rules"
    ctx["page_title"] = "Grant compliance rules"

    from tenant_finance.models import GrantComplianceRule

    qs = GrantComplianceRule.objects.using(tenant_db).select_related("donor", "grant", "project")

    status = (request.GET.get("status") or "").strip()
    donor_id = (request.GET.get("donor") or "").strip()
    grant_id = (request.GET.get("grant") or "").strip()
    eff_date = (request.GET.get("effective_date") or "").strip()

    if status:
        qs = qs.filter(status=status)
    if donor_id:
        qs = qs.filter(donor_id=donor_id)
    if grant_id:
        qs = qs.filter(grant_id=grant_id)
    if eff_date:
        from django.utils.dateparse import parse_date

        d = parse_date(eff_date)
        if d:
            qs = qs.filter(effective_from__lte=d, effective_to__gte=d)

    paginator = Paginator(qs.order_by("-effective_from", "name"), PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["rules_page"] = paginator.get_page(page_number)

    ctx["filter_status"] = status
    ctx["filter_donor"] = donor_id
    ctx["filter_grant"] = grant_id
    ctx["filter_effective_date"] = eff_date

    from tenant_grants.models import Donor, Grant

    ctx["status_choices"] = GrantComplianceRule.Status.choices
    ctx["mode_choices"] = GrantComplianceRule.Mode.choices
    ctx["donors"] = Donor.objects.using(tenant_db).all().order_by("name")
    ctx["grants"] = Grant.objects.using(tenant_db).all().order_by("code")

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/grant_compliance_rules_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_compliance_rules_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_grantcompliancerule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_grant_compliance_rules"
    ctx["page_title"] = "Add grant compliance rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage grant compliance rules.")
        return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    from tenant_finance.models import AccountCategory, GrantComplianceRule
    from tenant_grants.models import Donor, Grant, Project

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        donor_id = (request.POST.get("donor") or "").strip()
        grant_id = (request.POST.get("grant") or "").strip()
        effective_from = (request.POST.get("effective_from") or "").strip()
        effective_to = (request.POST.get("effective_to") or "").strip()
        reporting_start = (request.POST.get("reporting_period_start") or "").strip()
        reporting_end = (request.POST.get("reporting_period_end") or "").strip()
        reminder_days_raw = (request.POST.get("reminder_days_before_deadline") or "").strip()
        reporting_start = (request.POST.get("reporting_period_start") or "").strip()
        reporting_end = (request.POST.get("reporting_period_end") or "").strip()
        reminder_days_raw = (request.POST.get("reminder_days_before_deadline") or "").strip()
        max_admin_pct = (request.POST.get("maximum_admin_cost_percent") or "").strip()
        allowed_ids = request.POST.getlist("allowed_account_categories")
        disallowed_ids = request.POST.getlist("disallowed_account_categories")
        require_attachments = bool(request.POST.get("require_attachments"))
        require_procurement = bool(request.POST.get("require_procurement_compliance"))
        require_budget_check = bool(request.POST.get("require_budget_check"))
        allow_outside_period = bool(request.POST.get("allow_posting_outside_grant_period"))
        require_additional_approval = bool(request.POST.get("require_additional_approval"))
        additional_approval_role = (request.POST.get("additional_approval_role") or "").strip()
        mode = (request.POST.get("mode") or "").strip() or GrantComplianceRule.Mode.BLOCK
        status = (request.POST.get("status") or "").strip() or GrantComplianceRule.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []
        from django.utils.dateparse import parse_date

        if not name:
            errors.append("Rule name is required.")

        donor = None
        grant = None
        project = None
        if donor_id:
            from tenant_grants.models import Donor as DonorModel

            donor = DonorModel.objects.using(tenant_db).filter(pk=donor_id).first()
            if not donor:
                errors.append("Selected donor does not exist.")
        if grant_id:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            if not grant:
                errors.append("Selected grant does not exist.")
        project_id = (request.POST.get("project") or "").strip()
        if project_id:
            project = Project.objects.using(tenant_db).filter(pk=project_id).first()
            if not project:
                errors.append("Selected project does not exist.")

        start_date = parse_date(effective_from) if effective_from else None
        end_date = parse_date(effective_to) if effective_to else None
        if not start_date or not end_date:
            errors.append("Both effective from and effective to dates are required.")

        # Reporting period parsing
        reporting_start_date = parse_date(reporting_start) if reporting_start else None
        reporting_end_date = parse_date(reporting_end) if reporting_end else None

        max_admin = None
        if max_admin_pct:
            try:
                max_admin = Decimal(max_admin_pct)
            except Exception:
                errors.append("Maximum admin cost percentage must be a valid number.")

        # Reminder days parsing (default: 5 days)
        reminder_days = 5
        if reminder_days_raw:
            try:
                reminder_days = int(reminder_days_raw)
                if reminder_days <= 0:
                    errors.append("Reminder days before deadline must be a positive number.")
            except ValueError:
                errors.append("Reminder days before deadline must be a number.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rule = GrantComplianceRule(
                name=name,
                donor=donor,
                grant=grant,
                effective_from=start_date,
                effective_to=end_date,
                reporting_period_start=reporting_start_date,
                reporting_period_end=reporting_end_date,
                reminder_days_before_deadline=reminder_days,
                maximum_admin_cost_percent=max_admin,
                require_attachments=require_attachments,
                require_procurement_compliance=require_procurement,
                require_budget_check=require_budget_check,
                allow_posting_outside_grant_period=allow_outside_period,
                require_additional_approval=require_additional_approval,
                additional_approval_role=additional_approval_role,
                project=project,
                mode=mode,
                status=status,
                notes=notes,
            )
            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                if allowed_ids:
                    rule.allowed_account_categories.set(
                        AccountCategory.objects.using(tenant_db).filter(id__in=allowed_ids)
                    )
                if disallowed_ids:
                    rule.disallowed_account_categories.set(
                        AccountCategory.objects.using(tenant_db).filter(id__in=disallowed_ids)
                    )
                messages.success(request, "Grant compliance rule created.")
                return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    ctx["status_choices"] = GrantComplianceRule.Status.choices
    ctx["mode_choices"] = GrantComplianceRule.Mode.choices
    ctx["donors"] = Donor.objects.using(tenant_db).all().order_by("name")
    ctx["grants"] = Grant.objects.using(tenant_db).all().order_by("code")
    ctx["projects"] = Project.objects.using(tenant_db).all().order_by("code")
    ctx["account_categories"] = AccountCategory.objects.using(tenant_db).all().order_by("code")
    ctx["allowed_ids"] = []
    ctx["disallowed_ids"] = []
    ctx["overlap_categories"] = []
    ctx["form_initial"] = {
        "status": GrantComplianceRule.Status.ACTIVE,
        "require_budget_check": True,
        "reminder_days_before_deadline": 5,
    }
    return render(request, "tenant_portal/setup/grant_compliance_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_compliance_rules_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_grantcompliancerule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_grant_compliance_rules"
    ctx["page_title"] = "Edit grant compliance rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage grant compliance rules.")
        return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    from tenant_finance.models import AccountCategory, GrantComplianceRule
    from tenant_grants.models import Donor, Grant, Project

    rule = get_object_or_404(GrantComplianceRule.objects.using(tenant_db), pk=pk)
    ctx["rule"] = rule

    if request.method == "POST":
        from django.utils.dateparse import parse_date

        name = (request.POST.get("name") or "").strip()
        donor_id = (request.POST.get("donor") or "").strip()
        grant_id = (request.POST.get("grant") or "").strip()
        effective_from = (request.POST.get("effective_from") or "").strip()
        effective_to = (request.POST.get("effective_to") or "").strip()
        max_admin_pct = (request.POST.get("maximum_admin_cost_percent") or "").strip()
        allowed_ids = request.POST.getlist("allowed_account_categories")
        disallowed_ids = request.POST.getlist("disallowed_account_categories")
        require_attachments = bool(request.POST.get("require_attachments"))
        require_procurement = bool(request.POST.get("require_procurement_compliance"))
        require_budget_check = bool(request.POST.get("require_budget_check"))
        allow_outside_period = bool(request.POST.get("allow_posting_outside_grant_period"))
        require_additional_approval = bool(request.POST.get("require_additional_approval"))
        additional_approval_role = (request.POST.get("additional_approval_role") or "").strip()
        mode = (request.POST.get("mode") or "").strip() or GrantComplianceRule.Mode.BLOCK
        status = (request.POST.get("status") or "").strip() or GrantComplianceRule.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Rule name is required.")

        donor = None
        grant = None
        project = None
        if donor_id:
            donor = Donor.objects.using(tenant_db).filter(pk=donor_id).first()
            if not donor:
                errors.append("Selected donor does not exist.")
        if grant_id:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            if not grant:
                errors.append("Selected grant does not exist.")
        project_id = (request.POST.get("project") or "").strip()
        if project_id:
            project = Project.objects.using(tenant_db).filter(pk=project_id).first()
            if not project:
                errors.append("Selected project does not exist.")

        start_date = parse_date(effective_from) if effective_from else None
        end_date = parse_date(effective_to) if effective_to else None
        if not start_date or not end_date:
            errors.append("Both effective from and effective to dates are required.")

        max_admin = None
        if max_admin_pct:
            try:
                max_admin = Decimal(max_admin_pct)
            except Exception:
                errors.append("Maximum admin cost percentage must be a valid number.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rule.name = name
            rule.donor = donor
            rule.grant = grant
            rule.effective_from = start_date
            rule.effective_to = end_date
            rule.reporting_period_start = reporting_start_date
            rule.reporting_period_end = reporting_end_date
            rule.reminder_days_before_deadline = reminder_days
            rule.maximum_admin_cost_percent = max_admin
            rule.require_attachments = require_attachments
            rule.require_procurement_compliance = require_procurement
            rule.require_budget_check = require_budget_check
            rule.allow_posting_outside_grant_period = allow_outside_period
            rule.require_additional_approval = require_additional_approval
            rule.additional_approval_role = additional_approval_role
            rule.project = project
            rule.mode = mode
            rule.status = status
            rule.notes = notes

            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                rule.allowed_account_categories.set(
                    AccountCategory.objects.using(tenant_db).filter(id__in=allowed_ids)
                )
                rule.disallowed_account_categories.set(
                    AccountCategory.objects.using(tenant_db).filter(id__in=disallowed_ids)
                )
                messages.success(request, "Grant compliance rule updated.")
                return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    ctx["status_choices"] = GrantComplianceRule.Status.choices
    ctx["mode_choices"] = GrantComplianceRule.Mode.choices
    ctx["donors"] = Donor.objects.using(tenant_db).all().order_by("name")
    ctx["grants"] = Grant.objects.using(tenant_db).all().order_by("code")
    ctx["projects"] = Project.objects.using(tenant_db).all().order_by("code")
    account_categories = AccountCategory.objects.using(tenant_db).all().order_by("code")
    ctx["account_categories"] = account_categories
    allowed_ids = list(rule.allowed_account_categories.values_list("id", flat=True))
    disallowed_ids = list(rule.disallowed_account_categories.values_list("id", flat=True))
    ctx["allowed_ids"] = allowed_ids
    ctx["disallowed_ids"] = disallowed_ids
    overlap_ids = set(allowed_ids) & set(disallowed_ids)
    if overlap_ids:
        ctx["overlap_categories"] = [
            c for c in account_categories if c.id in overlap_ids
        ]
    else:
        ctx["overlap_categories"] = []
    ctx["form_initial"] = {
        "name": rule.name,
        "donor_id": rule.donor_id,
        "grant_id": rule.grant_id,
        "project_id": rule.project_id,
        "effective_from": rule.effective_from,
        "effective_to": rule.effective_to,
        "maximum_admin_cost_percent": rule.maximum_admin_cost_percent,
        "require_attachments": rule.require_attachments,
        "require_procurement_compliance": rule.require_procurement_compliance,
        "require_budget_check": rule.require_budget_check,
        "allow_posting_outside_grant_period": rule.allow_posting_outside_grant_period,
        "require_additional_approval": rule.require_additional_approval,
        "additional_approval_role": rule.additional_approval_role,
        "mode": rule.mode,
        "status": rule.status,
        "notes": rule.notes,
        "reporting_period_start": rule.reporting_period_start,
        "reporting_period_end": rule.reporting_period_end,
        "reminder_days_before_deadline": rule.reminder_days_before_deadline,
    }
    return render(request, "tenant_portal/setup/grant_compliance_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_grant_compliance_rules_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage grant compliance rules.")
        return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    from tenant_finance.models import GrantComplianceRule

    rule = get_object_or_404(GrantComplianceRule.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        if rule.status == GrantComplianceRule.Status.ACTIVE:
            rule.status = GrantComplianceRule.Status.INACTIVE
            msg = "Grant compliance rule deactivated."
        else:
            rule.status = GrantComplianceRule.Status.ACTIVE
            msg = "Grant compliance rule activated."
        try:
            rule.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            rule.save(using=tenant_db, update_fields=["status"])
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))

    return redirect(reverse("tenant_portal:setup_grant_compliance_rules_list"))


# ----- Inter-fund transfer rules -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_interfund_rules_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_interfundtransferrule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_interfund_rules"
    ctx["page_title"] = "Inter-fund transfer rules"

    from tenant_finance.models import InterFundTransferRule

    qs = InterFundTransferRule.objects.using(tenant_db).select_related("transfer_account")

    status = (request.GET.get("status") or "").strip()
    from_type = (request.GET.get("from_fund_type") or "").strip()
    to_type = (request.GET.get("to_fund_type") or "").strip()

    if status:
        qs = qs.filter(status=status)
    if from_type:
        qs = qs.filter(from_fund_type=from_type)
    if to_type:
        qs = qs.filter(to_fund_type=to_type)

    paginator = Paginator(qs.order_by("-effective_from", "name"), PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["rules_page"] = paginator.get_page(page_number)

    from tenant_finance.models import InterFundTransferRule as RuleModel

    ctx["status_choices"] = RuleModel.Status.choices
    ctx["fund_type_choices"] = RuleModel.FundType.choices

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/interfund_rules_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_interfund_rules_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_interfundtransferrule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_interfund_rules"
    ctx["page_title"] = "Add inter-fund transfer rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage inter-fund transfer rules.")
        return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    from decimal import Decimal
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, InterFundTransferRule

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        from_type = (request.POST.get("from_fund_type") or "").strip()
        to_type = (request.POST.get("to_fund_type") or "").strip()
        specific_from = (request.POST.get("specific_from_fund_code") or "").strip()
        specific_to = (request.POST.get("specific_to_fund_code") or "").strip()
        allow_transfer = bool(request.POST.get("allow_transfer"))
        require_approval = bool(request.POST.get("require_approval"))
        approval_role = (request.POST.get("approval_role") or "").strip()
        require_reason = bool(request.POST.get("require_reason"))
        max_amount_raw = (request.POST.get("maximum_transfer_amount") or "").strip()
        transfer_account_id = (request.POST.get("transfer_account") or "").strip()
        effective_from = (request.POST.get("effective_from") or "").strip()
        effective_to = (request.POST.get("effective_to") or "").strip()
        status = (request.POST.get("status") or "").strip() or InterFundTransferRule.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Rule name is required.")
        if not from_type or not to_type:
            errors.append("From fund type and To fund type are required.")
        if transfer_account_id:
            transfer_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=transfer_account_id).first()
            )
            if not transfer_account:
                errors.append("Selected transfer account does not exist.")
        else:
            transfer_account = None
            errors.append("Transfer / clearing account is required.")

        start_date = parse_date(effective_from) if effective_from else None
        end_date = parse_date(effective_to) if effective_to else None
        if not start_date or not end_date:
            errors.append("Both effective from and effective to dates are required.")

        max_amount = None
        if max_amount_raw:
            try:
                max_amount = Decimal(max_amount_raw)
            except Exception:
                errors.append("Maximum transfer amount must be a valid number.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rule = InterFundTransferRule(
                name=name,
                from_fund_type=from_type,
                to_fund_type=to_type,
                specific_from_fund_code=specific_from,
                specific_to_fund_code=specific_to,
                allow_transfer=allow_transfer,
                require_approval=require_approval,
                approval_role=approval_role,
                require_reason=require_reason,
                maximum_transfer_amount=max_amount,
                transfer_account=transfer_account,
                effective_from=start_date,
                effective_to=end_date,
                status=status,
                notes=notes,
            )
            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                messages.success(request, "Inter-fund transfer rule created.")
                return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    from tenant_finance.models import InterFundTransferRule as RuleModel, ChartAccount

    ctx["status_choices"] = RuleModel.Status.choices
    ctx["fund_type_choices"] = RuleModel.FundType.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "allow_transfer": True,
        "require_approval": True,
        "require_reason": True,
        "status": RuleModel.Status.ACTIVE,
    }
    return render(request, "tenant_portal/setup/interfund_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_interfund_rules_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    _ensure_interfundtransferrule_schema(tenant_db)
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_interfund_rules"
    ctx["page_title"] = "Edit inter-fund transfer rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage inter-fund transfer rules.")
        return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    from decimal import Decimal
    from django.utils.dateparse import parse_date

    from tenant_finance.models import ChartAccount, InterFundTransferRule

    rule = get_object_or_404(InterFundTransferRule.objects.using(tenant_db), pk=pk)
    ctx["rule"] = rule

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        from_type = (request.POST.get("from_fund_type") or "").strip()
        to_type = (request.POST.get("to_fund_type") or "").strip()
        specific_from = (request.POST.get("specific_from_fund_code") or "").strip()
        specific_to = (request.POST.get("specific_to_fund_code") or "").strip()
        allow_transfer = bool(request.POST.get("allow_transfer"))
        require_approval = bool(request.POST.get("require_approval"))
        approval_role = (request.POST.get("approval_role") or "").strip()
        require_reason = bool(request.POST.get("require_reason"))
        max_amount_raw = (request.POST.get("maximum_transfer_amount") or "").strip()
        transfer_account_id = (request.POST.get("transfer_account") or "").strip()
        effective_from = (request.POST.get("effective_from") or "").strip()
        effective_to = (request.POST.get("effective_to") or "").strip()
        status = (request.POST.get("status") or "").strip() or InterFundTransferRule.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Rule name is required.")
        if not from_type or not to_type:
            errors.append("From fund type and To fund type are required.")

        if transfer_account_id:
            transfer_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=transfer_account_id).first()
            )
            if not transfer_account:
                errors.append("Selected transfer account does not exist.")
        else:
            transfer_account = None
            errors.append("Transfer / clearing account is required.")

        start_date = parse_date(effective_from) if effective_from else None
        end_date = parse_date(effective_to) if effective_to else None
        if not start_date or not end_date:
            errors.append("Both effective from and effective to dates are required.")

        max_amount = None
        if max_amount_raw:
            try:
                max_amount = Decimal(max_amount_raw)
            except Exception:
                errors.append("Maximum transfer amount must be a valid number.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rule.name = name
            rule.from_fund_type = from_type
            rule.to_fund_type = to_type
            rule.specific_from_fund_code = specific_from
            rule.specific_to_fund_code = specific_to
            rule.allow_transfer = allow_transfer
            rule.require_approval = require_approval
            rule.approval_role = approval_role
            rule.require_reason = require_reason
            rule.maximum_transfer_amount = max_amount
            rule.transfer_account = transfer_account
            rule.effective_from = start_date
            rule.effective_to = end_date
            rule.status = status
            rule.notes = notes

            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                messages.success(request, "Inter-fund transfer rule updated.")
                return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    from tenant_finance.models import InterFundTransferRule as RuleModel, ChartAccount

    ctx["status_choices"] = RuleModel.Status.choices
    ctx["fund_type_choices"] = RuleModel.FundType.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "name": rule.name,
        "from_fund_type": rule.from_fund_type,
        "to_fund_type": rule.to_fund_type,
        "specific_from_fund_code": rule.specific_from_fund_code,
        "specific_to_fund_code": rule.specific_to_fund_code,
        "allow_transfer": rule.allow_transfer,
        "require_approval": rule.require_approval,
        "approval_role": rule.approval_role,
        "require_reason": rule.require_reason,
        "maximum_transfer_amount": rule.maximum_transfer_amount,
        "transfer_account_id": rule.transfer_account_id,
        "effective_from": rule.effective_from,
        "effective_to": rule.effective_to,
        "status": rule.status,
        "notes": rule.notes,
    }
    return render(request, "tenant_portal/setup/interfund_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_interfund_rules_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage inter-fund transfer rules.")
        return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    from tenant_finance.models import InterFundTransferRule

    rule = get_object_or_404(InterFundTransferRule.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        if rule.status == InterFundTransferRule.Status.ACTIVE:
            rule.status = InterFundTransferRule.Status.INACTIVE
            msg = "Inter-fund transfer rule deactivated."
        else:
            rule.status = InterFundTransferRule.Status.ACTIVE
            msg = "Inter-fund transfer rule activated."
        try:
            rule.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            rule.save(using=tenant_db, update_fields=["status"])
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_interfund_rules_list"))

    return redirect(reverse("tenant_portal:setup_interfund_rules_list"))


# ----- Voucher numbering -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_voucher_numbering_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_voucher_numbering"
    ctx["page_title"] = "Voucher numbering"

    from tenant_finance.models import DocumentSeries, FiscalYear
    from tenant_grants.models import Project, Grant

    qs = (
        DocumentSeries.objects.using(tenant_db)
        .select_related("fiscal_year", "project", "grant")
        .exclude(document_type=DocumentSeries.DocumentType.JOURNAL)
        .all()
    )

    doc_type = (request.GET.get("document_type") or "").strip()
    scope = (request.GET.get("scope") or "").strip()
    fy_id = (request.GET.get("fiscal_year") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if doc_type:
        qs = qs.filter(document_type=doc_type)
    if scope:
        qs = qs.filter(scope=scope)
    if fy_id:
        qs = qs.filter(fiscal_year_id=fy_id)
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("document_type", "scope", "prefix")

    paginator = Paginator(qs, PAGE_SIZE)
    ctx["series_page"] = paginator.get_page(request.GET.get("page", "1"))

    ctx["document_type_choices"] = [
        c for c in DocumentSeries.DocumentType.choices if c[0] != DocumentSeries.DocumentType.JOURNAL
    ]
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["scope_choices"] = DocumentSeries.Scope.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["projects"] = Project.objects.using(tenant_db).order_by("code")[:200]
    ctx["grants"] = Grant.objects.using(tenant_db).order_by("code")[:200]

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    ctx["filter_document_type"] = doc_type
    ctx["filter_scope"] = scope
    ctx["filter_fiscal_year"] = fy_id
    ctx["filter_status"] = status

    return render(request, "tenant_portal/setup/voucher_numbering_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_voucher_numbering_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_voucher_numbering"
    ctx["page_title"] = "Add voucher numbering series"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage voucher numbering.")
        return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))

    from tenant_finance.models import DocumentSeries, FiscalYear
    from tenant_grants.models import Project, Grant

    if request.method == "POST":
        document_type = (request.POST.get("document_type") or "").strip()
        prefix = (request.POST.get("prefix") or "").strip()
        number_format = (request.POST.get("number_format") or "").strip()
        start_number_raw = (request.POST.get("start_number") or "").strip()
        scope = (request.POST.get("scope") or "").strip() or DocumentSeries.Scope.GLOBAL
        project_id = (request.POST.get("project") or "").strip()
        grant_id = (request.POST.get("grant") or "").strip()
        fiscal_year_id = (request.POST.get("fiscal_year") or "").strip()
        reset_frequency = (request.POST.get("reset_frequency") or "").strip() or DocumentSeries.ResetFrequency.YEARLY
        status = (request.POST.get("status") or "").strip() or DocumentSeries.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []
        if not document_type:
            errors.append("Document type is required.")
        if document_type == DocumentSeries.DocumentType.JOURNAL:
            errors.append("Use Numbering & Document Series for journals. Voucher numbering excludes journals.")
        if not prefix:
            errors.append("Prefix is required.")

        try:
            start_number = int(start_number_raw or "1")
        except ValueError:
            errors.append("Starting number must be a valid integer.")
            start_number = 1

        fy = None
        if fiscal_year_id:
            fy = FiscalYear.objects.using(tenant_db).filter(pk=fiscal_year_id).first()
            if not fy:
                errors.append("Selected fiscal year does not exist.")

        project = None
        grant = None
        if scope == DocumentSeries.Scope.PROJECT:
            project = Project.objects.using(tenant_db).filter(pk=project_id).first()
            if not project:
                errors.append("Project is required for project-scoped numbering.")
        if scope == DocumentSeries.Scope.GRANT:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            if not grant:
                errors.append("Grant is required for grant-scoped numbering.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj = DocumentSeries(
                document_type=document_type,
                prefix=prefix,
                start_number=start_number,
                current_number=0,
                number_format=number_format or "{prefix}{year}-{seq:05d}",
                fiscal_year=fy,
                reset_frequency=reset_frequency,
                status=status,
                notes=notes,
                scope=scope,
                project=project,
                grant=grant,
            )
            try:
                obj.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                obj.save(using=tenant_db)
                from tenant_finance.models import AuditLog
                AuditLog.objects.using(tenant_db).create(
                    model_name="documentseries",
                    object_id=obj.id,
                    action=AuditLog.Action.CREATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
                    summary=f"Voucher numbering series created: {obj.get_document_type_display()} ({obj.scope}) {obj.prefix}",
                )
                messages.success(request, "Voucher numbering series created.")
                return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))

    ctx["document_type_choices"] = [
        c for c in DocumentSeries.DocumentType.choices if c[0] != DocumentSeries.DocumentType.JOURNAL
    ]
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["scope_choices"] = DocumentSeries.Scope.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["projects"] = Project.objects.using(tenant_db).order_by("code")[:200]
    ctx["grants"] = Grant.objects.using(tenant_db).order_by("code")[:200]
    ctx["form_initial"] = {
        "start_number": 1,
        "reset_frequency": DocumentSeries.ResetFrequency.YEARLY,
        "status": DocumentSeries.Status.ACTIVE,
        "number_format": "{prefix}{year}-{seq:05d}",
        "scope": DocumentSeries.Scope.GLOBAL,
    }
    return render(request, "tenant_portal/setup/voucher_numbering_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_voucher_numbering_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_voucher_numbering"
    ctx["page_title"] = "Edit voucher numbering series"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage voucher numbering.")
        return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))

    from tenant_finance.models import DocumentSeries, FiscalYear, AuditLog
    from tenant_grants.models import Project, Grant

    obj = get_object_or_404(DocumentSeries.objects.using(tenant_db), pk=pk)
    if obj.document_type == DocumentSeries.DocumentType.JOURNAL:
        messages.error(request, "Use Numbering & Document Series to edit journal series.")
        return redirect(reverse("tenant_portal:setup_numbering_edit", args=[pk]))

    before = {
        "document_type": obj.document_type,
        "prefix": obj.prefix,
        "number_format": obj.number_format,
        "start_number": obj.start_number,
        "current_number": obj.current_number,
        "fiscal_year_id": obj.fiscal_year_id,
        "reset_frequency": obj.reset_frequency,
        "status": obj.status,
        "scope": obj.scope,
        "project_id": obj.project_id,
        "grant_id": obj.grant_id,
    }

    if request.method == "POST":
        document_type = (request.POST.get("document_type") or "").strip()
        prefix = (request.POST.get("prefix") or "").strip()
        number_format = (request.POST.get("number_format") or "").strip()
        start_number_raw = (request.POST.get("start_number") or "").strip()
        current_number_raw = (request.POST.get("current_number") or "").strip()
        scope = (request.POST.get("scope") or "").strip() or DocumentSeries.Scope.GLOBAL
        project_id = (request.POST.get("project") or "").strip()
        grant_id = (request.POST.get("grant") or "").strip()
        fiscal_year_id = (request.POST.get("fiscal_year") or "").strip()
        reset_frequency = (request.POST.get("reset_frequency") or "").strip() or DocumentSeries.ResetFrequency.YEARLY
        status = (request.POST.get("status") or "").strip() or DocumentSeries.Status.ACTIVE
        notes = (request.POST.get("notes") or "").strip()

        errors: list[str] = []
        if not document_type:
            errors.append("Document type is required.")
        if document_type == DocumentSeries.DocumentType.JOURNAL:
            errors.append("Voucher numbering excludes journals.")
        if not prefix:
            errors.append("Prefix is required.")

        try:
            start_number = int(start_number_raw or "1")
        except ValueError:
            errors.append("Starting number must be a valid integer.")
            start_number = obj.start_number or 1

        try:
            current_number = int(current_number_raw or "0")
        except ValueError:
            errors.append("Current number must be a valid integer.")
            current_number = obj.current_number or 0

        fy = None
        if fiscal_year_id:
            fy = FiscalYear.objects.using(tenant_db).filter(pk=fiscal_year_id).first()
            if not fy:
                errors.append("Selected fiscal year does not exist.")

        project = None
        grant = None
        if scope == DocumentSeries.Scope.PROJECT:
            project = Project.objects.using(tenant_db).filter(pk=project_id).first()
            if not project:
                errors.append("Project is required for project-scoped numbering.")
        if scope == DocumentSeries.Scope.GRANT:
            grant = Grant.objects.using(tenant_db).filter(pk=grant_id).first()
            if not grant:
                errors.append("Grant is required for grant-scoped numbering.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            obj.document_type = document_type
            obj.prefix = prefix
            obj.number_format = number_format or "{prefix}{year}-{seq:05d}"
            obj.start_number = start_number
            obj.current_number = current_number
            obj.fiscal_year = fy
            obj.reset_frequency = reset_frequency
            obj.status = status
            obj.notes = notes
            obj.scope = scope
            obj.project = project
            obj.grant = grant
            try:
                obj.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                obj.save(using=tenant_db)
                AuditLog.objects.using(tenant_db).create(
                    model_name="documentseries",
                    object_id=obj.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
                    summary=f"Voucher numbering series updated: {obj.get_document_type_display()} ({obj.scope}) {obj.prefix}",
                    old_data=before,
                    new_data={
                        "document_type": obj.document_type,
                        "prefix": obj.prefix,
                        "number_format": obj.number_format,
                        "start_number": obj.start_number,
                        "current_number": obj.current_number,
                        "fiscal_year_id": obj.fiscal_year_id,
                        "reset_frequency": obj.reset_frequency,
                        "status": obj.status,
                        "scope": obj.scope,
                        "project_id": obj.project_id,
                        "grant_id": obj.grant_id,
                    },
                )
                messages.success(request, "Voucher numbering series updated.")
                return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))

    ctx["document_type_choices"] = [
        c for c in DocumentSeries.DocumentType.choices if c[0] != DocumentSeries.DocumentType.JOURNAL
    ]
    ctx["status_choices"] = DocumentSeries.Status.choices
    ctx["reset_choices"] = DocumentSeries.ResetFrequency.choices
    ctx["scope_choices"] = DocumentSeries.Scope.choices
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["projects"] = Project.objects.using(tenant_db).order_by("code")[:200]
    ctx["grants"] = Grant.objects.using(tenant_db).order_by("code")[:200]
    ctx["form_initial"] = {
        "document_type": obj.document_type,
        "prefix": obj.prefix,
        "start_number": obj.start_number,
        "current_number": obj.current_number,
        "number_format": obj.number_format,
        "fiscal_year_id": obj.fiscal_year_id,
        "reset_frequency": obj.reset_frequency,
        "status": obj.status,
        "notes": obj.notes,
        "scope": obj.scope,
        "project_id": obj.project_id,
        "grant_id": obj.grant_id,
    }
    ctx["series"] = obj
    return render(request, "tenant_portal/setup/voucher_numbering_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_voucher_numbering_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage voucher numbering.")
        return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))

    from tenant_finance.models import DocumentSeries, AuditLog

    obj = get_object_or_404(DocumentSeries.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if obj.status == DocumentSeries.Status.ACTIVE:
            obj.status = DocumentSeries.Status.INACTIVE
            msg = "Voucher numbering series deactivated."
        else:
            obj.status = DocumentSeries.Status.ACTIVE
            msg = "Voucher numbering series activated."
        obj.save(using=tenant_db, update_fields=["status"])
        AuditLog.objects.using(tenant_db).create(
            model_name="documentseries",
            object_id=obj.id,
            action=AuditLog.Action.UPDATE,
            user_id=request.tenant_user.id if request.tenant_user else None,
            username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
            summary=f"Voucher numbering series status changed: {obj.get_document_type_display()} → {obj.status}",
        )
        messages.success(request, msg)
    return redirect(reverse("tenant_portal:setup_voucher_numbering_list"))


# ----- Posting rules -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_rules_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_rules"
    ctx["page_title"] = "Posting rules"

    from tenant_finance.models import PostingRule

    qs = PostingRule.objects.using(tenant_db).select_related("debit_account", "credit_account")

    tx_type = (request.GET.get("transaction_type") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if tx_type:
        qs = qs.filter(transaction_type=tx_type)
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("transaction_type", "name")

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["rules_page"] = paginator.get_page(page_number)

    from tenant_finance.models import PostingRule as RuleModel

    ctx["transaction_type_choices"] = RuleModel.TransactionType.choices
    ctx["status_choices"] = RuleModel.Status.choices
    ctx["dimension_choices"] = RuleModel.Dimension.choices

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/posting_rules_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_rules_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_rules"
    ctx["page_title"] = "Add posting rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage posting rules.")
        return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    from tenant_finance.models import ChartAccount, PostingRule

    if request.method == "POST":
        import json

        name = (request.POST.get("name") or "").strip()
        transaction_type = (request.POST.get("transaction_type") or "").strip()
        debit_account_id = (request.POST.get("debit_account") or "").strip()
        credit_account_id = (request.POST.get("credit_account") or "").strip()
        apply_dimension = (request.POST.get("apply_dimension") or "").strip() or PostingRule.Dimension.NONE
        status = (request.POST.get("status") or "").strip() or PostingRule.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()
        priority_raw = (request.POST.get("priority") or "").strip() or "100"
        conditions_raw = (request.POST.get("conditions") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Rule name is required.")
        if not transaction_type:
            errors.append("Transaction type is required.")
        try:
            priority = int(priority_raw)
        except ValueError:
            errors.append("Priority must be a whole number.")
            priority = 100

        conditions = {}
        if conditions_raw:
            try:
                conditions = json.loads(conditions_raw)
                if not isinstance(conditions, dict):
                    errors.append("Conditions must be a JSON object.")
                    conditions = {}
            except Exception:
                errors.append("Conditions must be valid JSON (object).")
                conditions = {}

        debit_account = None
        credit_account = None
        if debit_account_id:
            debit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=debit_account_id).first()
            )
            if not debit_account:
                errors.append("Selected debit account does not exist.")
        else:
            errors.append("Debit account is required.")

        if credit_account_id:
            credit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=credit_account_id).first()
            )
            if not credit_account:
                errors.append("Selected credit account does not exist.")
        else:
            errors.append("Credit account is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            rule = PostingRule(
                name=name,
                transaction_type=transaction_type,
                debit_account=debit_account,
                credit_account=credit_account,
                apply_dimension=apply_dimension,
                status=status,
                description=description,
                priority=priority,
                conditions=conditions,
            )
            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                from tenant_finance.models import AuditLog
                AuditLog.objects.using(tenant_db).create(
                    model_name="postingrule",
                    object_id=rule.id,
                    action=AuditLog.Action.CREATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
                    summary=f"Posting rule created: {rule.name} ({rule.transaction_type})",
                    new_data={"priority": rule.priority, "conditions": rule.conditions},
                )
                messages.success(request, "Posting rule created.")
                return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    from tenant_finance.models import PostingRule as RuleModel

    ctx["transaction_type_choices"] = RuleModel.TransactionType.choices
    ctx["status_choices"] = RuleModel.Status.choices
    ctx["dimension_choices"] = RuleModel.Dimension.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "status": RuleModel.Status.ACTIVE,
        "apply_dimension": RuleModel.Dimension.NONE,
        "priority": 100,
        "conditions": "{}",
    }
    return render(request, "tenant_portal/setup/posting_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_rules_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_rules"
    ctx["page_title"] = "Edit posting rule"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage posting rules.")
        return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    from tenant_finance.models import ChartAccount, PostingRule

    rule = get_object_or_404(PostingRule.objects.using(tenant_db), pk=pk)
    ctx["rule"] = rule

    if request.method == "POST":
        import json

        name = (request.POST.get("name") or "").strip()
        transaction_type = (request.POST.get("transaction_type") or "").strip()
        debit_account_id = (request.POST.get("debit_account") or "").strip()
        credit_account_id = (request.POST.get("credit_account") or "").strip()
        apply_dimension = (request.POST.get("apply_dimension") or "").strip() or PostingRule.Dimension.NONE
        status = (request.POST.get("status") or "").strip() or PostingRule.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()
        priority_raw = (request.POST.get("priority") or "").strip() or str(rule.priority or 100)
        conditions_raw = (request.POST.get("conditions") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Rule name is required.")
        if not transaction_type:
            errors.append("Transaction type is required.")
        try:
            priority = int(priority_raw)
        except ValueError:
            errors.append("Priority must be a whole number.")
            priority = rule.priority or 100

        conditions = {}
        if conditions_raw:
            try:
                conditions = json.loads(conditions_raw)
                if not isinstance(conditions, dict):
                    errors.append("Conditions must be a JSON object.")
                    conditions = rule.conditions if isinstance(rule.conditions, dict) else {}
            except Exception:
                errors.append("Conditions must be valid JSON (object).")
                conditions = rule.conditions if isinstance(rule.conditions, dict) else {}

        debit_account = None
        credit_account = None
        if debit_account_id:
            debit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=debit_account_id).first()
            )
            if not debit_account:
                errors.append("Selected debit account does not exist.")
        else:
            errors.append("Debit account is required.")

        if credit_account_id:
            credit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=credit_account_id).first()
            )
            if not credit_account:
                errors.append("Selected credit account does not exist.")
        else:
            errors.append("Credit account is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            before = {
                "priority": rule.priority,
                "conditions": rule.conditions,
            }
            rule.name = name
            rule.transaction_type = transaction_type
            rule.debit_account = debit_account
            rule.credit_account = credit_account
            rule.apply_dimension = apply_dimension
            rule.status = status
            rule.description = description
            rule.priority = priority
            rule.conditions = conditions

            try:
                rule.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                rule.save(using=tenant_db)
                from tenant_finance.models import AuditLog
                AuditLog.objects.using(tenant_db).create(
                    model_name="postingrule",
                    object_id=rule.id,
                    action=AuditLog.Action.UPDATE,
                    user_id=request.tenant_user.id if request.tenant_user else None,
                    username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
                    summary=f"Posting rule updated: {rule.name} ({rule.transaction_type})",
                    old_data=before,
                    new_data={"priority": rule.priority, "conditions": rule.conditions},
                )
                messages.success(request, "Posting rule updated.")
                return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    from tenant_finance.models import PostingRule as RuleModel

    ctx["transaction_type_choices"] = RuleModel.TransactionType.choices
    ctx["status_choices"] = RuleModel.Status.choices
    ctx["dimension_choices"] = RuleModel.Dimension.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "name": rule.name,
        "transaction_type": rule.transaction_type,
        "debit_account_id": rule.debit_account_id,
        "credit_account_id": rule.credit_account_id,
        "apply_dimension": rule.apply_dimension,
        "status": rule.status,
        "description": rule.description,
        "priority": rule.priority or 100,
        "conditions": rule.conditions if isinstance(rule.conditions, dict) else {},
    }
    return render(request, "tenant_portal/setup/posting_rules_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_rules_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage posting rules.")
        return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    from tenant_finance.models import PostingRule

    rule = get_object_or_404(PostingRule.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        if rule.status == PostingRule.Status.ACTIVE:
            rule.status = PostingRule.Status.INACTIVE
            msg = "Posting rule deactivated."
        else:
            rule.status = PostingRule.Status.ACTIVE
            msg = "Posting rule activated."
        try:
            rule.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            rule.save(using=tenant_db, update_fields=["status"])
            from tenant_finance.models import AuditLog
            AuditLog.objects.using(tenant_db).create(
                model_name="postingrule",
                object_id=rule.id,
                action=AuditLog.Action.UPDATE,
                user_id=request.tenant_user.id if request.tenant_user else None,
                username=(request.tenant_user.full_name or request.tenant_user.email) if getattr(request, "tenant_user", None) else "",
                summary=f"Posting rule status changed: {rule.name} → {rule.status}",
            )
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_posting_rules_list"))

    return redirect(reverse("tenant_portal:setup_posting_rules_list"))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_permissions_list_view(request: HttpRequest) -> HttpResponse:
    from tenant_finance.models import PostingPermission

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_permissions"
    ctx["page_title"] = "Posting permissions"

    permissions = PostingPermission.objects.using(tenant_db).order_by("role_name")
    ctx["permissions"] = permissions
    return render(request, "tenant_portal/setup/posting_permissions_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_permissions_add_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal, InvalidOperation

    from tenant_finance.models import PostingPermission

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_permissions"
    ctx["page_title"] = "Add posting permission"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage posting permissions.")
        return redirect(reverse("tenant_portal:setup_posting_permissions_list"))

    initial = {
        "can_create_voucher": True,
        "can_approve_voucher": False,
        "can_post_to_ledger": False,
        "max_posting_amount": "0.00",
        "require_second_approval_above_amount": False,
    }

    if request.method == "POST":
        role_name = (request.POST.get("role_name") or "").strip()
        can_create = bool(request.POST.get("can_create_voucher"))
        can_approve = bool(request.POST.get("can_approve_voucher"))
        can_post = bool(request.POST.get("can_post_to_ledger"))
        max_amount_raw = (request.POST.get("max_posting_amount") or "").strip() or "0"
        require_second = bool(request.POST.get("require_second_approval_above_amount"))

        errors: list[str] = []
        if not role_name:
            errors.append("Role name is required.")

        try:
            max_amount = Decimal(max_amount_raw)
            if max_amount < 0:
                errors.append("Maximum posting amount cannot be negative.")
        except (InvalidOperation, ValueError):
            errors.append("Maximum posting amount must be a valid number.")
            max_amount = Decimal("0")

        if (
            role_name
            and PostingPermission.objects.using(tenant_db)
            .filter(role_name__iexact=role_name)
            .exists()
        ):
            errors.append("A posting permission for this role already exists.")

        if errors:
            for e in errors:
                messages.error(request, e)
            initial.update(
                {
                    "role_name": role_name,
                    "can_create_voucher": can_create,
                    "can_approve_voucher": can_approve,
                    "can_post_to_ledger": can_post,
                    "max_posting_amount": max_amount_raw,
                    "require_second_approval_above_amount": require_second,
                }
            )
        else:
            PostingPermission.objects.using(tenant_db).create(
                role_name=role_name,
                can_create_voucher=can_create,
                can_approve_voucher=can_approve,
                can_post_to_ledger=can_post,
                max_posting_amount=max_amount,
                require_second_approval_above_amount=require_second,
            )
            messages.success(request, "Posting permission created.")
            return redirect(reverse("tenant_portal:setup_posting_permissions_list"))

    ctx["form_initial"] = initial
    ctx["form_mode"] = "create"
    return render(request, "tenant_portal/setup/posting_permissions_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_permissions_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    from decimal import Decimal, InvalidOperation

    from tenant_finance.models import PostingPermission

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_posting_permissions"
    ctx["page_title"] = "Edit posting permission"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage posting permissions.")
        return redirect(reverse("tenant_portal:setup_posting_permissions_list"))

    perm = get_object_or_404(PostingPermission.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        role_name = (request.POST.get("role_name") or "").strip()
        can_create = bool(request.POST.get("can_create_voucher"))
        can_approve = bool(request.POST.get("can_approve_voucher"))
        can_post = bool(request.POST.get("can_post_to_ledger"))
        max_amount_raw = (request.POST.get("max_posting_amount") or "").strip() or "0"
        require_second = bool(request.POST.get("require_second_approval_above_amount"))

        errors: list[str] = []
        if not role_name:
            errors.append("Role name is required.")

        try:
            max_amount = Decimal(max_amount_raw)
            if max_amount < 0:
                errors.append("Maximum posting amount cannot be negative.")
        except (InvalidOperation, ValueError):
            errors.append("Maximum posting amount must be a valid number.")
            max_amount = Decimal("0")

        if (
            role_name
            and PostingPermission.objects.using(tenant_db)
            .filter(role_name__iexact=role_name)
            .exclude(pk=pk)
            .exists()
        ):
            errors.append("A posting permission for this role already exists.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            perm.role_name = role_name
            perm.can_create_voucher = can_create
            perm.can_approve_voucher = can_approve
            perm.can_post_to_ledger = can_post
            perm.max_posting_amount = max_amount
            perm.require_second_approval_above_amount = require_second
            perm.save(using=tenant_db)
            messages.success(request, "Posting permission updated.")
            return redirect(reverse("tenant_portal:setup_posting_permissions_list"))

    ctx["form_initial"] = {
        "role_name": perm.role_name,
        "can_create_voucher": perm.can_create_voucher,
        "can_approve_voucher": perm.can_approve_voucher,
        "can_post_to_ledger": perm.can_post_to_ledger,
        "max_posting_amount": perm.max_posting_amount,
        "require_second_approval_above_amount": perm.require_second_approval_above_amount,
    }
    ctx["form_mode"] = "edit"
    ctx["permission"] = perm
    return render(request, "tenant_portal/setup/posting_permissions_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_posting_permissions_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    from tenant_finance.models import PostingPermission

    tenant_db = request.tenant_db
    if not _setup_context(request)["can_manage"]:
        messages.error(request, "You do not have permission to manage posting permissions.")
        return redirect(reverse("tenant_portal:setup_posting_permissions_list"))

    perm = get_object_or_404(PostingPermission.objects.using(tenant_db), pk=pk)
    perm.delete(using=tenant_db)
    messages.success(request, "Posting permission deleted.")
    return redirect(reverse("tenant_portal:setup_posting_permissions_list"))


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_reversal_rules_view(request: HttpRequest) -> HttpResponse:
    from tenant_finance.models import TransactionReversalRule

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_reversal_rules"
    ctx["page_title"] = "Transaction reversal & correction rules"

    rule = TransactionReversalRule.objects.using(tenant_db).first()
    if not rule:
        rule = TransactionReversalRule.objects.using(tenant_db).create()

    if not ctx["can_manage"]:
        ctx["rule"] = rule
        ctx["read_only"] = True
        return render(request, "tenant_portal/setup/reversal_rules.html", ctx)

    if request.method == "POST":
        allow_reversal = bool(request.POST.get("allow_reversal"))
        allow_edit_before_posting = bool(request.POST.get("allow_edit_before_posting"))
        allow_delete_before_approval = bool(request.POST.get("allow_delete_before_approval"))
        require_reason_for_reversal = bool(request.POST.get("require_reason_for_reversal"))

        # Editing rules
        prevent_edit_after_posting = bool(request.POST.get("prevent_edit_after_posting"))

        # Deletion rules
        prevent_delete_after_approval = bool(
            request.POST.get("prevent_delete_after_approval")
        )

        # Reversal rules
        require_reversal_approval = bool(
            request.POST.get("require_reversal_approval")
        )
        prevent_reversal_if_period_closed = bool(
            request.POST.get("prevent_reversal_if_period_closed")
        )
        prevent_cross_period_reversal = bool(
            request.POST.get("prevent_cross_period_reversal")
        )
        authorized_roles_for_reversal = (
            request.POST.get("authorized_roles_for_reversal") or ""
        ).strip()
        authorized_roles_for_cross_period_reversal = (
            request.POST.get("authorized_roles_for_cross_period_reversal") or ""
        ).strip()

        rule.allow_reversal = allow_reversal
        rule.allow_edit_before_posting = allow_edit_before_posting
        rule.allow_delete_before_approval = allow_delete_before_approval
        rule.require_reason_for_reversal = require_reason_for_reversal
        rule.prevent_edit_after_posting = prevent_edit_after_posting
        rule.prevent_delete_after_approval = prevent_delete_after_approval
        rule.require_reversal_approval = require_reversal_approval
        rule.prevent_reversal_if_period_closed = prevent_reversal_if_period_closed
        rule.prevent_cross_period_reversal = prevent_cross_period_reversal
        rule.authorized_roles_for_reversal = authorized_roles_for_reversal
        rule.authorized_roles_for_cross_period_reversal = (
            authorized_roles_for_cross_period_reversal
        )
        rule.save(using=tenant_db)
        messages.success(request, "Reversal & correction rules updated.")
        return redirect(reverse("tenant_portal:setup_reversal_rules"))

    ctx["rule"] = rule
    ctx["read_only"] = False
    return render(request, "tenant_portal/setup/reversal_rules.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_audit_trail_settings_view(request: HttpRequest) -> HttpResponse:
    from decimal import Decimal

    from tenant_finance.models import AuditTrailSetting

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_audit_trail_settings"
    ctx["page_title"] = "Audit trail settings"

    setting = AuditTrailSetting.objects.using(tenant_db).first()
    if not setting:
        setting = AuditTrailSetting.objects.using(tenant_db).create()

    # Role list for multi-select RBAC field (tenant-scoped)
    try:
        from rbac.models import Role

        ctx["roles"] = list(Role.objects.using(tenant_db).order_by("name").values_list("name", flat=True))
    except Exception:
        ctx["roles"] = []

    selected_roles = [r.strip() for r in (setting.authorized_roles_for_audit_logs or "").split(",") if r.strip()]
    ctx["selected_roles_for_audit_logs"] = selected_roles

    if not ctx["can_manage"]:
        ctx["setting"] = setting
        ctx["read_only"] = True
        return render(request, "tenant_portal/setup/audit_trail_settings.html", ctx)

    if request.method == "POST":
        enable_audit_logging = bool(request.POST.get("enable_audit_logging"))
        track_voucher_edits = bool(request.POST.get("track_voucher_edits"))
        track_approvals = bool(request.POST.get("track_approvals"))
        track_posting_actions = bool(request.POST.get("track_posting_actions"))

        track_field_level_changes = bool(request.POST.get("track_field_level_changes"))

        # User activity
        track_logins = bool(request.POST.get("track_logins"))
        track_failed_logins = bool(request.POST.get("track_failed_logins"))
        track_user_changes = bool(request.POST.get("track_user_changes"))

        # Transaction protection
        prevent_hard_delete_transactions = bool(
            request.POST.get("prevent_hard_delete_transactions")
        )
        strict_posting_protection = bool(request.POST.get("strict_posting_protection"))
        require_reason_for_reversal = bool(
            request.POST.get("require_reason_for_reversal")
        )

        # Fraud / high-risk hooks
        track_high_risk_events = bool(request.POST.get("track_high_risk_events"))
        escalate_to_audit_risk = bool(request.POST.get("escalate_to_audit_risk"))
        risk_classification = (request.POST.get("risk_classification") or "").strip() or AuditTrailSetting.RiskClassification.MEDIUM

        # Access control
        authorized_roles_list = request.POST.getlist("authorized_roles_for_audit_logs")
        authorized_roles_list = [r.strip() for r in authorized_roles_list if (r or "").strip()]
        authorized_roles_raw = ",".join(authorized_roles_list)
        allow_users_see_own_activity = bool(
            request.POST.get("allow_users_see_own_activity")
        )

        retention_policy = (request.POST.get("retention_policy") or "").strip() or AuditTrailSetting.RetentionPolicy.DAYS_365
        auto_archive = bool(request.POST.get("auto_archive"))
        retention_raw = (request.POST.get("retention_days") or "").strip() or "0"
        receipt_approval_mode = (
            (request.POST.get("receipt_approval_mode") or "").strip()
            or AuditTrailSetting.ReceiptApprovalMode.NO_APPROVAL
        )
        receipt_approval_threshold_raw = (
            request.POST.get("receipt_approval_threshold") or ""
        ).strip() or "0"

        errors: list[str] = []
        try:
            retention_days = int(retention_raw)
            if retention_days < 0:
                errors.append("Retention period cannot be negative.")
        except ValueError:
            errors.append("Retention period must be a whole number of days.")
            retention_days = setting.retention_days

        if retention_policy != AuditTrailSetting.RetentionPolicy.CUSTOM:
            try:
                retention_days = int(retention_policy)
            except Exception:
                retention_days = setting.retention_days

        if not errors:
            try:
                receipt_approval_threshold = Decimal(receipt_approval_threshold_raw)
            except Exception:
                receipt_approval_threshold = Decimal("0")
                errors.append("Receipt approval threshold must be a valid amount.")

            if receipt_approval_threshold < 0:
                errors.append("Receipt approval threshold cannot be negative.")

            if receipt_approval_mode not in {
                AuditTrailSetting.ReceiptApprovalMode.NO_APPROVAL,
                AuditTrailSetting.ReceiptApprovalMode.ABOVE_AMOUNT,
                AuditTrailSetting.ReceiptApprovalMode.CASH_ONLY,
                AuditTrailSetting.ReceiptApprovalMode.DONOR_ONLY,
            }:
                errors.append("Invalid receipt approval mode selected.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            setting.enable_audit_logging = enable_audit_logging
            setting.track_voucher_edits = track_voucher_edits
            setting.track_approvals = track_approvals
            setting.track_posting_actions = track_posting_actions
            setting.track_field_level_changes = track_field_level_changes
            setting.retention_days = retention_days
            setting.retention_policy = retention_policy
            setting.auto_archive = auto_archive

            setting.track_logins = track_logins
            setting.track_failed_logins = track_failed_logins
            setting.track_user_changes = track_user_changes

            setting.prevent_hard_delete_transactions = prevent_hard_delete_transactions
            setting.strict_posting_protection = strict_posting_protection
            setting.require_reason_for_reversal = require_reason_for_reversal

            setting.track_high_risk_events = track_high_risk_events
            setting.risk_classification = risk_classification
            setting.escalate_to_audit_risk = escalate_to_audit_risk

            setting.authorized_roles_for_audit_logs = authorized_roles_raw
            setting.allow_users_see_own_activity = allow_users_see_own_activity
            setting.receipt_approval_mode = receipt_approval_mode
            setting.receipt_approval_threshold = receipt_approval_threshold

            setting.save(using=tenant_db)
            messages.success(request, "Audit trail settings updated.")
            return redirect(reverse("tenant_portal:setup_audit_trail_settings"))

    ctx["setting"] = setting
    ctx["read_only"] = False
    return render(request, "tenant_portal/setup/audit_trail_settings.html", ctx)


# ----- Default accounts -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_default_accounts_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_default_accounts"
    ctx["page_title"] = "Accounting Posting Rules – Transaction mappings"

    from tenant_finance.models import DefaultAccountMapping

    qs = DefaultAccountMapping.objects.using(tenant_db).select_related(
        "default_debit_account", "default_credit_account"
    )

    tx_type = (request.GET.get("transaction_type") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if tx_type:
        qs = qs.filter(transaction_type=tx_type)
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("transaction_type", "name")

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["mappings_page"] = paginator.get_page(page_number)

    from tenant_finance.models import DefaultAccountMapping as MappingModel

    ctx["transaction_type_choices"] = MappingModel.TransactionType.choices
    ctx["status_choices"] = MappingModel.Status.choices
    ctx["dimension_choices"] = MappingModel.Dimension.choices

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/default_accounts_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_default_accounts_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_default_accounts"
    ctx["page_title"] = "Add accounting posting rule (transaction mapping)"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage default account mappings.")
        return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    from tenant_finance.models import ChartAccount, DefaultAccountMapping

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        transaction_type = (request.POST.get("transaction_type") or "").strip()
        debit_account_id = (request.POST.get("default_debit_account") or "").strip()
        credit_account_id = (request.POST.get("default_credit_account") or "").strip()
        apply_dimension = (request.POST.get("apply_dimension") or "").strip() or DefaultAccountMapping.Dimension.NONE
        status = (request.POST.get("status") or "").strip() or DefaultAccountMapping.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Mapping name is required.")
        if not transaction_type:
            errors.append("Transaction type is required.")

        debit_account = None
        credit_account = None
        if debit_account_id:
            debit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=debit_account_id).first()
            )
            if not debit_account:
                errors.append("Selected default debit account does not exist.")
        else:
            errors.append("Default debit account is required.")

        if credit_account_id:
            credit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=credit_account_id).first()
            )
            if not credit_account:
                errors.append("Selected default credit account does not exist.")
        else:
            errors.append("Default credit account is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            mapping = DefaultAccountMapping(
                name=name,
                transaction_type=transaction_type,
                default_debit_account=debit_account,
                default_credit_account=credit_account,
                apply_dimension=apply_dimension,
                status=status,
                description=description,
            )
            try:
                mapping.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                mapping.save(using=tenant_db)
                messages.success(request, "Default account mapping created.")
                return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    from tenant_finance.models import DefaultAccountMapping as MappingModel

    ctx["transaction_type_choices"] = MappingModel.TransactionType.choices
    ctx["status_choices"] = MappingModel.Status.choices
    ctx["dimension_choices"] = MappingModel.Dimension.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "status": MappingModel.Status.ACTIVE,
        "apply_dimension": MappingModel.Dimension.NONE,
    }
    return render(request, "tenant_portal/setup/default_accounts_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_default_accounts_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_default_accounts"
    ctx["page_title"] = "Edit default account mapping"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage default account mappings.")
        return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    from tenant_finance.models import ChartAccount, DefaultAccountMapping

    mapping = get_object_or_404(DefaultAccountMapping.objects.using(tenant_db), pk=pk)
    ctx["mapping"] = mapping

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        transaction_type = (request.POST.get("transaction_type") or "").strip()
        debit_account_id = (request.POST.get("default_debit_account") or "").strip()
        credit_account_id = (request.POST.get("default_credit_account") or "").strip()
        apply_dimension = (request.POST.get("apply_dimension") or "").strip() or DefaultAccountMapping.Dimension.NONE
        status = (request.POST.get("status") or "").strip() or DefaultAccountMapping.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Mapping name is required.")
        if not transaction_type:
            errors.append("Transaction type is required.")

        debit_account = None
        credit_account = None
        if debit_account_id:
            debit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=debit_account_id).first()
            )
            if not debit_account:
                errors.append("Selected default debit account does not exist.")
        else:
            errors.append("Default debit account is required.")

        if credit_account_id:
            credit_account = (
                ChartAccount.objects.using(tenant_db).filter(pk=credit_account_id).first()
            )
            if not credit_account:
                errors.append("Selected default credit account does not exist.")
        else:
            errors.append("Default credit account is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            mapping.name = name
            mapping.transaction_type = transaction_type
            mapping.default_debit_account = debit_account
            mapping.default_credit_account = credit_account
            mapping.apply_dimension = apply_dimension
            mapping.status = status
            mapping.description = description

            try:
                mapping.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                mapping.save(using=tenant_db)
                messages.success(request, "Default account mapping updated.")
                return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    from tenant_finance.models import DefaultAccountMapping as MappingModel

    ctx["transaction_type_choices"] = MappingModel.TransactionType.choices
    ctx["status_choices"] = MappingModel.Status.choices
    ctx["dimension_choices"] = MappingModel.Dimension.choices
    ctx["accounts"] = (
        ChartAccount.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("code")
    )
    ctx["form_initial"] = {
        "name": mapping.name,
        "transaction_type": mapping.transaction_type,
        "default_debit_account_id": mapping.default_debit_account_id,
        "default_credit_account_id": mapping.default_credit_account_id,
        "apply_dimension": mapping.apply_dimension,
        "status": mapping.status,
        "description": mapping.description,
    }
    return render(request, "tenant_portal/setup/default_accounts_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_default_accounts_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage default account mappings.")
        return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    from tenant_finance.models import DefaultAccountMapping

    mapping = get_object_or_404(DefaultAccountMapping.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        if mapping.status == DefaultAccountMapping.Status.ACTIVE:
            mapping.status = DefaultAccountMapping.Status.INACTIVE
            msg = "Default account mapping deactivated."
        else:
            mapping.status = DefaultAccountMapping.Status.ACTIVE
            msg = "Default account mapping activated."
        try:
            mapping.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            mapping.save(using=tenant_db, update_fields=["status"])
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_default_accounts_list"))

    return redirect(reverse("tenant_portal:setup_default_accounts_list"))


# ----- Fiscal years -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_fiscal_years_list_view(request: HttpRequest) -> HttpResponse:
    from tenant_finance.models import FiscalYear

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_fiscal_years"
    ctx["page_title"] = "Fiscal years"

    years = FiscalYear.objects.using(tenant_db).order_by("-start_date")
    ctx["fiscal_years"] = years
    return render(request, "tenant_portal/setup/fiscal_years_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_fiscal_years_add_view(request: HttpRequest) -> HttpResponse:
    from datetime import date

    from tenant_finance.models import FiscalYear, FiscalPeriod

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_fiscal_years"
    ctx["page_title"] = "Add fiscal year"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage fiscal years.")
        return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        start = request.POST.get("start_date") or ""
        end = request.POST.get("end_date") or ""
        errors: list[str] = []

        try:
            start_date = date.fromisoformat(start)
        except ValueError:
            start_date = None
            errors.append("Start date is required and must be a valid date.")
        try:
            end_date = date.fromisoformat(end)
        except ValueError:
            end_date = None
            errors.append("End date is required and must be a valid date.")

        if not name:
            errors.append("Fiscal year name is required.")
        if start_date and end_date and end_date <= start_date:
            errors.append("End date must be after start date.")

        if start_date and end_date:
            overlap = FiscalYear.objects.using(tenant_db).filter(
                start_date__lte=end_date, end_date__gte=start_date
            ).exists()
            if overlap:
                errors.append("Fiscal year dates overlap with an existing fiscal year.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            fy = FiscalYear.objects.using(tenant_db).create(
                name=name,
                start_date=start_date,
                end_date=end_date,
                status=FiscalYear.Status.OPEN,
                created_by=request.tenant_user,
            )
            # Auto-generate 12 monthly periods
            current = start_date
            period_no = 1
            while current <= end_date and period_no <= 12:
                # month end is either end of month or fiscal year end
                if current.month == 12:
                    month_end = current.replace(day=31)
                else:
                    from calendar import monthrange

                    last_day = monthrange(current.year, current.month)[1]
                    month_end = current.replace(day=last_day)
                period_end = min(month_end, end_date)
                FiscalPeriod.objects.using(tenant_db).create(
                    fiscal_year=fy,
                    period_number=period_no,
                    name=current.strftime("%b %Y"),
                    period_name=current.strftime("%B %Y"),
                    start_date=current,
                    end_date=period_end,
                    status=FiscalPeriod.Status.OPEN,
                )
                current = period_end.replace(day=1)
                current = date(current.year + (1 if current.month == 12 else 0),
                               1 if current.month == 12 else current.month + 1,
                               1)
                period_no += 1

            messages.success(request, "Fiscal year created with monthly periods.")
            return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    return render(request, "tenant_portal/setup/fiscal_years_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_fiscal_years_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    from datetime import date

    from tenant_finance.models import FiscalYear

    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_fiscal_years"
    ctx["page_title"] = "Edit fiscal year"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage fiscal years.")
        return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    fy = get_object_or_404(FiscalYear.objects.using(tenant_db), pk=pk)
    ctx["fiscal_year"] = fy

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        start = request.POST.get("start_date") or ""
        end = request.POST.get("end_date") or ""
        errors: list[str] = []
        try:
            start_date = date.fromisoformat(start)
        except ValueError:
            start_date = None
            errors.append("Start date is required and must be a valid date.")
        try:
            end_date = date.fromisoformat(end)
        except ValueError:
            end_date = None
            errors.append("End date is required and must be a valid date.")

        if not name:
            errors.append("Fiscal year name is required.")
        if start_date and end_date and end_date <= start_date:
            errors.append("End date must be after start date.")

        if start_date and end_date:
            overlap = (
                FiscalYear.objects.using(tenant_db)
                .filter(start_date__lte=end_date, end_date__gte=start_date)
                .exclude(pk=pk)
                .exists()
            )
            if overlap:
                errors.append("Fiscal year dates overlap with an existing fiscal year.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            fy.name = name
            fy.start_date = start_date
            fy.end_date = end_date
            fy.save(using=tenant_db)
            messages.success(request, "Fiscal year updated.")
            return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    return render(request, "tenant_portal/setup/fiscal_years_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_fiscal_years_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to delete fiscal years.")
        return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    from tenant_finance.models import FiscalYear, FiscalPeriod

    fy = get_object_or_404(FiscalYear.objects.using(tenant_db), pk=pk)
    if request.method == "POST":
        if FiscalPeriod.objects.using(tenant_db).filter(fiscal_year=fy, is_closed=True).exists():
            messages.error(request, "Cannot delete a fiscal year that has closed periods.")
            return redirect(reverse("tenant_portal:setup_fiscal_years_list"))
        fy.delete(using=tenant_db)
        messages.success(request, "Fiscal year deleted.")
        return redirect(reverse("tenant_portal:setup_fiscal_years_list"))

    ctx["object"] = fy
    ctx["object_label"] = f"Fiscal year {fy.name}"
    ctx["cancel_url"] = reverse("tenant_portal:setup_fiscal_years_list")
    ctx["delete_url"] = reverse("tenant_portal:setup_fiscal_years_delete", args=[pk])
    return render(request, "tenant_portal/setup/confirm_delete.html", ctx)


# ----- Accounting periods -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_accounting_periods_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_accounting_periods"
    ctx["page_title"] = "Accounting periods"

    from tenant_finance.models import FiscalPeriod, FiscalYear
    from tenant_finance.services.accounting_periods import close_period, reopen_period
    from rbac.models import user_has_permission as _uhp

    cached = getattr(request, "rbac_permission_codes", None)

    def _has(code: str) -> bool:
        if isinstance(cached, set):
            return ("*" in cached) or (code in cached)
        return _uhp(request.tenant_user, code, using=tenant_db)

    if not _has("finance:periods.view"):
        return render(
            request,
            "tenant_portal/forbidden.html",
            {"tenant": request.tenant, "tenant_user": request.tenant_user, "reason": "You do not have permission to view accounting periods."},
            status=403,
        )

    fy_id = (request.GET.get("fiscal_year") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = FiscalPeriod.objects.using(tenant_db).select_related("fiscal_year").all()
    if fy_id:
        qs = qs.filter(fiscal_year_id=fy_id)
    if status == "open":
        qs = qs.filter(is_closed=False)
    elif status == "closed":
        qs = qs.filter(is_closed=True)

    qs = qs.order_by("-fiscal_year__start_date", "period_number")

    # Handle open/close action
    if request.method == "POST" and ctx.get("can_manage"):
        period_id = int(request.POST.get("period_id") or "0")
        action = (request.POST.get("action") or "").strip()
        reason = (request.POST.get("reason") or "").strip()
        try:
            if action == "soft_close":
                if not _has("finance:periods.close"):
                    raise ValueError("You do not have permission to close accounting periods.")
                close_period(using=tenant_db, period_id=period_id, close_type="soft", user=request.tenant_user, reason=reason)
                messages.success(request, "Period soft closed.")
            elif action == "hard_close":
                if not _has("finance:periods.close"):
                    raise ValueError("You do not have permission to close accounting periods.")
                close_period(using=tenant_db, period_id=period_id, close_type="hard", user=request.tenant_user, reason=reason)
                messages.success(request, "Period hard closed.")
            elif action == "reopen":
                if not _has("finance:periods.reopen"):
                    raise ValueError("You do not have permission to reopen accounting periods.")
                reopen_period(using=tenant_db, period_id=period_id, user=request.tenant_user, reason=reason)
                messages.success(request, "Period reopened.")
        except Exception as exc:
            messages.error(request, str(exc) or "Action failed.")

        return redirect(
            reverse("tenant_portal:setup_accounting_periods_list")
            + (f"?{request.GET.urlencode()}" if request.GET.urlencode() else "")
        )

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["periods_page"] = paginator.get_page(page_number)
    ctx["fiscal_years"] = FiscalYear.objects.using(tenant_db).all().order_by("-start_date")
    ctx["filter_fiscal_year"] = fy_id
    ctx["filter_status"] = status

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/accounting_periods_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_accounting_periods_add_view(request: HttpRequest) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_accounting_periods"
    ctx["page_title"] = "Add accounting period"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_accounting_periods_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_accounting_periods"
    ctx["page_title"] = "Edit accounting period"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_accounting_periods_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    return redirect(reverse("tenant_portal:setup_accounting_periods_list"))


# ----- Approval workflows -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_approval_workflows_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_approval_workflows"
    ctx["page_title"] = "Approval workflows"

    from tenant_finance.models import ApprovalWorkflow

    qs = ApprovalWorkflow.objects.using(tenant_db)

    tx_type = (request.GET.get("transaction_type") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if tx_type:
        qs = qs.filter(document_type=tx_type)
    if status:
        qs = qs.filter(status=status)

    qs = qs.order_by("document_type", "name")

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get("page", "1")
    ctx["workflows_page"] = paginator.get_page(page_number)

    from tenant_finance.models import ApprovalWorkflow as WFModel

    ctx["transaction_type_choices"] = WFModel.TransactionType.choices
    ctx["status_choices"] = WFModel.Status.choices

    get = request.GET.copy()
    if "page" in get:
        get.pop("page")
    ctx["base_query"] = get.urlencode()

    return render(request, "tenant_portal/setup/approval_workflows_list.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_approval_workflows_add_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_approval_workflows"
    ctx["page_title"] = "Add approval workflow"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage approval workflows.")
        return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    from tenant_finance.models import ApprovalWorkflow

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        document_type = (request.POST.get("document_type") or "").strip()
        status = (request.POST.get("status") or "").strip() or ApprovalWorkflow.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()

        # Levels and roles from the form
        level1_role = (request.POST.get("level1_role") or "").strip()
        level2_role = (request.POST.get("level2_role") or "").strip()
        level3_role = (request.POST.get("level3_role") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Workflow name is required.")
        if not document_type:
            errors.append("Transaction type is required.")

        steps = []
        if level1_role:
            steps.append({"role": level1_role, "order": 1})
        if level2_role:
            steps.append({"role": level2_role, "order": 2})
        if level3_role:
            steps.append({"role": level3_role, "order": 3})

        if not steps:
            errors.append("At least one approval level is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            wf = ApprovalWorkflow(
                name=name,
                document_type=document_type,
                steps=steps,
                status=status,
                description=description,
            )
            try:
                wf.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                wf.save(using=tenant_db)
                messages.success(request, "Approval workflow created.")
                return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    from tenant_finance.models import ApprovalWorkflow as WFModel

    ctx["transaction_type_choices"] = WFModel.TransactionType.choices
    ctx["status_choices"] = WFModel.Status.choices
    ctx["form_initial"] = {
        "status": WFModel.Status.ACTIVE,
    }
    return render(request, "tenant_portal/setup/approval_workflows_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_approval_workflows_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_approval_workflows"
    ctx["page_title"] = "Edit approval workflow"

    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage approval workflows.")
        return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    from tenant_finance.models import ApprovalWorkflow

    wf = get_object_or_404(ApprovalWorkflow.objects.using(tenant_db), pk=pk)
    ctx["workflow"] = wf

    # Extract existing levels
    steps_by_order = {int(s.get("order")): s.get("role") for s in (wf.steps or []) if s.get("order") is not None}

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        document_type = (request.POST.get("document_type") or "").strip()
        status = (request.POST.get("status") or "").strip() or ApprovalWorkflow.Status.ACTIVE
        description = (request.POST.get("description") or "").strip()

        level1_role = (request.POST.get("level1_role") or "").strip()
        level2_role = (request.POST.get("level2_role") or "").strip()
        level3_role = (request.POST.get("level3_role") or "").strip()

        errors: list[str] = []

        if not name:
            errors.append("Workflow name is required.")
        if not document_type:
            errors.append("Transaction type is required.")

        steps = []
        if level1_role:
            steps.append({"role": level1_role, "order": 1})
        if level2_role:
            steps.append({"role": level2_role, "order": 2})
        if level3_role:
            steps.append({"role": level3_role, "order": 3})

        if not steps:
            errors.append("At least one approval level is required.")

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            wf.name = name
            wf.document_type = document_type
            wf.status = status
            wf.description = description
            wf.steps = steps

            try:
                wf.full_clean()
            except ValidationError as exc:
                for field_errors in exc.message_dict.values():
                    for msg in field_errors:
                        messages.error(request, msg)
            else:
                wf.save(using=tenant_db)
                messages.success(request, "Approval workflow updated.")
                return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    from tenant_finance.models import ApprovalWorkflow as WFModel

    ctx["transaction_type_choices"] = WFModel.TransactionType.choices
    ctx["status_choices"] = WFModel.Status.choices
    ctx["form_initial"] = {
        "name": wf.name,
        "document_type": wf.document_type,
        "status": wf.status,
        "description": wf.description,
        "level1_role": steps_by_order.get(1, ""),
        "level2_role": steps_by_order.get(2, ""),
        "level3_role": steps_by_order.get(3, ""),
    }
    return render(request, "tenant_portal/setup/approval_workflows_form.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_approval_workflows_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    if not ctx["can_manage"]:
        messages.error(request, "You do not have permission to manage approval workflows.")
        return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    from tenant_finance.models import ApprovalWorkflow

    wf = get_object_or_404(ApprovalWorkflow.objects.using(tenant_db), pk=pk)

    if request.method == "POST":
        if wf.status == ApprovalWorkflow.Status.ACTIVE:
            wf.status = ApprovalWorkflow.Status.INACTIVE
            msg = "Approval workflow deactivated."
        else:
            wf.status = ApprovalWorkflow.Status.ACTIVE
            msg = "Approval workflow activated."
        try:
            wf.full_clean()
        except ValidationError as exc:
            for field_errors in exc.message_dict.values():
                for m in field_errors:
                    messages.error(request, m)
        else:
            wf.save(using=tenant_db, update_fields=["status"])
            messages.success(request, msg)
        return redirect(reverse("tenant_portal:setup_approval_workflows_list"))

    return redirect(reverse("tenant_portal:setup_approval_workflows_list"))


# ----- Budget control rules -----
@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_budget_control_rules_list_view(request: HttpRequest) -> HttpResponse:
    tenant_db = request.tenant_db
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_budget_control_rules"
    ctx["page_title"] = "Budget control rules"

    from decimal import Decimal
    from tenant_finance.models import BudgetControlRule

    rule = (
        BudgetControlRule.objects.using(tenant_db)
        .filter(is_active=True)
        .order_by("id")
        .first()
    )
    if not rule:
        rule = BudgetControlRule.objects.using(tenant_db).create(
            name="Default budget control",
            warn_at_percent=Decimal("80"),
            critical_at_percent=Decimal("90"),
            block_at_percent=Decimal("100"),
            is_active=True,
        )

    if request.method == "POST":
        warning = (request.POST.get("warning_threshold") or "").strip()
        critical = (request.POST.get("critical_threshold") or "").strip()
        block = (request.POST.get("block_threshold") or "").strip()
        allow_override = bool(request.POST.get("allow_override"))
        override_roles = (request.POST.get("override_roles") or "").strip()
        check_before_posting = bool(request.POST.get("check_before_posting"))
        include_commitments = bool(request.POST.get("include_commitments"))
        try:
            rule.warn_at_percent = Decimal(warning or rule.warn_at_percent)
            rule.critical_at_percent = Decimal(critical or rule.critical_at_percent)
            rule.block_at_percent = Decimal(block or rule.block_at_percent)
            rule.allow_override = allow_override
            rule.override_roles = override_roles
            rule.check_before_posting = check_before_posting
            rule.include_commitments = include_commitments
            rule.full_clean()
            rule.save(using=tenant_db)

            # Audit log for rule changes
            from tenant_finance.models import AuditLog

            AuditLog.objects.using(tenant_db).create(
                model_name="budgetcontrolrule",
                object_id=rule.id,
                action=AuditLog.Action.UPDATE,
                user_id=getattr(request.tenant_user, "id", None),
                username=getattr(request.tenant_user, "email", "") or "",
                summary=(
                    f"Updated budget control rules: warn {rule.warn_at_percent}%, "
                    f"critical {rule.critical_at_percent}%, block {rule.block_at_percent}%, "
                    f"override={'yes' if rule.allow_override else 'no'}."
                ),
            )

            messages.success(request, "Budget control rules updated.")
            return redirect(reverse("tenant_portal:setup_budget_control_rules_list"))
        except Exception as e:
            messages.error(request, str(e))

    ctx["rule"] = rule
    return render(request, "tenant_portal/setup/budget_control_rules.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_budget_control_rules_add_view(request: HttpRequest) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_budget_control_rules"
    ctx["page_title"] = "Add budget control rule"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_budget_control_rules_edit_view(request: HttpRequest, pk: int) -> HttpResponse:
    ctx = _setup_context(request)
    ctx["active_item"] = "setup_budget_control_rules"
    ctx["page_title"] = "Edit budget control rule"
    return render(request, "tenant_portal/setup/placeholder.html", ctx)


@tenant_view(require_module="finance_grants", require_perm="module:finance.view")
def setup_budget_control_rules_delete_view(request: HttpRequest, pk: int) -> HttpResponse:
    return redirect(reverse("tenant_portal:setup_budget_control_rules_list"))
